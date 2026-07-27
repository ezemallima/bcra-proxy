from flask import Flask, jsonify, send_from_directory, request, session, redirect, url_for
from functools import wraps
from flask_cors import CORS
import requests
import urllib3
import os
import io
import json
import sqlite3
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
import random
import traceback

# Módulos de scoring fiscal e integración ARCA (guards independientes:
# scoring_fiscal es Python puro; arca_ws requiere cryptography instalado).
# Se captura Exception y no solo ImportError: una instalación rota de
# cryptography (incompatibilidad binaria) puede fallar con otro tipo de error,
# y el arranque de la app nunca debe depender de estos módulos opcionales.
try:
    import scoring_fiscal
    SCORING_FISCAL_OK = True
except Exception as e:
    print(f"[init] scoring_fiscal no disponible ({type(e).__name__}): {e}", flush=True)
    SCORING_FISCAL_OK = False
try:
    import arca_ws
    _ARCA_MODULO_OK = True
    ARCA_DISPONIBLE = True   # se confirma en el init (necesita certificado)
except Exception as e:
    print(f"[init] arca_ws no disponible ({type(e).__name__}): {e}", flush=True)
    _ARCA_MODULO_OK = False
    ARCA_DISPONIBLE = False
try:
    import boto3
    BOTO3_OK = True
    from botocore.config import Config as _BotoCfg
    _LAMBDA_CFG = _BotoCfg(connect_timeout=10, read_timeout=30, retries={'max_attempts': 1})
except ImportError:
    BOTO3_OK = False
    _LAMBDA_CFG = None
    print("[aws] boto3 no instalado — usando Workers", flush=True)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__, static_folder='static')

# CORS: en producción restringir al dominio propio via ALLOWED_ORIGINS en Render.
# Ejemplo: ALLOWED_ORIGINS=https://vendeseguro.onrender.com,https://tudominio.com
_ALLOWED_ORIGINS = os.environ.get('ALLOWED_ORIGINS', '')
if _ALLOWED_ORIGINS:
    CORS(app, origins=[o.strip() for o in _ALLOWED_ORIGINS.split(',') if o.strip()])
else:
    CORS(app)  # desarrollo local: permite todos los orígenes

app.config['MAX_CONTENT_LENGTH'] = 512 * 1024 * 1024  # 512 MB — permite subir padrón mensual BCRA

_SECRET_KEY_DEFAULT = 'vs-artel-2026-key'
app.secret_key = os.environ.get('SECRET_KEY', _SECRET_KEY_DEFAULT)
if app.secret_key == _SECRET_KEY_DEFAULT:
    print('[SECURITY] SECRET_KEY usa valor default inseguro — configurar variable de entorno SECRET_KEY en Render', flush=True)

GEMINI_KEY      = os.environ.get('GEMINI_API_KEY', '')
OPENAI_KEY      = os.environ.get('OPENAI_API_KEY', '')
CUIT_API_KEY    = os.environ.get('API_KEY_CUIT', '')
CUIT_API_URL    = os.environ.get('API_SOLVENCY_URL', '')
SCRAPERAPI_KEY  = os.environ.get('SCRAPERAPI_KEY', '')

# ── Parámetros del motor de scoring ajustables sin redeploy ─────────────────
# MORA_TECNICA_UMBRAL_K: umbral en miles de ARS para clasificar mora como técnica (no comercial).
# Ajustar periódicamente según inflación. Default 200k ARS ≈ ~200 USD jun-2026.
# Ejemplo: en dic-2026 evaluar si subir a 300 o 400 según inflación acumulada.
MORA_TECNICA_UMBRAL_K = float(os.environ.get('MORA_TECNICA_UMBRAL_K', '200.0'))

# ── Apalancamiento (deuda bancaria / ingresos anuales estimados) ─────────────
# El chequeo venía comparando deuda en miles contra ingresos en pesos, así que
# el ratio daba 1000× por debajo del real y la penalización nunca se aplicaba.
# Corregidas las unidades, la penalización pasa a ser graduada en vez de un
# acantilado, y se escala por un factor de transición para no desplomar de golpe
# los scores de la cartera vigente.
#   _APAL_UMBRAL    : ratio a partir del cual empieza a penalizar (0.5 = deuda = medio ingreso anual)
#   _APAL_RANGO     : ancho del tramo hasta la penalización máxima (0.5 → 1.5 = tope)
#   _APAL_PENAL_MAX : penalización nominal plena, en puntos de score
#   _APAL_FACTOR    : calibración transicional 0..1 — subir a 1.0 aplica el rigor pleno
_APAL_UMBRAL    = float(os.environ.get('APALANCAMIENTO_UMBRAL', '0.5'))
_APAL_RANGO     = float(os.environ.get('APALANCAMIENTO_RANGO',  '1.0'))
_APAL_PENAL_MAX = float(os.environ.get('APALANCAMIENTO_PENAL_MAX', '200'))
_APAL_FACTOR    = float(os.environ.get('APALANCAMIENTO_FACTOR', '0.5'))

# ── Bright Data: RETIRADO de la cadena de proxies ────────────────────────────
# La IP de salida de Render quedó vetada en la zona residencial contratada y el
# proxy devolvía "407 Auth Failed (code: ip_forbidden)" en el 100% de las
# llamadas (BCRA, AFIP y scrapers), aportando solo latencia muerta antes del
# fallback. La cadena vigente es: Directo (BCRA) / ScraperAPI (no-BCRA).
# Las variables BRIGHTDATA_* pueden borrarse del entorno de Render: ya no se leen.

# ── Cloudflare R2 — bucket privado para bcra_nomdeu.db (padrón offline 24m) ────
# Si están las 4 variables configuradas, la descarga autenticada por R2 tiene
# prioridad sobre BCRA_NOMDEU_URL — evita exponer el archivo (datos de deuda de
# millones de CUITs) en una URL pública.
R2_ACCESS_KEY_ID     = os.environ.get('R2_ACCESS_KEY_ID', '').strip()
R2_SECRET_ACCESS_KEY = os.environ.get('R2_SECRET_ACCESS_KEY', '').strip()
R2_ENDPOINT_URL      = os.environ.get('R2_ENDPOINT_URL', '').strip()
R2_BUCKET_NAME       = os.environ.get('R2_BUCKET_NAME', '').strip()
_R2_CONFIGURADO      = bool(R2_ACCESS_KEY_ID and R2_SECRET_ACCESS_KEY and R2_ENDPOINT_URL and R2_BUCKET_NAME)

# ── Rate limiter global para API de BCRA — máx 2 llamadas directas simultáneas ─
# Evita que consultas concurrentes saturen la IP de Render y generen rate-limit.
_bcra_api_sem = threading.Semaphore(2)

# ── Rate limiter de login — protección contra fuerza bruta ───────────────────
# Máximo 5 intentos fallidos por IP en una ventana de 15 minutos.
_login_attempts: dict = {}   # ip → [timestamp, ...]
_login_attempts_lock = threading.Lock()
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW_SECS  = 900    # 15 minutos

def _login_rate_check(ip: str) -> bool:
    """True = permitido. False = bloqueado por exceso de intentos."""
    now = time.time()
    with _login_attempts_lock:
        hist = [t for t in _login_attempts.get(ip, []) if now - t < _LOGIN_WINDOW_SECS]
        if len(hist) >= _LOGIN_MAX_ATTEMPTS:
            return False
        hist.append(now)
        _login_attempts[ip] = hist
    return True

def _login_rate_reset(ip: str):
    """Limpia los intentos de una IP al loguearse con éxito."""
    with _login_attempts_lock:
        _login_attempts.pop(ip, None)

ADMIN_CUIT = '30710295022'
ADMIN_PASS = 'Artel2026'

DIRECTOR_USER = 'DIRECTORCOMERCIAL'
DIRECTOR_PASS = 'ARTEL2026'

# ── Fuentes BCRA externas ────────────────────────────────────────────────────
BCRA_WRAPPER_BASE = 'https://bcra-wrapper.vercel.app'   # proxy Vercel, sin rate-limit

# ── Caché macro ArgentinaDatos (24 h, una consulta diaria, no por CUIT) ─────
_macro_cache: dict = {'data': None, 'ts': 0.0}
_MACRO_TTL = 86400  # 24 horas

def _fetch_macro_data() -> dict:
    """Inflación interanual, riesgo país y dólar blue. Cachea 24h. Falla silenciosa."""
    global _macro_cache
    if _macro_cache['data'] and time.time() - _macro_cache['ts'] < _MACRO_TTL:
        return _macro_cache['data']
    result: dict = {}
    base = 'https://api.argentinadatos.com/v1'
    for url, key in [
        (base + '/finanzas/indices/inflacionInteranual',    'inflacion'),
        (base + '/finanzas/indices/riesgo-pais/ultimo',     'riesgo_pais'),
        (base + '/cotizaciones/dolares/blue',               'dolar_blue'),
    ]:
        try:
            r = requests.get(url, timeout=5, verify=False)
            if r.status_code == 200:
                d = r.json()
                if key == 'inflacion' and isinstance(d, list) and d:
                    result['inflacion'] = d[-1].get('valor')
                elif key == 'riesgo_pais' and isinstance(d, dict):
                    result['riesgo_pais'] = d.get('valor')
                elif key == 'dolar_blue':
                    if isinstance(d, list) and d:
                        u = d[-1]
                        result['dolar_blue_compra'] = u.get('compra')
                        result['dolar_blue_venta']  = u.get('venta')
                    elif isinstance(d, dict):
                        result['dolar_blue_compra'] = d.get('compra')
                        result['dolar_blue_venta']  = d.get('venta')
        except Exception:
            pass
    if result:
        _macro_cache['data'] = result
        _macro_cache['ts'] = time.time()
        print(f"[macro] inflacion={result.get('inflacion')} riesgo={result.get('riesgo_pais')} blue={result.get('dolar_blue_venta')}", flush=True)
    return result or (_macro_cache.get('data') or {})

# ── Startup: genera static/logo.png usando solo stdlib (sin PIL) ─────────────
def _generar_logo_png():
    import struct as _s, zlib as _z, math as _m
    W = H = 180; CX = CY = 90.0

    def _dseg(px, py, ax, ay, bx, by):
        dx, dy = bx - ax, by - ay
        l2 = dx*dx + dy*dy
        if l2 == 0:
            return _m.hypot(px - ax, py - ay)
        t = max(0.0, min(1.0, ((px - ax)*dx + (py - ay)*dy) / l2))
        return _m.hypot(px - ax - t*dx, py - ay - t*dy)

    def _chunk(tp, d):
        c = tp.encode('ascii') + d
        return _s.pack('>I', len(d)) + c + _s.pack('>I', _z.crc32(c) & 0xffffffff)

    rows = []
    for y in range(H):
        row = bytearray(1 + W * 3)
        for x in range(W):
            d = _m.hypot(x - CX, y - CY)
            r, g, b = 0x0b, 0x16, 0x28
            if d <= 76:  r, g, b = 0x25, 0x63, 0xeb
            if d <= 62:  r, g, b = 0x0b, 0x16, 0x28
            _e = min(
                _dseg(x, y,  54, 58, 126, 58),
                _dseg(x, y,  54, 58,  54, 96),
                _dseg(x, y, 126, 58, 126, 96),
                _dseg(x, y,  54, 96,  90, 133),
                _dseg(x, y, 126, 96,  90, 133),
            )
            if _e < 4.5: r, g, b = 0x60, 0xa5, 0xfa
            _v = min(_dseg(x, y, 72, 68, 90, 106), _dseg(x, y, 108, 68, 90, 106))
            if _v < 4.0: r, g, b = 0xff, 0xff, 0xff
            row[1 + x*3], row[1 + x*3+1], row[1 + x*3+2] = r, g, b
        rows.append(bytes(row))

    cmp  = _z.compress(b''.join(rows), 9)
    ihdr = _s.pack('>IIBBBBB', W, H, 8, 2, 0, 0, 0)
    return (b'\x89PNG\r\n\x1a\n' + _chunk('IHDR', ihdr)
            + _chunk('IDAT', cmp) + _chunk('IEND', b''))

_logo_path = os.path.join(app.static_folder, 'logo.png')
if not os.path.exists(_logo_path):
    try:
        with open(_logo_path, 'wb') as _lf:
            _lf.write(_generar_logo_png())
        print('[startup] logo.png generado OK', flush=True)
    except Exception as _le:
        print(f'[startup] logo.png error: {_le}', flush=True)

# Topes de facturación Monotributo 2026 — usados como ingreso estimado base
_MONOTRIB_INGRESOS = {
    'A':   3_500_000, 'B':   7_000_000, 'C':  11_500_000, 'D':  17_000_000,
    'E':  24_000_000, 'F':  34_000_000, 'G':  48_000_000, 'H':  67_000_000,
    'I':  93_000_000, 'J': 120_000_000, 'K': 155_000_000,
}

# User-Agents rotativos para evitar bloqueos de IP en scrapers públicos (AFIP, BORA)
_USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:125.0) Gecko/20100101 Firefox/125.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
]

# Estimación de ingresos anuales por rubro para Responsable Inscripto (2026)
# Lista ordenada por prioridad de match (más específico primero)
_ACTIVIDAD_INGRESOS_RI = [
    ('venta al por mayor',  180_000_000),
    ('bebida',              150_000_000),
    ('bodega',              130_000_000),
    ('vino',                120_000_000),
    ('distribucion',        120_000_000),
    ('licor',               100_000_000),
    ('construccion',        100_000_000),
    ('inmobilia',           100_000_000),
    ('agropec',              90_000_000),
    ('agricultura',          80_000_000),
    ('industria',           100_000_000),
    ('comercio',             80_000_000),
    ('transporte',           70_000_000),
    ('servicio',             55_000_000),
    ('venta al por menor',   40_000_000),
    ('restaurant',           45_000_000),
    ('gastronomia',          45_000_000),
]

GEMINI_MODEL = "gemini-2.0-flash"

# ── System prompt autoritativo — módulo de análisis de riesgo crediticio ──────
# Inyectado vía systemInstruction (Gemini) / role:system (OpenAI).
# Procesado por el modelo ANTES del prompt del usuario.
# v75.5: lógica basada en capacidad demostrada (sin multiplicadores), WhatsApp integrado.
CREDIT_ANALYSIS_SYSTEM_PROMPT = (
    "Sos un analista senior de riesgo crediticio de una distribuidora vitivinícola argentina. "
    "Tus dictámenes son técnicos, corporativos y fundados exclusivamente en los datos recibidos. "
    "Contexto operativo: una botella cuesta ~$15.000; un pedido mínimo de la bodega equivale "
    "a 5-10 cajas ($450.000-$900.000). El informe es leído por un comercial — "
    "debe ser limpio, ejecutivo y sin jerga técnica interna.\n\n"

    "REGLA 1 — TERMINOLOGÍA PROHIBIDA EN EL INFORME:\n"
    "El informe se titula siempre 'Análisis de Riesgo Crediticio'. "
    "Nunca uses en el texto de salida: 'forense', 'CRO', 'Director de Riesgos', "
    "'PICO_SIT1', 'piso operativo', 'multiplicador', 'factor de ajuste', "
    "'promedio mensual', 'promedio de deuda', 'DSO no disponible', 'DSO no reportado', "
    "'no hay datos de DSO', ni ninguna referencia a cómo se calculó el límite.\n\n"

    "REGLA 2 — LÍMITE DE CRÉDITO (DEFINITIVO — NO RECALCULAR):\n"
    "El prompt incluye una sección 'RANGO OFICIAL DE CRÉDITO' con el límite ya determinado "
    "por el sistema crediticio. Este valor es DEFINITIVO e INAPELABLE — el sistema ya aplicó "
    "todos los ajustes de riesgo (score, historial BCRA, mora, cheques, degradaciones).\n"
    "Tu tarea es VALIDAR ese rango con los datos del cliente y EXPLICAR por qué es adecuado. "
    "En RECOMENDACIÓN ESTRATÉGICA debés citar EXACTAMENTE la frase del rango oficial "
    "tal como aparece en el prompt, sin modificarla en ningún aspecto.\n"
    "PROHIBIDO: calcular, inferir, sugerir o mencionar cualquier monto de crédito diferente "
    "al oficial. Si detectás riesgos adicionales, usá 'se recomienda seguimiento activo' — "
    "nunca reduciendo ni alterando el límite.\n\n"

    "REGLA 3 — VERACIDAD Y FOCO DE CIFRAS:\n"
    "El historial incluye una sección 'MONTO VIGENTE POR ENTIDAD' con el valor actual "
    "de cada entidad. Usá EXCLUSIVAMENTE esos valores al mencionar montos por entidad. "
    "PROHIBIDO calcular o reportar rangos históricos de montos ('fluctuó de X a Y', "
    "'entre X y Y', 'desde X hasta Y'). Si la situación fue siempre Sit.1, el monto "
    "histórico es irrelevante — solo importa el vigente. "
    "Ejemplo correcto: 'Banco Macro: $5.254.000 en Sit.1 sostenida.' "
    "Ejemplo incorrecto: 'montos que fluctuaron desde $5.254.000 hasta $382.000.' "
    "Solo mencioná variación de monto si va acompañada de degradación de situación "
    "(Sit.2 o peor) — y en ese caso, citá el monto del período afectado, no el rango.\n\n"

    "REGLA 4 — HISTORIAL INSUFICIENTE:\n"
    "Si no hay montos concretos en el historial BCRA, señalá que el cliente es nuevo "
    "en el sistema financiero. No inventes cifras. "
    "El rango oficial del sistema ya contempla esta condición.\n\n"

    "REGLA 5 — OBSERVACIONES DE MERCADO:\n"
    "El prompt puede incluir mensajes bajo 'Menciones grupo bodegas'. "
    "Integrá los hallazgos relevantes como una nota breve bajo el título "
    "'OBSERVACIONES DE MERCADO', sin revelar la fuente ni el mecanismo de análisis. "
    "Si no hay alertas, escribí: 'Sin observaciones de mercado en los últimos 6 meses.'\n\n"

    "REGLA 6 — CLÁUSULA DE SUSPENSIÓN (OBLIGATORIA):\n"
    "Toda recomendación DEBE cerrar con: 'Ante cualquier nueva degradación bancaria "
    "(Situación >= 2), suspender venta a crédito automáticamente.'\n\n"

    "REGLA 7B — COMUNICACIÓN DE MONTOS (lenguaje ejecutivo):\n"
    "Al mencionar deuda en el sistema financiero: usá siempre pesos con formato claro. "
    "Para montos < $1.000.000: escribí '$XX.XXX' (p.ej. '$60.000', '$500.000'). "
    "Para montos >= $1.000.000: escribí '$X,XM' (p.ej. '$1,2M', '$3,5M'). "
    "NUNCA uses '$0.1M', '$0.5M' ni notaciones fraccionarias para montos en miles. "
    "Deuda baja en el sistema (< $500.000) es indicador de bajo apalancamiento — "
    "no lo interpretes como 'capacidad restringida'. Interpretalo como perfil conservador.\n\n"

    "REGLA 7 — FORMATO Y EXTENSIÓN:\n"
    "Español corporativo. Sin markdown, sin asteriscos. Máximo 280 palabras. "
    "Estructura de salida OBLIGATORIA (respetar exactamente estos títulos y orden):\n"
    "ANÁLISIS DE RIESGO CREDITICIO: "
    "[comportamiento 24m por entidad + situaciones observadas + pagos internos + tendencia]\n"
    "OBSERVACIONES DE MERCADO: "
    "[hallazgos en red de bodegas o 'Sin observaciones de mercado en los últimos 6 meses.']\n"
    "DIAGNÓSTICO FINANCIERO: "
    "[evaluación de salud crediticia consolidada del cliente]\n"
    "RECOMENDACIÓN ESTRATÉGICA: "
    "[citar el rango oficial exacto del prompt + validar con perfil observado + cláusula de suspensión]"
)

DATA_DIR      = '/data' if os.path.exists('/data') else os.getcwd()
PADRON_DB_PATH  = os.path.join(DATA_DIR, 'bcra_padron.db')
NOMDEU_DB_PATH  = os.path.join(DATA_DIR, 'bcra_nomdeu.db')
MIPYME_DB_PATH  = os.path.join(DATA_DIR, 'mipyme_padron.db')
MIPYME_CSV_URL  = (
    'https://datos.produccion.gob.ar/dataset/registro-mipyme/'
    'archivo/bd407e64-0f11-44a2-b1d6-9a7a05700d73'
)
# Topes de facturación anual MiPyME — Resolución SEyPyME 1/2026 (en ARS)
TOPES_FACTURACION_ANUAL: dict = {
    'Comercio':            {'Micro': 1_738_060_000, 'Pequeña': 12_380_800_000, 'Mediana_T1': 57_922_750_000, 'Mediana_T2': 84_070_280_000},
    'Servicios':           {'Micro':   374_060_000, 'Pequeña':  2_666_040_000, 'Mediana_T1': 12_470_690_000, 'Mediana_T2': 18_097_990_000},
    'Industria y Minería': {'Micro': 1_097_270_000, 'Pequeña':  7_820_750_000, 'Mediana_T1': 36_594_360_000, 'Mediana_T2': 53_083_920_000},
    'Construcción':        {'Micro':   583_520_000, 'Pequeña':  4_158_750_000, 'Mediana_T1': 19_459_900_000, 'Mediana_T2': 28_238_310_000},
    'Agropecuario':        {'Micro':   831_000_000, 'Pequeña':  5_924_300_000, 'Mediana_T1': 27_726_150_000, 'Mediana_T2': 40_237_220_000},
}
# Rango de empleados estimado por categoría MiPyME (fuente: SEPYME estándar)
_EMPLEADOS_RANGO: dict = {
    'Micro':      '1-10',
    'Pequeña':   '11-50',
    'Mediana_T1': '51-200',
    'Mediana_T2': '51-200',
}
ALERTAS_FILE        = os.path.join(DATA_DIR, 'db_v17_final.json')
ALERTAS_BCRA_FILE   = os.path.join(DATA_DIR, 'alertas_bcra.json')
DATOS_FILE          = os.path.join(DATA_DIR, 'datos_bodega.json')
SCORE_CACHE_FILE    = os.path.join(DATA_DIR, 'score_cache.json')
NOMBRES_CUSTOM_FILE = os.path.join(DATA_DIR, 'nombres_custom.json')
print(f"[init] Almacenamiento en: {DATA_DIR}", flush=True)
print(
    f"[init] ScraperAPI: {'ACTIVO — proxy rotativo habilitado' if SCRAPERAPI_KEY else 'no configurado — modo directo legacy'}",
    flush=True,
)
WSP_FILE = os.path.join(os.getcwd(), 'whatsapp_index.json')

# ── Estado local de facturas: pendiente_validacion + enviado_whatsapp ─────────
_FACTURAS_ESTADO_FILE = os.path.join(DATA_DIR, 'facturas_estado.json')
_facturas_estado_lock = threading.Lock()

def _fac_key(cuit_limpio: str, nro: str) -> str:
    return f"{cuit_limpio}__{nro}"

def _fac_estado_load() -> dict:
    try:
        if os.path.exists(_FACTURAS_ESTADO_FILE):
            with open(_FACTURAS_ESTADO_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"[fac_estado] Error cargando: {e}", flush=True)
    return {}

def _fac_estado_save(estado: dict):
    try:
        with open(_FACTURAS_ESTADO_FILE, 'w', encoding='utf-8') as f:
            json.dump(estado, f, ensure_ascii=False)
    except Exception as e:
        print(f"[fac_estado] Error guardando: {e}", flush=True)

def _fac_anotar_estado(enriched: list, cuit_limpio: str) -> list:
    """Agrega _cobrada y _enviado_whatsapp a cada factura desde el estado persistido."""
    estado = _fac_estado_load()
    for f in enriched:
        key = _fac_key(cuit_limpio, str(f.get('nroFactura', '')))
        e = estado.get(key, {})
        f['_cobrada'] = e.get('estado') == 'pendiente_validacion'
        f['_enviado_whatsapp'] = bool(e.get('enviado_whatsapp', False))
    return enriched

bcra_cache = {}
CACHE_TTL = 60 * 60 * 24   # 24 horas — reduce consultas al BCRA y mejora latencia
CACHE_TTL_ERROR = 300
BCRA_VACIO = {"results": None, "sin_deudas": None, "error_bcra": None}

# ── Nombres personalizados (admin puede fijar nombre manualmente cuando todas las fuentes fallan) ──
_nombres_custom: dict = {}
try:
    if os.path.exists(NOMBRES_CUSTOM_FILE):
        with open(NOMBRES_CUSTOM_FILE, 'r', encoding='utf-8') as _nc_f:
            _nombres_custom = json.load(_nc_f)
        print(f"[nombres_custom] {len(_nombres_custom)} nombres cargados", flush=True)
except Exception as _nc_e:
    print(f"[nombres_custom] Error cargando: {_nc_e}", flush=True)

# ── Cartera comercial ──
_cartera_comercial = []
_CC_FILE = os.path.join(DATA_DIR, 'cartera_comercial.json')
try:
    # DATA_DIR first (persistent on Render), fallback to repo copy
    _cc_path = _CC_FILE if os.path.exists(_CC_FILE) else os.path.join(os.getcwd(), 'cartera_comercial.json')
    with open(_cc_path, encoding='utf-8') as f:
        _cartera_comercial = json.load(f)
    print(f"[comercial] {len(_cartera_comercial)} clientes cargados desde {_cc_path}", flush=True)
except Exception as e:
    print(f"[comercial] Error cargando cartera: {e}", flush=True)


def _cc_desde_disco() -> list:
    """Lee cartera_comercial.json desde disco — necesario en entornos multi-worker
    (gunicorn) donde cada worker tiene su propio espacio de memoria y el upload
    puede haber ocurrido en un worker distinto al que inicia el proceso."""
    global _cartera_comercial
    try:
        _p = _CC_FILE if os.path.exists(_CC_FILE) else os.path.join(os.getcwd(), 'cartera_comercial.json')
        with open(_p, 'r', encoding='utf-8') as _f:
            fresh = json.load(_f)
        if len(fresh) != len(_cartera_comercial):
            print(f"[cc-reload] disco={len(fresh)} mem={len(_cartera_comercial)} → actualizando", flush=True)
            _cartera_comercial = fresh
        return fresh
    except Exception as _e:
        print(f"[cc-reload] Error leyendo disco ({_e}), usando memoria", flush=True)
        return _cartera_comercial

def cache_get(cuit):
    try:
        cf = os.path.join(DATA_DIR, 'bcra_cache.json') if os.path.exists(DATA_DIR) else '/tmp/bcra_cache.json'
        if os.path.exists(cf):
            with open(cf, 'r') as f:
                cache = json.load(f)
            if cuit in cache:
                entry = cache[cuit]
                ttl = 300 if entry.get('error') else CACHE_TTL
                if time.time() - entry.get('ts', 0) < ttl:
                    return entry.get('data'), entry.get('error')
    except: pass
    return None, None

def cache_set(cuit, data, error=None):
    try:
        cf = os.path.join(DATA_DIR, 'bcra_cache.json') if os.path.exists(DATA_DIR) else '/tmp/bcra_cache.json'
        cache = {}
        if os.path.exists(cf):
            with open(cf, 'r') as f:
                cache = json.load(f)
        cache[cuit] = {'data': data, 'error': error, 'ts': time.time()}
        ahora = time.time()
        cache = {k: v for k, v in cache.items() if ahora - v.get('ts', 0) < CACHE_TTL * 2}
        with open(cf, 'w') as f:
            json.dump(cache, f)
    except: pass

# ── Padrón local BCRA (SQLite offline) ─────────────────────────────────────

def _init_padron_db():
    """Crea la tabla e índice único del padrón local si no existen.

    Ejecuta una purga única (versionada) de entradas corruptas generadas por el
    pre-cacheo masivo: respuestas vacías de la API legacy que no reporta clientes
    en Sit 1 sin mora. La purga solo corre una vez por versión de esquema.
    """
    _SCHEMA_VER = 'purge_v2'
    try:
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS bcra_padron_local (
                cuit          TEXT PRIMARY KEY,
                denominacion  TEXT,
                periodo       TEXT,
                sit_max       INTEGER,
                monto_total   INTEGER,
                num_entidades INTEGER,
                detalle       TEXT,
                importado_en  TEXT
            )
        """)
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_padron_cuit ON bcra_padron_local(cuit)"
        )
        # Tabla de metadatos para control de versión de schema/purgas
        conn.execute("""
            CREATE TABLE IF NOT EXISTS _padron_meta (
                key   TEXT PRIMARY KEY,
                valor TEXT
            )
        """)
        # Tabla de cheques rechazados BCRA (snapshot diario bulk file)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cheques_bcra (
                cuit          TEXT NOT NULL,
                nro_cheque    TEXT NOT NULL,
                fecha_rechazo TEXT NOT NULL,
                monto         REAL NOT NULL DEFAULT 0,
                estado        TEXT,
                tipo          TEXT,
                cuit_entidad  TEXT,
                PRIMARY KEY (cuit, nro_cheque)
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cheques_cuit ON cheques_bcra(cuit)"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS _cheques_meta (
                key   TEXT PRIMARY KEY,
                valor TEXT
            )
        """)
        conn.commit()

        # Purga única: eliminar todas las entradas sin entidades reales.
        # Estas fueron guardadas por el pre-cacheo masivo cuando la API legacy
        # devolvió 404 para clientes en Sit 1 que no tienen mora (falso negativo).
        ya_purgado = conn.execute(
            "SELECT valor FROM _padron_meta WHERE key = ?", (_SCHEMA_VER,)
        ).fetchone()
        if ya_purgado is None:
            cur = conn.execute(
                "DELETE FROM bcra_padron_local WHERE num_entidades = 0 OR detalle = '[]' OR detalle IS NULL OR detalle = ''"
            )
            eliminados = cur.rowcount
            conn.execute(
                "INSERT OR REPLACE INTO _padron_meta (key, valor) VALUES (?, ?)",
                (_SCHEMA_VER, time.strftime('%Y-%m-%dT%H:%M:%S'))
            )
            conn.commit()
            print(f"[padron] Purga {_SCHEMA_VER}: {eliminados} entradas corruptas eliminadas", flush=True)

        conn.close()
        print(f"[padron] DB inicializada en {PADRON_DB_PATH}", flush=True)
    except Exception as e:
        print(f"[padron] Error al inicializar DB: {e}", flush=True)


def consultar_padron_local(cuit_limpio):
    """Busca el CUIT en el padrón mensual local. Retorna respuesta BCRA-compatible o None."""
    try:
        if not os.path.exists(PADRON_DB_PATH):
            return None
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM bcra_padron_local WHERE cuit = ?", (cuit_limpio,)
        ).fetchone()
        conn.close()
        if row is None:
            return None
        detalle = json.loads(row['detalle'] or '[]')
        # Entradas sin entidades son basura de pre-cacheo con 404 falso (Sit 1 sin mora).
        # Retornar None obliga a una consulta en vivo en lugar de servir datos vacíos.
        if not detalle:
            return None
        # sin_deudas solo es True cuando realmente no hay entidades reportantes.
        # Sit 1 con monto > 0 NO es "sin deudas": tiene crédito activo al día.
        sin_deudas = len(detalle) == 0
        return {
            "results": {
                "periodos":     [{"entidades": detalle}],
                "denominacion": row['denominacion'] or "",
            },
            "sin_deudas":     sin_deudas,
            "bcra_disponible": True,
            "fuente":          "padron_local",
            "periodo_padron":  row['periodo'],
        }
    except Exception as e:
        print(f"[padron] Error consultando {cuit_limpio}: {e}", flush=True)
        return None


# ── Cheques rechazados — DB local (bulk file diario BCRA) ─────────────────────
# El BCRA publica cada día en https://www.bcra.gob.ar/archivos/zips/cheques/YYYYMMDD.zip
# un snapshot completo de TODOS los cheques rechazados del sistema financiero.
# Formato fijo 81 chars/línea:
#   cols  0-10  CUIT librador (11 chars)
#   cols 11-20  NRO_CHEQUE (10 chars, right-justified)
#   cols 21-28  FECHA_RECHAZO (8 chars, YYYYMMDD)
#   cols 29-43  MONTO (15 chars, centavos enteros → ÷100 = pesos)
#   cols 44-52  FECHA_MULTA (9 chars, YYYYMMDD o espacios)
#   cols 53-54  TIPO (2 chars: 'SF' = sin fondos)
#   cols 55-65  CUIT_ENTIDAD bancaria (11 chars)
#   cols 66-75  ESTADO (10 chars: 'IMPAGA    ' o 'DD/MM/YYYY' = fecha de pago)
#   cols 76-80  padding (5 chars)

_cheques_db_estado = {
    "corriendo":       False,
    "progreso":        0,
    "total":           0,
    "ultimo_paso":     "",
    "fecha_importada": "",
}

# Backup en R2 del último ZIP de cheques importado con éxito — permite
# restaurar la DB si BCRA/proxies fallan y el disco local está vacío
# (ej: redeploy en Render sin volumen persistente).
_CHEQUES_ZIP_R2_KEY  = 'cheques_bcra_ultimo.zip'
_CHEQUES_ZIP_R2_META = 'cheques_bcra_ultimo_meta.json'


def get_cheques_local(cuit: str):
    """Consulta cheques rechazados del CUIT en la DB local (snapshot diario BCRA).

    Retorna dict compatible con la respuesta BCRA o None si la DB no está
    disponible / aún no fue importada.
    """
    if not os.path.exists(PADRON_DB_PATH):
        return None
    try:
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        meta_row = conn.execute(
            "SELECT valor FROM _cheques_meta WHERE key = 'last_import_date'"
        ).fetchone()
        if not meta_row:
            conn.close()
            return None  # DB vacía: no se ha importado aún
        rows = conn.execute(
            "SELECT nro_cheque, fecha_rechazo, monto, estado, tipo, cuit_entidad "
            "FROM cheques_bcra WHERE cuit = ? ORDER BY fecha_rechazo DESC",
            (cuit,)
        ).fetchall()
        conn.close()

        if not rows:
            return {
                "results":    {"causales": []},
                "sin_deudas": True,
                "error_bcra": None,
                "source":     "local_db",
            }

        detalles = []
        for nro, fecha_r, monto, estado, tipo, cuit_ent in rows:
            estado_s = (estado or '').strip()
            if estado_s == 'IMPAGA':
                fecha_pago   = None
                estado_multa = 'IMPAGA'
            else:
                # BCRA almacena la fecha de pago como 'DD/MM/YYYY'
                try:
                    d, m, y  = estado_s.split('/')
                    fecha_pago = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                except Exception:
                    fecha_pago = estado_s or None
                estado_multa = 'LEVANTADA'

            fecha_r_fmt = fecha_r
            if len(fecha_r) == 8:
                fecha_r_fmt = f"{fecha_r[:4]}-{fecha_r[4:6]}-{fecha_r[6:8]}"

            detalles.append({
                "numeroCheque":       (nro       or '').strip(),
                "fechaRechazo":       fecha_r_fmt,
                "monto":              monto,
                "fechaPago":          fecha_pago,
                "estadoMulta":        estado_multa,
                "fechaLevantamiento": fecha_pago,
                "tipo":               (tipo      or '').strip(),
                "cuitEntidad":        (cuit_ent  or '').strip(),
            })

        return {
            "results": {
                "causales": [{"entidades": [{"detalle": detalles}]}]
            },
            "sin_deudas": False,
            "error_bcra": None,
            "source":     "local_db",
        }
    except Exception as e:
        print(f"[cheques_local] Error consultando {cuit}: {e}", flush=True)
        return None


def _cheques_activos_de(cheq_data: dict):
    """Extrae (cantidad_activos, total, detalle) de una respuesta de cheques
    (en vivo, caché o DB local) — activo = sin fecha de pago o IMPAGA."""
    if not cheq_data or cheq_data.get('sin_deudas') or not isinstance(cheq_data.get('results'), dict):
        return 0, 0, []
    causales = cheq_data['results'].get('causales') or []
    detalle = []
    for cau in causales:
        if isinstance(cau, dict):
            for ent in (cau.get('entidades') or []):
                if isinstance(ent, dict):
                    detalle.extend(x for x in (ent.get('detalle') or []) if isinstance(x, dict))
    activos = sum(1 for d in detalle if not d.get('fechaPago') or d.get('estadoMulta') == 'IMPAGA')
    return activos, len(detalle), detalle


def _import_cheques_zip(date_str: str = None) -> bool:
    """Descarga e importa el snapshot diario de cheques rechazados del BCRA.

    El archivo 'al{YYMMDD}.txt' dentro del ZIP es un reemplazo total del día:
    contiene TODOS los cheques activos del sistema financiero argentino (~700k).
    La tabla cheques_bcra se trunca y se reconstruye en cada importación.

    Args:
        date_str: Fecha en formato 'YYYYMMDD'. None = fecha de hoy.
    Returns:
        True si OK, False si falló.
    """
    import zipfile as _zipfile

    if not date_str:
        date_str = time.strftime('%Y%m%d')

    _cheques_db_estado['corriendo']   = True
    _cheques_db_estado['progreso']    = 0
    _cheques_db_estado['ultimo_paso'] = f"Iniciando descarga para {date_str}"

    # BCRA publica cheques rechazados en dos URLs conocidas (la segunda es el fallback
    # por si BCRA cambia la ubicación del archivo, como ocurrió al mover a /actualiza/).
    _URL_CANDIDATAS = [
        f"https://www.bcra.gob.ar/archivos/zips/cheques/{date_str}.zip",
        f"https://www.bcra.gob.ar/actualiza/{date_str}.zip",
    ]
    url      = _URL_CANDIDATAS[0]
    zip_path = os.path.join(DATA_DIR, f"cheques_{date_str}.zip")
    al_path  = None

    try:
        # 1. Descarga streaming del ZIP (~10-15 MB comprimido, ~58 MB expandido)
        # Estrategia: probar URLs candidatas en orden, Bright Data como último recurso.
        _browser_hdrs = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
            'Accept': 'application/zip,application/octet-stream,*/*;q=0.8',
            'Accept-Language': 'es-AR,es;q=0.9,en;q=0.8',
        }

        resp = None
        _status_directos = []   # códigos HTTP de los intentos directos

        # Intento 1 y 1b: URLs candidatas directas (sin proxy)
        for _url_c in _URL_CANDIDATAS:
            _cheques_db_estado['ultimo_paso'] = f"Descargando {_url_c}"
            print(f"[cheques_db] Intentando {_url_c}", flush=True)
            try:
                _r = requests.get(_url_c, headers=_browser_hdrs, timeout=90, verify=False, stream=True)
                _status_directos.append(_r.status_code)
                if _r.status_code == 200:
                    resp = _r
                    url  = _url_c
                    print(f"[cheques_db] OK en {_url_c}", flush=True)
                    break
                else:
                    print(f"[cheques_db] HTTP {_r.status_code} en {_url_c}", flush=True)
            except Exception as e_d:
                print(f"[cheques_db] Error en {_url_c}: {e_d}", flush=True)

        # Si TODAS las respuestas directas fueron 404, el archivo no existe todavía
        # (BCRA no publica fines de semana/feriados) — un proxy no cambia un 404,
        # así que se saltean para no quemar créditos ni demorar el loop.
        _archivo_inexistente = bool(_status_directos) and all(s == 404 for s in _status_directos)

        if resp is None and not _archivo_inexistente:
            print(f"[cheques_db] Todas las URLs directas fallaron — intentando vía proxy", flush=True)

        # Intento 2: ScraperAPI — proxy rotativo con salida en Argentina.
        if resp is None and not _archivo_inexistente and SCRAPERAPI_KEY:
            print(f"[cheques_db] Descarga vía ScraperAPI", flush=True)
            for _url_c3 in _URL_CANDIDATAS:
                try:
                    _r3 = requests.get(
                        'http://api.scraperapi.com',
                        params={'api_key': SCRAPERAPI_KEY, 'url': _url_c3,
                                'country_code': 'ar', 'binary_target': 'binary'},
                        timeout=180, stream=True)
                    if _r3.status_code == 200:
                        resp = _r3
                        url  = _url_c3
                        print(f"[cheques_db] ScraperAPI OK en {_url_c3}", flush=True)
                        break
                    else:
                        print(f"[cheques_db] ScraperAPI HTTP {_r3.status_code} en {_url_c3}", flush=True)
                except Exception as e_s:
                    print(f"[cheques_db] ScraperAPI error en {_url_c3}: {e_s}", flush=True)

        # Sin descarga posible → resiliencia:
        #   a) si la DB vigente tiene datos, se mantiene activa tal cual (el
        #      truncado solo ocurre con un ZIP nuevo validado) y se corta acá;
        #   b) si está vacía (disco nuevo post-redeploy), se restaura el último
        #      ZIP válido desde R2 y se importa ese.
        _desde_r2 = False
        if resp is None:
            _regs_actuales = 0
            try:
                _c_chk = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
                _regs_actuales = _c_chk.execute("SELECT COUNT(*) FROM cheques_bcra").fetchone()[0]
                _c_chk.close()
            except Exception:
                pass

            if _regs_actuales > 0:
                _motivo = "aún no publicado (404)" if _archivo_inexistente else "descarga fallida"
                _cheques_db_estado['ultimo_paso'] = (
                    f"Sin archivo para {date_str} ({_motivo}) — "
                    f"DB vigente se mantiene ({_regs_actuales:,} registros)"
                )
                print(f"[cheques_db] {date_str} {_motivo} — DB actual sigue activa "
                      f"({_regs_actuales:,} registros)", flush=True)
                return False

            _r2_zip = _r2_download_bytes(_CHEQUES_ZIP_R2_KEY)
            if _r2_zip:
                # La fecha real del backup viene en el JSON compañero (el ZIP
                # puede ser de días atrás — no registrar la fecha de hoy)
                try:
                    _meta_raw = _r2_download_bytes(_CHEQUES_ZIP_R2_META)
                    if _meta_raw:
                        date_str = json.loads(_meta_raw.decode()).get('date') or date_str
                except Exception:
                    pass
                with open(zip_path, 'wb') as zf:
                    zf.write(_r2_zip)
                _desde_r2 = True
                print(f"[cheques_db] DB vacía — restaurando último ZIP válido desde R2 "
                      f"({len(_r2_zip) // 1048576} MB, fecha {date_str})", flush=True)
            else:
                _cheques_db_estado['ultimo_paso'] = "ERROR: descarga imposible y sin backup R2"
                return False

        if not _desde_r2:
            with open(zip_path, 'wb') as zf:
                for chunk in resp.iter_content(chunk_size=65536):
                    if chunk:
                        zf.write(chunk)

        tam_mb = os.path.getsize(zip_path) / (1024 * 1024)

        # Validar magic bytes: ZIP comienza con PK\x03\x04
        with open(zip_path, 'rb') as _f:
            _magic = _f.read(4)
        if _magic[:2] != b'PK':
            with open(zip_path, 'r', encoding='utf-8', errors='replace') as _f:
                _snippet = _f.read(200)
            _cheques_db_estado['ultimo_paso'] = f"ERROR: descarga no es ZIP ({tam_mb:.1f}MB). Contenido: {_snippet[:100]}"
            print(f"[cheques_db] No es ZIP. Primeros bytes: {repr(_magic)} | {_snippet[:150]}", flush=True)
            return False

        print(f"[cheques_db] ZIP válido descargado: {tam_mb:.1f} MB", flush=True)

        # Backup a R2 del ZIP recién descargado (background) — es la fuente de
        # restauración si mañana falla la descarga y el disco local está vacío.
        # No re-subir si este mismo ZIP ya vino de R2 (_desde_r2).
        if not _desde_r2:
            with open(zip_path, 'rb') as _f:
                _zip_bytes_bk = _f.read()

            def _bg_backup_cheques(data=_zip_bytes_bk, fecha=date_str):
                try:
                    ok1 = _r2_upload_bytes(_CHEQUES_ZIP_R2_KEY, data, 'application/zip')
                    ok2 = _r2_upload_bytes(
                        _CHEQUES_ZIP_R2_META,
                        json.dumps({'date': fecha, 'ts': time.time()}).encode(),
                        'application/json',
                    )
                    print(f"[cheques_db] Backup R2 {'OK' if ok1 and ok2 else 'FALLÓ'} "
                          f"({len(data) // 1048576} MB, fecha {fecha})", flush=True)
                except Exception as _e_bk:
                    print(f"[cheques_db] Backup R2 error: {_e_bk}", flush=True)
            threading.Thread(target=_bg_backup_cheques, daemon=True).start()

        # 2. Extraer solo el archivo 'al*' (snapshot completo)
        # El archivo puede estar en la raíz o dentro de un subdirectorio del ZIP.
        _cheques_db_estado['ultimo_paso'] = "Extrayendo ZIP"
        with _zipfile.ZipFile(zip_path, 'r') as zf:
            names   = zf.namelist()
            print(f"[cheques_db] Contenido ZIP: {names}", flush=True)
            # Buscar por basename para manejar subdirectorios dentro del ZIP
            al_name = next(
                (n for n in names if os.path.basename(n).lower().startswith('al') and not n.endswith('/')),
                None
            )
            if not al_name:
                _cheques_db_estado['ultimo_paso'] = f"ERROR: sin 'al*' en ZIP. Contenido: {names}"
                print(f"[cheques_db] No se encontró 'al*'. Nombres: {names}", flush=True)
                return False
            zf.extract(al_name, DATA_DIR)
            # zf.extract preserva la estructura de directorios del ZIP
            al_path = os.path.join(DATA_DIR, al_name.replace('\\', os.sep).replace('/', os.sep))
        print(f"[cheques_db] Extraído: {al_name} → {al_path}", flush=True)

        # 3. Parsear el archivo de ancho fijo e importar a SQLite en lotes
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA cache_size=-32768")  # 32 MB de cache SQLite

        _cheques_db_estado['ultimo_paso'] = "Truncando tabla anterior"
        conn.execute("DELETE FROM cheques_bcra")
        conn.commit()

        lote    = []
        lineas  = 0
        errores = 0
        t0      = time.time()
        _cheques_db_estado['ultimo_paso'] = "Importando registros"

        with open(al_path, 'r', encoding='latin-1', errors='replace') as f:
            for linea in f:
                if len(linea) < 29:
                    continue
                try:
                    cuit       = linea[0:11].strip()
                    nro_cheque = linea[11:21].strip()
                    fecha_r    = linea[21:29].strip()
                    monto_str  = linea[29:44].strip()
                    tipo       = linea[53:55].strip() if len(linea) > 55 else ''
                    cuit_ent   = linea[55:66].strip() if len(linea) > 66 else ''
                    estado     = linea[66:76].strip() if len(linea) > 76 else 'IMPAGA'

                    if len(cuit) != 11 or not cuit.isdigit():
                        continue
                    if not nro_cheque or not fecha_r:
                        continue

                    try:
                        monto = round(int(monto_str or '0') / 100, 2)
                    except (ValueError, TypeError):
                        monto = 0.0

                    lote.append((cuit, nro_cheque, fecha_r, monto, estado, tipo, cuit_ent))
                    lineas += 1

                    if len(lote) >= 5000:
                        conn.executemany(
                            "INSERT OR REPLACE INTO cheques_bcra "
                            "(cuit, nro_cheque, fecha_rechazo, monto, estado, tipo, cuit_entidad) "
                            "VALUES (?,?,?,?,?,?,?)",
                            lote
                        )
                        conn.commit()
                        lote.clear()
                        _cheques_db_estado['progreso'] = lineas
                        if lineas % 100_000 == 0:
                            elapsed = time.time() - t0
                            print(f"[cheques_db]   {lineas:,} líneas — {elapsed:.0f}s", flush=True)
                except Exception as erow:
                    errores += 1
                    if errores <= 5:
                        print(f"[cheques_db] Error fila: {erow} | {repr(linea[:40])}", flush=True)

        if lote:
            conn.executemany(
                "INSERT OR REPLACE INTO cheques_bcra "
                "(cuit, nro_cheque, fecha_rechazo, monto, estado, tipo, cuit_entidad) "
                "VALUES (?,?,?,?,?,?,?)",
                lote
            )
            conn.commit()

        # 4. Registrar metadatos de la importación
        ts_now = time.strftime('%Y-%m-%dT%H:%M:%S')
        for key, val in [
            ('last_import_date', date_str),
            ('last_import_ts',   ts_now),
            ('total_registros',  str(lineas)),
        ]:
            conn.execute(
                "INSERT OR REPLACE INTO _cheques_meta (key, valor) VALUES (?, ?)",
                (key, val)
            )
        conn.commit()
        conn.close()

        elapsed = time.time() - t0
        print(
            f"[cheques_db] IMPORTACIÓN COMPLETA: {lineas:,} registros en {elapsed:.1f}s "
            f"(errores: {errores})",
            flush=True,
        )
        _cheques_db_estado['total']           = lineas
        _cheques_db_estado['fecha_importada'] = date_str
        _cheques_db_estado['ultimo_paso']     = f"OK — {lineas:,} registros"
        return True

    except Exception as e:
        print(f"[cheques_db] Error en importación: {e}", flush=True)
        _cheques_db_estado['ultimo_paso'] = f"ERROR: {e}"
        return False
    finally:
        _cheques_db_estado['corriendo'] = False
        for tmp in filter(None, [zip_path, al_path]):
            try:
                if os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                pass


def _padron_contar_registros():
    """Devuelve (total_cuits, periodo_mas_reciente) del padrón local."""
    try:
        if not os.path.exists(PADRON_DB_PATH):
            return 0, None
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        row = conn.execute(
            "SELECT COUNT(*) as total, MAX(periodo) as periodo FROM bcra_padron_local"
        ).fetchone()
        conn.close()
        return (row[0] or 0), (row[1] or None)
    except Exception:
        return 0, None


# Estado global del proceso de importación (un import a la vez)
_padron_import_estado: dict = {
    "corriendo":  False,
    "lineas":     0,
    "insertados": 0,
    "mensaje":    "Sin importación activa",
    "error":      None,
}


def _importar_padron_worker(ruta_archivo: str, borrar_fuente: bool = True):
    """Importa el padrón BCRA desde un archivo de texto.

    Diseñado para correr en hilo de fondo. Lee línea a línea (streaming,
    <120 MB RAM) e inserta en SQLite en lotes de 5.000.

    Soporta delimitadores `;`, `|`, `,` y detección automática de columnas
    por cabecera. El archivo original se borra al finalizar si `borrar_fuente`
    es True (libera disco en Render).
    """
    global _padron_import_estado

    # Mapeo flexible de nombres de columnas → campos internos
    _ALIAS = {
        'cuit':              'cuit',
        'identificacion':    'cuit',
        'id':                'cuit',
        'denominacion':      'denominacion',
        'razon_social':      'denominacion',
        'nombre':            'denominacion',
        'denom':             'denominacion',
        'entidad':           'entidad',
        'nom_entidad':       'entidad',
        'nombre_entidad':    'entidad',
        'banco':             'entidad',
        'cod_entidad':       'entidad',
        'situacion':         'situacion',
        'sit':               'situacion',
        'calificacion':      'situacion',
        'monto':             'monto',
        'monto_miles':       'monto',
        'saldo':             'monto',
        'periodo':           'periodo',
        'periodo_informado': 'periodo',
        'periodoInformado':  'periodo',
        'fecha':             'periodo',
        'dias_atraso':       'dias_atraso',
        'diasatraso':        'dias_atraso',
        'dias':              'dias_atraso',
    }

    _padron_import_estado = {
        "corriendo": True, "lineas": 0, "insertados": 0,
        "mensaje": "Iniciando...", "error": None,
    }

    conn = None
    try:
        # ── Detectar encoding y delimitador ─────────────────────────────────
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                with open(ruta_archivo, 'r', encoding=enc, errors='strict') as _f:
                    primera = _f.readline().strip()
                encoding = enc
                break
            except UnicodeDecodeError:
                encoding = 'latin-1'
                with open(ruta_archivo, 'r', encoding='latin-1') as _f:
                    primera = _f.readline().strip()

        if '|' in primera:
            delimitador = '|'
        elif ';' in primera:
            delimitador = ';'
        elif '\t' in primera:
            delimitador = '\t'
        else:
            delimitador = ','

        # Detectar si la primera línea es cabecera o dato
        tokens = [t.strip().lower() for t in primera.split(delimitador)]
        tiene_header = any(t in _ALIAS for t in tokens)

        if tiene_header:
            idx_map = {}
            for i, t in enumerate(tokens):
                campo = _ALIAS.get(t)
                if campo and campo not in idx_map:
                    idx_map[campo] = i
        else:
            # Asumir formato posicional estándar BCRA: CUIT|DENOM|PERIODO|ENTIDAD|SIT|MONTO
            idx_map = {'cuit': 0, 'denominacion': 1, 'periodo': 2, 'entidad': 3, 'situacion': 4, 'monto': 5}

        _padron_import_estado['mensaje'] = f"Formato detectado: delim='{delimitador}', enc={encoding}"

        # ── Preparar DB: tabla staging ───────────────────────────────────────
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS _padron_staging (
                cuit TEXT, denominacion TEXT, periodo TEXT,
                entidad TEXT, situacion INTEGER, monto INTEGER, dias_atraso INTEGER
            )
        """)
        conn.execute("DELETE FROM _padron_staging")
        conn.commit()

        # ── Lectura streaming + bulk insert ──────────────────────────────────
        lote = []
        lineas = 0

        with open(ruta_archivo, 'r', encoding=encoding, errors='replace') as f:
            if tiene_header:
                next(f)  # saltar cabecera

            for linea in f:
                linea = linea.strip()
                if not linea:
                    continue
                partes = linea.split(delimitador)

                def _get(campo, default=''):
                    idx = idx_map.get(campo)
                    if idx is None or idx >= len(partes):
                        return default
                    return partes[idx].strip()

                cuit_raw = _get('cuit').replace('-', '').replace(' ', '')
                if len(cuit_raw) != 11 or not cuit_raw.isdigit():
                    continue  # línea inválida

                try:
                    sit = int(_get('situacion', '0') or '0')
                    monto = int(float(_get('monto', '0') or '0'))
                    dias  = int(_get('dias_atraso', '0') or '0')
                except (ValueError, TypeError):
                    sit, monto, dias = 0, 0, 0

                lote.append((
                    cuit_raw,
                    _get('denominacion')[:120],
                    _get('periodo')[:6],
                    _get('entidad')[:100],
                    sit,
                    monto,
                    dias,
                ))
                lineas += 1

                if len(lote) >= 5000:
                    conn.executemany(
                        "INSERT INTO _padron_staging VALUES (?,?,?,?,?,?,?)", lote
                    )
                    conn.commit()
                    lote.clear()
                    _padron_import_estado['lineas'] = lineas
                    _padron_import_estado['mensaje'] = f"Leyendo... {lineas:,} líneas procesadas"

        if lote:
            conn.executemany("INSERT INTO _padron_staging VALUES (?,?,?,?,?,?,?)", lote)
            conn.commit()

        _padron_import_estado['lineas'] = lineas
        _padron_import_estado['mensaje'] = "Agregando por CUIT y guardando..."

        # ── Agregación SQL: staging → tabla final ────────────────────────────
        conn.execute("DELETE FROM bcra_padron_local")
        conn.execute("""
            INSERT INTO bcra_padron_local
                (cuit, denominacion, periodo, sit_max, monto_total, num_entidades, detalle, importado_en)
            SELECT
                cuit,
                MAX(denominacion),
                MAX(periodo),
                MAX(situacion),
                SUM(monto),
                COUNT(*),
                json_group_array(
                    json_object(
                        'entidad',    entidad,
                        'situacion',  situacion,
                        'monto',      monto,
                        'diasAtraso', dias_atraso
                    )
                ),
                strftime('%Y-%m-%dT%H:%M:%S', 'now')
            FROM _padron_staging
            GROUP BY cuit
        """)
        conn.execute("DELETE FROM _padron_staging")
        conn.commit()

        total_cuits, _ = _padron_contar_registros()
        _padron_import_estado['insertados'] = total_cuits
        _padron_import_estado['mensaje']    = f"Importación completa: {total_cuits:,} CUITs cargados"

    except Exception as e:
        _padron_import_estado['error']   = str(e)
        _padron_import_estado['mensaje'] = f"Error durante importación: {e}"
        print(f"[padron] Error importación: {e}", flush=True)
    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass
        if borrar_fuente and os.path.exists(ruta_archivo):
            try:
                os.remove(ruta_archivo)
                print(f"[padron] Archivo fuente borrado: {ruta_archivo}", flush=True)
            except Exception:
                pass
        _padron_import_estado['corriendo'] = False


def _guardar_en_padron_local(cuit: str, data: dict):
    """Convierte una respuesta BCRA en vivo y la persiste en bcra_padron_local.

    Llamado automáticamente tras cada consulta exitosa para que futuras búsquedas
    del mismo CUIT respondan desde la base local (<1ms, sin red).
    """
    try:
        results    = data.get('results') or {}
        periodos   = results.get('periodos') or []
        denom      = results.get('denominacion') or ''
        # Tomar el período más reciente (viene ordenado del más nuevo al más viejo)
        entidades  = []
        periodo_id = ''
        if periodos:
            primer_p   = periodos[0] if isinstance(periodos[0], dict) else {}
            entidades  = primer_p.get('entidades') or []
            periodo_id = str(primer_p.get('periodo') or '')

        # No persistir si no hay entidades: evita cachear respuestas vacías (404 falsos
        # de clientes en Sit 1 sin mora que solo aparecen en CDI v1.0, no en legacy).
        if not entidades:
            print(f"[padron] {cuit} no guardado — respuesta sin entidades (posible 404 falso de API legacy)", flush=True)
            return

        if not periodo_id:
            # Sin período explícito: usar AAAAMM del momento actual
            periodo_id = time.strftime('%Y%m')

        sit_max    = max((int(e.get('situacion') or 0) for e in entidades), default=0)
        # Los montos del BCRA CDI vienen en miles de pesos — se guardan en miles
        # para consistencia con el motor de scoring. El frontend multiplica ×1000 al mostrar.
        monto_tot  = sum(int(e.get('monto') or 0)     for e in entidades)
        detalle_js = json.dumps([{
            'entidad':    str(e.get('entidad') or ''),
            'situacion':  int(e.get('situacion') or 0),
            'monto':      int(e.get('monto') or 0),
            'diasAtraso': int(e.get('diasAtraso') or e.get('diasAtrasoPago') or 0),
        } for e in entidades], ensure_ascii=False)

        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        conn.execute("""
            INSERT OR REPLACE INTO bcra_padron_local
                (cuit, denominacion, periodo, sit_max, monto_total, num_entidades, detalle, importado_en)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cuit, denom[:120], periodo_id, sit_max, monto_tot,
            len(entidades), detalle_js, time.strftime('%Y-%m-%dT%H:%M:%S'),
        ))
        conn.commit()
        conn.close()
        print(f"[padron] {cuit} guardado en padrón local (sit={sit_max}, monto={monto_tot})", flush=True)
    except Exception as e:
        print(f"[padron] Error al guardar {cuit}: {e}", flush=True)


def consultar_bcra_cached(cuit, skip_padron=False):
    # 1. Padrón local indexado — respuesta instantánea sin red
    # Se omite cuando skip_padron=True (fresh=1) para forzar consulta en vivo y
    # actualizar el padrón local con datos frescos del BCRA.
    if not skip_padron:
        local = consultar_padron_local(cuit)
        if local is not None:
            print(f"[bcra] {cuit} desde padrón local (offline)", flush=True)
            return local, None

    # 2. Caché de disco (24 h) — evita re-consultas recientes
    cached_data, cached_error = cache_get(cuit)
    if cached_data is not None:
        origen = "cache-error" if cached_error else "caché"
        print(f"[bcra] {cuit} desde {origen}", flush=True)
        return _norm_bcra_resp(cached_data), cached_error

    # 3. Bulk offline (historial_detalle 12m, ~25M CUITs) — <100ms, sin red.
    # Tiene prioridad sobre BCRA live en consultas normales: el 95%+ de CUITs con
    # actividad bancaria están aquí. BCRA live queda solo para CUITs genuinamente
    # nuevos en el sistema (sin historial en los últimos 12 meses).
    # Se omite con skip_padron=True (fresh=1) para forzar dato actualizado del día.
    if not skip_padron:
        _nomdeu_bulk = _nomdeu_build_deudas_resp(cuit)
        if _nomdeu_bulk:
            print(f"[bcra] {cuit} desde bulk offline (historial_detalle 12m)", flush=True)
            _guardar_en_padron_local(cuit, _nomdeu_bulk)
            return _nomdeu_bulk, None

    # 4. Consulta en vivo — solo para CUITs sin historial en el bulk,
    # o cuando el usuario forzó refresco (skip_padron=True).
    print(f"[bcra] {cuit} consultando BCRA en vivo...", flush=True)
    data, error = consultar_bcra(cuit)
    if error or not data:
        # 4.5. Último recurso: nomdeu (alcanzado solo en fresh=1 si BCRA falla)
        _nomdeu_fb = _nomdeu_build_deudas_resp(cuit)
        if _nomdeu_fb:
            print(f"[bcra] {cuit} fallback nomdeu offline (BCRA saturado — historial_detalle)", flush=True)
            return _nomdeu_fb, None
        data_cache = {
            "results": None, "sin_deudas": None,
            "error_bcra": "bcra_saturado",
            "bcra_disponible": False,
        }
        cache_set(cuit, data_cache, error)
        print(f"[bcra] {cuit} sin respuesta — saturado", flush=True)
        return data_cache, error
    data['bcra_disponible'] = True
    cache_set(cuit, data)
    # 5. Auto-guardar en padrón local para servir sin red la próxima vez
    _guardar_en_padron_local(cuit, data)
    print(f"[bcra] {cuit} OK en vivo → guardado en padrón local", flush=True)
    return data, None

verificacion_estado = {
    "corriendo": False,
    "progreso": 0,
    "total": 0,
    "cliente_actual": "",
    "mensaje": ""
}

_proceso_integral_estado: dict = {
    "corriendo": False, "total": 0, "procesados": 0,
    "errores": 0, "cliente_actual": "", "mensaje": "Listo",
    "iniciado_en": None, "log_errores": []
}
# Locks para evitar escrituras concurrentes sobre archivos compartidos
_score_cache_lock   = threading.Lock()
_alertas_file_lock  = threading.Lock()
_proceso_lock       = threading.Lock()


def _pi_safe(v, cast=None, default=None):
    """Convierte v al tipo `cast`; devuelve `default` si None, vacío o no convertible."""
    if v is None or v == "":
        return default
    if cast is None:
        return v
    try:
        return cast(v)
    except (TypeError, ValueError):
        return default


def _norm_bcra_resp(data) -> dict:
    """Normaliza cualquier variante de respuesta BCRA a un dict seguro.
    Capas normalizadas:
      1. Nivel raíz lista → toma [0]
      2. results lista    → toma [0]
      3. results no-dict  → {}
      4. periodos: items no-dict → descartados; entidades anidadas → aplanadas
    Llamado en todos los puntos de inyección del codebase."""
    # Capa 1: raíz
    if isinstance(data, list):
        data = data[0] if data else {}
    if not isinstance(data, dict):
        return {}

    # Capa 2: results
    r = data.get('results')
    if isinstance(r, list):
        r = r[0] if r else {}
        data = {**data, 'results': r}
    if r is not None and not isinstance(r, dict):
        data = {**data, 'results': {}}
        return data
    if not isinstance(r, dict):
        return data

    # Capa 3: periodos y entidades dentro de results
    periodos = r.get('periodos')
    if isinstance(periodos, list):
        limpios = []
        for p in periodos:
            if isinstance(p, list):
                # Worker devolvió periodo como lista de entidades → envolverlo
                p = {'entidades': [e for e in p if isinstance(e, dict)]}
            if not isinstance(p, dict):
                continue
            ents = p.get('entidades')
            if isinstance(ents, list):
                # Aplanar si hay entidades envueltas en lista extra [[{...}]]
                planas = []
                for e in ents:
                    if isinstance(e, list):
                        planas.extend(x for x in e if isinstance(x, dict))
                    elif isinstance(e, dict):
                        planas.append(e)
                p = {**p, 'entidades': planas}
            limpios.append(p)
        data = {**data, 'results': {**r, 'periodos': limpios}}
    return data


def _bcra_get(url: str, timeout: int = 0) -> requests.Response:
    """Transporte HTTP unificado para consultas al BCRA y scrapers públicos.

    Cadena de prioridad:
      1. Directo    — siempre el primer intento (cero latencia añadida)
      2. ScraperAPI — solo si el destino bloquea a la IP de Render, y solo en
                      dominios no-BCRA (bcra.gob.ar responde 403 a ScraperAPI,
                      así que ahí el directo es la única vía posible)

    Bright Data se retiró de la cadena: la IP de Render quedó vetada en la zona
    residencial contratada y devolvía 407 ip_forbidden en el 100% de las
    llamadas, sumando un timeout muerto a cada consulta.
    """
    _t      = timeout if timeout > 0 else 12
    _is_bcra = 'bcra.gob.ar' in url

    # ── BCRA: directo con semáforo, sin proxy posible ────────────────────────
    if _is_bcra:
        with _bcra_api_sem:
            try:
                return requests.get(url, timeout=_t, verify=False)
            except requests.exceptions.ConnectionError:
                # BCRA resetea la conexión de forma esporádica bajo carga, no
                # necesariamente porque el dato no exista. Un reintento corto
                # recupera la mayoría de estos casos sin penalizar la latencia
                # cuando el endpoint está realmente caído (ahí el 2do también falla).
                print(f"[bcra_get] conexión reseteada en {url[:60]}... — reintentando en 0.6s", flush=True)
                time.sleep(0.6)
                return requests.get(url, timeout=_t, verify=False)

    # ── No-BCRA (AFIP, Infocred, etc.): directo primero ──────────────────────
    # Timeout acotado: si el sitio bloquea a Render conviene enterarse rápido y
    # saltar a ScraperAPI, en vez de agotar los 12s completos.
    _STATUS_BLOQUEO = (401, 403, 407, 409, 429, 451, 502, 503)
    _resp_directa   = None
    try:
        _resp_directa = requests.get(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "es-AR,es;q=0.9,en;q=0.8",
            },
            timeout=min(_t, 8), verify=False,
        )
        if _resp_directa.status_code not in _STATUS_BLOQUEO:
            return _resp_directa
        print(f"[proxy] directo HTTP {_resp_directa.status_code} en {url[:60]}... "
              f"— cayendo a ScraperAPI", flush=True)
    except requests.RequestException as _e_dir:
        print(f"[proxy] directo falló ({type(_e_dir).__name__}) en {url[:60]}... "
              f"— cayendo a ScraperAPI", flush=True)

    # ── ScraperAPI: solo ante bloqueo comprobado ─────────────────────────────
    if SCRAPERAPI_KEY:
        return requests.get(
            'http://api.scraperapi.com',
            params={'api_key': SCRAPERAPI_KEY, 'url': url, 'country_code': 'ar'},
            timeout=_t,
        )

    # Sin ScraperAPI configurado: devolver la respuesta bloqueada para que el
    # caller decida (mejor que lanzar y perder el código de estado real).
    if _resp_directa is not None:
        return _resp_directa
    return requests.get(url, timeout=_t, verify=False)


def _map_detalle_bcra(raw: dict) -> dict:
    """Convierte respuesta CentralDeInformacion v1.0 al formato interno de periodos.

    Entrada (CDI v1.0):
      { results: { denominacion, detalle: [{periodo, entidad, situacion, monto, ...}] } }

    Salida (formato interno del motor de scoring):
      { results: { denominacion, periodos: [{periodo, entidades: [{entidad, situacion, monto}]}] },
        sin_deudas: bool }

    Los periodos se ordenan con el más reciente primero (coincide con la expectativa
    de periodos_curr = periodos[0] en calcular_rating_predictivo).
    """
    if not isinstance(raw, dict):
        return {}
    results = raw.get('results') or {}
    if not isinstance(results, dict):
        return {}

    denominacion = str(results.get('denominacion') or results.get('identificacion') or '')
    detalle      = results.get('detalle') or []

    if not isinstance(detalle, list) or not detalle:
        return {
            'results': {'denominacion': denominacion, 'periodos': []},
            'sin_deudas': True,
        }

    # Agrupar entidades por periodo — dict {periodo_str: [entidad_dict, ...]}
    grupos: dict = {}
    for item in detalle:
        if not isinstance(item, dict):
            continue
        periodo = str(item.get('periodo') or item.get('fechaPeriodo') or '')
        entidad = {
            'entidad':    str(item.get('entidad') or item.get('codigoEntidad') or ''),
            'situacion':  int(item.get('situacion') or 1),
            'monto':      float(item.get('monto') or item.get('deuda') or 0),
            'diasAtraso': int(item.get('diasAtraso') or 0),
        }
        grupos.setdefault(periodo, []).append(entidad)

    # Periodos ordenados descendente → más reciente primero
    periodos = [
        {'periodo': p, 'entidades': grupos[p]}
        for p in sorted(grupos.keys(), reverse=True)
    ]

    print(
        f"[bcra_cdi] mapeado: {len(periodos)} periodos, "
        f"max_sit={max((e['situacion'] for ents in grupos.values() for e in ents), default=1)}",
        flush=True,
    )
    return {
        'results': {'denominacion': denominacion, 'periodos': periodos},
        'sin_deudas': len(periodos) == 0,
    }


def _consultar_bcra_directo(cuit: str, tipo: str = 'deudas', timeout_per_req: int = 20, max_intentos: int = 2):
    """Consulta api.bcra.gob.ar con hasta 4 intentos distribuidos entre dos endpoints.

    Estrategia de waterfall:
      1. CentralDeInformacion v1.0 (nuevo oficial) — max_intentos via _bcra_get
      2. centraldedeudores v1.0 (legacy)           — max_intentos via _bcra_get (fallback)

    Detección automática de formato:
      - Respuesta con 'detalle' en results → _map_detalle_bcra (CDI v1.0)
      - Respuesta con 'periodos' en results → _norm_bcra_resp (legacy)

    tipo: 'deudas' | 'historial' | 'cheques'
    timeout_per_req: segundos por intento (default 20; usar 8 desde calcular_score_servidor)
    max_intentos: intentos por endpoint (default 2; usar 1 desde calcular_score_servidor)
    """
    _urls_cdi = {
        'deudas':    f'https://api.bcra.gob.ar/CentralDeInformacion/v1.0/Deudas/{cuit}',
        'historial': f'https://api.bcra.gob.ar/CentralDeInformacion/v1.0/Deudas/Historicas/{cuit}',
        'cheques':   f'https://api.bcra.gob.ar/CentralDeInformacion/v1.0/ChequesRechazados/{cuit}',
    }
    _urls_legacy = {
        'deudas':    f'https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/{cuit}',
        'historial': f'https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/Historicas/{cuit}',
        'cheques':   f'https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/ChequesRechazados/{cuit}',
    }

    via          = 'directo'   # _bcra_get omite ScraperAPI para bcra.gob.ar
    ultimo_error = 'sin_respuesta'
    _sleep       = 1.5

    for url, api_ver in [
        (_urls_cdi.get(tipo,    _urls_cdi['deudas']),    'cdi_v1'),
        (_urls_legacy.get(tipo, _urls_legacy['deudas']), 'legacy'),
    ]:
        for intento in range(max_intentos):
            try:
                r = _bcra_get(url, timeout=timeout_per_req)
                if r.status_code == 404:
                    if tipo == 'historial':
                        # Para historial 404 puede ser rate-limit — probar endpoint legacy antes de rendirse
                        ultimo_error = 'http_404'
                        print(f"[bcra_directo] {cuit}/{tipo} {api_ver} 404 — probando siguiente endpoint", flush=True)
                        break
                    return {'results': {'denominacion': '', 'periodos': []}, 'sin_deudas': True}, None
                if r.status_code == 200 and len(r.text.strip()) > 10:
                    raw   = r.json()
                    _res  = raw.get('results') if isinstance(raw, dict) else None
                    # Elegir mapper según formato detectado en la respuesta
                    if isinstance(_res, dict) and 'detalle' in _res:
                        data = _map_detalle_bcra(raw)
                    else:
                        data = _norm_bcra_resp(raw)
                    if not data.get('error') and data.get('results') is not None:
                        print(
                            f"[bcra_directo] {cuit}/{tipo} OK via {via}/{api_ver} intento {intento+1}",
                            flush=True,
                        )
                        return data, None
                ultimo_error = f'http_{r.status_code}'
                print(
                    f"[bcra_directo] {cuit}/{tipo} {via}/{api_ver} intento {intento+1}: {ultimo_error}",
                    flush=True,
                )
            except requests.exceptions.Timeout:
                ultimo_error = 'timeout'
                print(
                    f"[bcra_directo] {cuit}/{tipo} {via}/{api_ver} Timeout intento {intento+1}",
                    flush=True,
                )
            except Exception as e:
                ultimo_error = str(e)[:80]
                print(
                    f"[bcra_directo] {cuit}/{tipo} {via}/{api_ver} error intento {intento+1}: {e}",
                    flush=True,
                )
            if intento < 1:
                time.sleep(_sleep)
        print(
            f"[bcra_directo] {cuit}/{tipo} {api_ver} agotó intentos ({ultimo_error}) — probando siguiente",
            flush=True,
        )

    print(f"[bcra_directo] {cuit}/{tipo} todos los endpoints agotados", flush=True)
    return None, f'bcra_no_disponible:{ultimo_error}'

def gemini_request(payload, timeout=250, system_prompt=None):
    if GEMINI_KEY:
        url = "https://generativelanguage.googleapis.com/v1beta/models/" + GEMINI_MODEL + ":generateContent?key=" + GEMINI_KEY
        for intento in range(2):
            try:
                r = requests.post(url, headers={"Content-Type": "application/json"}, json=payload, timeout=timeout)
                print(f"[gemini] Intento {intento+1} status {r.status_code}", flush=True)
                if r.status_code == 200:
                    data = r.json()
                    if 'candidates' in data:
                        print("[gemini] OK", flush=True)
                        return data['candidates'][0]['content']['parts'][0]['text'], None
                    if 'error' in data:
                        msg = data['error'].get('message', 'Error')
                        print(f"[gemini] Error: {msg[:80]}", flush=True)
                        if 'demand' in msg.lower() or 'demanda' in msg.lower():
                            if intento < 1:
                                time.sleep(20)
                                continue
                        break
                else:
                    print(f"[gemini] HTTP {r.status_code}", flush=True)
                    break
            except Exception as e:
                print(f"[gemini] Excepcion: {e}", flush=True)
                if intento < 1:
                    time.sleep(10)
        print("[gemini] Fallando a OpenAI...", flush=True)

    if OPENAI_KEY:
        try:
            partes = payload.get('contents', [{}])[0].get('parts', [])
            prompt_text = ''
            for parte in partes:
                if 'text' in parte:
                    prompt_text += parte['text']
            headers_oai = {
                "Content-Type": "application/json",
                "Authorization": "Bearer " + OPENAI_KEY
            }
            _msgs_oai = []
            if system_prompt:
                _msgs_oai.append({"role": "system", "content": system_prompt})
            _msgs_oai.append({"role": "user", "content": prompt_text})
            body_oai = {
                "model": "gpt-4o-mini",
                "messages": _msgs_oai,
                "max_tokens": 2000,
                "temperature": 0.3
            }
            r2 = requests.post("https://api.openai.com/v1/chat/completions",
                headers=headers_oai, json=body_oai, timeout=60)
            print(f"[openai] Status {r2.status_code}", flush=True)
            if r2.status_code == 200:
                data2 = r2.json()
                texto = data2['choices'][0]['message']['content']
                print("[openai] OK", flush=True)
                return texto, None
            else:
                msg = f"OpenAI HTTP {r2.status_code}: {r2.text[:100]}"
                print(f"[openai] {msg}", flush=True)
                return None, msg
        except Exception as e:
            print(f"[openai] Excepcion: {e}", flush=True)
            return None, str(e)

    return None, "No hay APIs de IA disponibles."

# Workers de Cloudflare eliminados: BCRA recomienda IP controlada (Render) + caché intensivo.
# Fuente de datos crediticios: Banco Central de la República Argentina (BCRA) — api.bcra.gob.ar

# ── API de respaldo — se activa solo si todos los endpoints directos fallan ───
# Cargar en Render: RESPALDO_API_URL=https://proveedor.com/bcra  RESPALDO_API_KEY=xxx
RESPALDO_API_URL = os.environ.get('RESPALDO_API_URL', '').rstrip('/')
RESPALDO_API_KEY = os.environ.get('RESPALDO_API_KEY', '')

AWS_LAMBDA_FUNCTION = os.environ.get('AWS_LAMBDA_FUNCTION', 'vende-seguro-bcra')
AWS_REGION = os.environ.get('AWS_REGION', 'us-east-1')

def consultar_bcra_lambda(cuit):
    """Consulta deudas + historial + cheques via AWS Lambda en una sola llamada.
    Devuelve (deudas, historial, cheques) o None si falla."""
    if not BOTO3_OK:
        return None
    aws_key = os.environ.get('AWS_ACCESS_KEY_ID')
    aws_secret = os.environ.get('AWS_SECRET_ACCESS_KEY')
    if not aws_key or not aws_secret:
        return None
    try:
        client = boto3.client(
            'lambda',
            region_name=AWS_REGION,
            aws_access_key_id=aws_key,
            aws_secret_access_key=aws_secret,
            **({'config': _LAMBDA_CFG} if _LAMBDA_CFG else {})
        )
        payload = json.dumps({'cuit': cuit}).encode('utf-8')
        response = client.invoke(
            FunctionName=AWS_LAMBDA_FUNCTION,
            InvocationType='RequestResponse',
            Payload=payload
        )
        result = json.loads(response['Payload'].read())
        if response.get('FunctionError') or result.get('statusCode') != 200:
            print(f"[aws] Error Lambda {cuit}: {result}", flush=True)
            return None
        body = json.loads(result['body'])
        # Validar que tenga datos reales
        deudas = body.get('deudas')
        historial = body.get('historial')
        cheques = body.get('cheques')
        if not deudas or deudas.get('error'):
            print(f"[aws] Sin datos BCRA para {cuit}", flush=True)
            return None
        print(f"[aws] {cuit} OK via Lambda", flush=True)
        return deudas, historial, cheques
    except Exception as e:
        print(f"[aws] Fallo Lambda {cuit}: {e}", flush=True)
        return None

def _consultar_respaldo(cuit: str):
    """Failover BCRA: llama a RESPALDO_API_URL cuando todos los workers fallan.
    Mapea cualquier formato de respuesta al esquema interno BCRA que usa el Score.
    Retorna (data_dict, None) en éxito, o (None, motivo) si no hay respaldo configurado
    o la llamada falla."""
    if not RESPALDO_API_URL or not RESPALDO_API_KEY:
        return None, "respaldo_no_configurado"

    url = f"{RESPALDO_API_URL}/{cuit}"
    headers = {
        'Authorization': f'Bearer {RESPALDO_API_KEY}',
        'x-api-key': RESPALDO_API_KEY,
        'Accept': 'application/json',
    }
    try:
        r = requests.get(url, headers=headers, timeout=12, verify=False)
        print(f"[respaldo] {cuit} HTTP {r.status_code}", flush=True)
        if r.status_code == 404:
            return {"results": {"denominacion": "", "periodos": []}, "sin_deudas": True}, None
        if r.status_code != 200:
            return None, f"respaldo_http_{r.status_code}"

        raw = _norm_bcra_resp(r.json())

        # ── Caso A: el proveedor ya devuelve formato BCRA nativo ─────────────
        if raw.get('results') and isinstance(raw['results'], dict):
            periodos = raw['results'].get('periodos') or []
            raw['sin_deudas'] = len(periodos) == 0
            print(f"[respaldo] {cuit} OK (formato BCRA nativo)", flush=True)
            return raw, None

        # ── Caso B: formato alternativo → mapear a esquema interno ───────────
        # Campos comunes en Apidata, RapidAPI BCRA, etc.
        denominacion = (raw.get('denominacion') or raw.get('razon_social') or
                        raw.get('nombre') or '')
        deudas_raw   = (raw.get('deudas') or raw.get('entidades') or
                        raw.get('periodos') or [])

        # Normalizar cada deuda al formato entidad BCRA
        entidades = []
        for d in (deudas_raw if isinstance(deudas_raw, list) else []):
            sit = int(d.get('situacion') or d.get('situation') or d.get('sit') or 1)
            mon = float(d.get('monto') or d.get('amount') or d.get('deuda') or 0)
            entidades.append({
                'entidad':   str(d.get('entidad') or d.get('banco') or d.get('bank') or ''),
                'situacion': max(1, min(6, sit)),
                'monto':     mon,
            })

        data = {
            'results': {
                'denominacion': denominacion,
                'periodos': [{'entidades': entidades}] if entidades else [],
            },
            'sin_deudas': len(entidades) == 0,
        }
        print(f"[respaldo] {cuit} OK (mapeado) — {len(entidades)} entidades", flush=True)
        return data, None

    except Exception as e:
        print(f"[respaldo] Error {cuit}: {e}", flush=True)
        return None, str(e)


def consultar_bcra(cuit, reintentos=3):
    """Consulta deudas BCRA usando la IP directa de Render (canal oficial recomendado por BCRA).
    Fuente: Banco Central de la República Argentina (BCRA) — api.bcra.gob.ar"""

    def _parse_bcra(raw):
        _res = raw.get('results') if isinstance(raw, dict) else None
        d = (_map_detalle_bcra(raw)
             if isinstance(_res, dict) and 'detalle' in _res
             else _norm_bcra_resp(raw))
        if not d.get('error') and d.get('results') is not None:
            d['sin_deudas'] = len((d.get('results') or {}).get('periodos') or []) == 0
            return d
        return None

    def _fetch(url, tmt, via):
        try:
            r = _bcra_get(url, timeout=tmt)
            if r.status_code == 404:
                return 'NOT_FOUND', via
            if r.status_code == 200 and len(r.text.strip()) > 10:
                d = _parse_bcra(r.json())
                if d:
                    return d, via
        except Exception as e:
            print(f"[bcra] {cuit} {via} error: {e}", flush=True)
        return None, via

    endpoints = [
        (BCRA_WRAPPER_BASE + '/central-deudores/' + cuit, 12, 'bcra_wrapper'),
        (f"https://api.bcra.gob.ar/CentralDeInformacion/v1.0/Deudas/{cuit}", 12, 'bcra_cdi'),
        (f"https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/{cuit}",    12, 'bcra_legacy'),
    ]

    got_404 = False
    _best = [None, None]
    with ThreadPoolExecutor(max_workers=len(endpoints)) as ex:
        futs = {ex.submit(_fetch, url, tmt, via): via for url, tmt, via in endpoints}
        try:
            for fut in as_completed(futs, timeout=20):
                result, via = fut.result()
                if result == 'NOT_FOUND':
                    got_404 = True
                elif result:
                    _best[0], _best[1] = result, via
                    print(f"[bcra] {cuit} OK via {via}", flush=True)
                    break
        except Exception:
            pass

    if _best[0]:
        # Mismo criterio de seguridad que para el 404: una respuesta "exitosa" pero
        # vacía (sin_deudas=True) también puede ser un falso negativo de la API
        # legacy. Cross-check contra historial_detalle antes de confiar — si el padrón
        # offline tiene antecedentes reales, esos mandan sobre un 200 vacío.
        if _best[0].get('sin_deudas'):
            _nomdeu_vacio = _nomdeu_build_deudas_resp(cuit)
            if _nomdeu_vacio:
                _sit_off = _nomdeu_vacio['results']['periodos'][0]['entidades'][0]['situacion']
                print(f"[bcra] {cuit} respuesta en vivo vacía pero historial_detalle tiene antecedentes — bulk manda (sit_max={_sit_off})", flush=True)
                return _nomdeu_vacio, None
        return _best[0], None
    if got_404:
        # CRÍTICO (riesgo crediticio): 404 en TODOS los endpoints en vivo no es
        # prueba confiable de "sin deudas" — BCRA puede responder 404 durante un
        # bloqueo/rate-limit en vez de un 5xx. Antes de asumir Situación 1, el
        # padrón offline (historial_detalle, _HIST_DETALLE_MESES meses reales, R2)
        # tiene la última palabra: si existe con sit_max>1, ese dato manda y NUNCA
        # se subestima el riesgo de un deudor real disfrazado de "cliente nuevo".
        _nomdeu_404 = _nomdeu_build_deudas_resp(cuit)
        if _nomdeu_404:
            _sit_off = _nomdeu_404['results']['periodos'][0]['entidades'][0]['situacion']
            print(f"[bcra] {cuit} 404 en todos los endpoints en vivo — historial_detalle offline manda (sit_max={_sit_off})", flush=True)
            return _nomdeu_404, None
        return {"results": {"denominacion": "", "periodos": []}, "sin_deudas": True}, None
    data_rb, _ = _consultar_respaldo(cuit)
    if data_rb is not None:
        return data_rb, None
    print(f"[bcra] {cuit} sin respuesta en todos los endpoints", flush=True)
    return None, "sin_respuesta"

def analizar_bodegas_server(cuit, nombre, mensajes):
    if not mensajes:
        return False, ""
    try:
        mensajes_texto = "\n".join(["- " + m for m in mensajes[:20]])
        prompt = (
            "Sos un Analista de Riesgo Crediticio experto en el sector vitinicola argentino.\n"
            "Analiza estos mensajes del grupo de bodegas sobre " + nombre + " (CUIT: " + cuit + ").\n\n"
            "DICCIONARIO DE TERMINOS (OBLIGATORIO USAR):\n"
            "- LC: Limite de Credito\n- MM: Millones de pesos\n- s/ CP: Segun condiciones de pago pactadas\n"
            "- fct: Facturas\n- opera con...: Relacion comercial activa\n"
            "- contado anticipado: Paga antes de recibir mercaderia (mejor escenario)\n"
            "- pagar con +X dias: Cliente se financia con la bodega (riesgo de flujo)\n"
            "- cheque reemplazado / repuesto: Problema resuelto, NO es negativo\n\n"
            "REGLAS:\n"
            "- Priorizá el chat sobre el reporte financiero. El chat es la realidad operativa.\n"
            "- Solo negativo si hay deudas impagas NO resueltas, estafas o desaparicion.\n"
            "- Si distintas bodegas dicen cosas contradictorias, marcalo como comportamiento_inconsistente=true.\n"
            "- Cheques rechazados pero reemplazados = NO negativo.\n"
            "- Mensaje sobre OTRO CUIT diferente = NO negativo para este cliente.\n"
            "- NUNCA digas que no hay antecedentes si el chat tiene mensajes.\n\n"
            "MENSAJES:\n" + mensajes_texto + "\n\n"
            'Responde SOLO con este JSON sin markdown: {"es_negativo": false, "motivo": "texto descriptivo", "comportamiento_inconsistente": false}'
        )
        payload = {"contents": [{"parts": [{"text": prompt}]}]}
        texto, error = gemini_request(payload, timeout=45)
        if error or not texto:
            return False, ""
        import re as _re_mod
        # Limpiar delimitadores markdown y extraer primer objeto JSON
        texto_limpio = texto.strip()
        texto_limpio = texto_limpio.replace("```json", "").replace("```", "").strip()
        _match = _re_mod.search(r'\{[\s\S]+\}', texto_limpio)
        if _match:
            texto_limpio = _match.group(0)
        try:
            resultado = json.loads(texto_limpio)
        except json.JSONDecodeError as _je:
            print(f"[bodegas] JSONDecodeError ({_je}) — raw: {texto_limpio[:200]}", flush=True)
            return False, ""
        motivo = resultado.get("motivo", "")
        if resultado.get("comportamiento_inconsistente"):
            motivo = "⚠ Comportamiento Inconsistente: " + motivo
        return resultado.get("es_negativo", False), motivo
    except Exception:
        return False, ""

def _clean_text(s):
    """Limpia texto: decodifica HTML entities, normaliza unicode, elimina chars raros."""
    import unicodedata, html as _html_mod, re
    if not s:
        return ''
    s = _html_mod.unescape(str(s))
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[^\w\s\.,\-\/\(\)]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _inferir_ingresos_afip(cat, tipo, actividad, es_empleador):
    """Retorna (ingresos_estimados: int, fuente: str) basado en categoría AFIP 2026."""
    cat_u  = (cat      or '').strip().upper()
    tipo_u = (tipo     or '').strip().upper()
    act_l  = (actividad or '').lower()

    if cat_u in _MONOTRIB_INGRESOS:
        return _MONOTRIB_INGRESOS[cat_u], 'monotrib_cap_2026'

    if 'JURIDICA' in tipo_u or 'EMPLEADOR' in tipo_u or es_empleador:
        for kw, ing in _ACTIVIDAD_INGRESOS_RI:
            if kw in act_l:
                return ing, 'ri_empleador_rubro'
        return 150_000_000, 'ri_empleador_default'

    if any(t in tipo_u for t in ('INSCRIPTO', 'RI', 'IVA', 'RESPONSABLE', 'AUTONOMO')):
        for kw, ing in _ACTIVIDAD_INGRESOS_RI:
            if kw in act_l:
                return ing, 'ri_rubro'
        return 60_000_000, 'ri_default'

    return 0, 'sin_datos'


def _scrape_tangofactura_full(cuit, ua):
    """TangoFactura GetContribuyenteFull — extrae TODOS los campos disponibles."""
    import datetime, re
    try:
        r = requests.get(
            f"https://afip.tangofactura.com/Rest/GetContribuyenteFull?cuitContribuyente={cuit}",
            headers={'User-Agent': ua, 'Accept': 'application/json'},
            timeout=12, verify=False)
        if r.status_code != 200:
            return None
        contrib = (_norm_bcra_resp(r.json()).get('Contribuyente') or {})
        if not contrib:
            return None

        cat          = _clean_text(contrib.get('categMonotrib') or '').upper()
        tipo         = _clean_text(contrib.get('tipoPersona')   or '').upper()
        es_empleador = bool(contrib.get('empleador'))

        # Actividad principal — el de menor orden es la principal
        actividad = ''
        acts = contrib.get('actividades') or []
        if acts:
            acts_sorted = sorted(acts, key=lambda a: int(a.get('orden') or 999))
            actividad   = _clean_text(acts_sorted[0].get('descripcion', ''))

        # Antigüedad: fecha más temprana entre impuestos, regímenes y contrato social
        fechas = []
        if contrib.get('fechaContratoSocial'):
            fechas.append(str(contrib['fechaContratoSocial'])[:10])
        for imp in (contrib.get('impuestos') or []):
            if imp.get('desde'):
                fechas.append(str(imp['desde'])[:10])
        for reg in (contrib.get('regimenes') or []):
            if reg.get('desde'):
                fechas.append(str(reg['desde'])[:10])
        fechas = [f for f in fechas if re.match(r'^\d{4}', f)]
        fecha_inicio    = min(fechas) if fechas else ''
        antiguedad_anos = 0
        if fecha_inicio:
            try:
                antiguedad_anos = max(0, datetime.datetime.now().year - int(fecha_inicio[:4]))
            except: pass

        ingresos, fuente_ing = _inferir_ingresos_afip(cat, tipo, actividad, es_empleador)
        return {
            'tipo_persona':        tipo,
            'categoria_monotrib':  cat,
            'actividad_principal': actividad,
            'es_empleador':        es_empleador,
            'antiguedad_anos':     antiguedad_anos,
            'fecha_inicio':        fecha_inicio,
            'ingresos_anuales':    ingresos,
            'fuente_ingresos':     fuente_ing,
            'fuente':              'tangofactura',
        }
    except Exception as e:
        print(f"[solvency] TangoFactura error {cuit}: {e}", flush=True)
        return None


def _scrape_afip_html(cuit, ua):
    """Scraper HTML de endpoints públicos AFIP/ARCA como fallback a TangoFactura.
    Con SCRAPERAPI_KEY: usa proxy rotativo (evita bloqueos por IP en Render).
    Sin SCRAPERAPI_KEY: directo con headers UA personalizados."""
    import re
    _headers_direct = {
        'User-Agent': ua,
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
        'Accept-Language': 'es-AR,es;q=0.9',
        'Referer': 'https://www.afip.gob.ar/',
    }
    endpoints = [
        f"https://serviciosweb.afip.gov.ar/publico/empresa/data.aspx?cuit={cuit}",
        f"https://www.afip.gob.ar/cadena-de-valor/consulta-cadena-valor/?type=1&cuit={cuit}",
    ]
    for url in endpoints:
        try:
            if SCRAPERAPI_KEY:
                r = _bcra_get(url, timeout=12)  # ScraperAPI gestiona UA y bloqueos
            else:
                r = requests.get(url, headers=_headers_direct, timeout=10, verify=False, allow_redirects=True)
            if r.status_code != 200 or len(r.text) < 100:
                continue
            html = r.text
            cat_m  = re.search(r'(?i)Categor[ií]a\s*[^:]*:\s*([A-K])\b', html)
            tipo_m = re.search(
                r'(?i)(Monotribut\w*|Responsable\s+Inscripto|Aut[oó]nomo|'
                r'Empleador|Persona\s+Jur[ií]dica)', html)
            act_m  = re.search(
                r'(?i)Actividad\s+[Pp]rincipal[^:]*:[^\n]*\n([^\n<]{10,80})', html)
            cat       = _clean_text(cat_m.group(1)).upper()  if cat_m  else ''
            tipo      = _clean_text(tipo_m.group(1)).upper() if tipo_m else ''
            actividad = _clean_text(act_m.group(1))          if act_m  else ''
            if cat or tipo:
                ingresos, fuente_ing = _inferir_ingresos_afip(
                    cat, tipo, actividad, 'EMPLEADOR' in tipo)
                return {
                    'tipo_persona':        tipo,
                    'categoria_monotrib':  cat,
                    'actividad_principal': actividad,
                    'es_empleador':        'EMPLEADOR' in tipo,
                    'antiguedad_anos':     0,
                    'ingresos_anuales':    ingresos,
                    'fuente_ingresos':     fuente_ing,
                    'fuente':              'afip_html',
                }
        except Exception as e:
            print(f"[solvency] AFIP HTML {url}: {e}", flush=True)
    return None


# Scraper de ANSES (certificación negativa) retirado del proyecto.
# Motivos: (a) apuntaba a endpoints internos que hoy están detrás del código de
# verificación del trámite público, por lo que no devolvía señal;
# (b) la certificación negativa expone datos previsionales de personas físicas
# (beneficios sociales, jubilación) cuyo tratamiento sin consentimiento del
# titular no encuadra en los datos patrimoniales de solvencia del art. 26 de la
# Ley 25.326. La sustancia económica se evalúa ahora con el Padrón A13 de ARCA,
# que es canal oficial y trata datos fiscales del contribuyente consultado.


_BORA_CACHE: dict = {}   # {cuit_11d: razon_social | None}  — in-process, reset en restart


def _scrape_bora_razon_social(cuit: str) -> str:
    """
    Recupera Razón Social desde Boletín Oficial de la República Argentina (BORA),
    Sección II — Sociedades Civiles y Comerciales.

    Única finalidad: rescatar el nombre cuando AFIP/TangoFactura devuelven None.
    No extrae capital social (dato históricamente obsoleto por inflación).

    Solo aplica a personas jurídicas (prefijo CUIT 30/33/34).
    Retorna str con la razón social normalizada, o None si no encuentra publicación.
    """
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    if len(cuit_limpio) != 11 or cuit_limpio[:2] not in ('30', '33', '34'):
        return None   # solo jurídicas — personas físicas no publican en Sección II
    if cuit_limpio in _BORA_CACHE:
        return _BORA_CACHE[cuit_limpio]

    # BORA indexa con guiones (formato estándar de los edictos)
    cuit_fmt = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10]}"
    import re as _re_bora
    result = None
    try:
        _hdrs = {
            'User-Agent': 'Mozilla/5.0 (compatible; VendeSeguro/1.0; +https://vendeseguro.ar)',
            'Accept': 'application/json, text/html, */*',
            'Referer': 'https://www.boletinoficial.gob.ar/',
        }
        # Intentamos el endpoint de búsqueda rápida (GET) de la Sección II
        _url_get = (
            'https://www.boletinoficial.gob.ar/norma/busquedaRapida'
            f'?textoBusqueda={cuit_fmt}&tipoBusqueda=0&seccion=2'
        )
        r = requests.get(_url_get, headers=_hdrs, timeout=12, verify=True)
        if r.status_code == 200:
            _body = r.text
            # Los edictos de Sección II usan patrones consistentes para la razón social
            _PATS = [
                r'[Rr]az[oó]n\s+[Ss]ocial[:\s]+([A-ZÁÉÍÓÚÜÑ][^\n<]{4,80}?)[\.<\n]',
                r'[Dd]enominaci[oó]n\s+[Ss]ocial[:\s]+([A-ZÁÉÍÓÚÜÑ][^\n<]{4,80}?)[\.<\n]',
                r'[Dd]enominaci[oó]n[:\s]+([A-ZÁÉÍÓÚÜÑ][^\n<]{4,80}?)[\.<\n]',
                r'"denominacion"\s*:\s*"([^"]{4,120})"',
                r'"razonSocial"\s*:\s*"([^"]{4,120})"',
                r'"nombre"\s*:\s*"([^"]{4,120})"',
            ]
            for pat in _PATS:
                m = _re_bora.search(pat, _body)
                if m:
                    candidate = m.group(1).strip().rstrip('.,;')
                    # Filtrar resultados que son solo números o muy cortos
                    if len(candidate) >= 4 and not candidate[:4].isdigit():
                        result = candidate
                        break

        # Si el GET no devolvió resultado, intentar POST (formato API avanzada)
        if result is None:
            _url_post = 'https://www.boletinoficial.gob.ar/norma/busquedaAvanzadaResultado'
            _payload  = {
                'textoBusqueda': cuit_fmt,
                'tipoBusqueda': 0,
                'seccion': 2,
                'pagina': 1,
            }
            r2 = requests.post(_url_post, json=_payload, headers=_hdrs, timeout=12, verify=True)
            if r2.status_code == 200:
                try:
                    _json = r2.json()
                    # Iterar resultados buscando el campo de denominación
                    for _item in (_json.get('normas') or _json.get('results') or []):
                        for _fld in ('denominacion', 'razonSocial', 'titulo', 'nombre'):
                            val = (_item.get(_fld) or '').strip()
                            if val and len(val) >= 4 and not val[:4].isdigit():
                                result = val
                                break
                        if result:
                            break
                except Exception:
                    pass

    except Exception as e:
        print(f"[bora] {cuit_limpio} error: {e}", flush=True)

    _BORA_CACHE[cuit_limpio] = result
    if result:
        print(f"[bora] {cuit_limpio} → '{result}'", flush=True)
    else:
        print(f"[bora] {cuit_limpio} sin coincidencia en Sección II", flush=True)
    return result


def _inferir_desde_bcra(cuit):
    """Fallback definitivo: infiere ingresos desde crédito bancario activo en BCRA.
    Fundamento: el banco validó capacidad de pago antes de otorgar el crédito.
    Ratio conservador: deuda activa ≈ 25% del ingreso anual (factor 4x)."""
    try:
        bcra_path = os.path.join(DATA_DIR, 'bcra_cache.json')
        if not os.path.exists(bcra_path):
            return None
        with open(bcra_path, 'r') as f:
            cache = json.load(f)
        entry = cache.get(cuit) or cache.get(
            cuit.replace('-', '').replace(' ', '').strip())
        if not entry:
            return None
        periodos = (
            (entry.get('data') or {}).get('results') or {}
        ).get('periodos') or []
        if not periodos:
            return None
        monto_total = sum(
            (e.get('monto') or 0) for e in periodos[0].get('entidades', []))
        if monto_total <= 0:
            return None
        return {
            'tipo_persona':        'DESCONOCIDO',
            'categoria_monotrib':  '',
            'actividad_principal': '',
            'es_empleador':        False,
            'antiguedad_anos':     0,
            'ingresos_anuales':    round(monto_total * 1_000 * 3),
            'fuente_ingresos':     'bcra_inferido',
            'fuente':              'bcra_fallback',
        }
    except Exception as e:
        print(f"[solvency] BCRA inference {cuit}: {e}", flush=True)
    return None


def get_solvency_data(cuit):
    """
    Solvencia multi-fuente con cadena de fallback activo. Caché 24h.
      0. ARCA oficial (WSAA + Padrón A13, con A5 de respaldo) — canal autorizado
      1. API configurada (env var)
      2. TangoFactura AFIP JSON — extrae cat, actividad, empleador, antigüedad
      3. AFIP HTML scraper — endpoints públicos con UA rotativo
      4. Inferencia desde deuda BCRA — si el banco prestó $X, el cliente tiene ingresos
    """
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    cache_path  = os.path.join(DATA_DIR, f'solvency_{cuit_limpio}.json')
    try:
        if os.path.exists(cache_path):
            with open(cache_path, 'r') as f:
                cached = json.load(f)
            if time.time() - cached.get('ts', 0) < 86400:
                cached_data = cached.get('data') or {}
                # Normalizar: caché antiguo pudo guardar lista en lugar de dict
                if isinstance(cached_data, list):
                    cached_data = cached_data[0] if cached_data else {}
                if not isinstance(cached_data, dict):
                    cached_data = {}
                # Si el caché no tiene tipo_persona, inferirlo desde prefijo CUIT
                if not cached_data.get('tipo_persona'):
                    _pref = cuit_limpio[:2]
                    if _pref in ('20', '23', '24', '27'):
                        cached_data['tipo_persona'] = 'FISICA'
                    elif _pref in ('30', '33', '34'):
                        cached_data['tipo_persona'] = 'JURIDICA'
                # Sanity: re-inferir si el caché guardó ingresos=0 con tipo/cat válidos
                if not cached_data.get('ingresos_anuales') and (
                    cached_data.get('categoria_monotrib') or cached_data.get('tipo_persona')
                ):
                    _ing, _fi = _inferir_ingresos_afip(
                        cached_data.get('categoria_monotrib', ''),
                        cached_data.get('tipo_persona', ''),
                        cached_data.get('actividad_principal', ''),
                        cached_data.get('es_empleador', False),
                    )
                    if _ing > 0:
                        cached_data['ingresos_anuales'] = _ing
                        cached_data['fuente_ingresos']  = _fi
                        try:
                            with open(cache_path, 'w') as _fw:
                                json.dump({'data': cached_data,
                                           'ts': cached.get('ts', time.time())},
                                          _fw, ensure_ascii=False)
                        except: pass
                return cached_data or None
    except: pass

    ua   = random.choice(_USER_AGENTS)
    data = None

    # ── Fuente 0: ARCA oficial (WSAA + padrón A5) — canal autorizado ───────
    # Prioridad máxima: dato oficial sin riesgo de bloqueo. Si falla por
    # cualquier motivo retorna None y la cadena sigue con las fuentes fallback.
    if ARCA_DISPONIBLE:
        arca_data = arca_ws.obtener_datos_fiscales_arca(cuit_limpio)
        if arca_data:
            if not arca_data.get('ingresos_anuales'):
                ing, fi = _inferir_ingresos_afip(
                    arca_data.get('categoria_monotrib', ''), arca_data.get('tipo_persona', ''),
                    arca_data.get('actividad_principal', ''), arca_data.get('es_empleador', False))
                arca_data['ingresos_anuales'] = ing
                arca_data['fuente_ingresos']  = fi
            data = arca_data
            print(
                f"[solvency] {cuit_limpio} ARCA oficial "
                f"cat={data.get('categoria_monotrib')} empl={data.get('es_empleador')} "
                f"clae={data.get('clae_actividad')} ant={data.get('antiguedad_anos')}a",
                flush=True)

    # ── Fuente 1: API configurada ──────────────────────────────────────────
    if data is None and CUIT_API_URL and CUIT_API_KEY:
        try:
            r = requests.get(
                f"{CUIT_API_URL.rstrip('/')}/{cuit_limpio}",
                headers={'Authorization': f'Bearer {CUIT_API_KEY}',
                         'x-api-key': CUIT_API_KEY, 'User-Agent': ua},
                timeout=8, verify=False)
            if r.status_code == 200:
                raw = _norm_bcra_resp(r.json())
                if not raw.get('ingresos_anuales'):
                    ing, fi = _inferir_ingresos_afip(
                        raw.get('categoria_monotrib', ''), raw.get('tipo_persona', ''),
                        raw.get('actividad_principal', ''), raw.get('es_empleador', False))
                    raw['ingresos_anuales'] = ing
                    raw['fuente_ingresos']  = fi
                data = raw
                print(f"[solvency] {cuit_limpio} OK via API configurada", flush=True)
        except Exception as e:
            print(f"[solvency] API externa: {e}", flush=True)

    # ── Fuente 2: TangoFactura AFIP (JSON completo) ────────────────────────
    if data is None:
        data = _scrape_tangofactura_full(cuit_limpio, ua)
        if data:
            print(
                f"[solvency] {cuit_limpio} TangoFactura "
                f"cat={data.get('categoria_monotrib')} empl={data.get('es_empleador')} "
                f"act='{data.get('actividad_principal','')[:30]}' "
                f"ant={data.get('antiguedad_anos')}a ing≈{data.get('ingresos_anuales')}",
                flush=True)

    # ── Fuente 2.5: Padrón MiPyME — categoría, sector, empleados, tope ──────
    if data is not None and _mipyme_conn is not None:
        _mp = _mipyme_get(cuit_limpio)
        if _mp:
            cat_mp = _mp.get('categoria')
            sec_mp = _mp.get('sector')
            data['categoria_mipyme'] = cat_mp
            data['sector_mipyme']    = sec_mp
            data['empleados_rango']  = _mp.get('empleados_rango')
            _tope = _mipyme_tope(sec_mp or '', cat_mp or '')
            data['tope_mipyme'] = _tope if _tope else None
            # Cap conservador: el tope regulatorio es el máximo posible de facturación
            if _tope and data.get('ingresos_anuales') and data['ingresos_anuales'] > _tope:
                data['ingresos_anuales']  = _tope
                data['ingresos_capeados'] = True
            print(
                f"[solvency] {cuit_limpio} MiPyME cat={cat_mp} sec={sec_mp} "
                f"empl={data.get('empleados_rango')} tope={_tope}",
                flush=True)

    # ── Fuente 3: AFIP HTML scraper ────────────────────────────────────────
    if data is None:
        data = _scrape_afip_html(cuit_limpio, ua)
        if data:
            print(f"[solvency] {cuit_limpio} AFIP HTML "
                  f"cat={data.get('categoria_monotrib')} tipo={data.get('tipo_persona')}",
                  flush=True)

    # ── Fuente 4: Inferencia BCRA — fallback completo y piso obligatorio ─────
    if data is None or not (data or {}).get('ingresos_anuales'):
        _bcra_inf = _inferir_desde_bcra(cuit_limpio)
        if _bcra_inf:
            if data is None:
                data = _bcra_inf
                print(f"[solvency] {cuit_limpio} BCRA fallback completo "
                      f"ing≈{data.get('ingresos_anuales')}", flush=True)
            else:
                data['ingresos_anuales'] = _bcra_inf.get('ingresos_anuales', 0)
                data['fuente_ingresos']  = 'bcra_piso'
                print(f"[solvency] {cuit_limpio} BCRA piso aplicado "
                      f"ing≈{data['ingresos_anuales']}", flush=True)

    # ── Post-proceso: campos enriquecidos ─────────────────────────────────────
    if data is not None:
        # tipo_persona desde prefijo CUIT si TangoFactura no lo aportó.
        # CUIT 20/23/24/27 = Persona Física; 30/33/34 = Persona Jurídica.
        if not data.get('tipo_persona'):
            _pref = cuit_limpio[:2]
            if _pref in ('20', '23', '24', '27'):
                data['tipo_persona'] = 'FISICA'
            elif _pref in ('30', '33', '34'):
                data['tipo_persona'] = 'JURIDICA'

        # antiguedad_fiscal: años desde inicio de actividades (0 si no disponible)
        data.setdefault('antiguedad_fiscal', data.get('antiguedad_anos') or 0)

        # estado_empleo: clasificación laboral basada en fuentes ya consultadas
        if not data.get('estado_empleo'):
            _es_emp  = data.get('es_empleador')
            _tipo    = (data.get('tipo_persona') or '').upper()
            _cat     = data.get('categoria_monotrib') or ''
            if _es_emp or any(k in _tipo for k in ('JURIDICA', 'S.A.', 'S.R.L.', 'S.A.S')):
                data['estado_empleo'] = 'activo'
            elif _cat:
                data['estado_empleo'] = 'monotrib'
            else:
                data['estado_empleo'] = None

        # juicios_comerciales: desde API de respaldo si viene, default 0
        data.setdefault('juicios_comerciales', data.get('juicios') or 0)

        # ── Fuente 6: BORA Sección II — Razón Social como último recurso ─────
        # Solo corre si ninguna fuente anterior pudo aportar el nombre.
        # Aplica únicamente a personas jurídicas (CUIT 30/33/34).
        if not (data.get('razon_social') or data.get('nombre') or '').strip():
            _bora_rs = _scrape_bora_razon_social(cuit_limpio)
            if _bora_rs:
                data['razon_social'] = _bora_rs
                data['fuente_nombre'] = 'bora_seccion2'

    try:
        with open(cache_path, 'w') as f:
            json.dump({'data': data, 'ts': time.time()}, f, ensure_ascii=False)
    except: pass
    return data


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  MODELO NACIONAL DE RIESGO VENDE SEGURO v10.0  (Anti-Videla)            ║
# ║  Capa A (40%): Estabilidad Bancaria BCRA  | Capa B (40%): Conducta Odoo ║
# ║  Capa C (20%): Comunidad Chat             | Liquidez: bonus cheques      ║
# ║  Prospectos: BCRA+AFIP(80%) / Comunidad(20%) — sin historial Odoo        ║
# ╚══════════════════════════════════════════════════════════════════════════╝

_SCORE_VERSION          = "21.0"   # pipeline unificado BCRA(24m) + ARCA oficial en paralelo
_MOTOR_VERSION_CARTERA  = "v18.1"   # bump aquí cada vez que cambie la lógica del motor

# Keywords para NLP de chat de bodegas (Capa C — Comunidad)
_KW_NEG = [
    "atraso", "atrasado", "atrasada", "moroso", "morosa",
    "cheque rechazado", "cheque rebotado", "reboto", "rebotó",
    "devolvio", "devolvió", "deuda", "impago", "no paga",
    "mal pagador", "incobrable", "incobrable", "pésimo", "pesimo",
    "problema", "cuidado", "precaucion", "precaución",
]
_KW_POS = [
    "contado", "paga bien", "cumple", "buen cliente", "puntual",
    "sin problemas", "recomendado", "excelente", "confiable",
    "siempre paga", "buen pagador", "cliente confiable",
]

# Session-level cache — se limpia al inicio de cada verificación
_score_session_cache: dict = {}

# Mapa de riesgo logístico por zona geográfica
_GEO_RIESGO_LOGISTICO: dict = {
    'salta': 'medio', 'jujuy': 'medio', 'tucuman': 'medio', 'catamarca': 'medio',
    'la rioja': 'medio', 'santiago del estero': 'medio', 'formosa': 'medio',
    'chaco': 'medio', 'misiones': 'medio', 'corrientes': 'medio',
    'mendoza': 'medio', 'san juan': 'medio', 'san luis': 'medio',
    'neuquen': 'alto', 'neuquén': 'alto', 'rio negro': 'alto', 'río negro': 'alto',
    'chubut': 'alto', 'santa cruz': 'alto', 'tierra del fuego': 'alto',
    'la pampa': 'medio',
}

def _alerta_logistica(ciudad: str) -> str:
    """Devuelve '' | 'medio' | 'alto' según riesgo logístico de incobrabilidad."""
    c = str(ciudad or '').lower()
    for zona, nivel in _GEO_RIESGO_LOGISTICO.items():
        if zona in c:
            return nivel
    return ''


def _layer1_estabilidad_bancaria(
    max_sit: int,
    n_periodos_h: int,
    monto_real: float,
    periodos_hist: list,
    periodos_curr: list,
    sit_pond: float = None,
    mora_administrativa: bool = False,
) -> tuple:
    """
    Capa A: Estabilidad Bancaria — 40% del score (0-400 pts).
    Situación BCRA (0-280) + Tendencia 24m (0-120).
    sit_pond: situación ponderada por monto; None → usa max_sit binario.
    mora_administrativa: True → scoring como sit_efectivo=round(sp), sin
      double-penalty, tendencia forzada a neutral (criterio humano).
    Returns (pts: int, tendencia: str, alerta_deuda_creciente: bool)
    """
    sp = sit_pond if sit_pond is not None else float(max_sit)

    # Criterio Humano: para mora administrativa, el tier se calcula sobre
    # la situación ponderada redondeada — no sobre el peor outlier.
    sp_tier = float(round(sp)) if mora_administrativa else sp

    if sp_tier < 1.15:
        # Cartera virtualmente íntegra en Sit.1 → tratar como Sit.1 pura
        if   monto_real == 0:         pts_sit = 160
        elif monto_real < 500_000:    pts_sit = 200
        elif monto_real < 2_500_000:  pts_sit = 240
        elif n_periodos_h >= 12:      pts_sit = 280
        elif n_periodos_h >= 6:       pts_sit = 260
        elif n_periodos_h >= 2:       pts_sit = 240
        else:                         pts_sit = 210
    elif sp_tier < 1.5:  pts_sit = 180
    elif sp_tier < 2.0:  pts_sit = 130
    elif sp_tier < 2.1:  pts_sit = 100   # ≈ Sit.2 puro
    elif sp_tier < 2.5:  pts_sit = 70
    elif sp_tier < 3.0:  pts_sit = 40
    elif sp_tier < 4.0:  pts_sit = 15
    else:                pts_sit = 5

    pts_tend = 55
    tendencia = 'neutral'

    if mora_administrativa:
        # La "deterioración" viene de la entidad administrativa, no del cliente.
        # Forzar tendencia neutral para no destruir el score por un outlier.
        pts_tend = 65
        tendencia = 'estable'
    else:
        pool = _safe_periodos(periodos_hist if periodos_hist else periodos_curr)
        if pool and len(pool) >= 4:
            def _ms(p):
                try:
                    if not isinstance(p, dict): return 1.0
                    ents   = [e for e in (p.get('entidades') or []) if isinstance(e, dict)]
                    montos = [float(e.get('monto') or 0) for e in ents]
                    sits   = [float(e.get('situacion') or 1) for e in ents]
                    total  = sum(montos)
                    if total > 0:
                        return sum(s * m for s, m in zip(sits, montos)) / total
                    return max(sits, default=1.0)
                except Exception:
                    return 1.0
            r3 = [_ms(p) for p in pool[:3]]
            a9 = [_ms(p) for p in pool[3:12]]
            pr = sum(r3) / len(r3)
            pa = sum(a9) / len(a9) if a9 else pr
            if   pr < pa - 0.3:  pts_tend = 120; tendencia = 'mejorando'
            elif pr <= pa + 0.1: pts_tend = 65;  tendencia = 'estable'
            elif pr <= pa + 0.5: pts_tend = 10;  tendencia = 'deteriorando_leve'
            else:                pts_tend = 0;   tendencia = 'deteriorando'

        if tendencia in ('deteriorando_leve', 'deteriorando') and sp <= 2.1:
            pts_sit = pts_sit // 2   # penalización doble solo para deterioro real

    alerta_creciente = False
    pool_a = _safe_periodos(periodos_curr if periodos_curr else periodos_hist)
    if pool_a and len(pool_a) >= 2:
        def _m(p):
            try:
                if not isinstance(p, dict): return 0
                ents = [e for e in (p.get('entidades') or []) if isinstance(e, dict)]
                return sum((e.get('monto') or 0) for e in ents)
            except Exception:
                return 0
        m0, m1 = _m(pool_a[0]), _m(pool_a[1])
        if m1 > 0 and m0 / m1 > 1.30:
            alerta_creciente = True

    return (min(400, pts_sit + pts_tend), tendencia, alerta_creciente)


def _layer2_solvencia_federal(solvency_data: dict) -> tuple:
    """
    Capa 2: Solvencia Federal — 30% del score (0-300 pts).
    AFIP: tipo persona + categoría Monotributo.
    Empleador/Jurídica → 300 pts. Monotrib A/B → flag cap externo 600.
    Returns (pts: int, es_empleador: bool, es_monotrib_bajo: bool, indice: float)
    """
    if not solvency_data or not isinstance(solvency_data, dict):
        return 120, False, False, 0.40   # graceful degradation: neutral degradado

    cat  = (solvency_data.get('categoria_monotrib') or '').strip().upper()
    tipo = (solvency_data.get('tipo_persona') or '').strip().upper()

    # Usar el campo explícito es_empleador si está disponible (TangoFactura lo provee)
    es_empleador     = bool(solvency_data.get('es_empleador')) or 'JURIDICA' in tipo or 'EMPLEADOR' in tipo
    es_monotrib_bajo = cat in ('A', 'B')

    if es_empleador:
        pts_base = 300
    elif 'INSCRIPTO' in tipo or tipo in ('RI', 'IVA'):
        pts_base = 220
    elif cat:
        CAT_PTS = {
            'K': 200, 'J': 175, 'I': 150, 'H': 130,
            'G': 110, 'F': 90,  'E': 70,  'D': 50,
            'C': 35,  'B': 20,  'A': 10,
        }
        pts_base = CAT_PTS.get(cat, 80)
    else:
        pts_base = 120

    indice = round(pts_base / 300, 3)
    return (min(300, pts_base), es_empleador, es_monotrib_bajo, indice)


def _layer3_comportamiento_interno(
    cuit_limpio: str,
    en_mora: bool,
    wsp_index: dict,
    en_cartera: bool,
) -> tuple:
    """
    Capa 3: Comportamiento Interno — 30% del score (0-300 pts).
    Mora Piattelli (Hard Block, 0-150) + Chat Bodegas (0-100) + Relación (0-50).
    hard_block=True → score final no puede superar 400 pts.
    Returns (pts: int, hard_block: bool)
    """
    if en_mora:
        pts_mora   = 0
        hard_block = True
    else:
        pts_mora   = 150
        hard_block = False

    wsp_entry = wsp_index.get(cuit_limpio, {})
    if not wsp_entry:
        pts_wsp = 50
    elif wsp_entry.get('es_negativo'):
        pts_wsp = 0
    else:
        pts_wsp = 100

    pts_rel = 50 if en_cartera else 0
    return (pts_mora + pts_wsp + pts_rel, hard_block)


# ─────────────────────────────────────────────────────────────────────────────
#  NUEVAS CAPAS v10.0
# ─────────────────────────────────────────────────────────────────────────────

def _evaluar_intencionalidad_mora(
    periodos_hist: list,
    periodos_curr: list,
) -> tuple:
    """
    Clasifica mora BCRA como Administrativa vs Default Real.
    Administrativa: la entidad NUNCA reportó Sit.1 previo → −15%, sin Hard Block.
    Default Real: ≥3 meses en Sit.1 y luego cayó → Hard Block D2 ($0).
    Returns: (tipo, pct_adm, aviso)
    """
    _pc0 = periodos_curr[0] if periodos_curr and isinstance(periodos_curr[0], dict) else {}
    curr_ents = [e for e in (_pc0.get('entidades') or []) if isinstance(e, dict)]
    ents_mora = [
        (str(e.get('entidad') or '').strip().upper(),
         int(e.get('situacion') or 1),
         float(e.get('monto') or 0))
        for e in curr_ents if (e.get('situacion') or 1) >= 2
    ]
    if not ents_mora:
        return ('limpio', 0.0, '')

    # Construir historial de (situacion, monto) por entidad (todos los períodos)
    hist_sit: dict = {}
    todos = _safe_periodos(periodos_hist if periodos_hist else periodos_curr)
    for p in todos:
        if not isinstance(p, dict): continue
        for e in [e for e in (p.get('entidades') or []) if isinstance(e, dict)]:
            n = str(e.get('entidad') or '').strip().upper()
            if n:
                hist_sit.setdefault(n, []).append(
                    (int(e.get('situacion') or 1), float(e.get('monto') or 0))
                )

    monto_total_mora = sum(m for _, _, m in ents_mora)
    monto_adm = 0.0
    tipo_final = 'mora_administrativa'

    for nombre, _sit, monto in ents_mora:
        hist = hist_sit.get(nombre, [])
        # Solo cuenta meses con monto > 0 para evitar que líneas inactivas
        # (saldo cero) activen 'default_real' incorrectamente.
        meses_sit1 = sum(1 for sit, m in hist if sit == 1 and m > 0)
        print(
            f"[intencionalidad] {nombre}: sit_actual={_sit} monto={monto} "
            f"meses_sit1_con_saldo={meses_sit1} hist={hist}",
            flush=True
        )
        if meses_sit1 >= 3:
            tipo_final = 'default_real'
            print(f"[intencionalidad] → default_real por {nombre}", flush=True)
            break
        monto_adm += monto

    pct_adm = round(monto_adm / monto_total_mora, 3) if monto_total_mora > 0 else 0.0
    if tipo_final == 'mora_administrativa':
        aviso = 'Mora administrativa: banco reporta atraso sin historial previo de incumplimiento.'
    else:
        aviso = 'Default real: el cliente tenía historial limpio y cayó en mora.'
    return (tipo_final, pct_adm, aviso)


def _layer_conducta_interna(
    cuit_limpio: str,
    saldos_data: list,
    en_mora: bool,
) -> tuple:
    """
    Capa B: Conducta Interna Odoo — 40% del score (0-400 pts).
    Regularidad pago (0-200) + Volumen relación (0-100) +
    Mora interna (0-100) − Penalidad DSO (−80 si DSO↑>15% en 60d).
    Returns: (pts, dso_individual, dso_deteriorando, sin_historial,
              promedio_mensual, hard_block_mora)
    """
    _FMTS = ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y')

    def _parse(s):
        if not s:
            return None
        s = str(s)[:10]
        for fmt in _FMTS:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    # ── Filtrar facturas: CUIT directo → nombre vía _cartera_comercial ──
    facturas = [
        f for f in saldos_data
        if isinstance(f, dict) and str(f.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio
    ]
    if not facturas:
        nombre_cliente = next(
            (str(c.get('nombre', '')).strip().upper()
             for c in _cartera_comercial
             if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio),
            None
        )
        if nombre_cliente:
            # Match exacto primero; fallback: nombre_cliente como prefijo completo de palabra
            # (no substring libre — evita que "WINE BAR" matchee "VINOTECAS ROMA WINE BAR SRL")
            _nc_norm = _norm_nombre(nombre_cliente)
            facturas = [
                f for f in saldos_data
                if isinstance(f, dict) and _norm_nombre(str(f.get('cliente', ''))) == _nc_norm
            ]

    if not facturas:
        return (120, 0.0, False, True, 0.0, False, False, 0.0)

    # Fecha de corte = última fechaFactura del cliente (no la fecha del sistema).
    # Así el score usa el mes del upload, no el día de hoy.
    _fechas_cliente = [_parse(f.get('fechaFactura')) for f in facturas]
    _fechas_cliente = [d for d in _fechas_cliente if d]
    hoy = max(_fechas_cliente) if _fechas_cliente else datetime.now()

    # ── DSO individual: condición de pago vs fecha de corte ──────────────
    cutoff_rec = hoy - timedelta(days=60)
    cutoff_ant = hoy - timedelta(days=120)
    dsos_rec, dsos_ant = [], []
    for f in facturas:
        ff = _parse(f.get('fechaFactura'))
        fp = _parse(f.get('fechaPago'))
        if not ff or not fp:
            continue
        dso = (fp - ff).days
        if not (0 <= dso <= 365):
            continue
        if ff >= cutoff_rec:
            dsos_rec.append(dso)
        elif ff >= cutoff_ant:
            dsos_ant.append(dso)

    dso_individual  = round(sum(dsos_rec) / len(dsos_rec), 1) if dsos_rec else 0.0
    dso_deteriorando = False
    if dsos_rec and dsos_ant:
        dso_ant_avg = sum(dsos_ant) / len(dsos_ant)
        if dso_ant_avg > 0 and dso_individual / dso_ant_avg > 1.15:
            dso_deteriorando = True

    # ── Regularidad de pago (0-200 pts) ──────────────────────────────────
    # Solo tenemos facturas abiertas (saldo>0). Métrica correcta:
    # % de facturas cuya fecha de vencimiento (fechaPago) NO ha llegado → "al día".
    total_f = len(facturas)
    al_dia  = sum(1 for f in facturas if _parse(f.get('fechaPago')) and _parse(f.get('fechaPago')) >= hoy)
    vencidas = sum(1 for f in facturas
                   if float(f.get('saldo') or 0) > 0
                   and _parse(f.get('fechaPago'))
                   and _parse(f.get('fechaPago')) < hoy)

    ratio = al_dia / total_f if total_f > 0 else 0.0
    if   ratio >= 0.95: pts_reg = 200
    elif ratio >= 0.85: pts_reg = 160
    elif ratio >= 0.70: pts_reg = 120
    elif ratio >= 0.50: pts_reg = 80
    elif ratio >= 0.30: pts_reg = 40
    else:               pts_reg = 10

    if vencidas > 3:
        pts_reg = max(0, pts_reg - 40)

    # ── Volumen relación (0-100 pts) ──────────────────────────────────────
    vol_total = sum(float(f.get('totalFactura') or 0) for f in facturas)
    if   vol_total >= 5_000_000: pts_vol = 100
    elif vol_total >= 2_000_000: pts_vol = 80
    elif vol_total >= 500_000:   pts_vol = 60
    elif vol_total >= 100_000:   pts_vol = 40
    elif vol_total > 0:          pts_vol = 20
    else:                        pts_vol = 10

    # ── Mora interna (0-100 pts) + Hard block ────────────────────────────
    if en_mora:
        pts_mora_int    = 0
        hard_block_mora = True
    else:
        pts_mora_int    = 100
        hard_block_mora = False

    # ── Penalidad DSO (−80 si deterioró >15%) ────────────────────────────
    pen_dso = -80 if dso_deteriorando else 0

    # ── Promedio mensual de compras (para límite dinámico — Tarea 2) ─────
    fechas = [_parse(f.get('fechaFactura')) for f in facturas]
    fechas = [d for d in fechas if d]
    promedio_mensual = 0.0
    if fechas:
        meses = max(1, (hoy - min(fechas)).days / 30)
        promedio_mensual = round(vol_total / meses, 2)

    pts = max(0, min(400, pts_reg + pts_vol + pts_mora_int + pen_dso))

    # ── Deuda interna +90 días ─────────────────────────────────────────────
    deuda_90d      = False
    monto_deuda_90d = 0.0
    for f in facturas:
        saldo = float(f.get('saldo') or 0)
        if saldo > 0:
            ff = _parse(f.get('fechaFactura'))
            if ff and (hoy - ff).days > 90:
                deuda_90d       = True
                monto_deuda_90d += saldo
    if deuda_90d:
        print(f"[conducta] {cuit_limpio} deuda_90d monto={monto_deuda_90d:.0f}", flush=True)

    return (pts, dso_individual, dso_deteriorando, False, promedio_mensual, hard_block_mora,
            deuda_90d, monto_deuda_90d)


def _evaluar_comunidad(cuit_limpio: str, wsp_index: dict) -> tuple:
    """
    Capa C: Comunidad Chat Bodegas — 20% del score (0-200 pts).
    NLP sobre menciones en WhatsApp indexadas por CUIT — últimos 6 meses.
    Soporta formato array [{fecha, mensajes:[{texto}]}] y formato legacy dict.
    Returns: (pts, es_negativo, menciones_neg, menciones_pos)
    """
    if not isinstance(wsp_index, dict):
        return (100, False, 0, 0)
    raw = wsp_index.get(cuit_limpio)
    if not raw:
        return (100, False, 0, 0)

    hace_6m = datetime.now() - timedelta(days=180)

    def _parse_fecha_wsp(s):
        if not s: return None
        try:
            partes = str(s).strip().split('/')
            if len(partes) == 3:
                return datetime(int(partes[2]), int(partes[1]), int(partes[0]))
        except: pass
        return None

    if isinstance(raw, list):
        # Formato actual: [{fecha, cuit_mencionado, mensajes:[{autor, texto}]}]
        # Aplicar filtro 6 meses sobre la fecha del thread
        partes_texto = []
        for t in raw:
            if not isinstance(t, dict): continue
            fecha_t = _parse_fecha_wsp(t.get('fecha') or
                       (t.get('mensajes') or [{}])[0].get('fecha'))
            if fecha_t and fecha_t < hace_6m:
                continue  # thread fuera de ventana
            for msg in (t.get('mensajes') or []):
                if isinstance(msg, dict):
                    partes_texto.append(str(msg.get('texto', '')))
        texto = ' '.join(partes_texto).lower()
    elif isinstance(raw, dict):
        # Formato legacy: {texto: "...", mensajes: [...]}
        texto = str(raw.get('texto') or raw.get('mensajes') or '').lower()
    else:
        return (100, False, 0, 0)

    if not texto.strip():
        return (100, False, 0, 0)

    neg = sum(1 for kw in _KW_NEG if kw in texto)
    pos = sum(1 for kw in _KW_POS if kw in texto)
    es_negativo = neg > pos or (neg > 0 and pos == 0)

    if   neg == 0 and pos == 0: pts = 100
    elif neg > 0  and pos == 0: pts = max(0,   100 - neg * 25)
    elif pos > 0  and neg == 0: pts = min(200, 100 + pos * 25)
    else:
        balance = pos - neg
        pts = max(0, min(200, 100 + balance * 20))

    print(f"[wsp] {cuit_limpio} neg={neg} pos={pos} pts={pts}", flush=True)
    return (pts, es_negativo, neg, pos)


def _detectar_degradacion(score_history: list) -> tuple:
    """
    Anti-Videla: detecta caída sostenida en el historial de scores (ventana 12 meses).
    Umbral: ≥15% de caída vs promedio últimos 30d → 'degradacion_moderada'.
             ≥20% o ≥150 pts absolutas → 'degradacion_severa'.
    Returns: (tipo, delta, mensaje)
    """
    if not score_history or len(score_history) < 2:
        return ('', 0, '')

    try:
        hist = sorted(score_history, key=lambda x: x.get('fecha', '') if isinstance(x, dict) else '', reverse=True)
    except Exception:
        hist = list(reversed(score_history))

    hist = [h for h in hist if isinstance(h, dict)]
    if not hist:
        return ('', 0, '')

    score_actual = int(hist[0].get('score') or 0)
    if not score_actual:
        return ('', 0, '')

    corte_30d = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    recientes  = [h for h in hist[1:] if isinstance(h, dict) and (h.get('fecha') or '') >= corte_30d]
    comparar   = recientes[:4] if recientes else hist[1:5]
    if not comparar:
        return ('', 0, '')

    promedio   = sum(int((h.get('score') if isinstance(h, dict) else 0) or 0) for h in comparar) / len(comparar)
    delta      = round(promedio - score_actual)
    pct_caida  = delta / promedio if promedio > 0 else 0.0

    if delta >= 150 or pct_caida >= 0.20:
        return ('degradacion_severa', delta,
                f'Caída severa de {delta} pts ({round(pct_caida*100)}%) vs promedio histórico.')
    elif delta >= 80 or pct_caida >= 0.15:
        return ('degradacion_moderada', delta,
                f'Deterioro de {delta} pts ({round(pct_caida*100)}%) vs promedio últimos 30 días.')
    return ('', 0, '')


def _layer_liquidez(cheq_data: dict, max_sit: int, n_periodos_h: int) -> tuple:
    """
    Liquidez (cheques rechazados) — bonus/penalidad (0-100 pts).
    rechazar=True → score = 1 (rechazo definitivo por cheques críticos).
    Returns (pts: int, rechazar: bool)
    """
    if not cheq_data:
        return 0, False
    if cheq_data.get('sin_deudas'):
        if   n_periodos_h >= 12: pts = 100
        elif n_periodos_h >= 6:  pts = 80
        elif n_periodos_h >= 2:  pts = 60
        elif n_periodos_h >= 1:  pts = 40
        else:                    pts = 20
    else:
        res_c    = cheq_data.get('results') or {}
        causales = res_c.get('causales', []) if isinstance(res_c, dict) else []
        detalles: list = []
        for causal in causales:
            if not isinstance(causal, dict):
                continue
            for ent in (causal.get('entidades') or []):
                if not isinstance(ent, dict):
                    continue
                detalles.extend(x for x in (ent.get('detalle') or []) if isinstance(x, dict))
        total_ch   = len(detalles)
        activos_ch = sum(1 for d in detalles
                         if not d.get('fechaPago') or d.get('estadoMulta') == 'IMPAGA')
        if activos_ch > 5 or total_ch >= 113:
            return 0, True
        if   total_ch == 0:   pts = 100 if n_periodos_h > 6 else (50 if n_periodos_h >= 1 else 20)
        elif activos_ch == 0: pts = 50
        else:                 pts = 0
    if   max_sit >= 3:               pts = 0
    elif max_sit == 2 and pts > 50:  pts = 50
    return pts, False


def _safe_periodos(lst) -> list:
    """Garantiza que cada elemento de una lista de periodos sea un dict
    con clave 'entidades' que contenga solo dicts. Aplanado total."""
    if not isinstance(lst, list):
        return []
    result = []
    for p in lst:
        if isinstance(p, list):
            ents = [e for e in p if isinstance(e, dict)]
            if ents:
                result.append({'entidades': ents})
            continue
        if not isinstance(p, dict):
            continue
        ents = p.get('entidades')
        if not isinstance(ents, list):
            p = {**p, 'entidades': []}
        else:
            flat = []
            for e in ents:
                if isinstance(e, list):
                    flat.extend(x for x in e if isinstance(x, dict))
                elif isinstance(e, dict):
                    flat.append(e)
            p = {**p, 'entidades': flat}
        result.append(p)
    return result


# ── Detección de falsas deudas ──────────────────────────────────────────────
# Umbral y ventana configurables sin reescribir código.
FALSE_DEBT_CONFIG: dict = {
    'umbral_monto_ars':          200_000,  # Monto máx para considerar deuda no representativa (reducido de 500k — evita ocultar deudas reales en contexto inflacionario)
    'ventana_degradacion_meses': 6,        # Meses máx desde sit.1 → sit≥2 para validar degradación
    'min_entidades_normales':    2,        # Entidades en sit.1 requeridas para aislar la anomalía (aumentado de 1 → más evidencia de perfil limpio requerida)
}


def detectar_falsas_deudas(
    bcra_data: dict,
    historial_data: dict = None,
    cheques_data: dict   = None,
    cuit: str            = '',
    config: dict         = None,
) -> dict:
    """
    Detecta entidades con deuda no representativa del perfil real del cliente.
    Una entidad se excluye cuando se cumplen TODOS estos criterios:
      1. Situación >= 2 (irregular) en el período actual
      2. Monto <= umbral_monto_ars (deuda simbólica — default $500k)
      3. Al menos 1 otra entidad en Sit.1 en el mismo período (perfil global limpio)
      4. La entidad estuvo en Sit.1 dentro de los últimos ventana_degradacion_meses
         (degradación reciente, no mora estructural)
      5. Sin cheques rechazados activos
      6. Anomalía aislada: exactamente 1 entidad irregular en el período actual

    No modifica datos originales del BCRA; solo informa qué excluir del cómputo.

    Returns:
        {'excluidas': [str], 'razones': {str: {'sit': int, 'monto': int, 'criterio': str}}}
    """
    cfg = {**FALSE_DEBT_CONFIG, **(config or {})}
    excluidas: list = []
    razones:   dict = {}

    if not bcra_data:
        return {'excluidas': excluidas, 'razones': razones}

    periodos = _safe_periodos((bcra_data.get('results') or {}).get('periodos') or [])
    if not periodos or not isinstance(periodos[0], dict):
        return {'excluidas': excluidas, 'razones': razones}

    ents_actual = [e for e in periodos[0].get('entidades', []) if isinstance(e, dict)]
    if not ents_actual:
        return {'excluidas': excluidas, 'razones': razones}

    # Criterio 5: sin cheques impagos activos — si los hay, ninguna exclusión aplica
    if cheques_data:
        cheq_n = _norm_bcra_resp(cheques_data)
        for _ca in (cheq_n.get('results') or {}).get('causales', []):
            for _en in (_ca.get('entidades') or []):
                for _d in (_en.get('detalle') or []):
                    if isinstance(_d, dict) and not _d.get('fechaLevantamiento'):
                        return {'excluidas': excluidas, 'razones': razones}

    ents_irreg = [
        e for e in ents_actual
        if (e.get('situacion') or 1) >= 2 and (e.get('monto') or 0) > 0
    ]
    ents_norm = [e for e in ents_actual if (e.get('situacion') or 1) == 1]

    # Criterio 6: exactamente 1 entidad irregular (anomalía aislada)
    if len(ents_irreg) != 1:
        return {'excluidas': excluidas, 'razones': razones}

    # Criterio 3: al menos min_entidades_normales en Sit.1
    if len(ents_norm) < cfg['min_entidades_normales']:
        return {'excluidas': excluidas, 'razones': razones}

    ent = ents_irreg[0]
    nombre    = ent.get('entidad', '?')
    monto_k   = float(ent.get('monto') or 0)   # miles de pesos (formato BCRA)
    monto_ars = monto_k * 1000
    sit       = int(ent.get('situacion') or 2)

    # Criterio 2: monto <= umbral
    if monto_ars > cfg['umbral_monto_ars']:
        return {'excluidas': excluidas, 'razones': razones}

    # Criterio 4: la entidad estuvo en Sit.1 dentro de la ventana de degradación
    degradacion_verificada = False
    hist_periodos: list = []
    if historial_data:
        hist_n = _norm_bcra_resp(historial_data)
        hist_periodos = _safe_periodos((hist_n.get('results') or {}).get('periodos') or [])

    # Si no hay historial separado, usar períodos anteriores del bcra_data
    ventana_src = hist_periodos[:cfg['ventana_degradacion_meses']] or periodos[1:cfg['ventana_degradacion_meses']+1]
    for ph in ventana_src:
        if not isinstance(ph, dict):
            continue
        for eh in ph.get('entidades', []):
            if isinstance(eh, dict) and eh.get('entidad') == nombre and (eh.get('situacion') or 1) == 1:
                degradacion_verificada = True
                break
        if degradacion_verificada:
            break

    if not degradacion_verificada:
        return {'excluidas': excluidas, 'razones': razones}

    # Todos los criterios cumplidos — marcar como falsa deuda
    criterio = (
        f"deuda aislada ${round(monto_ars):,} ARS en Sit.{sit} con degradación reciente "
        f"dentro de {cfg['ventana_degradacion_meses']}m — "
        f"{len(ents_norm)} entidad(es) en Sit.1"
    )
    excluidas.append(nombre)
    razones[nombre] = {'sit': sit, 'monto': round(monto_ars), 'criterio': criterio}
    print(
        f"[falsas_deudas] {cuit} excluida: {nombre} | Sit.{sit} | "
        f"${round(monto_ars):,} | {criterio}",
        flush=True
    )
    return {'excluidas': excluidas, 'razones': razones}


def calcular_rating_predictivo(
    cuit: str,
    bcra_data: dict,
    hist_data: dict     = None,
    cheq_data: dict     = None,
    en_mora: bool       = None,
    solvency_data: dict = None,
    ciudad: str         = '',
) -> dict:
    """
    Modelo Nacional de Riesgo Vende Seguro v20.0 (Anti-Videla)

    Capa A — Estabilidad Bancaria   40%  (0-400 pts): BCRA + Tendencia 24m
    Capa B — Conducta Interna       40%  (0-400 pts): Odoo DSO/regularidad/volumen
    Capa C — Comunidad              20%  (0-200 pts): NLP Chat Bodegas
    Liquidez                              (0-100 pts): Cheques bonus

    Prospectos (sin historial Odoo): AFIP proxy llena Capa B (80% BCRA+AFIP / 20% Com.)
    Intencionalidad: mora administrativa (nunca Sit.1 previo) → −15%, sin Hard Block.
    Default real (≥3m en Sit.1 luego Sit.2+) → Hard Block D2 ($0).
    Anti-Videla: scoreHistory[] 12 meses; degradación ≥15% → alerta.
    """
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    # Limpieza extra: elimina cualquier carácter no numérico residual
    id_cliente  = ''.join(c for c in cuit_limpio if c.isdigit())
    print(f">>> ENTRANDO AL MOTOR - CUIT: {cuit_limpio} | id_cliente: {id_cliente}", flush=True)

    # Log mínimo de entrada al motor — sin datos financieros para proteger privacidad en logs
    print(
        f"[score] entrada cuit={cuit_limpio} "
        f"bcra_type={type(bcra_data).__name__} "
        f"has_results={isinstance((bcra_data or {}).get('results'), dict)}",
        flush=True
    )

    # Blindaje total: normaliza lista, results-lista, periodos y entidades anidadas
    bcra_data = _norm_bcra_resp(bcra_data)
    if hist_data  is not None: hist_data  = _norm_bcra_resp(hist_data)
    if cheq_data  is not None: cheq_data  = _norm_bcra_resp(cheq_data)

    # Guardia post-norm: si results sigue sin ser dict, forzar vacío
    for _d_name, _d_ref in [('bcra', bcra_data), ('hist', hist_data), ('cheq', cheq_data)]:
        if _d_ref is None:
            continue
        _r = _d_ref.get('results')
        if _r is not None and not isinstance(_r, dict):
            print(f"[NORM WARN] {cuit_limpio} {_d_name}.results={type(_r).__name__} forzado a {{}}", flush=True)
            _d_ref['results'] = {}

    if cuit_limpio in _score_session_cache:
        print(f">>> CACHE HIT - CUIT: {cuit_limpio}", flush=True)
        return _score_session_cache[cuit_limpio]

    # ── Parsear BCRA ──────────────────────────────────────────────────────
    sin_deudas_real = bcra_data.get('sin_deudas', False)
    periodos_curr   = _safe_periodos((bcra_data.get('results') or {}).get('periodos') or [])
    max_sit = 1; nro_entidades = 0; monto_total_m = 0.0
    try:
        if periodos_curr:
            ents = periodos_curr[0].get('entidades', []) if isinstance(periodos_curr[0], dict) else []
            ents = [e for e in ents if isinstance(e, dict)]
            nro_entidades = len(ents)
            if ents:
                max_sit       = max((e.get('situacion', 1) or 1) for e in ents)
                monto_total_m = sum((e.get('monto', 0) or 0) for e in ents) / 1000
        elif sin_deudas_real:
            max_sit = 1
    except Exception as _pe:
        print(f"[score parse_err] {cuit_limpio} periodos_curr: {_pe}", flush=True)
        max_sit = 1; nro_entidades = 0; monto_total_m = 0.0
    monto_real = monto_total_m * 1000

    # ── Detección de falsas deudas — recalibrar score antes de ponderar ───
    # Criterios: deuda aislada, monto<=500k, degradación reciente, sin cheques.
    # No modifica bcra_data original; filtra solo periodos_curr[0] para el cómputo.
    _fd: dict = detectar_falsas_deudas(bcra_data, hist_data, cheq_data, cuit_limpio)
    if _fd['excluidas'] and periodos_curr and isinstance(periodos_curr[0], dict):
        _ents_fd_limpias = [
            e for e in periodos_curr[0].get('entidades', [])
            if e.get('entidad') not in _fd['excluidas']
        ]
        periodos_curr[0] = dict(periodos_curr[0], entidades=_ents_fd_limpias)
        if _ents_fd_limpias:
            max_sit       = max((e.get('situacion', 1) or 1) for e in _ents_fd_limpias)
            nro_entidades = len(_ents_fd_limpias)
            monto_total_m = sum((e.get('monto', 0) or 0) for e in _ents_fd_limpias) / 1000
        else:
            max_sit = 1; nro_entidades = 0; monto_total_m = 0.0
        monto_real = monto_total_m * 1000
        print(
            f"[falsas_deudas] {cuit_limpio} recalibrado: max_sit={max_sit} "
            f"monto={monto_real:.0f}ARS | excluidas={_fd['excluidas']}",
            flush=True
        )

    # ── Ponderación de mora por monto ─────────────────────────────────────
    _ents_curr    = (periodos_curr[0].get('entidades', []) if isinstance(periodos_curr[0], dict) else []) if periodos_curr else []
    _ents_curr    = [e for e in _ents_curr if isinstance(e, dict)]
    _raw_total_m  = monto_total_m * 1000           # miles de pesos (raw BCRA)
    monto_mora_k  = 0.0                            # miles de pesos en Sit.>1
    monto_sit1_k  = 0.0                            # miles de pesos en Sit.1 (limpio)
    sit_ponderada = float(max_sit)
    if _ents_curr and _raw_total_m > 0:
        monto_mora_k = sum(
            (e.get('monto', 0) or 0) for e in _ents_curr
            if (e.get('situacion', 1) or 1) > 1
        )
        monto_sit1_k = sum(
            (e.get('monto', 0) or 0) for e in _ents_curr
            if (e.get('situacion', 1) or 1) == 1
        )
        sit_ponderada = sum(
            (e.get('situacion', 1) or 1) * (e.get('monto', 0) or 0)
            for e in _ents_curr
        ) / _raw_total_m
    pct_mora = monto_mora_k / _raw_total_m if _raw_total_m > 0 else (0.0 if max_sit == 1 else 1.0)

    print(
        f">>> DIAG {cuit_limpio}: max_sit={max_sit} sit_pond={sit_ponderada:.2f} "
        f"monto_sit1k={monto_sit1_k} monto_morak={monto_mora_k} pct_mora={pct_mora:.3f}",
        flush=True
    )

    # Mora Técnica v11.0: solo max_sit == 2 Y monto en mora < umbral configurable (MORA_TECNICA_UMBRAL_K).
    # Umbral en miles ARS, ajustable via env var según inflación anual.
    es_mora_tecnica = (max_sit == 2 and monto_mora_k < MORA_TECNICA_UMBRAL_K)
    es_mora_comercial_activa = (max_sit > 1 and not es_mora_tecnica)
    if es_mora_tecnica:
        print(
            f"[mora_tec v11] {cuit_limpio} mora={monto_mora_k:.0f}k sit={max_sit} → Mora Técnica",
            flush=True
        )
    elif es_mora_comercial_activa:
        print(
            f"[mora_com v11] {cuit_limpio} mora={monto_mora_k:.0f}k sit={max_sit} → Mora Comercial Activa",
            flush=True
        )

    # ── Historial 24m ─────────────────────────────────────────────────────
    periodos_hist = _safe_periodos((hist_data.get('results') or {}).get('periodos') or []) if hist_data else []
    if not periodos_hist:
        periodos_hist = periodos_curr
    n_periodos_h = n_periodos_recientes = meses_malos = 0
    sit_grave_6m = False
    _mm_recientes = 0
    _mm_antiguos  = 0
    for idx_p, p in enumerate(periodos_hist[:24]):
        try:
            if not isinstance(p, dict):
                continue
            ents_h = [e for e in (p.get('entidades') or []) if isinstance(e, dict)]
            smax        = max((e.get('situacion') or 1 for e in ents_h), default=1)
            tiene_deuda = any((e.get('monto') or 0) > 0 for e in ents_h)
        except Exception as _ph:
            print(f"[score parse_err] {cuit_limpio} p{idx_p} hist: {_ph}", flush=True)
            smax = 1; tiene_deuda = False
        if tiene_deuda:
            n_periodos_h += 1
            if idx_p < 6: n_periodos_recientes += 1
        if smax > 1:
            meses_malos += 1
            if idx_p < 6: _mm_recientes += 1
            else:          _mm_antiguos  += 1
        if idx_p < 6 and smax >= 3: sit_grave_6m = True
    n_periodos_h = min(n_periodos_h, n_periodos_recientes * 4)

    # ── Deterioro Estructural v11.0: transición Sit.1 → Sit.3/4/5 sostenida ─
    deterioro_estructural = False
    _periodos_rev = periodos_hist[:24]
    if len(_periodos_rev) >= 6:
        def _smax_p(p):
            try:
                if not isinstance(p, dict): return 1
                ents = [e for e in (p.get('entidades') or []) if isinstance(e, dict)]
                return max((e.get('situacion') or 1 for e in ents), default=1)
            except Exception: return 1
        _max_per_mes = [_smax_p(p) for p in _periodos_rev]
        _sit_reciente = _max_per_mes[:6]
        _sit_anterior = _max_per_mes[6:]
        _estaba_estable = (
            len(_sit_anterior) > 0 and
            sum(1 for s in _sit_anterior if s == 1) >= max(1, len(_sit_anterior) * 0.6)
        )
        _deterioro_sostenido = sum(1 for s in _sit_reciente if s >= 3) >= 3
        if _estaba_estable and _deterioro_sostenido:
            deterioro_estructural = True
            print(
                f"[deterioro v11] {cuit_limpio}: transición Sit.1→Sit.3+ sostenida "
                f"ant={_sit_anterior} rec={_sit_reciente}",
                flush=True
            )

    # ── Mora Piattelli ────────────────────────────────────────────────────
    moras_norm: dict = {}
    _mp = os.path.join(DATA_DIR, 'moras_piattelli.json')
    if not os.path.exists(_mp):
        _mp = os.path.join(os.getcwd(), 'moras_piattelli.json')
    try:
        with open(_mp, 'r', encoding='utf-8') as _mf:
            _md = json.load(_mf)
        moras_norm = {str(k).replace('-','').replace(' ','').strip(): v for k, v in _md.items()}
    except: pass
    if en_mora is None:
        en_mora = cuit_limpio in moras_norm

    # ── WhatsApp Bodegas ──────────────────────────────────────────────────
    wsp_index: dict = {}
    try:
        with open(WSP_FILE, 'r', encoding='utf-8') as _wf:
            wsp_index = json.load(_wf)
    except: pass
    if not isinstance(wsp_index, dict):
        wsp_index = {}

    # ── Solvencia AFIP (graceful degradation) ────────────────────────────
    if solvency_data is None:
        solvency_data = get_solvency_data(cuit_limpio)
    if not isinstance(solvency_data, dict):
        solvency_data = None

    # ── Intencionalidad de mora BCRA (debe ir ANTES de _layer1) ──────────
    tipo_mora_bcra, pct_mora_adm, aviso_mora = _evaluar_intencionalidad_mora(
        periodos_hist, periodos_curr
    )
    # Proporcionalidad: < 15% mora Y entidad principal en Sit.1 → no bloquear
    # "Entidad principal" = mayor exposición por monto, NO el primer elemento del
    # array — BCRA no garantiza que las entidades vengan ordenadas por monto, así
    # que _ents_curr[0] podía ser cualquier banco y dar un falso "banco sucio"
    # incluso cuando el acreedor dominante real estaba en Sit.1.
    banco_principal_limpio = False
    if _ents_curr and pct_mora < 0.15:
        _banco_ppal = max(_ents_curr, key=lambda e: float(e.get('monto') or 0))
        banco_principal_limpio = int(_banco_ppal.get('situacion') or 1) == 1

    # Criterio Humano: para mora administrativa, toda la lógica de caps usa
    # sit_efectivo = round(sit_ponderada) en lugar del max_sit del outlier.
    # La materialidad ($50k ARS o <5%) también clasifica como administrativa.
    es_mora_administrativa = (
        tipo_mora_bcra == 'mora_administrativa' or
        banco_principal_limpio or
        es_mora_tecnica   # materialidad: deuda baja no refleja insolvencia
    )
    sit_efectivo = max(1, round(sit_ponderada)) if es_mora_administrativa else max_sit
    print(
        f"[OVERRIDE DIAG] {cuit_limpio}: tipo_mora={tipo_mora_bcra} "
        f"banco_ppal_limpio={banco_principal_limpio} mora_tec={es_mora_tecnica} "
        f"→ es_mora_adm={es_mora_administrativa} sit_ef={sit_efectivo}",
        flush=True
    )

    # Default Real + Sit.2+ + no mora técnica + no proporcional → Hard Block D2
    hard_block_bcra = (
        max_sit >= 2 and
        tipo_mora_bcra == 'default_real' and
        not es_mora_tecnica and
        not banco_principal_limpio
    )

    # ═══ CAPA A: Estabilidad Bancaria (0-400 pts) ════════════════════════
    # Para mora administrativa: pts_sit usa sp_tier=round(sp), tendencia=estable,
    # sin double-penalty (criterio humano — el banco principal manda).
    pts_c1, tendencia, alerta_creciente = _layer1_estabilidad_bancaria(
        sit_efectivo, n_periodos_h, monto_real, periodos_hist, periodos_curr,
        sit_ponderada, mora_administrativa=es_mora_administrativa
    )

    # ═══ CAPA B: Conducta Interna Odoo (0-400 pts) ═══════════════════════
    pts_c2, es_empleador, es_monotrib_bajo, indice_solv = _layer2_solvencia_federal(solvency_data)

    # ── Cerebro Fiscal Deductivo (0-400) — se calcula SIEMPRE ─────────────
    # Paridad con BCRA: el perfil fiscal (antigüedad impositiva + escala
    # impositiva/MiPyME + riesgo sectorial CLAE) participa en todo score, no
    # solo cuando el BCRA viene vacío. Requiere al menos una señal fiscal real
    # para no ponderar datos inventados.
    pts_fiscal: int | None = None
    dbg_fiscal: dict       = {}
    _fiscal_es_capa_b      = False   # True si el perfil fiscal reemplaza la Capa B
                                     # (evita contarlo dos veces en el bonus)
    if SCORING_FISCAL_OK and solvency_data and (
        solvency_data.get('antiguedad_anos')
        or solvency_data.get('clae_actividad')
        or solvency_data.get('actividad_principal')
        or solvency_data.get('categoria_mipyme')
        or solvency_data.get('estado_clave')
        or solvency_data.get('n_impuestos_activos') is not None
    ):
        try:
            pts_fiscal, dbg_fiscal = scoring_fiscal.puntaje_perfil_fiscal(
                antiguedad_anos=solvency_data.get('antiguedad_anos'),
                categoria_mipyme=solvency_data.get('categoria_mipyme') or '',
                categoria_monotrib=solvency_data.get('categoria_monotrib') or '',
                clae_actividad=(solvency_data.get('clae_actividad')
                                or solvency_data.get('actividad_principal') or ''),
                es_empleador=solvency_data.get('es_empleador') is True,
                tipo_persona=solvency_data.get('tipo_persona') or '',
                # ── Padrón A13 (neutros si la fuente fue A5 o un fallback) ──
                estado_clave=solvency_data.get('estado_clave') or '',
                domicilios=solvency_data.get('domicilios'),
                n_impuestos_activos=solvency_data.get('n_impuestos_activos'),
                tiene_iva=solvency_data.get('tiene_iva') is True,
                tiene_ganancias=solvency_data.get('tiene_ganancias') is True,
                tiene_monotributo=solvency_data.get('tiene_monotributo') is True,
            )
            print(f"[cerebro-fiscal] {cuit_limpio} {dbg_fiscal.get('componentes', '')}", flush=True)
            for _alerta in dbg_fiscal.get('alertas', []):
                print(f"[cerebro-fiscal][ALERTA] {cuit_limpio} {_alerta}", flush=True)
        except Exception as _e_fisc:
            print(f"[cerebro-fiscal] {cuit_limpio} error: {_e_fisc} — se ignora el perfil fiscal", flush=True)
            pts_fiscal, dbg_fiscal = None, {}

    (pts_cb, dso_individual, dso_deteriorando,
     sin_historial_interno, promedio_mensual, hard_block_mora,
     deuda_90d_interna, monto_deuda_90d_interna) = \
        _layer_conducta_interna(cuit_limpio, _saldos_facturas, en_mora)

    if sin_historial_interno:
        # Score de Prospección: AFIP solvencia como proxy de Capa B (0-400)
        pts_cb = min(400, round(pts_c2 * 400 / 300))

        # Sin historial interno NI historial bancario estamos "a ciegas": el
        # perfil fiscal completo es mejor evidencia que el proxy de categoría
        # sola, así que pasa a ser la Capa B.
        if pts_fiscal is not None and n_periodos_h == 0 and not periodos_curr:
            pts_cb = pts_fiscal
            _fiscal_es_capa_b = True
            print(f"[cerebro-fiscal] {cuit_limpio} BCRA vacío → Capa B fiscal = {pts_fiscal}", flush=True)

    # ═══ CAPA C: Comunidad Chat Bodegas (0-200 pts) ══════════════════════
    pts_cc, comunidad_negativa, _neg_count, _pos_count = _evaluar_comunidad(
        cuit_limpio, wsp_index
    )

    # ═══ LIQUIDEZ: Cheques (0-100 pts) ═══════════════════════════════════
    # Para mora administrativa: sit_efectivo (no max_sit) para no zerear el bonus
    pts_liq, rechazar = _layer_liquidez(cheq_data or {}, sit_efectivo, n_periodos_h)
    if rechazar:
        resultado = {
            'score': 1, 'rango': 'Rechazar', 'color': '#7f1d1d', 'emoji': '⛔',
            'alerta_temprana': False, 'bloquear_oportunidad': True,
            'alerta_logistica': _alerta_logistica(ciudad),
            'componentes': {'capaA': pts_c1, 'capaB': pts_cb, 'capaC': pts_cc, 'liquidez': 0},
            'tendencia': tendencia, 'es_empleador': es_empleador,
            'indice_solvencia': indice_solv, 'version': _SCORE_VERSION,
            'max_sit': max_sit,
        }
        _score_session_cache[cuit_limpio] = resultado
        print(f"[score v{_SCORE_VERSION}] {cuit_limpio} RECHAZAR — cheques críticos", flush=True)
        return resultado

    # ── Suma bruta (A + B + C + Liquidez) ────────────────────────────────
    puntos = pts_c1 + pts_cb + pts_cc + pts_liq

    # ── Bonus estructura MiPyME — proxy empleados (padrón Min. Producción) ───
    # Fuente: categoria_mipyme enriquecida por get_solvency_data → mipyme_padron.db.
    # Refleja sustancia operativa declarada ante SEPYME/AFIP: nómina real implícita.
    # Gran empresa (jurídica sin registro MiPyME): escala corporativa, bonus fijo.
    _bonus_estructura = 0
    if solvency_data:
        _cat_mp   = (solvency_data.get('categoria_mipyme') or '').strip()
        # Requiere confirmación explícita de AFIP/TangoFactura — no asumimos por tipo_persona
        _es_empl2 = solvency_data.get('es_empleador') is True
        _BONUS_MP = {'Micro': 10, 'Pequeña': 25, 'Mediana_T1': 45, 'Mediana_T2': 45}
        if _cat_mp in _BONUS_MP:
            _bonus_estructura = _BONUS_MP[_cat_mp]
        elif _es_empl2:
            _bonus_estructura = 60  # empleador confirmado por AFIP sin registro MiPyME

        # Bonus fiscal unificado: cuando el perfil fiscal NO es ya la Capa B
        # (es decir, cuando sí hay historial BCRA o interno), entra acá para que
        # el score combine mora de 24 meses + deducción fiscal. Escala 0-90 pts
        # sobre 400 del perfil. Nunca por debajo del bonus MiPyME histórico:
        # ningún cliente pierde puntos respecto de la calibración anterior.
        if pts_fiscal is not None and not _fiscal_es_capa_b:
            _bonus_fiscal     = round(pts_fiscal / 400 * 90)
            _bonus_estructura = max(_bonus_estructura, _bonus_fiscal)

        if _bonus_estructura:
            puntos += _bonus_estructura
            _cat_log = _cat_mp if _cat_mp else 'gran_empresa'
            print(f"[score] {cuit_limpio} bonus_estructura cat={_cat_log} "
                  f"fiscal={pts_fiscal} → +{_bonus_estructura}", flush=True)

    # ── Piso v25.1: Sit.1 + deuda BCRA $0 + historial bancario real ──────────
    # Solo aplica si el cliente tiene períodos BCRA reportados (fue cliente de algún banco).
    # Si no hay historial en absoluto (CUIT sin actividad bancaria nunca), el score raw
    # es más honesto: 495 ≠ 650 porque no sabemos nada positivo de él, solo que no es moroso.
    # Diferencia: "limpio con track record" (650+) vs "desconocido sin historial" (raw~495).
    _tiene_historial_bancario = len(periodos_hist) > 0
    _cliente_sin_deuda = (
        max_sit == 1 and monto_real == 0
        and not en_mora and not hard_block_mora
        and _tiene_historial_bancario
    )
    if _cliente_sin_deuda:
        puntos = max(puntos, 650)
        print(f"[score v{_SCORE_VERSION}] {cuit_limpio} sin_deuda_sit1 → piso 650", flush=True)
    elif max_sit == 1 and monto_real == 0 and not en_mora and not _tiene_historial_bancario:
        print(f"[score v{_SCORE_VERSION}] {cuit_limpio} sin_historial_bancario → score raw {round(puntos)}", flush=True)

    # ── Ajuste: concentración de deuda ────────────────────────────────────
    if   nro_entidades == 0 or sin_deudas_real:     puntos += 25
    elif nro_entidades == 1 and monto_total_m < 50: puntos += 18
    elif nro_entidades <= 2 and monto_total_m < 100:puntos += 12

    # ── Ajuste: ratio de apalancamiento BCRA/AFIP ─────────────────────────
    if solvency_data:
        try:
            _ing_afip = float(solvency_data.get('ingresos_anuales') or 0)
        except (TypeError, ValueError):
            _ing_afip = 0.0
        if es_mora_administrativa and monto_sit1_k > 0:
            # Criterio Humano: ingreso presunto = deuda Sit.1 × 3 (bancos limpios).
            # Solo evaluar la deuda limpia contra ese ingreso — el outlier no cuenta.
            _ing     = max(_ing_afip, int(monto_sit1_k * 3))
            _deu_chk = monto_sit1_k
        else:
            if not _ing_afip and monto_total_m > 0:
                solvency_data = dict(solvency_data)
                solvency_data['ingresos_anuales'] = round(monto_total_m * 1_000 * 3)
                solvency_data['fuente_ingresos']  = 'bcra_floor_scorer'
            try:
                _ing = float(solvency_data.get('ingresos_anuales') or 0)
            except (TypeError, ValueError):
                _ing = 0.0
            _deu_chk = monto_total_m * 1000
        # Piso de ingreso: si AFIP reporta menos de $100k anuales para alguien
        # con deuda bancaria real, el dato AFIP es incompleto → usar deuda × 3.
        if 0 < _ing < 100_000 and monto_total_m > 0:
            _ing_floor = round(monto_total_m * 1_000 * 3)
            print(
                f"[score v{_SCORE_VERSION}] {cuit_limpio} ingreso_afip={_ing} < 100k → "
                f"floor a deuda×3={_ing_floor}",
                flush=True
            )
            _ing = _ing_floor
        # Fix de unidades: los montos del BCRA vienen en MILES de pesos, mientras
        # que ingresos_anuales está en pesos. El ratio se venía calculando
        # mezclando ambas escalas, quedando 1000× por debajo del real — la
        # penalización era, en los hechos, código muerto.
        _deu_pesos = _deu_chk * 1_000
        _ratio_apal = (_deu_pesos / _ing) if _ing > 0 else 0.0
        print(
            f"[score v{_SCORE_VERSION}] {cuit_limpio} ing={_ing} deu_pesos={_deu_pesos} "
            f"ratio={round(_ratio_apal, 3) if _ing else 'inf'}",
            flush=True
        )
        # Penalización graduada + factor de transición.
        # Antes era un acantilado: 0 puntos hasta 0.5 y −200 un centavo después.
        # Con el ratio ya corregido eso produciría saltos bruscos de rango en la
        # cartera, así que la penalización crece de forma lineal entre el umbral
        # y el techo, y se escala por APALANCAMIENTO_FACTOR (0.5 = mitad de la
        # penalización nominal). Subir esa variable a 1.0 en Render aplica el
        # rigor pleno sin necesidad de redeploy.
        if _ing > 0 and _ratio_apal > _APAL_UMBRAL:
            _exceso = min(1.0, (_ratio_apal - _APAL_UMBRAL) / _APAL_RANGO)
            _penal  = round(_APAL_PENAL_MAX * _exceso * _APAL_FACTOR)
            if _penal > 0:
                puntos -= _penal
                print(
                    f"[score v{_SCORE_VERSION}] {cuit_limpio} apalancamiento "
                    f"ratio={_ratio_apal:.2f} exceso={_exceso:.2f} "
                    f"factor={_APAL_FACTOR} → -{_penal}",
                    flush=True
                )

    # ── Deuda interna +90 días: penaliza aunque BCRA no lo vea aún ──────────
    if deuda_90d_interna:
        puntos -= 200
        print(f"[score v{_SCORE_VERSION}] {cuit_limpio} deuda_90d_interna → -200", flush=True)

    # ── DSO v20.0: bono pago rápido / penalidad pago lento ───────────────────
    if dso_individual > 0:
        if dso_individual < 45:
            puntos += 50
            print(f"[score v{_SCORE_VERSION}] {cuit_limpio} DSO={dso_individual:.0f}d<45 → +50", flush=True)
        elif dso_individual > 90:
            puntos -= 100
            print(f"[score v{_SCORE_VERSION}] {cuit_limpio} DSO={dso_individual:.0f}d>90 → -100", flush=True)

    # ── Time Decay BINARIO: si hoy está en Sit.1, penalidad histórica >6m = CERO ──
    # "Un cliente que hoy cumple no puede ser castigado eternamente por el pasado."
    # Regla: meses malos ocurridos hace >6 meses se eliminan completamente cuando
    # la situación efectiva actual es 1. Los meses recientes (0-6m) siguen contando.
    if sit_efectivo == 1 and _mm_antiguos > 0:
        meses_malos_td = _mm_recientes   # antiguo contribuye CERO
        print(
            f"[time-decay] {cuit_limpio}: sit_ef=1 "
            f"mm_rec={_mm_recientes} mm_ant={_mm_antiguos}→0 (eliminado) "
            f"meses_malos {meses_malos}→{meses_malos_td}",
            flush=True
        )
        meses_malos = meses_malos_td

    # ── Penalidades históricas ────────────────────────────────────────────
    if 2 <= meses_malos <= 5 and not es_mora_tecnica and not es_mora_administrativa:
        puntos = round(puntos * 0.75)
    # sit_grave_6m: solo aplica cuando max_sit < 3 (max_sit >= 3 → elastic bounding)
    if sit_grave_6m and not es_mora_tecnica and sit_efectivo >= 3 and max_sit < 3:
        puntos = min(puntos, 150)
    elif sit_grave_6m and es_mora_tecnica and sit_efectivo >= 3 and max_sit < 3:
        puntos = min(puntos, 350)

    # ── Hard Block D2: Default Real BCRA → cap en rango Rechazar ────────
    # max_sit>=3: puntos=0 (el elastic bounding lo lleva a [300-550] abajo).
    # max_sit==2: score=1 era demasiado drástico para un cliente que cayó a sit≥4
    # y ya está recuperando a sit=2. Se diferencia por historial:
    #   - deterioro_estructural (sit≥3 sostenido en últimos 6m): cap ~80
    #   - default_real moderado (sin historial sit≥3 reciente):  cap ~150
    # Score=1 queda reservado para cheques críticos y sit=5 activo.
    if hard_block_bcra:
        if max_sit >= 3:
            puntos = 0  # Elastic bounding aplica debajo
        elif deterioro_estructural:
            puntos = min(puntos, 80)
        else:
            puntos = min(puntos, 150)

    # ── Elastic Bounding v12.1: max_sit >= 3 → penalización dinámica ─────
    # Aplica SIEMPRE para max_sit >= 3, incluso si hard_block_bcra es True.
    # hard_block_bcra = True es común (default real = tuvo Sit.1 y cayó) y
    # no justifica un score de 1. El rango elástico ya captura la severidad.
    # max_sit 3 → [460, 550] | max_sit >= 4 → [300, 460] (pisa 200 con historial severo)
    if max_sit >= 3:
        _pit     = min(1.0, pct_mora)                    # 0=todo Sit.1, 1=todo en mora
        _sit_mul = 1.0 + (max_sit - 3) * 0.25           # Sit3→1.0, Sit4→1.25, Sit5→1.50
        _pit_adj = min(1.0, _pit * _sit_mul)
        if max_sit == 3:
            _lo, _hi = 460, 550
        else:                                             # max_sit >= 4
            _lo, _hi = 300, 460
            if sit_grave_6m and meses_malos >= 3:        # historial severo → pisa piso 300
                _lo = max(200, _lo - 100)
        _puntos_eb = round(_hi - _pit_adj * (_hi - _lo))
        puntos     = max(_lo, min(_hi, _puntos_eb))
        print(
            f"[elastic v12] {cuit_limpio} max_sit={max_sit} pct_mora={pct_mora:.2f} "
            f"pit_adj={_pit_adj:.3f} box=[{_lo},{_hi}] → {puntos}",
            flush=True
        )

    # ── Piso estructura empresarial (solo max_sit==1, BCRA limpia) ───────────
    # Un empleador/PyME con BCRA Sit.1 tiene sustancia crediticia mínima verificada.
    # El padrón MiPyME actúa como proxy de nómina declarada ante AFIP/SEPYME.
    # Los caps posteriores (mora Odoo, comunidad negativa) siguen teniendo prioridad:
    # la estructura reduce riesgo base pero no protege de incumplimiento real.
    if max_sit == 1 and not hard_block_bcra and solvency_data:
        _cat_mp_p  = (solvency_data.get('categoria_mipyme') or '').strip()
        # Requiere confirmación explícita de AFIP/TangoFactura — no asumimos por tipo_persona
        _es_empl3  = solvency_data.get('es_empleador') is True
        _PISO_MP   = {'Pequeña': 450, 'Mediana_T1': 550, 'Mediana_T2': 600}
        _piso_empl = _PISO_MP.get(_cat_mp_p, 0)
        if not _cat_mp_p and _es_empl3:
            _piso_empl = 650   # empleador confirmado por AFIP sin registro MiPyME
        if _piso_empl and puntos < _piso_empl:
            _cat_log2 = _cat_mp_p if _cat_mp_p else 'gran_empresa'
            print(
                f"[score] {cuit_limpio} piso_estructura cat={_cat_log2} "
                f"→ {round(puntos)}→{_piso_empl}",
                flush=True
            )
            puntos = _piso_empl

    # ── Hard Block: mora interna Odoo → score ≤ 400 ──────────────────────
    if hard_block_mora:
        puntos = min(puntos, 400)

    # ── Cap: Monotrib A/B → score ≤ 600 ──────────────────────────────────
    if es_monotrib_bajo:
        puntos = min(puntos, 600)

    # ── Cap preventivo por irregularidad fiscal (Padrón A13) ─────────────────
    # Clave fiscal dada de baja o CUIT sin ningún impuesto activo: el cliente no
    # está en condiciones de facturar legalmente. No se rechaza automáticamente
    # —el dato puede estar desactualizado— pero se fuerza revisión humana
    # llevando el score a la banda "Revisar".
    _irregular_fiscal = bool(dbg_fiscal.get('cuit_fantasma')) or (
        dbg_fiscal.get('factor_estado_clave') is not None
        and dbg_fiscal.get('factor_estado_clave') <= 0.5
    )
    if _irregular_fiscal and puntos > 400:
        print(
            f"[score v{_SCORE_VERSION}] {cuit_limpio} irregularidad fiscal ARCA "
            f"({'; '.join(dbg_fiscal.get('alertas', [])) or 'sin detalle'}) "
            f"→ cap 400", flush=True
        )
        puntos = min(puntos, 400)

    # ── Cap v20.0: comunidad negativa → score ≤ 600 (salvo mora técnica) ────
    if comunidad_negativa and not es_mora_tecnica:
        puntos = min(puntos, 600)
        print(f"[score v{_SCORE_VERSION}] {cuit_limpio} comunidad_negativa → cap 600", flush=True)

    # ── Piso mora técnica (no aplica si hay Default Real) ────────────────
    if es_mora_tecnica and not hard_block_bcra:
        puntos = max(puntos, 700)

    # ── Penalización diasAtraso (CDI v1.0) ────────────────────────────────
    if _ents_curr:
        _max_dias = max((e.get('diasAtraso', 0) or 0) for e in _ents_curr)
        if _max_dias > 180:
            puntos -= 100
            print(f"[score] {cuit_limpio} diasAtraso={_max_dias} → −100pts", flush=True)
        elif _max_dias > 90:
            puntos -= 50
            print(f"[score] {cuit_limpio} diasAtraso={_max_dias} → −50pts", flush=True)

    score = max(1, min(999, round(puntos)))

    if   score >= 750: rango, color, emoji = 'Excelente',   '#16a34a', '🟢'
    elif score >= 600: rango, color, emoji = 'Bueno',       '#ca8a04', '🟡'
    elif score >= 400: rango, color, emoji = 'Revisar',     '#ea580c', '🟠'
    elif score >= 200: rango, color, emoji = 'Alto riesgo', '#dc2626', '🔴'
    else:              rango, color, emoji = 'Rechazar',    '#7f1d1d', '⛔'

    # ── Cap: sin actividad bancaria (fantasma crediticio) → max 350 ──────────
    # Solo aplica cuando NO tenemos ninguna confirmación BCRA del cliente.
    # Si BCRA respondió con results válidos (aunque sea sin_deudas=True o datos bulk),
    # el cliente existe y está registrado — no es un fantasma crediticio.
    # Fallar en el endpoint de historial NO es evidencia de fantasma; es un problema
    # de conectividad BCRA que no debe penalizar al cliente.
    _bcra_confirmo = bool(bcra_data and bcra_data.get('results') is not None and not bcra_data.get('error_bcra'))
    _sin_actividad_bancaria = (not _ents_curr) and n_periodos_h == 0 and monto_real == 0 and not _bcra_confirmo
    if _sin_actividad_bancaria:
        score = min(score, 350)
        rango, color, emoji = 'Alto riesgo', '#dc2626', '🔴'
        print(f"[score] {cuit_limpio} sin actividad bancaria → cap 350", flush=True)

    alerta_temprana      = alerta_creciente or es_monotrib_bajo or indice_solv < 0.40
    bloquear_oportunidad = (
        (hard_block_mora or hard_block_bcra or (en_mora and score > 700)) and
        not es_mora_tecnica and
        not es_mora_administrativa   # belt-and-suspenders: administrativa ≠ bloqueo
    )
    alerta_log = _alerta_logistica(ciudad)

    if es_mora_tecnica:
        color = '#ca8a04'
        rango = rango if score >= 750 else ('Revisar' if rango in ('Rechazar', 'Alto riesgo') else rango)

    # Criterio Humano final: mora administrativa no puede salir como 'Rechazar'
    if es_mora_administrativa and rango in ('Rechazar', 'Alto riesgo'):
        rango = 'VENTA RESTRINGIDA'
        color = '#7c3aed'    # violeta = requiere revisión humana
        emoji = '⚠️'

    print(
        f"[score v{_SCORE_VERSION}] {cuit_limpio} sit={max_sit} sp={sit_ponderada:.2f} "
        f"mt={es_mora_tecnica} tend={tendencia} prosp={sin_historial_interno} "
        f"cA={pts_c1} cB={pts_cb} cC={pts_cc} liq={pts_liq} "
        f"mora_bcra={tipo_mora_bcra} sit_ef={sit_efectivo} "
        f"mc={es_mora_comercial_activa} det={deterioro_estructural} "
        f"→ {score} {rango} | at={alerta_temprana} bloq={bloquear_oportunidad} geo={alerta_log or '-'}",
        flush=True
    )

    resultado = {
        'score':                    score,
        'rango':                    rango,
        'color':                    color,
        'emoji':                    emoji,
        'alerta_temprana':          alerta_temprana,
        'bloquear_oportunidad':     bloquear_oportunidad,
        'alerta_logistica':         alerta_log,
        'componentes': {
            'capaA': pts_c1, 'capaB': pts_cb,
            'capaC': pts_cc, 'liquidez': pts_liq,
            'fiscal': pts_fiscal, 'fiscal_es_capaB': _fiscal_es_capa_b,
            'bonus_estructura': _bonus_estructura,
        },
        'perfil_fiscal':            dbg_fiscal or None,
        'alertas_fiscales':         dbg_fiscal.get('alertas') or [],
        'cuit_fantasma':            bool(dbg_fiscal.get('cuit_fantasma')),
        'tendencia':                tendencia,
        'es_empleador':             es_empleador,
        'indice_solvencia':         indice_solv,
        'version':                  _SCORE_VERSION,
        'mora_tecnica':             es_mora_tecnica,
        'mora_comercial_activa':    es_mora_comercial_activa,
        'deterioro_estructural':    deterioro_estructural,
        'sit_ponderada':            round(sit_ponderada, 3),
        'pct_mora':                 round(pct_mora, 4),
        'max_sit':                  max_sit,
        'sit_efectivo':             sit_efectivo,
        'mora_administrativa':      es_mora_administrativa,
        # Campos v10.0 / v20.0
        'sin_historial_interno':    sin_historial_interno,
        'dso_individual':           dso_individual,
        'dso_deteriorando':         dso_deteriorando,
        'promedio_mensual':         promedio_mensual,
        'comunidad_negativa':       comunidad_negativa,
        'deuda_90d_interna':        deuda_90d_interna,
        'monto_deuda_90d':          round(monto_deuda_90d_interna, 2),
        'tipo_mora_bcra':           tipo_mora_bcra,
        'degradacion_bcra_reciente':hard_block_bcra,
        'aviso_mora':               aviso_mora or None,
        'limite_dinamico_sugerido': round(monto_sit1_k * 1000 * 0.30) if (es_mora_administrativa and monto_sit1_k > 0) else None,
        'nota_mora_tecnica': (
            f"Atención: Se detecta una mora técnica por monto menor que no afecta "
            f"la solvencia general. Deuda en mora: ${round(monto_mora_k):,} miles ARS "
            f"({round(pct_mora*100, 1)}% del total), situación ponderada {sit_ponderada:.2f}. "
            f"El {round((1-pct_mora)*100, 1)}% de la cartera bancaria se encuentra en Situación 1."
        ) if es_mora_tecnica else None,
        'razonamiento_score': (
            'Score preventivo: El cliente mantiene situación normal sin registros de deuda bancaria activa.'
        ) if _cliente_sin_deuda else (
            f"Score preservado: mora técnica de baja materialidad "
            f"(${round(monto_mora_k * 1000):,} ARS, {round(pct_mora*100, 1)}% del total)"
        ) if es_mora_tecnica else (
            "Score preservado: mora clasificada como administrativa (patrón crediticio limpio)"
        ) if es_mora_administrativa else None,
        # Campos v20.0
        'semaforo': 'verde' if score >= 700 else ('amarillo' if score >= 400 else 'rojo'),
        'falsas_deudas': _fd,   # entidades excluidas + razones para UI y análisis IA
    }
    _score_session_cache[cuit_limpio] = resultado
    return resultado


# v20.0 alias para compatibilidad con callers externos
calcular_vende_score_pro = calcular_rating_predictivo


# ══════════════════════════════════════════════════════════════════════════════
# PADRÓN LOCAL DE IDENTIDADES ARCA
# Nomdeu (BCRA) solo conoce a quien está bancarizado. Para el cliente nuevo sin
# historial bancario, ARCA es la única fuente de identidad real — se persiste
# acá para que la app pueda nombrarlo sin volver a consultar el WS.
# ══════════════════════════════════════════════════════════════════════════════

_IDENT_ARCA_FILE   = os.path.join(DATA_DIR, 'identidades_arca.json')
_IDENT_ARCA_R2_KEY = 'identidades_arca.json'
_ident_arca_lock   = threading.Lock()
_ident_arca_cache: dict | None = None   # lazy load, se mantiene en memoria
_ident_arca_r2_ts  = [0.0]              # último upload a R2 (debounce, mutable p/ closure)


def _ident_arca_load() -> dict:
    """Carga el padrón de identidades ARCA (memoria → disco → R2)."""
    global _ident_arca_cache
    if _ident_arca_cache is not None:
        return _ident_arca_cache
    try:
        with open(_IDENT_ARCA_FILE, 'r', encoding='utf-8') as f:
            _ident_arca_cache = json.load(f)
    except Exception:
        _ident_arca_cache = {}
        raw = _r2_download_bytes(_IDENT_ARCA_R2_KEY)
        if raw:
            try:
                _ident_arca_cache = json.loads(raw.decode('utf-8'))
                print(f"[ident-arca] Restaurado desde R2: {len(_ident_arca_cache)} identidades", flush=True)
            except Exception:
                _ident_arca_cache = {}
    if not isinstance(_ident_arca_cache, dict):
        _ident_arca_cache = {}
    return _ident_arca_cache


def _ident_arca_registrar(cuit: str, solvency: dict) -> bool:
    """Da de alta / actualiza la identidad real de un CUIT desde datos ARCA.

    Solo persiste cuando la fuente es el canal oficial y hay razón social:
    no queremos ensuciar el padrón con inferencias de scrapers de terceros.
    Retorna True si escribió (alta o cambio), False si no había nada nuevo.
    """
    if not isinstance(solvency, dict):
        return False
    if 'arca' not in str(solvency.get('fuente') or '').lower():
        return False
    razon = str(solvency.get('razon_social') or '').strip()
    if not razon:
        return False

    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    registro = {
        'razon_social':       razon,
        'tipo_persona':       solvency.get('tipo_persona') or '',
        'clae_actividad':     solvency.get('clae_actividad') or solvency.get('actividad_principal') or '',
        'categoria_monotrib': solvency.get('categoria_monotrib') or '',
        'es_empleador':       solvency.get('es_empleador') is True,
        'antiguedad_anos':    solvency.get('antiguedad_anos'),
        'estado_afip':        solvency.get('estado_afip') or '',
        'ts':                 time.time(),
    }

    with _ident_arca_lock:
        ident = _ident_arca_load()
        previo = ident.get(cuit_limpio) or {}
        # Comparar solo campos de negocio: el ts siempre difiere y provocaría
        # una escritura a disco + upload R2 en cada consulta del mismo CUIT.
        _campos = ('razon_social', 'tipo_persona', 'clae_actividad',
                   'categoria_monotrib', 'es_empleador', 'antiguedad_anos', 'estado_afip')
        if all(previo.get(k) == registro[k] for k in _campos):
            return False
        ident[cuit_limpio] = registro
        try:
            with open(_IDENT_ARCA_FILE, 'w', encoding='utf-8') as f:
                json.dump(ident, f, ensure_ascii=False)
        except Exception as e:
            print(f"[ident-arca] Error guardando {cuit_limpio}: {e}", flush=True)
            return False

        # Backup R2 con debounce: en el primer barrido de cartera se dan de alta
        # cientos de CUITs seguidos; subir el JSON completo en cada uno sería
        # cuadrático. El disco es la fuente de verdad, R2 solo el respaldo.
        _snapshot = None
        if time.time() - _ident_arca_r2_ts[0] > 120:
            _ident_arca_r2_ts[0] = time.time()
            _snapshot = json.dumps(ident, ensure_ascii=False).encode('utf-8')

    if _snapshot is not None:
        def _bg(data=_snapshot):
            _r2_upload_bytes(_IDENT_ARCA_R2_KEY, data, 'application/json')
        threading.Thread(target=_bg, daemon=True).start()

    print(f"[ident-arca] {cuit_limpio} alta identidad ARCA: {razon[:50]}", flush=True)
    return True


def _ident_arca_get(cuit: str) -> dict | None:
    """Identidad ARCA registrada de un CUIT, o None."""
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    return _ident_arca_load().get(cuit_limpio)


def _denominacion_local(cuit: str) -> str | None:
    """Nombre del CUIT sin tocar la red: padrón BCRA (Nomdeu) → identidad ARCA.

    Cubre al cliente no bancarizado, que nunca aparece en Nomdeu. Pensada para
    bucles calientes (barrido de cartera): nunca hace llamadas de red.
    """
    nombre = _nomdeu_get_nombre(cuit)
    if nombre:
        return nombre
    ident = _ident_arca_get(cuit)
    return (ident or {}).get('razon_social') or None


def _denominacion_arca(cuit: str) -> str | None:
    """Denominación oficial desde ARCA, dando de alta la identidad al pasar.

    Resuelve el caso que el padrón BCRA no puede cubrir: un CUIT nuevo, o de
    alguien que nunca operó en el sistema financiero, no figura en Nomdeu — pero
    sí existe en el padrón fiscal desde el día que se inscribió.

    Consulta el registro local primero (costo cero); solo va al web service si
    el CUIT nunca se resolvió antes. El resultado queda persistido, así que
    cada CUIT se consulta una única vez.

    Cubre tanto personas jurídicas (razón social) como físicas (apellido y
    nombre), que es más de lo que puede aportar cualquiera de los scrapers.
    """
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()

    # 1. Registro local — ya resuelto en una consulta anterior
    ident = _ident_arca_get(cuit_limpio)
    if ident and ident.get('razon_social'):
        return ident['razon_social']

    # 2. Canal oficial
    if not ARCA_DISPONIBLE:
        return None
    try:
        datos = arca_ws.obtener_datos_fiscales_arca(cuit_limpio)
    except Exception as e:
        print(f"[afip] {cuit_limpio} ARCA oficial falló: {e}", flush=True)
        return None
    if not datos:
        return None

    razon = str(datos.get('razon_social') or '').strip()
    if not razon or razon.isdigit():
        return None

    # Persistir la identidad completa (nombre, CLAE, estado de clave, etc.)
    # para que toda la app la aproveche sin volver a consultar el WS.
    _ident_arca_registrar(cuit_limpio, datos)
    return razon


def calcular_score_servidor(cuit: str, bcra_data: dict, en_mora=None, ciudad: str = '') -> dict:
    """
    Wrapper de calcular_rating_predictivo v9.0.
    Carga historial y cheques desde caché local (graceful degradation).
    Mantiene compatibilidad con todos los callers existentes.

    Paridad BCRA/ARCA: el historial de 24 meses y la consulta al canal oficial
    ARCA arrancan en paralelo — ninguna espera a la otra, y el score se calcula
    combinando ambas.
    """
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()

    # ── ARCA/solvencia en paralelo con la extracción BCRA ──────────────────
    # get_solvency_data cachea 24h en disco, así que en consultas repetidas el
    # thread termina de inmediato. Nunca propaga excepción: el score debe poder
    # calcularse aunque el canal fiscal esté caído.
    _solv_box: dict = {}

    def _fetch_solvencia():
        try:
            _sv = get_solvency_data(cuit_limpio)
            if isinstance(_sv, dict):
                _solv_box['data'] = _sv
        except Exception as _e_sv:
            print(f"[score] {cuit_limpio} solvencia paralela falló: {_e_sv}", flush=True)

    _th_solv = threading.Thread(target=_fetch_solvencia, daemon=True)
    _th_solv.start()

    def _cache_load(fname):
        p = os.path.join(DATA_DIR, fname)
        try:
            if os.path.exists(p):
                with open(p, 'r') as f:
                    return json.load(f).get('payload')
        except: pass
        return None

    hist_data = _cache_load(f'historial_{cuit_limpio}.json')
    cheq_data = _cheques_cache_get(cuit_limpio)   # usa TTL diferenciado: 1h "sin cheques", 24h con datos

    # Historial vacío = fetch fallido anterior — siempre re-intentar (directo, sin ScraperAPI)
    if hist_data and not (hist_data.get('results') or {}).get('periodos'):
        hist_data = None

    if not hist_data:
        # Timeout corto (8s × 1 intento por endpoint = 16s máx) para que el endpoint
        # /fetch-score responda siempre dentro del timeout de 40s del frontend.
        _hd, _ = _consultar_bcra_directo(cuit_limpio, 'historial', timeout_per_req=8, max_intentos=1)
        if _hd:
            hist_data = _hd
            # Solo escribir a disco si hay periodos reales — evita cachear resultados vacíos
            if (_hd.get('results') or {}).get('periodos'):
                try:
                    with open(os.path.join(DATA_DIR, f'historial_{cuit_limpio}.json'), 'w') as f:
                        json.dump({'payload': hist_data, 'ts': time.time()}, f)
                except: pass

    # BCRA en vivo caído/sin periodos — usar padrón offline (historial_detalle, cargado
    # desde R2 al arrancar) en vez de degradar la tendencia a un solo período (periodos_curr).
    if not hist_data or not (hist_data.get('results') or {}).get('periodos'):
        try:
            _deuda_bulk = _nomdeu_get_deuda(cuit_limpio)
            if _deuda_bulk:
                hist_data = _bulk_to_hist_data(cuit_limpio)
                print(f"[score] {cuit_limpio} hist_data offline desde historial_detalle ({_HIST_DETALLE_MESES}m reales, R2)", flush=True)
        except Exception as _e_bulk:
            print(f"[score] {cuit_limpio} historial_detalle fallback error: {_e_bulk}", flush=True)

    # Módulo cheques — aislado con fallback absoluto.
    # Un timeout o error en este módulo NO debe abortar el cálculo del score.
    try:
        if not cheq_data:
            # DB local (snapshot diario BCRA) — cero latencia de red
            cheq_data = get_cheques_local(cuit_limpio)
            if cheq_data is not None:
                _cheques_cache_set(cuit_limpio, cheq_data)
                print(f"[score] cheq {cuit_limpio} desde DB local", flush=True)
        if not cheq_data:
            # Timeout corto para no bloquear el endpoint
            _cd, _ = _consultar_bcra_directo(cuit_limpio, 'cheques', timeout_per_req=8, max_intentos=1)
            if _cd:
                cheq_data = _cd
                try:
                    with open(os.path.join(DATA_DIR, f'cheques_{cuit_limpio}.json'), 'w') as f:
                        json.dump({'payload': cheq_data, 'ts': time.time()}, f)
                except: pass
    except Exception as _cheq_err:
        print(f"[cheques][FALLBACK] {cuit_limpio}: error en módulo cheques ({_cheq_err}) — score continúa sin antecedentes", flush=True)
        cheq_data = None

    if not isinstance(bcra_data, dict):
        bcra_data = _norm_bcra_resp(bcra_data) if bcra_data else {}
    _bcra_disponible = bcra_data.get('bcra_disponible', not bool(bcra_data.get('error_bcra')))

    # ── Reunir la rama fiscal (ya venía corriendo en paralelo) ─────────────
    # Cota superior generosa pero acotada: si el canal fiscal se cuelga, el
    # score sale igual con los datos BCRA y el thread termina de poblar la
    # caché en background para la próxima consulta.
    _th_solv.join(timeout=25)
    if _th_solv.is_alive():
        # Pasar {} y no None: None haría que el motor dispare una SEGUNDA
        # consulta fiscal sobre la misma cadena colgada y el request se pasaría
        # del timeout del frontend.
        solvency_data = {}
        print(f"[score] {cuit_limpio} solvencia aún en curso a los 25s — score con BCRA solamente", flush=True)
    else:
        # None solo si el fetch falló: el motor reintenta (barato, con caché 24h),
        # que es exactamente el comportamiento previo a la paralelización.
        solvency_data = _solv_box.get('data')

    # ── Alta de identidad para el cliente no bancarizado ───────────────────
    # BCRA sin historial pero ARCA respondió: es un cliente real que todavía no
    # pisó el sistema financiero. Su identidad oficial se registra en el padrón
    # local para que la app pueda nombrarlo sin volver a consultar el WS.
    if solvency_data:
        _res_hist     = (hist_data or {}).get('results')
        _hay_periodos = bool(_res_hist.get('periodos')) if isinstance(_res_hist, dict) else False
        if not _hay_periodos:
            _ident_arca_registrar(cuit_limpio, solvency_data)

    resultado = calcular_rating_predictivo(
        cuit=cuit_limpio, bcra_data=bcra_data,
        hist_data=hist_data, cheq_data=cheq_data,
        en_mora=en_mora, ciudad=ciudad,
        solvency_data=solvency_data,
    )
    resultado['bcra_disponible'] = _bcra_disponible
    return resultado


# Alias para mantener compatibilidad con código legado
_calcular_score = calcular_score_servidor


def _actualizar_score_en_cartera(cuit_limpio: str, score_data: dict, solvency: dict = None):
    """Merge atómico: actualiza / inserta el score de un CUIT en alertas_cartera.json.
    Persiste scoreHistory[] (ventana 12 meses) y detecta degradación Anti-Videla."""
    try:
        try:
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
                existente = json.load(_f)
        except:
            existente = {"alertas": [], "ultima_verif": "", "cartera": []}
        cartera = existente.get('cartera', [])
        nc = str(cuit_limpio).replace('-', '').replace(' ', '').strip()

        # Recuperar historial previo del cliente
        entrada_prev = next(
            (c for c in cartera
             if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == nc),
            {}
        )
        hist_prev = list(entrada_prev.get('scoreHistory') or [])

        # Agregar punto actual y recortar a 24 entradas (ventana 24 meses = 2 ciclos anuales)
        # 24 meses cubre ciclos económicos completos en Argentina y mejora detección Anti-Videla.
        hist_prev.append({
            'score':    score_data.get('score'),
            'fecha':    datetime.now().strftime('%Y-%m-%d'),
            'sit_bcra': score_data.get('max_sit', 1),
        })
        score_history = hist_prev[-24:]

        # Detección de degradación Anti-Videla
        deg_tipo, deg_delta, deg_msg = _detectar_degradacion(score_history)

        patch = {
            'scoreCompleto':        score_data.get('score'),
            'scoreRango':           score_data.get('rango'),
            'scoreColor':           score_data.get('color'),
            'scoreEmoji':           score_data.get('emoji'),
            # max_sit: situación BCRA real (sin ajustes de mora técnica/administrativa).
            # sit_efectivo puede quedar en 1 por lógica de scoring, pero lo que el banco
            # reporta es max_sit — ese es el dato que debe ver el vendedor.
            'ultimaSit':            score_data.get('max_sit', 1),
            'alerta_temprana':      score_data.get('alerta_temprana', False),
            'bloquear_oportunidad': score_data.get('bloquear_oportunidad', False),
            'alerta_logistica':     score_data.get('alerta_logistica', ''),
            'inferencia_ingresos':  (solvency if isinstance(solvency, dict) else {}).get('ingresos_anuales'),
            'fuente_ingresos':      (solvency if isinstance(solvency, dict) else {}).get('fuente_ingresos'),
            'actividad_principal':  (solvency if isinstance(solvency, dict) else {}).get('actividad_principal'),
            'ultimaVerif':          time.strftime('%d/%m/%Y'),
            'score_ts':             time.time(),
            'pendiente':            False,
            'scoreHistory':         score_history,
            'degradacion_tipo':     deg_tipo,
            'degradacion_delta':    deg_delta,
            'degradacion_msg':      deg_msg or None,
        }
        found = False
        for i, c in enumerate(cartera):
            if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == nc:
                cartera[i].update(patch)
                found = True
                break
        if not found:
            cartera.append({'cuit': cuit_limpio, 'nombre': '', 'ciudad': '', **patch})
        existente['cartera'] = cartera
        _tmp_af = ALERTAS_FILE + '.tmp'
        with open(_tmp_af, 'w', encoding='utf-8') as _f:
            json.dump(existente, _f, ensure_ascii=False, default=str)
            _f.flush()
            os.fsync(_f.fileno())
        os.replace(_tmp_af, ALERTAS_FILE)
        if deg_tipo:
            print(f"[anti-videla] {cuit_limpio} → {deg_tipo} (−{deg_delta} pts)", flush=True)
    except Exception as _e:
        print(f"[score-update] Error persistiendo {cuit_limpio}: {_e}", flush=True)
        try:
            os.remove(ALERTAS_FILE + '.tmp')
        except OSError:
            pass


def _analizar_bodegas_batch(clientes_batch):
    """Analiza mensajes de bodegas para un lote de clientes en UNA sola llamada a Gemini.
    Reduce llamadas IA de N (una por cliente) a N/8 (una por lote).
    Args: clientes_batch — lista de {cuit, nombre, mensajes: [str]}
    Returns: {cuit: (es_negativo: bool, motivo: str)}
    """
    if not clientes_batch:
        return {}
    secciones = []
    for cli in clientes_batch:
        msgs_txt = "\n".join(f"- {m}" for m in cli.get('mensajes', [])[:10])
        secciones.append(f"CUIT: {cli['cuit']} | {cli['nombre']}\n{msgs_txt}")
    bloque = "\n\n---\n\n".join(secciones)
    prompt = (
        "Sos un Analista de Riesgo Crediticio experto en el sector vitivinícola argentino.\n"
        "Analizá los mensajes de grupo de bodegas para CADA cliente listado.\n\n"
        "REGLAS:\n"
        "- Solo negativo si hay deudas impagas NO resueltas, estafas o desaparición del deudor.\n"
        "- Cheques rechazados pero reemplazados = NO negativo.\n"
        "- Si distintas bodegas dicen cosas contradictorias → comportamiento_inconsistente: true.\n"
        "- NUNCA respondas 'sin antecedentes' si el chat tiene mensajes.\n\n"
        "CLIENTES A ANALIZAR:\n\n" + bloque + "\n\n"
        "Respondé SOLO con JSON sin markdown. Una clave por CUIT exacto:\n"
        '{"CUIT1": {"es_negativo": false, "motivo": "...", "comportamiento_inconsistente": false}, '
        '"CUIT2": {"es_negativo": false, "motivo": "...", "comportamiento_inconsistente": false}}'
    )
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    fallback = {cli['cuit']: (False, "") for cli in clientes_batch}
    texto, error = gemini_request(payload, timeout=90)
    if error or not texto:
        print(f"[bodegas-batch] Sin respuesta IA — fallback no-negativo para {len(clientes_batch)} clientes", flush=True)
        return fallback
    try:
        import re as _re
        texto_limpio = texto.strip().replace("```json", "").replace("```", "").strip()
        _m = _re.search(r'\{[\s\S]+\}', texto_limpio)
        if not _m:
            return fallback
        data = json.loads(_m.group(0))
        result = {}
        for cli in clientes_batch:
            cuit = cli['cuit']
            entrada = data.get(cuit, {})
            motivo = entrada.get("motivo", "")
            if entrada.get("comportamiento_inconsistente"):
                motivo = "⚠ Comportamiento Inconsistente: " + motivo
            result[cuit] = (entrada.get("es_negativo", False), motivo)
        print(f"[bodegas-batch] Lote OK — {len(result)} clientes analizados", flush=True)
        return result
    except Exception as _e:
        print(f"[bodegas-batch] Parse error: {_e} — raw: {texto[:200]}", flush=True)
        return fallback


# ─── HELPERS DE VERIFICACIÓN BULK ────────────────────────────────────────────
# Estas funciones permiten que la verificación masiva consulte primero las bases
# locales (bcra_nomdeu.db + cheques_bcra) antes de ir a BCRA live.
# Resultado: 500 clientes → consulta bulk < 2s, live BCRA solo para ~100-150.

_PERIODO_BASE_BULK = 202605  # Mayo 2026 — último período del archivo histórico


def _mes_anterior(yyyymm: int, n: int) -> int:
    """Retrocede N meses desde un período YYYYMM."""
    año, mes = yyyymm // 100, yyyymm % 100
    mes -= n
    while mes <= 0:
        mes += 12
        año -= 1
    return año * 100 + mes


# Cantidad de meses reales guardados en historial_detalle (mes_01 = más
# reciente / período _PERIODO_BASE_BULK). Si en el futuro se reconstruye la
# base con más o menos meses, solo hay que ajustar esta constante.
_HIST_DETALLE_MESES = 12


def _historial_detalle_rows(cuit: str) -> list:
    """Filas crudas de historial_detalle para un CUIT — una por entidad
    financiera, con situación y monto reales de cada uno de los
    _HIST_DETALLE_MESES meses (None donde no hay dato ese mes)."""
    if _nomdeu_conn is None:
        return []
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    if not cuit_limpio:
        return []
    try:
        cur = _nomdeu_conn.execute(
            "SELECT * FROM historial_detalle WHERE cuit = ?", (cuit_limpio,)
        )
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]
    except Exception:
        return []


def _nomdeu_agregar_filas(filas: list) -> dict:
    """
    Agrega las filas de historial_detalle de un CUIT (una por entidad) a un
    resumen compatible con el resto del motor de scoring.

    - sit_max / meses_en_mora / monto_max: peor situación, cantidad total de
      meses en mora (situación > 1) y monto máximo vistos en CUALQUIER
      entidad durante los _HIST_DETALLE_MESES meses reales — ya no son
      estimados ni distribuidos sintéticamente, son los datos tal cual
      vinieron del archivo 24DSF del BCRA.
    - sit_padron / monto_total / entidades_cod: snapshot del mes más
      reciente (mes_01) únicamente — reemplaza a deudas_resumen.
    """
    sit_max = 1
    meses_en_mora = 0
    monto_max = 0.0
    sit_actual = 1
    monto_actual_total = 0.0
    ent_codigos_actuales = []

    for fila in filas:
        # BCRA usa el dígito "0" (no NULL/blanco) para marcar "sin información
        # ese mes en esa entidad" — no es una situación válida (los códigos
        # reales son 1-5 y 11), así que se descarta igual que None.
        sit_01 = fila.get('sit_01')
        if sit_01:
            ent_codigos_actuales.append(fila['entidad'])
            monto_actual_total += (fila.get('monto_01') or 0) / 10.0
            if sit_01 > sit_actual:
                sit_actual = sit_01
        for i in range(1, _HIST_DETALLE_MESES + 1):
            sit_i = fila.get(f'sit_{i:02d}')
            if not sit_i:
                continue
            if sit_i > sit_max:
                sit_max = sit_i
            if sit_i > 1:
                meses_en_mora += 1
            monto_i = (fila.get(f'monto_{i:02d}') or 0) / 10.0
            if monto_i > monto_max:
                monto_max = monto_i

    return {
        'sit_max':       max(sit_max, sit_actual),
        'monto_total':   round(monto_actual_total, 1),
        'entidades_cod': ','.join(ent_codigos_actuales),
        'periodo':       str(_PERIODO_BASE_BULK),
        'sit_padron':    sit_actual,
        'sit_hist_12m':  sit_max,
        'meses_en_mora': meses_en_mora,
        'monto_max':     round(monto_max, 1),
    }


def _nomdeu_batch(cuits: list) -> dict:
    """
    Consulta masiva de historial_detalle para N CUITs en una sola query.
    Mucho más eficiente que N llamadas individuales a _nomdeu_get_deuda().
    Retorna {cuit: {'sit_max': N, 'meses_en_mora': N, 'monto_max': N, ...}}.
    """
    if _nomdeu_conn is None or not cuits:
        return {}
    cuits_norm = [str(c).replace('-', '').replace(' ', '').strip() for c in cuits]
    cuits_norm = [c for c in cuits_norm if c]
    if not cuits_norm:
        return {}
    try:
        placeholders = ','.join('?' * len(cuits_norm))
        cur = _nomdeu_conn.execute(
            f"SELECT * FROM historial_detalle WHERE cuit IN ({placeholders})", cuits_norm
        )
        cols = [d[0] for d in cur.description]
        filas_por_cuit: dict = {}
        for row in cur.fetchall():
            fila = dict(zip(cols, row))
            filas_por_cuit.setdefault(fila['cuit'], []).append(fila)
    except Exception as _e:
        print(f"[bulk_batch] historial_detalle: {_e}", flush=True)
        return {}
    return {c: _nomdeu_agregar_filas(filas) for c, filas in filas_por_cuit.items()}


# ── Módulo MiPyME — padrón de empresas PyME inscriptas (Min. Producción) ─────

def _mipyme_get(cuit: str):
    """Ficha MiPyME para un CUIT. Retorna dict con categoria/sector/empleados o None."""
    if _mipyme_conn is None:
        return None
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    try:
        row = _mipyme_conn.execute(
            "SELECT razon_social, categoria, sector, provincia "
            "FROM mipyme_padron WHERE cuit = ?", (cuit_limpio,)
        ).fetchone()
        if not row:
            return None
        cat = row[1] or ''
        return {
            'razon_social':   row[0],
            'categoria':      cat,
            'sector':         row[2],
            'provincia':      row[3],
            'empleados_rango': _EMPLEADOS_RANGO.get(cat),
        }
    except Exception:
        return None


def _mipyme_batch(cuits: list) -> dict:
    """Consulta masiva del padrón MiPyME para N CUITs. Retorna {cuit: ficha_dict}."""
    if _mipyme_conn is None or not cuits:
        return {}
    cuits = [str(c).replace('-', '').replace(' ', '').strip() for c in cuits]
    cuits = [c for c in cuits if c]
    if not cuits:
        return {}
    try:
        placeholders = ','.join('?' * len(cuits))
        rows = _mipyme_conn.execute(
            f"SELECT cuit, razon_social, categoria, sector, provincia "
            f"FROM mipyme_padron WHERE cuit IN ({placeholders})", cuits
        ).fetchall()
        result = {}
        for row in rows:
            cat = row[2] or ''
            result[str(row[0])] = {
                'razon_social':   row[1],
                'categoria':      cat,
                'sector':         row[3],
                'provincia':      row[4],
                'empleados_rango': _EMPLEADOS_RANGO.get(cat),
            }
        return result
    except Exception as e:
        print(f"[mipyme_batch] {e}", flush=True)
        return {}


def _mipyme_tope(sector: str, categoria: str) -> int:
    """Tope de facturación anual (ARS) según sector y categoría. 0 si no se encuentra."""
    if not sector or not categoria:
        return 0
    # Normalización defensiva: acentos y variantes tipográficas del CSV oficial
    _SEC_ALIAS = {
        'industria y mineria':  'Industria y Minería',
        'industria y minería':  'Industria y Minería',
        'construccion':         'Construcción',
        'construcción':         'Construcción',
        'comercio':             'Comercio',
        'servicios':            'Servicios',
        'agropecuario':         'Agropecuario',
    }
    sec_norm = _SEC_ALIAS.get(sector.lower().strip(), sector.strip())
    cat_norm = categoria.strip()
    return (TOPES_FACTURACION_ANUAL.get(sec_norm) or {}).get(cat_norm, 0)


def _mipyme_empleados_rango(categoria: str):
    """Rango de empleados estimado según categoría MiPyME. Retorna str o None."""
    if not categoria:
        return None
    return _EMPLEADOS_RANGO.get(str(categoria).strip())


def _cheques_local_batch(cuits: list) -> dict:
    """
    Consulta masiva de cheques rechazados para N CUITs en una sola SELECT.
    Retorna {cuit: cheq_dict} en formato compatible con calcular_rating_predictivo.
    """
    if not os.path.exists(PADRON_DB_PATH) or not cuits:
        return {}
    # Normalizar a 11 dígitos sin guiones — formato exacto de cheques_bcra
    cuits = [str(c).replace('-', '').replace(' ', '').strip() for c in cuits]
    cuits = [c for c in cuits if c]
    if not cuits:
        return {}
    try:
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        try:
            meta = conn.execute(
                "SELECT valor FROM _cheques_meta WHERE key = 'last_import_date'"
            ).fetchone()
        except Exception:
            conn.close(); return {}
        if not meta:
            conn.close(); return {}
        placeholders = ','.join('?' * len(cuits))
        rows = conn.execute(
            f"SELECT cuit, nro_cheque, fecha_rechazo, monto, estado, tipo, cuit_entidad "
            f"FROM cheques_bcra WHERE cuit IN ({placeholders}) "
            f"ORDER BY cuit, fecha_rechazo DESC", cuits
        ).fetchall()
        conn.close()
        by_cuit: dict = {}
        for row in rows:
            c = row[0]
            if c not in by_cuit:
                by_cuit[c] = []
            estado_s = (row[4] or '').strip()
            if estado_s == 'IMPAGA':
                fecha_pago, estado_multa = None, 'IMPAGA'
            else:
                try:
                    d, m, y = estado_s.split('/')
                    fecha_pago = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
                except Exception:
                    fecha_pago = estado_s or None
                estado_multa = 'LEVANTADA'
            fecha_r = row[2]
            if fecha_r and len(fecha_r) == 8:
                fecha_r = f"{fecha_r[:4]}-{fecha_r[4:6]}-{fecha_r[6:8]}"
            by_cuit[c].append({
                'numeroCheque': (row[1] or '').strip(), 'fechaRechazo': fecha_r,
                'monto': row[3], 'fechaPago': fecha_pago, 'estadoMulta': estado_multa,
                'fechaLevantamiento': fecha_pago,
                'tipo': (row[5] or '').strip(), 'cuitEntidad': (row[6] or '').strip(),
            })
        result = {}
        for c, detalles in by_cuit.items():
            result[c] = {
                'results': {'causales': [{'entidades': [{'detalle': detalles}]}]},
                'sin_deudas': False, 'error_bcra': None, 'source': 'local_db_batch',
            }
        for c in cuits:
            if c not in result:
                result[c] = {
                    'results': {'causales': []},
                    'sin_deudas': True, 'error_bcra': None, 'source': 'local_db_batch',
                }
        return result
    except Exception as e:
        print(f"[cheques_batch] Error: {e}", flush=True)
        return {}


def _bulk_bcra_from_historial(cuit: str, nombre: str = '') -> dict:
    """BCRA data del periodo más reciente desde historial_detalle (datos reales per-entidad).
    Usa sit_01/monto_01 de cada entidad — misma fuente que _bulk_to_hist_data() pero
    devuelve solo el periodo actual en formato compatible con /deudas/{cuit}.
    A diferencia de _bulk_to_bcra_data(), NO distribuye sit_max a todas las entidades:
    cada entidad tiene su situación y monto real del último periodo del bulk."""
    filas = _historial_detalle_rows(cuit)
    if not filas:
        return {}
    entidades = []
    for fila in filas:
        sit = fila.get('sit_01')
        if not sit:
            continue
        monto = (fila.get('monto_01') or 0) / 10.0
        entidades.append({
            'entidad':   _nomdeu_get_entidad(fila['entidad']) or f"Entidad {fila['entidad']}",
            'situacion': sit,
            'monto':     round(monto, 1),
        })
    if not entidades:
        return {}
    periodo = _mes_anterior(_PERIODO_BASE_BULK, 0)
    denom = _denominacion_local(cuit) or nombre or ''
    return {
        'results': {
            'denominacion': denom,
            'periodos': [{'periodo': periodo, 'entidades': entidades}],
        },
        'sin_deudas': all(e['situacion'] <= 1 for e in entidades),
        'bcra_disponible': False,
        'fuente_offline': 'historial_detalle',
    }


def _bulk_to_bcra_data(nombre: str, deuda: dict) -> dict:
    """Respuesta BCRA sintética desde datos del padrón offline."""
    sit    = deuda.get('sit_max', 1)
    monto  = deuda.get('monto_total') or deuda.get('monto_max') or 0
    ent_cods = [c.strip() for c in (deuda.get('entidades_cod') or '').split(',') if c.strip()]
    n_ents = max(1, len(ent_cods)) if ent_cods else 1
    ents = []
    for cod in ent_cods:
        ents.append({'entidad': _nomdeu_get_entidad(cod) or f"Entidad {cod}",
                     'situacion': sit, 'monto': round(monto / n_ents, 1)})
    if not ents:
        ents = [{'entidad': 'Padron_BCRA', 'situacion': sit, 'monto': float(monto)}]
    periodo_str = str(deuda.get('periodo') or _PERIODO_BASE_BULK)
    periodo_int = int(periodo_str) if periodo_str.isdigit() else _PERIODO_BASE_BULK
    return {
        'results': {'denominacion': nombre or '', 'periodos': [{'periodo': periodo_int, 'entidades': ents}]},
        'sin_deudas': sit <= 1, 'bcra_disponible': False, 'fuente_offline': 'bcra_nomdeu_local',
    }


def _bulk_to_hist_data(cuit: str) -> dict:
    """
    Historial REAL de los últimos _HIST_DETALLE_MESES meses, leído directo
    de historial_detalle: situación y monto reales de cada entidad en cada
    mes, tal cual figuran en el archivo 24DSF del BCRA — ya no son
    estimados ni distribuidos sintéticamente en los períodos más recientes.
    """
    filas = _historial_detalle_rows(cuit)
    if not filas:
        return {
            'results': {'denominacion': '', 'periodos': []},
            'sin_deudas': True, 'fuente_offline': 'historial_detalle',
        }

    periodos = []
    sit_max_total = 1
    for i in range(1, _HIST_DETALLE_MESES + 1):
        periodo = _mes_anterior(_PERIODO_BASE_BULK, i - 1)
        entidades = []
        for fila in filas:
            # "0" (no None) marca "sin informacion ese mes" — no es una
            # situacion valida, se descarta igual que en _nomdeu_agregar_filas.
            sit = fila.get(f'sit_{i:02d}')
            if not sit:
                continue
            monto = (fila.get(f'monto_{i:02d}') or 0) / 10.0
            entidades.append({
                'entidad': _nomdeu_get_entidad(fila['entidad']) or f"Entidad {fila['entidad']}",
                'situacion': sit,
                'monto': round(monto, 1),
            })
            if sit > sit_max_total:
                sit_max_total = sit
        if entidades:
            periodos.append({'periodo': periodo, 'entidades': entidades})

    return {
        'results': {'denominacion': '', 'periodos': periodos},
        'sin_deudas': sit_max_total <= 1, 'fuente_offline': 'historial_detalle',
    }


def _clasificar_bulk(deuda) -> str:
    """
    Clasifica un cliente según datos bulk para decidir si necesita BCRA live.
    'alto_riesgo' — riesgo confirmado por bulk (sit≥4 o sit=3 ≥2 meses) → no ir live
    'zona_gris'   — mora ambigua (sit=2-3, pocos meses) → confirmar live
    'limpio_bulk' — figura en padrón pero sit=1 actual → no ir live
    'nuevo'       — ausente del bulk completamente → ir live (puede ser limpio o nuevo deudor)
    """
    if deuda is None:
        return 'nuevo'
    sit   = deuda.get('sit_max', 1)
    meses = int(deuda.get('meses_en_mora') or 0)
    if sit >= 4:
        return 'alto_riesgo'          # Concurso/quiebra — irrecuperable
    if sit == 3 and meses >= 2:
        return 'alto_riesgo'          # Mora grave sostenida
    if sit >= 2:
        return 'zona_gris'            # Mora técnica o leve — confirmar
    return 'limpio_bulk'              # Presente en padrón, sit=1


def ejecutar_verificacion(cartera_data):
    global verificacion_estado
    _score_session_cache.clear()   # reset session cache para esta verificación
    verificacion_estado["corriendo"] = True
    verificacion_estado["progreso"] = 0
    verificacion_estado["total"] = len(cartera_data)
    verificacion_estado["mensaje"] = "Iniciando verificacion..."

    # Purgar solo entradas del padrón > 72h — purgar TODAS provocaba rate-limit en cascada
    # durante Phase 1 (500 clientes × 12 workers todos fresheando simultáneamente).
    # Datos de las últimas 72h son suficientemente frescos para detectar deterioro.
    _cartera_cuits_v = {
        str(c.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
        for c in cartera_data if isinstance(c, dict)
    }
    try:
        if os.path.exists(PADRON_DB_PATH) and _cartera_cuits_v:
            conn_p = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
            placeholders = ','.join('?' * len(_cartera_cuits_v))
            cur_p = conn_p.execute(
                f"""DELETE FROM bcra_padron_local
                    WHERE cuit IN ({placeholders})
                    AND importado_en < datetime('now', '-72 hours')""",
                list(_cartera_cuits_v)
            )
            conn_p.commit(); conn_p.close()
            print(f"[verif] Padrón local purgado: {cur_p.rowcount} entradas > 72h eliminadas", flush=True)
    except Exception as _ep:
        print(f"[verif] Advertencia purga padrón: {_ep}", flush=True)
    # NO se invalida bcra_cache — misma razón que proceso_integral:
    # la invalidación masiva provoca rate-limiting en BCRA desde el cliente 4.

    palabras_riesgo = [
        'rechaz', 'no paga', 'cuidado', 'mora', 'deuda', 'incobrable',
        'estafa', 'desapareci', 'fuga', 'impago', 'quiebra', 'concurso',
        'sin fondos', 'rebotado', 'mal pagador', 'no responde', 'no contesta',
        'bloqueado', 'vencid', 'no cancel', 'no liquido', 'no abono',
        'atencion', 'ojo', 'problema', 'judicial', 'cobrar', 'nos debe', 'debia'
    ]

    wsp_index = {}
    try:
        with open(WSP_FILE, 'r', encoding='utf-8') as f:
            wsp_index = json.load(f)
    except Exception:
        pass
    if not isinstance(wsp_index, dict):
        wsp_index = {}

    nuevas_alertas = []
    cartera_actualizada = []

    # ── Pre-poblar alertas_cartera.json con stubs para todos los clientes ─────
    # Esto garantiza que la App Comercial vea todos los clientes desde el inicio,
    # con scores llenándose a medida que el robot avanza.
    def _nc_v(x):
        return str(x or '').replace('-', '').replace(' ', '').strip()

    try:
        stubs = [{
            "cuit": c.get("cuit"), "nombre": c.get("nombre", ""),
            "ultimaSit": c.get("ultimaSit", 1), "ultimaVerif": None,
            "scoreCompleto": None, "scoreRango": None, "scoreColor": None, "scoreEmoji": None,
            "pendiente": True,
        } for c in cartera_data if c.get("cuit")]
        with open(ALERTAS_FILE, 'w', encoding='utf-8') as _f:
            json.dump({
                "alertas": [],
                "ultima_verif": f"En progreso — {time.strftime('%d/%m/%Y %H:%M')}",
                "cartera": stubs,
            }, _f, ensure_ascii=False, indent=2)
        print(f"[verif] Pre-poblado: {len(stubs)} stubs guardados en {ALERTAS_FILE}", flush=True)
    except Exception as _ep:
        print(f"[verif] Error pre-poblar: {_ep}", flush=True)

    try:
        cache_file = os.path.join(DATA_DIR, 'bcra_cache.json')
        if os.path.exists(cache_file):
            os.remove(cache_file)
            print("[verif] Caché BCRA limpiado para verificación fresca", flush=True)
    except: pass
    # Eliminar solo caché de historial/cheques STALE (> 24h) — los frescos se reutilizan
    try:
        import glob as _glob
        ahora_v = time.time()
        elim = 0
        for _fp in _glob.glob(os.path.join(DATA_DIR, 'historial_*.json')) + \
                   _glob.glob(os.path.join(DATA_DIR, 'cheques_*.json')):
            try:
                with open(_fp, 'r') as _ff:
                    _ts = json.load(_ff).get('ts', 0)
                if ahora_v - _ts > 86400:
                    os.remove(_fp); elim += 1
            except:
                try: os.remove(_fp); elim += 1
                except: pass
        print(f"[verif] Caché stale eliminado: {elim} archivos (frescos conservados)", flush=True)
    except: pass

    def _guardar_alertas(nuevas_alertas, cartera_actualizada, parcial=False):
        sufijo = ' (parcial)' if parcial else ''

        def _entry(c):
            return {
                "cuit":                 c.get('cuit'),
                "nombre":               c.get('nombre', ''),
                "ciudad":               c.get('ciudad', ''),
                "ultimaSit":            c.get('ultimaSit'),
                "ultimaVerif":          c.get('ultimaVerif'),
                "scoreCompleto":        c.get('scoreCompleto'),
                "scoreRango":           c.get('scoreRango'),
                "scoreColor":           c.get('scoreColor'),
                "scoreEmoji":           c.get('scoreEmoji'),
                "alerta_temprana":      c.get('alerta_temprana', False),
                "bloquear_oportunidad": c.get('bloquear_oportunidad', False),
                "alerta_logistica":     c.get('alerta_logistica', ''),
                "inferencia_ingresos":  c.get('inferencia_ingresos'),
                "fuente_ingresos":      c.get('fuente_ingresos'),
                "actividad_principal":  c.get('actividad_principal'),
                "score_ts":             c.get('score_ts', 0),
                "pendiente":            c.get('pendiente', False),
            }

        if parcial and os.path.exists(ALERTAS_FILE):
            # Merge: preservar stubs no procesados aún, actualizar los ya procesados
            try:
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as _fr:
                    existente = json.load(_fr)
                base = {_nc_v(c.get('cuit', '')): _entry(c) for c in existente.get('cartera', [])}
                for c in cartera_actualizada:
                    nc = _nc_v(c.get('cuit', ''))
                    if nc:
                        base[nc] = _entry(c)
                cartera_final = list(base.values())
            except Exception:
                cartera_final = [_entry(c) for c in cartera_actualizada]
        else:
            cartera_final = [_entry(c) for c in cartera_actualizada]

        datos = {
            "motor_version": _MOTOR_VERSION_CARTERA,
            "alertas": nuevas_alertas,
            "ultima_verif": time.strftime('%d/%m/%Y %H:%M') + sufijo,
            "cartera": cartera_final,
        }
        with open(ALERTAS_FILE, 'w', encoding='utf-8') as _f:
            json.dump(datos, _f, ensure_ascii=False, indent=None if parcial else 2)

    total = len(cartera_data)
    try:
        # ═══════════════════════════════════════════════════════════════════
        # FASE 0 — Consulta bulk local instantánea (< 2s para 500+ clientes)
        # Fuentes: bcra_nomdeu.db (historial 24m + padrón) + cheques_bcra SQLite
        # Clasifica cada cliente: alto_riesgo | zona_gris | limpio_bulk | nuevo
        # ═══════════════════════════════════════════════════════════════════
        verificacion_estado["mensaje"] = "Fase 0: Consultando bases locales (bulk offline)..."
        print(f"[verif] FASE 0: Bulk batch — {total} CUITs...", flush=True)
        _nc_v2 = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()
        _cuits_lista = [_nc_v2(c.get('cuit')) for c in cartera_data if c.get('cuit')]
        _cuits_lista = [c for c in _cuits_lista if c]
        _t0_bulk = time.time()
        _bulk_deudas  = _nomdeu_batch(_cuits_lista)
        _bulk_cheques = _cheques_local_batch(_cuits_lista)
        _bulk_mipyme  = _mipyme_batch(_cuits_lista)   # padrón PyME: categoría + sector
        # Pre-calcular datos sintéticos por cliente
        bulk_prefetch = {}  # {cuit: {deuda, cheques, categoria, bcra_bulk, hist_bulk, mipyme}}
        _cat_count = {'alto_riesgo': 0, 'zona_gris': 0, 'limpio_bulk': 0, 'nuevo': 0}
        for _c0 in cartera_data:
            _cuit0 = _nc_v2(_c0.get('cuit', ''))
            if not _cuit0:
                continue
            _deuda0  = _bulk_deudas.get(_cuit0)
            _cheq0   = _bulk_cheques.get(_cuit0)
            _mipyme0 = _bulk_mipyme.get(_cuit0)
            _cat0    = _clasificar_bulk(_deuda0)
            _cat_count[_cat0] = _cat_count.get(_cat0, 0) + 1
            _nom0    = str(_c0.get('nombre', '') or '') or (_denominacion_local(_cuit0) or _cuit0)
            bulk_prefetch[_cuit0] = {
                'deuda':    _deuda0,
                'cheques':  _cheq0,
                'mipyme':   _mipyme0,
                'categoria': _cat0,
                'bcra_bulk': _bulk_to_bcra_data(_nom0, _deuda0) if _deuda0 else None,
                'hist_bulk': _bulk_to_hist_data(_cuit0) if _deuda0 else None,
            }
        print(
            f"[verif] FASE 0 OK ({time.time()-_t0_bulk:.1f}s) — "
            f"alto_riesgo={_cat_count['alto_riesgo']} | zona_gris={_cat_count['zona_gris']} | "
            f"limpio_bulk={_cat_count['limpio_bulk']} | nuevo={_cat_count['nuevo']}",
            flush=True,
        )

        # ═══════════════════════════════════════════════════════════════════
        # FASE 1 — Live BCRA SOLO para clientes que necesitan datos frescos
        # Prioridad 1: zona_gris  — mora ambigua en bulk → confirmar
        # Prioridad 2: nuevo      — sin antecedentes bulk → puede estar en mora
        # Prioridad 3: limpio_bulk con deuda interna >90d — señal Odoo indica
        #              riesgo aunque el bulk (puede tener 2 meses de antigüedad)
        #              los marque como limpios. Triage basado en datos propios.
        # Cap: LIVE_BCRA_MAX (default 150) — controla rate-limiting BCRA
        # alto_riesgo: confirmado por bulk, no requiere BCRA live
        # ═══════════════════════════════════════════════════════════════════
        _LIVE_BCRA_MAX = int(os.environ.get('LIVE_BCRA_MAX', '150'))
        _zona_gris_v = [c for c in cartera_data
                        if bulk_prefetch.get(str(c.get('cuit','')).strip(), {}).get('categoria') == 'zona_gris']
        _nuevos_v    = [c for c in cartera_data
                        if bulk_prefetch.get(str(c.get('cuit','')).strip(), {}).get('categoria') == 'nuevo']

        # ── Triage limpio_bulk: detectar los que tienen señales internas de riesgo ──
        # Un cliente con deuda interna vencida >90 días ya está dando señales antes
        # que BCRA lo registre. Esos van al check en vivo aunque el bulk los marque sit=1.
        _saldos_idx_nb: dict = {}   # nombre_normalizado → [facturas]
        for _sf_t in (_saldos_facturas or []):
            _nb_t = str(_sf_t.get('cliente') or '').upper().strip()
            if _nb_t:
                _saldos_idx_nb.setdefault(_nb_t, []).append(_sf_t)

        _FMTS_T = ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y')
        def _parse_f(s):
            if not s: return None
            for _fmt in _FMTS_T:
                try: return datetime.strptime(str(s)[:10], _fmt)
                except Exception: pass
            return None

        _hoy_t = datetime.now()
        _limpio_sospechosos = []
        for _c_t in cartera_data:
            _cuit_t = str(_c_t.get('cuit', '') or '').strip()
            if not _cuit_t: continue
            if bulk_prefetch.get(_cuit_t, {}).get('categoria') != 'limpio_bulk': continue
            _nb_t2 = str(_c_t.get('nombre') or _c_t.get('cliente') or '').upper().strip()
            _facts_t = _saldos_idx_nb.get(_nb_t2, [])
            _tiene_90d = any(
                _parse_f(_f_t.get('fechaPago')) and
                float(_f_t.get('saldo') or 0) > 0 and
                (_hoy_t - _parse_f(_f_t.get('fechaPago'))).days > 90
                for _f_t in _facts_t
                if _parse_f(_f_t.get('fechaPago'))
            )
            if _tiene_90d:
                _limpio_sospechosos.append(_c_t)

        if _limpio_sospechosos:
            print(
                f"[verif] Triage DSO: {len(_limpio_sospechosos)} limpio_bulk con deuda >90d → agregar a live",
                flush=True
            )

        # Orden de prioridad: zona_gris → nuevo → limpio sospechoso por señal interna
        _para_live = _zona_gris_v + _nuevos_v + _limpio_sospechosos
        if len(_para_live) > _LIVE_BCRA_MAX:
            _para_live = _para_live[:_LIVE_BCRA_MAX]
            print(f"[verif] FASE 1: limitado a {_LIVE_BCRA_MAX} clientes (LIVE_BCRA_MAX)", flush=True)

        bcra_prefetch = {}  # {cuit: (lambda_result, bcra_data, cheques_data)}

        if _para_live:
            def _fetch_cliente_bcra(cliente_f):
                cuit_f = str(cliente_f.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
                try:
                    lr = consultar_bcra_lambda(cuit_f)
                    if lr:
                        return cuit_f, (lr, lr[0], lr[2])
                    bd, _ = consultar_bcra_cached(cuit_f)
                    cheques_d = None
                    if bd and bd.get('results') is not None and not bd.get('error_bcra'):
                        try:
                            cheques_d, _ = _consultar_bcra_directo(
                                cuit_f, 'cheques', timeout_per_req=8, max_intentos=1
                            )
                        except Exception:
                            pass
                    return cuit_f, (None, bd, cheques_d)
                except Exception as _ef:
                    print(f"[verif-p1] {cuit_f} error: {type(_ef).__name__}", flush=True)
                    return cuit_f, (None, None, None)

            _N_WORKERS = 8
            _n_live = len(_para_live)
            verificacion_estado["mensaje"] = f"Fase 1/3: BCRA live para {_n_live} clientes ({_N_WORKERS} workers)..."
            print(f"[verif] FASE 1: Live BCRA — {_n_live}/{total} clientes (zona_gris + nuevo)", flush=True)
            with ThreadPoolExecutor(max_workers=_N_WORKERS) as _pool:
                _futures = {_pool.submit(_fetch_cliente_bcra, c): c for c in _para_live}
                _done = 0
                for _fut in as_completed(_futures, timeout=1500):
                    try:
                        _cuit_r, _data_r = _fut.result(timeout=40)
                        bcra_prefetch[_cuit_r] = _data_r
                    except Exception:
                        _c = _futures[_fut]
                        bcra_prefetch[str(_c.get('cuit', '')).strip()] = (None, None, None)
                    _done += 1
                    verificacion_estado["progreso"] = _done
                    if _done % 20 == 0:
                        print(f"[verif-p1] {_done}/{_n_live} live fetched", flush=True)
            _con_live   = sum(1 for v in bcra_prefetch.values() if v and v[1] and v[1].get('results') is not None and not v[1].get('error_bcra'))
            _con_cheq_l = sum(1 for v in bcra_prefetch.values() if v and v[2])
            print(f"[verif] FASE 1 OK — {_con_live}/{_n_live} con BCRA live · {_con_cheq_l} con cheques live", flush=True)
        else:
            print("[verif] FASE 1: Sin clientes para live BCRA — todo resuelto por bulk", flush=True)

        # ═══════════════════════════════════════════════════════════════════
        # FASE 2 — Análisis bodegas en LOTES (8 clientes por llamada Gemini)
        # Reduce llamadas IA de N individuales a ceil(N/8) lotes
        # ═══════════════════════════════════════════════════════════════════
        verificacion_estado["mensaje"] = "Fase 2/3: Analizando bodegas en lotes..."
        _BATCH = 8
        bodegas_prefetch = {}  # {cuit: (es_negativo, motivo)}
        clientes_para_bodegas = []
        hace_6m = datetime.now() - timedelta(days=180)
        for _cli in cartera_data:
            _cuit_b  = str(_cli.get('cuit', '') or '').strip()
            _nom_b   = str(_cli.get('nombre', '') or '').strip()
            _threads = wsp_index.get(_cuit_b, [])
            _trec = []
            for _t in _threads:
                _fs = _t.get('fecha') or (_t.get('mensajes', [{}])[0].get('fecha') if _t.get('mensajes') else None)
                if _fs:
                    try:
                        if datetime.fromisoformat(str(_fs)[:10]) >= hace_6m:
                            _trec.append(_t)
                    except Exception:
                        pass
            if _trec:
                _tmsgs, _sosp = [], False
                for _t in _trec:
                    for _m in _t.get('mensajes', []):
                        _txt = _m.get('texto', '')
                        _tmsgs.append(_m.get('autor', '') + ': ' + _txt)
                        if any(_p in _txt.lower() for _p in palabras_riesgo):
                            _sosp = True
                if _sosp:
                    clientes_para_bodegas.append({'cuit': _cuit_b, 'nombre': _nom_b, 'mensajes': _tmsgs})
        if clientes_para_bodegas:
            _n_lotes = (len(clientes_para_bodegas) + _BATCH - 1) // _BATCH
            print(f"[verif] FASE 2: {len(clientes_para_bodegas)} clientes sospechosos → {_n_lotes} lote(s) de {_BATCH}", flush=True)
            for _idx_l in range(0, len(clientes_para_bodegas), _BATCH):
                _lote = clientes_para_bodegas[_idx_l:_idx_l + _BATCH]
                try:
                    bodegas_prefetch.update(_analizar_bodegas_batch(_lote))
                    print(f"[verif-p2] Lote {_idx_l // _BATCH + 1}/{_n_lotes} OK", flush=True)
                except Exception as _e_lote:
                    print(f"[verif-p2] Error lote: {_e_lote}", flush=True)
                    for _cl in _lote:
                        bodegas_prefetch[_cl['cuit']] = (False, "")
        else:
            print("[verif] FASE 2: Sin mensajes sospechosos — skip", flush=True)
        print(f"[verif] FASE 2 OK — {len(bodegas_prefetch)} resultados bodegas pre-cacheados", flush=True)

        # ═══════════════════════════════════════════════════════════════════
        # FASE 3 — Calcular scores con datos pre-fetched (sin I/O BCRA bloqueante)
        # ═══════════════════════════════════════════════════════════════════
        verificacion_estado["progreso"] = 0
        verificacion_estado["mensaje"] = f"Fase 3/3: Calculando scores ({total} clientes)..."
        print(f"[verif] FASE 3: Scoring con datos pre-fetched...", flush=True)

        for i, cliente in enumerate(cartera_data):
            cuit         = str(cliente.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
            nombre       = str(cliente.get('nombre', '') or '').strip()
            sit_anterior = cliente.get('ultimaSit', 1) or 1
            tag          = f"[verif {i+1}/{total} {cuit}]"

            verificacion_estado["progreso"]       = i + 1
            verificacion_estado["cliente_actual"] = nombre
            verificacion_estado["mensaje"]        = f"Fase 3/3: Score {i+1}/{total}: {nombre}"

            cliente_actualizado = dict(cliente)

            # Recuperar datos: live BCRA (Fase 1) con fallback a bulk offline (Fase 0)
            _pf            = bcra_prefetch.get(cuit, (None, None, None))
            lambda_result  = _pf[0] if _pf else None
            bcra_data      = _pf[1] if _pf else None
            _cheq_prefetch = _pf[2] if (_pf and len(_pf) > 2) else None
            _bp            = bulk_prefetch.get(cuit, {})
            _cat_c         = _bp.get('categoria', 'nuevo')
            _hist_para_score = None   # hist_data que se pasa al motor de scoring
            _uso_bulk      = False

            # Si no hay datos live válidos, usar datos bulk sintéticos de Fase 0
            _live_ok = bool(bcra_data and bcra_data.get('results') is not None and not bcra_data.get('error_bcra'))
            if not _live_ok:
                _bcra_bulk_data = _bp.get('bcra_bulk')
                if _bcra_bulk_data:
                    bcra_data       = _bcra_bulk_data
                    _hist_para_score = _bp.get('hist_bulk')
                    _uso_bulk       = True
                    print(f"{tag} usando datos bulk [{_cat_c}]", flush=True)

            # Cheques: live (Fase 1) → bulk local batch (Fase 0) → disco → sqlite individual
            if not _cheq_prefetch:
                _cheq_prefetch = _bp.get('cheques')  # del batch de Fase 0

            bcra_ok = bool(
                bcra_data
                and bcra_data.get('results') is not None
                and not bcra_data.get('error_bcra')
            )

            # Persistir caché historial/cheques si vienen de Fase 1 (Lambda o BCRA directo).
            # NO escribir payload null — "envenena" el disco y bloquea el fallback local.
            if lambda_result:
                try:
                    if lambda_result[1]:
                        with open(os.path.join(DATA_DIR, f'historial_{cuit}.json'), 'w') as _f:
                            json.dump({'payload': lambda_result[1], 'ts': time.time()}, _f)
                except Exception as _e:
                    print(f"{tag} Advertencia caché historial: {_e}", flush=True)
            if _cheq_prefetch:  # persiste cheques sin importar la fuente (Lambda o BCRA directo)
                try:
                    with open(os.path.join(DATA_DIR, f'cheques_{cuit}.json'), 'w') as _f:
                        json.dump({'payload': _cheq_prefetch, 'ts': time.time()}, _f)
                except Exception as _e:
                    print(f"{tag} Advertencia caché cheques: {_e}", flush=True)

            # Score — datos ya en memoria (live de Fase 1 o bulk sintético de Fase 0)
            score_data = None
            _ciudad = str(cliente.get('ciudad', '') or '')
            try:
                if lambda_result and not _uso_bulk:
                    # Lambda: trae bcra+hist+cheques en una sola llamada — máxima calidad
                    score_data = calcular_rating_predictivo(
                        cuit=cuit, bcra_data=bcra_data or {},
                        hist_data=lambda_result[1], cheq_data=_cheq_prefetch,
                        en_mora=None, ciudad=_ciudad,
                    )
                elif bcra_data and not _uso_bulk:
                    # BCRA live directo (sin Lambda): calcular_score_servidor carga hist del disco
                    score_data = calcular_score_servidor(
                        cuit, bcra_data, en_mora=None, ciudad=_ciudad
                    )
                elif _uso_bulk:
                    # Bulk offline: bcra_data sintético + hist 24m sintético (ningún call de red)
                    score_data = calcular_rating_predictivo(
                        cuit=cuit, bcra_data=bcra_data or {},
                        hist_data=_hist_para_score, cheq_data=_cheq_prefetch,
                        en_mora=None, ciudad=_ciudad,
                    )
                    if score_data:
                        score_data['fuente_score'] = f'bulk_{_cat_c}'
                if score_data:
                    cliente_actualizado['scoreCompleto']        = score_data['score']
                    cliente_actualizado['scoreRango']           = score_data['rango']
                    cliente_actualizado['scoreColor']           = score_data['color']
                    cliente_actualizado['scoreEmoji']           = score_data['emoji']
                    cliente_actualizado['alerta_temprana']      = score_data.get('alerta_temprana', False)
                    cliente_actualizado['bloquear_oportunidad'] = score_data.get('bloquear_oportunidad', False)
                    cliente_actualizado['alerta_logistica']     = score_data.get('alerta_logistica', '')
                    _sv = get_solvency_data(cuit)
                    if _sv:
                        cliente_actualizado['inferencia_ingresos'] = _sv.get('ingresos_anuales')
                        cliente_actualizado['fuente_ingresos']     = _sv.get('fuente_ingresos')
                        cliente_actualizado['actividad_principal'] = _sv.get('actividad_principal')
                    cliente_actualizado['score_ts'] = time.time()
                    print(f"{tag} score={score_data['score']}", flush=True)
            except Exception as e_sc:
                print(f"{tag} ERROR score: {type(e_sc).__name__}: {e_sc}", flush=True)

            # Situación BCRA — fuente de verdad en orden de prioridad:
            # 1. score_data['max_sit']: calculado por calcular_rating_predictivo con los mismos
            #    datos BCRA, incluyendo toda la lógica de normalización. Más confiable.
            # 2. Lectura directa de periodos[0] del raw bcra_data (fallback).
            # 3. Sin datos: conservar sit_anterior (marcar como fallida si no hubo fetch).
            max_sit = sit_anterior  # default conservador
            if score_data and score_data.get('max_sit') is not None and bcra_ok:
                # Fuente primaria: max_sit del motor de scoring — solo si BCRA real disponible.
                # Si bcra_ok=False (saturado/error) no sobreescribir ultimaSit con max_sit=1
                # de un score calculado sin datos, que ocultaría el riesgo real del cliente.
                max_sit = score_data['max_sit']
                cliente_actualizado['ultimaSit']   = max_sit
                cliente_actualizado['ultimaVerif'] = time.strftime('%d/%m/%Y')
                print(f"{tag} ultimaSit={max_sit} (desde score_data)", flush=True)
            elif bcra_ok and bcra_data and bcra_data.get('results') is not None:
                # Fuente secundaria: leer periodos directamente del bcra_data
                periodos  = (bcra_data.get('results') or {}).get('periodos') or []
                entidades = periodos[0].get('entidades', []) if periodos else []
                max_sit   = max((e.get('situacion', 1) or 1) for e in entidades) if entidades else 1
                cliente_actualizado['ultimaSit']   = max_sit
                cliente_actualizado['ultimaVerif'] = time.strftime('%d/%m/%Y')
                print(f"{tag} ultimaSit={max_sit} (desde periodos raw)", flush=True)
            else:
                cliente_actualizado['ultimaVerif'] = time.strftime('%d/%m/%Y')
                if not bcra_ok:
                    # Ni live ni bulk disponibles — cliente completamente fuera de cualquier fuente
                    # (muy raro: solo ocurre si _nomdeu_conn es None y BCRA live falló)
                    _deuda_fallback = _bp.get('deuda')
                    if _deuda_fallback:
                        _bulk_sit = _deuda_fallback.get('sit_max') or 1
                        _bulk_per = _deuda_fallback.get('periodo', '')
                        if _bulk_sit >= (sit_anterior or 1):
                            max_sit = _bulk_sit
                            cliente_actualizado['ultimaSit'] = max_sit
                        cliente_actualizado['fuente_sit'] = f'padron_bulk_{_bulk_per}'
                        print(f"{tag} fallback dict bulk: sit={_bulk_sit}", flush=True)
                    else:
                        cliente_actualizado['verificacion_fallida'] = True
                        print(f"{tag} Sin datos BCRA ni bulk — conserva sit_anterior={sit_anterior}", flush=True)

            # Generar alerta si la situación empeoró o es grave
            if max_sit > sit_anterior or max_sit >= 3:
                alerta = {
                    "nombre": nombre, "cuit": cuit,
                    "sitAnterior": sit_anterior, "sitActual": max_sit,
                    "fecha": time.strftime('%d/%m/%Y'), "tipo": "bcra",
                    "fuente": cliente_actualizado.get('fuente_sit', 'bcra_live'),
                }
                if score_data:
                    alerta.update({
                        "scoreCompleto": score_data["score"], "scoreRango": score_data["rango"],
                        "scoreColor": score_data["color"], "scoreEmoji": score_data["emoji"]
                    })
                nuevas_alertas.append(alerta)

            # Detectar cheques rechazados activos — 3 niveles de fallback:
            # 1) Fase 1 prefetch (Lambda o BCRA directo — paralelo, ya resuelto)
            # 2) Disco caché fresco (< 24h)
            # 3) SQLite local (snapshot diario BCRA, cero latencia de red)
            # Fallback 4 (BCRA live en Fase 3) eliminado: causaba 500 llamadas seriales
            # de 12s cada una = 6000s adicionales que mataban la verificación.
            # Los cheques se traen en Fase 1 para todos los clientes con BCRA accesible.
            try:
                _cheq_v = _cheq_prefetch  # Fase 1: Lambda[2] o BCRA directo (paralelo)
                if not _cheq_v:
                    _cheq_path_v = os.path.join(DATA_DIR, f'cheques_{cuit}.json')
                    if os.path.exists(_cheq_path_v):
                        try:
                            with open(_cheq_path_v, 'r', encoding='utf-8') as _cvf:
                                _loaded_cheq = json.load(_cvf).get('payload')
                                if _loaded_cheq:  # no usar null de caché envenenado
                                    _cheq_v = _loaded_cheq
                        except Exception:
                            pass
                _act_live_v, _, _det_live_v = _cheques_activos_de(_cheq_v)
                # Cruzar SIEMPRE contra el bulk local (snapshot diario BCRA), no solo
                # cuando la fuente en vivo/caché está vacía — un "sin_deudas=True" en
                # vivo puede ser un falso negativo, igual que con deudas/historial.
                # Nunca subestimar el riesgo: usar la fuente con más cheques activos.
                _act_bulk_v, _, _det_bulk_v = _cheques_activos_de(get_cheques_local(cuit))
                if _act_bulk_v > _act_live_v:
                    _activos_v, _det_v = _act_bulk_v, _det_bulk_v
                else:
                    _activos_v, _det_v = _act_live_v, _det_live_v
                if _activos_v > 0 and not any(
                    a.get('cuit') == cuit and a.get('tipo') == 'cheque'
                    for a in nuevas_alertas
                ):
                    alerta_ch = {
                        'nombre':       nombre,
                        'cuit':         cuit,
                        'tipo':         'cheque',
                        'nroCheques':   _activos_v,
                        'totalCheques': len(_det_v),
                        'fecha':        time.strftime('%d/%m/%Y'),
                    }
                    if score_data:
                        alerta_ch.update({
                            'scoreCompleto': score_data.get('score'),
                            'scoreRango':    score_data.get('rango'),
                            'scoreColor':    score_data.get('color'),
                            'scoreEmoji':    score_data.get('emoji'),
                        })
                    nuevas_alertas.append(alerta_ch)
                    print(f"{tag} ALERTA CHEQUES: {_activos_v}/{len(_det_v)} cheque(s) activo(s)", flush=True)
            except Exception as _cheq_v_e:
                print(f"{tag} Cheques alert parse fallo: {_cheq_v_e}", flush=True)

            # Bodegas: resultado pre-fetched en Fase 2 (sin llamada IA individual por cliente)
            _es_neg, _motivo = bodegas_prefetch.get(cuit, (False, ""))
            if _es_neg and not any(a['cuit'] == cuit and a['tipo'] == 'bodegas' for a in nuevas_alertas):
                nuevas_alertas.append({"nombre": nombre, "cuit": cuit,
                    "fecha": time.strftime('%d/%m/%Y'), "tipo": "bodegas", "mensajes": [_motivo]})

            cartera_actualizada.append(cliente_actualizado)

            # Guardado parcial cada 10 clientes
            if (i + 1) % 10 == 0:
                try:
                    _guardar_alertas(nuevas_alertas, cartera_actualizada, parcial=True)
                    print(f"[verif] Parcial guardado — {i+1}/{total}", flush=True)
                except Exception as e_sv:
                    print(f"[verif] Error guardado parcial: {e_sv}", flush=True)

            # Sin delay inter-cliente — ScraperAPI rota IPs automáticamente

        # ── Guardado final ────────────────────────────────────────────────────
        _guardar_alertas(nuevas_alertas, cartera_actualizada, parcial=False)
        ok_count    = sum(1 for c in cartera_actualizada if c.get('scoreCompleto'))
        err_count   = sum(1 for c in cartera_actualizada if c.get('verificacion_fallida'))
        _n_bcra_v   = sum(1 for a in nuevas_alertas if a.get('tipo') == 'bcra')
        _n_cheq_v   = sum(1 for a in nuevas_alertas if a.get('tipo') == 'cheque')
        _n_bodeg_v  = sum(1 for a in nuevas_alertas if a.get('tipo') == 'bodegas')
        print(
            f"[verif] FIN: {ok_count}/{total} con score, {err_count} fallidos"
            f" | {_n_bcra_v} BCRA · {_n_cheq_v} cheques · {_n_bodeg_v} bodegas",
            flush=True,
        )
        verificacion_estado["mensaje"] = (
            f"Completado: {ok_count}/{total} verificados, {err_count} fallidos"
            f" | {_n_bcra_v} alerta(s) BCRA · {_n_cheq_v} cheques · {_n_bodeg_v} bodegas."
        )
        verificacion_estado["progreso"] = total

    except BaseException as e_fatal:
        print(f"[verif] ERROR FATAL — {type(e_fatal).__name__}: {e_fatal}", flush=True)
        print(f"[verif] Traceback:\n{traceback.format_exc()}", flush=True)
        verificacion_estado["mensaje"] = f"Error crítico: {type(e_fatal).__name__}: {e_fatal}"
        # Guardar lo que se procesó hasta el momento
        if cartera_actualizada:
            try:
                _guardar_alertas(nuevas_alertas, cartera_actualizada, parcial=True)
                print(f"[verif] Guardado de emergencia: {len(cartera_actualizada)} clientes", flush=True)
            except Exception:
                pass
    finally:
        verificacion_estado["corriendo"] = False
        print("[verif] Flag liberado.", flush=True)

# ─── AUTH ─────────────────────────────────────────────────

def require_login(f):
    @wraps(f)
    def _wrapped(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return _wrapped

def require_director(f):
    @wraps(f)
    def _wrapped(*args, **kwargs):
        if not session.get('director_logged_in'):
            return redirect(url_for('director_login_page'))
        return f(*args, **kwargs)
    return _wrapped

# ─── ENDPOINTS ───────────────────────────────────────────

@app.route("/")
@require_login
def index():
    resp = send_from_directory('static', 'index.html')
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@app.route("/ping")
def ping():
    return jsonify({"ok": True, "ts": time.time()})


# ── Configuración de usuario — persiste en memoria del proceso ───────────────
# Incluye: umbrales DSO, nombre de bodega, preferencias UI.
# Los valores sobreviven reinicios de tab (no de deploy). Para persistencia
# total entre deploys, sincronizar con Cloudflare R2 en una iteración futura.
_user_config: dict = {
    'dso_umbral_bajo':   45,
    'dso_umbral_alto':   65,
    'nombre_bodega':     '',
}
_user_config_lock = threading.Lock()
_USER_CONFIG_FILE = os.path.join(os.environ.get('DATA_DIR', os.getcwd()), 'user_config.json')

def _load_user_config():
    global _user_config
    try:
        if os.path.exists(_USER_CONFIG_FILE):
            with open(_USER_CONFIG_FILE, 'r', encoding='utf-8') as _f:
                stored = json.load(_f)
            with _user_config_lock:
                _user_config.update({k: v for k, v in stored.items() if k in _user_config})
    except Exception as _e:
        print(f'[config] Error cargando user_config: {_e}', flush=True)

_load_user_config()

@app.route('/api/config', methods=['GET', 'POST'])
@require_login
def api_user_config():
    """GET: devuelve configuración actual. POST: actualiza y persiste campos permitidos."""
    global _user_config
    _ALLOWED_KEYS = {'dso_umbral_bajo', 'dso_umbral_alto', 'nombre_bodega'}
    if request.method == 'GET':
        with _user_config_lock:
            return jsonify(dict(_user_config))
    data = request.get_json(silent=True) or {}
    updates = {k: v for k, v in data.items() if k in _ALLOWED_KEYS}
    if not updates:
        return jsonify({'error': 'Sin campos válidos para actualizar'}), 400
    with _user_config_lock:
        _user_config.update(updates)
        snapshot = dict(_user_config)
    try:
        with open(_USER_CONFIG_FILE, 'w', encoding='utf-8') as _f:
            json.dump(snapshot, _f, ensure_ascii=False, indent=2)
    except Exception as _e:
        print(f'[config] Error guardando user_config: {_e}', flush=True)
    return jsonify({'ok': True, 'config': snapshot})


@app.route("/api/macro-context")
@require_login
def api_macro_context():
    """Datos macro en caché (inflación, riesgo país, dólar blue). Actualiza si venció TTL."""
    data = _fetch_macro_data()
    return jsonify({
        "ok": bool(data),
        "data": data,
        "cache_age_min": round((time.time() - _macro_cache['ts']) / 60, 1) if _macro_cache['ts'] else None,
    })

@app.route("/comercial")
def comercial():
    resp = send_from_directory('static', 'comercial.html')
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@app.route("/director-login", methods=["GET", "POST"])
def director_login_page():
    if request.method == "POST":
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown').split(',')[0].strip()
        _rate_key = f'dir:{ip}'
        if not _login_rate_check(_rate_key):
            print(f'[SECURITY] Director-login bloqueado por rate limit — IP: {ip}', flush=True)
            return jsonify({"ok": False, "error": "Demasiados intentos. Esperá 15 minutos."}), 429
        data = request.get_json(silent=True) or {}
        usuario = str(data.get('usuario', '')).strip().upper()
        clave   = str(data.get('clave', '')).strip()
        if usuario == DIRECTOR_USER and clave == DIRECTOR_PASS:
            _login_rate_reset(_rate_key)
            session['director_logged_in'] = True
            session.permanent = True
            return jsonify({"ok": True})
        print(f'[SECURITY] Director-login fallido — IP: {ip}', flush=True)
        return jsonify({"ok": False, "error": "Usuario o clave incorrectos"}), 401
    # GET → mostrar pantalla de login
    return send_from_directory('static', 'director_login.html')

@app.route("/director-logout")
def director_logout():
    session.pop('director_logged_in', None)
    return redirect(url_for('director_login_page'))

@app.route("/director")
@require_director
def director():
    resp = send_from_directory('static', 'director.html')
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@app.route("/api/director-data")
@require_director
def api_director_data():
    """Panel Dirección Comercial: saldos, aging, score y DSO por cliente.
    Aging: días desde fecha de emisión hasta HOY (no desde fechaPago)."""
    def _parse_f(s):
        if not s:
            return None
        s = str(s).strip()
        if '/' in s:                          # DD/MM/YYYY
            p = s.split('/')
            try: return datetime(int(p[2]), int(p[1]), int(p[0]))
            except Exception: return None
        if '-' in s and len(s) >= 10:         # YYYY-MM-DD (ISO)
            p = s.split('-')
            try: return datetime(int(p[0]), int(p[1]), int(p[2]))
            except Exception: return None
        return None

    _nc = lambda x: str(x).replace('-','').replace(' ','').strip()
    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    # Aging por días desde fechaFactura hasta hoy (no usa fechaPago)
    def _bucket(dias):
        if dias <= 30:  return 'd30'
        if dias <= 60:  return 'd60'
        if dias <= 90:  return 'd90'
        if dias <= 120: return 'd120'
        return 'd120plus'

    # Índice nombre→cuit desde _cartera_comercial para cruzar cuando saldos no trae CUIT
    import re as _re
    def _norm_nombre(s):
        """Normaliza nombre: mayúsculas, sin puntuación, sin espacios extra.
        'LOYDIS S.A.' → 'LOYDIS SA' | 'Abdenur C.' → 'ABDENUR C'"""
        return _re.sub(r'\s+', ' ', _re.sub(r'[^A-Z0-9 ]', '', str(s).upper())).strip()

    # Índice nombre_normalizado → CUIT (doble clave: con y sin puntuación)
    nombre_a_cuit: dict = {}
    for _cc in _cartera_comercial:
        _nombre_raw = str(_cc.get('nombre') or '').strip()
        _ck = _nc(str(_cc.get('cuit') or ''))
        if _nombre_raw and _ck:
            nombre_a_cuit[_nombre_raw.upper()] = _ck       # exacto
            nombre_a_cuit[_norm_nombre(_nombre_raw)] = _ck  # normalizado sin puntuación

    # Índice CUIT → ciudad, limiteCredito, vendedor desde _cartera_comercial
    cc_info_map: dict = {}
    for _cc in _cartera_comercial:
        _ck = _nc(str(_cc.get('cuit') or ''))
        if _ck:
            cc_info_map[_ck] = {
                'ciudad':        str(_cc.get('ciudad') or '').strip(),
                'limiteCredito': float(_cc.get('limiteCredito') or 0),
                'vendedor':      str(_cc.get('vendedor') or '').strip(),
            }

    # Cargar scores desde alertas_cartera.json (CUIT como clave)
    scores_map: dict = {}
    try:
        if os.path.exists(ALERTAS_FILE):
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as _fa:
                _ad = json.load(_fa)
            for _ce in (_ad.get('cartera') or []):
                _cuit_e = _nc(str(_ce.get('cuit') or ''))
                if _cuit_e and _ce.get('scoreCompleto'):
                    scores_map[_cuit_e] = {
                        'score':          _ce.get('scoreCompleto'),
                        'rango':          _ce.get('scoreRango') or '—',
                        'color':          _ce.get('scoreColor') or '#6b7280',
                        'bloquear':       bool(_ce.get('bloquear_oportunidad')),
                        'ultimaSit':      int(_ce.get('ultimaSit') or 1),
                        'alerta_temprana': bool(_ce.get('alerta_temprana')),
                    }
    except Exception: pass

    # También cargar desde score_cache.json como segunda fuente
    try:
        sc_cache_path = os.path.join(DATA_DIR, 'score_cache.json')
        if os.path.exists(sc_cache_path):
            with open(sc_cache_path, 'r', encoding='utf-8') as _scf:
                _scc = json.load(_scf)
            for _cuit_sc, _sc_v in _scc.items():
                if _cuit_sc not in scores_map and isinstance(_sc_v, dict) and _sc_v.get('score'):
                    scores_map[_nc(_cuit_sc)] = {
                        'score':  _sc_v.get('score'),
                        'rango':  _sc_v.get('rango') or '—',
                        'color':  _sc_v.get('color') or '#6b7280',
                        'bloquear': bool(_sc_v.get('bloquear_oportunidad')),
                    }
    except Exception: pass

    # Sincronizar desde disco: otro worker puede haber subido saldos más recientes.
    _saldos_gestion_desde_disco()

    # Índice vendedor desde _saldos_gestion (siempre tiene campo vendedor).
    _vend_enrich: dict = {}
    for _fv in (_saldos_gestion if _saldos_gestion else _saldos_facturas):
        _cn = _norm_nombre(str(_fv.get('cliente') or ''))
        _cv = str(_fv.get('vendedor') or '').strip()
        if _cn and _cv:
            _vend_enrich[_cn] = _cv

    # Fuente de facturas para el panel Director:
    # 1. _saldos_gestion (upload comercial, más frecuente) — tiene precedencia si está cargado.
    # 2. dso_saldos_actual.json — fallback solo si aún no se subió reporte de gestión.
    # El indicador DSO global del header usa SIEMPRE /dso-global-saldos (solo DSO mensual).
    _fuente_director = []
    if _saldos_gestion:
        _fuente_director = _saldos_gestion
    else:
        _dso_path = os.path.join(DATA_DIR, 'dso_saldos_actual.json')
        if os.path.exists(_dso_path):
            try:
                with open(_dso_path, 'r', encoding='utf-8') as _fdso:
                    _raw_dso = json.load(_fdso).get('saldos', [])
                for _s in _raw_dso:
                    _fuente_director.append({
                        'cliente':      _s.get('cliente', ''),
                        'cuit':         _s.get('cuit', ''),
                        'vendedor':     _s.get('vendedor', ''),
                        'fechaFactura': _s.get('fecha_factura', _s.get('fechaFactura', '')),
                        'fechaPago':    _s.get('fecha_pago', _s.get('fechaPago', '')),
                        'saldo':        _s.get('saldo', 0),
                        'nroFactura':   _s.get('nroFactura', _s.get('nro_factura', '')),
                        'totalFactura': _s.get('totalFactura', _s.get('total_factura', 0)),
                    })
            except Exception:
                pass
        if not _fuente_director:
            _fuente_director = _saldos_facturas
    # Enriquecer vendedor: 1) _vend_enrich (saldos_gestion) → 2) cartera_comercial
    for _fd in _fuente_director:
        if not str(_fd.get('vendedor') or '').strip():
            _cli_norm_fd = _norm_nombre(str(_fd.get('cliente') or ''))
            _cuit_fd     = _nc(str(_fd.get('cuit') or ''))
            # Si tampoco viene CUIT, intentar resolverlo por nombre (igual que abajo)
            if not _cuit_fd:
                _cuit_fd = nombre_a_cuit.get(str(_fd.get('cliente') or '').upper(), '') \
                        or nombre_a_cuit.get(_cli_norm_fd, '')
            _fd['vendedor'] = (
                _vend_enrich.get(_cli_norm_fd, '')
                or cc_info_map.get(_cuit_fd, {}).get('vendedor', '')
            )

    # Agrupar facturas por cliente
    clientes_map: dict = {}
    for f in _fuente_director:
        saldo = float(f.get('saldo') or 0)
        if saldo <= 0:
            continue
        nombre   = str(f.get('cliente') or '').strip()
        cuit_raw = _nc(str(f.get('cuit') or ''))
        # Si no viene CUIT, buscar por nombre exacto primero, luego normalizado
        if not cuit_raw:
            cuit_raw = (nombre_a_cuit.get(nombre.upper())
                        or nombre_a_cuit.get(_norm_nombre(nombre), ''))
        vendedor = str(f.get('vendedor') or '').strip()
        key = cuit_raw if cuit_raw else nombre.upper()
        if key not in clientes_map:
            clientes_map[key] = {
                'nombre': nombre, 'cuit': cuit_raw, 'vendedor': vendedor,
                'facturas': [], 'saldo_total': 0.0,
                'buckets': {'d30': 0.0, 'd60': 0.0, 'd90': 0.0, 'd120': 0.0, 'd120plus': 0.0},
            }
        ff = _parse_f(f.get('fechaFactura', ''))
        dias = max(0, (hoy - ff).days) if ff else 0
        bucket = _bucket(dias)
        clientes_map[key]['saldo_total']     += saldo
        clientes_map[key]['buckets'][bucket] += saldo
        fp       = _parse_f(f.get('fechaPago', ''))
        dias_venc = max(0, (hoy - fp).days) if fp else 0
        # Formato DD/MM/YYYY para presentación en el frontend
        def _iso_to_dmy(s):
            if not s: return ''
            try:
                y, m, d = str(s).split('-')
                return f"{d}/{m}/{y}"
            except Exception:
                return str(s)
        fecha_fac_fmt = _iso_to_dmy(f.get('fechaFactura', ''))
        fecha_pago_fmt = _iso_to_dmy(f.get('fechaPago', ''))
        clientes_map[key]['facturas'].append({
            'nro':           str(f.get('nroFactura') or ''),
            'fecha_factura': fecha_fac_fmt,
            'fecha_pago':    fecha_pago_fmt,
            'total':         float(f.get('totalFactura') or 0),
            'saldo':         saldo,
            'dias':          dias,
            'dias_venc':     dias_venc,
            'bucket':        bucket,
        })

    # Construir lista de clientes
    clientes_list = []
    for key, c in clientes_map.items():
        sc = scores_map.get(c['cuit']) if c['cuit'] else None
        saldo_total = c['saldo_total']
        # DSO individual viene EXCLUSIVAMENTE del módulo DSO (dso_individual_actual.json).
        # El reporte de "Actualizar Saldos Comerciales" no debe afectar este valor.
        # El enriquecimiento post-loop lo setea; si no hay dato DSO, queda en 0.
        dso = 0
        _ci = cc_info_map.get(c['cuit'], {})
        clientes_list.append({
            'nombre':          c['nombre'],
            'cuit':            c['cuit'],
            'vendedor':        c['vendedor'],
            'ciudad':          _ci.get('ciudad', ''),
            'limiteCredito':   _ci.get('limiteCredito', 0),
            'saldo_total':     round(saldo_total),
            'dso':             dso,
            'score':           sc['score']           if sc else None,
            'rango':           sc['rango']            if sc else '—',
            'score_color':     sc['color']            if sc else '#6b7280',
            'bloquear':        sc['bloquear']         if sc else False,
            'ultimaSit':       sc['ultimaSit']        if sc else 1,
            'alerta_temprana': sc['alerta_temprana']  if sc else False,
            'buckets':         {k: round(v) for k, v in c['buckets'].items()},
            'facturas':        sorted(c['facturas'], key=lambda x: x['dias'], reverse=True),
        })

    # ── Reemplazar DSO con valor congelado del reporte mensual (dso_individual_actual.json) ──
    # _get_dso_individual devuelve el DSO calculado una sola vez al momento del upload mensual.
    # No varía día a día. Si no hay dato estático, conserva el DSO calculado desde saldos de gestión.
    _enriq_count = 0
    for _cl in clientes_list:
        _dso_m = _get_dso_individual(
            str(_cl.get('cuit') or '').replace('-', '').replace(' ', '').strip(),
            str(_cl.get('nombre') or '')
        )
        if _dso_m is not None:
            _cl['dso'] = _dso_m
            _enriq_count += 1
    print(f"[director-dso] DSO estático: {_enriq_count}/{len(clientes_list)} via dso_individual_actual.json", flush=True)

    clientes_list.sort(key=lambda x: x['saldo_total'], reverse=True)

    # Totales globales
    total_saldo = sum(c['saldo_total'] for c in clientes_list)
    total_b: dict = {'d30': 0, 'd60': 0, 'd90': 0, 'd120': 0, 'd120plus': 0}
    for c in clientes_list:
        for bk in total_b:
            total_b[bk] += c['buckets'].get(bk, 0)
    total_b = {k: round(v) for k, v in total_b.items()}

    suma_pond_g = sum(f['saldo'] * f['dias'] for c in clientes_list for f in c['facturas'])
    dso_global  = round(suma_pond_g / total_saldo) if total_saldo > 0 else 0

    n_con_score = sum(1 for c in clientes_list if c.get('score'))
    n_riesgo    = sum(1 for c in clientes_list if c.get('score') and c['score'] < 400)
    n_vencido   = sum(1 for c in clientes_list
                      if sum(c['buckets'].get(bk, 0) for bk in ['d60','d90','d120','d120plus']) > 0)
    n_critico   = sum(1 for c in clientes_list
                      if c['buckets'].get('d90', 0) + c['buckets'].get('d120', 0)
                         + c['buckets'].get('d120plus', 0) > 0)

    # ── DSO global ponderado: agotamiento real por cliente (= app comercial) ────
    # Se calcula aquí mismo para no depender de caché ni de una llamada paralela.
    _dso_g_pond = None
    try:
        _vp = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
        _vgl: dict = {}   # (year,month) → total empresa
        _vpc: dict = {}   # cliente_norm → {(year,month) → total}
        if os.path.exists(_vp):
            with open(_vp, 'r', encoding='utf-8') as _fv:
                _vh = json.load(_fv)
            for _ym, _t in _vh.get('meses', {}).items():
                try: _vgl[(int(_ym[:4]), int(_ym[5:7]))] = float(_t)
                except: pass
            for _cli_v, _mc in _vh.get('por_cliente', {}).items():
                _cn = _norm_nombre(_cli_v)
                _vpc[_cn] = {}
                for _ym, _t in _mc.items():
                    try: _vpc[_cn][(int(_ym[:4]), int(_ym[5:7]))] = float(_t)
                    except: pass
        _ts = total_saldo if total_saldo > 0 else 1
        _sp = 0.0; _ss = 0.0
        for _c in clientes_list:
            _s = float(_c['saldo_total'])
            if _s <= 0:
                continue
            _cn = _norm_nombre(_c['nombre'])
            # Ventas propias del cliente; si no hay, distribución proporcional al saldo
            _vc = _vpc.get(_cn) or {_k: _v * _s / _ts for _k, _v in _vgl.items()}
            _r = _dso_exhaustion(_s, _vc, hoy)
            if _r.get('dso'):
                _sp += _r['dso'] * _s
                _ss += _s
        if _ss > 0:
            _dso_g_pond = round(_sp / _ss)
    except Exception as _ex:
        print(f"[director-data] DSO ponderado error: {_ex}", flush=True)

    return jsonify({
        'fecha_hoy':            hoy.strftime('%d/%m/%Y'),
        'total_saldo':          round(total_saldo),
        'total_buckets':        total_b,
        'dso_global':           dso_global,
        'dso_global_ponderado': _dso_g_pond,
        'n_clientes':           len(clientes_list),
        'n_con_score':          n_con_score,
        'n_riesgo':             n_riesgo,
        'n_vencido':            n_vencido,
        'n_critico':            n_critico,
        'clientes':             clientes_list,
    })

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown').split(',')[0].strip()
        if not _login_rate_check(ip):
            print(f'[SECURITY] Login bloqueado por rate limit — IP: {ip}', flush=True)
            return jsonify({"error": "Demasiados intentos fallidos. Esperá 15 minutos."}), 429
        data = request.get_json(silent=True) or {}
        cuit = str(data.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        pwd  = str(data.get('password', '')).strip()
        if cuit == ADMIN_CUIT and pwd == ADMIN_PASS:
            _login_rate_reset(ip)
            session['logged_in'] = True
            return jsonify({"ok": True})
        print(f'[SECURITY] Login fallido — IP: {ip}', flush=True)
        return jsonify({"error": "Credenciales incorrectas"}), 401
    return send_from_directory('static', 'login.html')

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route("/supabase-session.js")
def supabase_session_js():
    return send_from_directory('static', 'supabase-session.js')


# ── Admin: Padrón Local BCRA ────────────────────────────────────────────────

def _admin_auth(req) -> bool:
    """Autentica la solicitud admin por header X-Admin-Pass o JSON body."""
    pwd = req.headers.get('X-Admin-Pass', '')
    if not pwd:
        body = req.get_json(silent=True) or {}
        pwd = str(body.get('admin_pass', '') or body.get('password', ''))
    return pwd == ADMIN_PASS


@app.route("/admin/padron-info")
def admin_padron_info():
    """Devuelve estadísticas del padrón local: total CUITs, período, estado import."""
    if not _admin_auth(request):
        return jsonify({"error": "no_autorizado"}), 403
    total, periodo = _padron_contar_registros()
    return jsonify({
        "total_cuits":   total,
        "periodo":       periodo,
        "db_path":       PADRON_DB_PATH,
        "db_existe":     os.path.exists(PADRON_DB_PATH),
        "import_estado": _padron_import_estado,
    })


@app.route("/admin/padron-progreso")
def admin_padron_progreso():
    """Endpoint de polling para monitorear el progreso de la importación."""
    if not _admin_auth(request):
        return jsonify({"error": "no_autorizado"}), 403
    return jsonify(_padron_import_estado)


@app.route("/admin/importar-padron", methods=["POST"])
def admin_importar_padron():
    """Inicia la importación del padrón mensual BCRA en un hilo de fondo.

    Modos de operación:
      A) Archivo subido (form-data campo 'archivo'): acepta hasta 512 MB.
      B) URL remota (JSON campo 'url'): el servidor descarga el archivo directamente.
      C) Ruta en disco (JSON campo 'ruta'): usa un archivo ya presente en /data.

    Retorna inmediatamente con {"estado": "iniciado"}.
    Consultar /admin/padron-progreso para seguimiento.
    """
    if not _admin_auth(request):
        return jsonify({"error": "no_autorizado"}), 403

    if _padron_import_estado.get('corriendo'):
        return jsonify({"error": "importacion_en_curso", "estado": _padron_import_estado}), 409

    ruta_tmp = None
    borrar   = True

    # Modo A: archivo subido via multipart
    if 'archivo' in request.files:
        f = request.files['archivo']
        if not f.filename:
            return jsonify({"error": "archivo_vacio"}), 400
        ruta_tmp = os.path.join(DATA_DIR, f'padron_upload_{int(time.time())}.txt')
        f.save(ruta_tmp)

    # Modo B: URL para descargar
    elif request.is_json and request.json.get('url'):
        url_padron = request.json['url']
        ruta_tmp   = os.path.join(DATA_DIR, f'padron_download_{int(time.time())}.txt')
        try:
            with requests.get(url_padron, stream=True, timeout=300, verify=False) as r:
                r.raise_for_status()
                with open(ruta_tmp, 'wb') as out:
                    for chunk in r.iter_content(chunk_size=1024 * 256):
                        out.write(chunk)
        except Exception as e:
            return jsonify({"error": f"descarga_fallida: {e}"}), 502

    # Modo C: ruta en disco
    elif request.is_json and request.json.get('ruta'):
        ruta_tmp = request.json['ruta']
        borrar   = bool(request.json.get('borrar', False))
        if not os.path.exists(ruta_tmp):
            return jsonify({"error": "archivo_no_encontrado", "ruta": ruta_tmp}), 404
    else:
        return jsonify({
            "error": "sin_fuente",
            "modos": ["archivo (multipart)", "url (json)", "ruta (json)"],
        }), 400

    # Iniciar importación en hilo de fondo
    t = threading.Thread(
        target=_importar_padron_worker,
        args=(ruta_tmp, borrar),
        daemon=True,
    )
    t.start()

    return jsonify({
        "estado":  "iniciado",
        "archivo": os.path.basename(ruta_tmp),
        "progreso_url": "/admin/padron-progreso",
    })


# ── Admin: Pre-cacheo de la cartera actual ──────────────────────────────────

_precacheo_estado: dict = {
    "corriendo":   False,
    "total":       0,
    "procesados":  0,
    "exitosos":    0,
    "saltados":    0,
    "errores":     0,
    "cliente_actual": "",
    "mensaje":     "Sin pre-cacheo activo",
}


def _precachear_cartera_worker(delay_seg: float = 1.0):
    """Recorre _cartera_comercial, consulta BCRA por cada CUIT y guarda en padrón local.

    Salta CUITs que ya existen en bcra_padron_local para no repetir trabajo.
    `delay_seg` entre clientes evita saturar el proxy en el burst inicial.
    """
    global _precacheo_estado

    pendientes = [
        str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        for c in _cartera_comercial
        if c.get('cuit')
    ]
    pendientes = [c for c in pendientes if len(c) == 11 and c.isdigit()]
    pendientes = list(dict.fromkeys(pendientes))  # deduplicar manteniendo orden

    _precacheo_estado = {
        "corriendo": True, "total": len(pendientes), "procesados": 0,
        "exitosos": 0, "saltados": 0, "errores": 0,
        "cliente_actual": "", "mensaje": f"Iniciando: {len(pendientes)} CUITs en cartera",
    }

    for cuit in pendientes:
        _precacheo_estado['cliente_actual'] = cuit

        # Saltar si ya está en padrón local
        if consultar_padron_local(cuit) is not None:
            _precacheo_estado['saltados']   += 1
            _precacheo_estado['procesados'] += 1
            _precacheo_estado['mensaje'] = f"Saltado (ya en padrón): {cuit}"
            continue

        _precacheo_estado['mensaje'] = f"Consultando BCRA: {cuit}..."
        try:
            data, error = consultar_bcra(cuit)
            if data and not error:
                _guardar_en_padron_local(cuit, data)
                _precacheo_estado['exitosos'] += 1
                sit = (data.get('results') or {})
                _precacheo_estado['mensaje'] = f"OK: {cuit}"
            else:
                _precacheo_estado['errores'] += 1
                _precacheo_estado['mensaje'] = f"Sin respuesta: {cuit}"
        except Exception as e:
            _precacheo_estado['errores'] += 1
            _precacheo_estado['mensaje'] = f"Error en {cuit}: {e}"

        _precacheo_estado['procesados'] += 1
        time.sleep(delay_seg)

    total_ok = _precacheo_estado['exitosos']
    _precacheo_estado['corriendo'] = False
    _precacheo_estado['mensaje']   = (
        f"Pre-cacheo completo: {total_ok} CUITs guardados, "
        f"{_precacheo_estado['saltados']} ya estaban, "
        f"{_precacheo_estado['errores']} errores"
    )
    print(f"[padron] Pre-cacheo finalizado: {_precacheo_estado['mensaje']}", flush=True)


@app.route("/admin/precachear-cartera", methods=["POST"])
def admin_precachear_cartera():
    """Inicia el pre-cacheo de todos los CUITs de la cartera en un hilo de fondo.

    Parámetros opcionales (JSON):
      - delay: float — segundos entre consultas (default 1.5, mínimo 0.5)
    """
    if not _admin_auth(request):
        return jsonify({"error": "no_autorizado"}), 403
    if _precacheo_estado.get('corriendo'):
        return jsonify({"error": "precacheo_en_curso", "estado": _precacheo_estado}), 409

    body  = request.get_json(silent=True) or {}
    delay = max(0.5, float(body.get('delay', 1.0)))

    t = threading.Thread(target=_precachear_cartera_worker, args=(delay,), daemon=True)
    t.start()

    return jsonify({
        "estado":        "iniciado",
        "total_cuits":   len(_cartera_comercial),
        "delay_seg":     delay,
        "progreso_url":  "/admin/precacheo-progreso",
    })


@app.route("/admin/precacheo-progreso")
def admin_precacheo_progreso():
    """Progreso en tiempo real del pre-cacheo de la cartera."""
    if not _admin_auth(request):
        return jsonify({"error": "no_autorizado"}), 403
    return jsonify(_precacheo_estado)


@app.route("/todos-los-clientes")
def get_todos_los_clientes():
    """Devuelve la cartera completa con CUITs para que el admin pueda verificar sin localStorage."""
    resp = jsonify(_cartera_comercial)
    resp.headers['Cache-Control'] = 'no-store'
    return resp

@app.route("/vendedores")
def get_vendedores():
    vs = sorted({(c.get('vendedor') or '').strip() for c in _cartera_comercial if (c.get('vendedor') or '').strip()})
    resp = jsonify(vs)
    resp.headers['Cache-Control'] = 'no-store'
    return resp

def _cascade_cuit_rename(cuit_old: str, cuit_new: str) -> dict:
    """Propagación en cascada de un cambio de CUIT a todos los archivos y cachés.

    Archivos actualizados:
      db_v17_final.json     — alertas[].cuit  +  cartera[].cuit
      alertas_bcra.json     — clave plana o arrays (ambas ubicaciones)
      score_cache.json      — clave directa
      bcra_cache.json       — clave directa
      historial_{c}.json, cheques_{c}.json, solvency_{c}.json — renombrar

    Memoria actualizada:
      bcra_cache dict, _score_session_cache dict — clave renombrada

    Todas las escrituras a disco son atómicas: tmp → fsync → os.replace.
    Usa _alertas_file_lock y _score_cache_lock para evitar race conditions.
    """
    global bcra_cache, _score_session_cache

    if cuit_old == cuit_new:
        return {'ok': True, 'cambios': [], 'errores': []}

    cambios: list = []
    errores: list = []
    _nc = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()

    # ── 1. db_v17_final.json  (alertas[] + cartera[]) ────────────────────────
    try:
        with _alertas_file_lock:
            if os.path.exists(ALERTAS_FILE):
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
                    _d = json.load(_f)
                n_alr = n_car = 0
                for _a in _d.get('alertas', []):
                    if _nc(_a.get('cuit')) == cuit_old:
                        _a['cuit'] = cuit_new
                        n_alr += 1
                for _c in _d.get('cartera', []):
                    if _nc(_c.get('cuit')) == cuit_old:
                        _c['cuit'] = cuit_new
                        n_car += 1
                if n_alr or n_car:
                    _tmp = ALERTAS_FILE + '.tmp'
                    with open(_tmp, 'w', encoding='utf-8') as _f:
                        json.dump(_d, _f, ensure_ascii=False, default=str)
                        _f.flush(); os.fsync(_f.fileno())
                    os.replace(_tmp, ALERTAS_FILE)
                    cambios.append(
                        f'db_v17_final.json: {n_alr} alerta(s) + {n_car} entrada(s) cartera'
                    )
    except Exception as _e:
        errores.append(f'db_v17_final.json: {_e}')

    # ── 2. alertas_bcra.json  (clave plana o arrays) ─────────────────────────
    for _bp in list(dict.fromkeys([
        ALERTAS_BCRA_FILE,
        os.path.join(os.getcwd(), 'alertas_bcra.json'),
    ])):
        try:
            if not os.path.exists(_bp):
                continue
            with open(_bp, 'r', encoding='utf-8') as _f:
                _bd = json.load(_f)
            _mod = False
            # Formato plano: { "30123456789": { score… }, … }
            if cuit_old in _bd and isinstance(_bd.get(cuit_old), dict):
                _bd[cuit_new] = _bd.pop(cuit_old)
                _mod = True
            # Formato con arrays (alertas[] / cartera[])
            for _a in _bd.get('alertas', []):
                if _nc(_a.get('cuit')) == cuit_old:
                    _a['cuit'] = cuit_new; _mod = True
            for _c in _bd.get('cartera', []):
                if _nc(_c.get('cuit')) == cuit_old:
                    _c['cuit'] = cuit_new; _mod = True
            if _mod:
                _tmp = _bp + '.tmp'
                with open(_tmp, 'w', encoding='utf-8') as _f:
                    json.dump(_bd, _f, ensure_ascii=False, default=str)
                    _f.flush(); os.fsync(_f.fileno())
                os.replace(_tmp, _bp)
                cambios.append(f'{os.path.basename(_bp)}: CUIT actualizado')
        except Exception as _e:
            errores.append(f'{os.path.basename(_bp)}: {_e}')

    # ── 3. score_cache.json ──────────────────────────────────────────────────
    try:
        with _score_cache_lock:
            _sc = _score_cache_read()
            if cuit_old in _sc:
                _sc[cuit_new] = _sc.pop(cuit_old)
                _score_cache_write(_sc)
                cambios.append('score_cache.json: clave renombrada')
    except Exception as _e:
        errores.append(f'score_cache.json: {_e}')

    # ── 4. bcra_cache.json ───────────────────────────────────────────────────
    try:
        _bc_path = os.path.join(DATA_DIR, 'bcra_cache.json')
        if os.path.exists(_bc_path):
            with open(_bc_path, 'r', encoding='utf-8') as _f:
                _bc = json.load(_f)
            if cuit_old in _bc:
                _bc[cuit_new] = _bc.pop(cuit_old)
                _tmp = _bc_path + '.tmp'
                with open(_tmp, 'w', encoding='utf-8') as _f:
                    json.dump(_bc, _f, ensure_ascii=False, default=str)
                    _f.flush(); os.fsync(_f.fileno())
                os.replace(_tmp, _bc_path)
                cambios.append('bcra_cache.json: clave renombrada')
    except Exception as _e:
        errores.append(f'bcra_cache.json: {_e}')

    # ── 5. Archivos per-CUIT (historial, cheques, solvency) ──────────────────
    for _pfx in ('historial', 'cheques', 'solvency'):
        _old_f = os.path.join(DATA_DIR, f'{_pfx}_{cuit_old}.json')
        _new_f = os.path.join(DATA_DIR, f'{_pfx}_{cuit_new}.json')
        try:
            if os.path.exists(_old_f):
                os.replace(_old_f, _new_f)
                cambios.append(f'{_pfx}_{cuit_old}.json → {_pfx}_{cuit_new}.json')
        except Exception as _e:
            errores.append(f'{_pfx}_{cuit_old}.json rename: {_e}')

    # ── 6. Cachés en memoria ─────────────────────────────────────────────────
    try:
        if cuit_old in bcra_cache:
            bcra_cache[cuit_new] = bcra_cache.pop(cuit_old)
            cambios.append('bcra_cache (RAM): clave renombrada')
    except Exception as _e:
        errores.append(f'bcra_cache RAM: {_e}')

    try:
        if cuit_old in _score_session_cache:
            _score_session_cache[cuit_new] = _score_session_cache.pop(cuit_old)
            cambios.append('_score_session_cache (RAM): clave renombrada')
    except Exception as _e:
        errores.append(f'_score_session_cache RAM: {_e}')

    _tag = f'[cascade-cuit] {cuit_old} → {cuit_new}'
    print(f'{_tag}: {len(cambios)} cambio(s), {len(errores)} error(es)', flush=True)
    for _ch in cambios: print(f'  ✓ {_ch}', flush=True)
    for _er in errores: print(f'  ✗ {_er}', flush=True)
    return {'ok': len(errores) == 0, 'cambios': cambios, 'errores': errores}


@app.route("/cliente/guardar", methods=["POST"])
def guardar_cliente():
    global _cartera_comercial
    try:
        data = request.get_json(force=True)
        cuit = str(data.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        nombre = str(data.get('nombre', '')).strip()
        if not cuit or not nombre:
            return jsonify({"ok": False, "error": "CUIT y nombre son obligatorios"}), 400

        cuit_orig = str(data.get('_cuitOriginal', cuit)).replace('-', '').replace(' ', '').strip()

        cliente = {
            'nombre': nombre,
            'cuit': cuit,
            'ciudad': str(data.get('ciudad', '')).strip(),
            'vendedor': str(data.get('vendedor', '')).strip(),
            'email': str(data.get('email', '')).strip(),
            'limiteCredito': float(data.get('limiteCredito', 0) or 0),
            'plazoDias': int(data.get('plazoDias', 0) or 0),
        }

        idx = next((i for i, c in enumerate(_cartera_comercial)
                    if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_orig), None)

        if idx is not None:
            # Preserve fields not managed here (e.g. plazoDias)
            merged = dict(_cartera_comercial[idx])
            merged.update(cliente)
            _cartera_comercial[idx] = merged
            accion = "actualizado"
        else:
            _cartera_comercial.append(cliente)
            accion = "agregado"

        # Escritura atómica de cartera_comercial.json
        _cc_tmp = _CC_FILE + '.tmp'
        with open(_cc_tmp, 'w', encoding='utf-8') as f:
            json.dump(_cartera_comercial, f, ensure_ascii=False, indent=2)
            f.flush(); os.fsync(f.fileno())
        os.replace(_cc_tmp, _CC_FILE)

        # ── Propagación en cascada si el CUIT cambió ─────────────────────────
        cascade = None
        if cuit_orig != cuit:
            cascade = _cascade_cuit_rename(cuit_orig, cuit)

        resp_data = {"ok": True, "accion": accion, "total": len(_cartera_comercial)}
        if cascade:
            resp_data['cascade'] = cascade.get('cambios', [])
            if cascade.get('errores'):
                resp_data['cascade_errores'] = cascade['errores']
        return jsonify(resp_data)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/cartera-por-vendedor/<vendedor>")
def get_cartera_por_vendedor(vendedor):
    """Fuente: cartera_comercial.json (todos los clientes del vendedor).
    Saldo viene del cruce con _saldos_gestion. Nunca limita a clientes con saldo."""
    _saldos_gestion_desde_disco()   # sincroniza si otro worker subió datos
    from urllib.parse import unquote
    v = unquote(vendedor).strip().lower()

    # Base: todos los clientes del vendedor en cartera_comercial
    if v in ('todos', 'all', ''):
        base = _cartera_comercial
    else:
        base = [c for c in _cartera_comercial if (c.get('vendedor') or '').strip().lower() == v]

    # Fallback: si no hay clientes en cartera_comercial, construir desde saldos
    # (cubre vendedores nuevos que todavía no están en el JSON de cartera)
    if not base and v not in ('todos', 'all', ''):
        fuente_fb = _saldos_gestion if _saldos_gestion else _saldos_facturas
        clientes_visto = set()
        for f in fuente_fb:
            vend_fb = (f.get('vendedor') or '').strip().lower()
            if vend_fb != v:
                continue
            cli = (f.get('cliente') or '').strip()
            if not cli or cli in clientes_visto:
                continue
            clientes_visto.add(cli)
            base.append({
                'nombre':   cli,
                'cuit':     '',   # CUIT desconocido hasta que se consulte BCRA
                'vendedor': (f.get('vendedor') or '').strip(),
                'ciudad':   '',
                'total_saldo': float(f.get('saldo') or 0),
            })
        if base:
            print(f"[cartera] {v}: {len(base)} clientes desde saldos (no estaba en cartera_comercial)", flush=True)

    # Mapa de saldos desde gestión (o auditoría como fallback)
    fuente_s = _saldos_gestion if _saldos_gestion else _saldos_facturas
    saldo_map = {}
    for f in fuente_s:
        cli = (f.get('cliente') or '').strip()
        if not cli:
            continue
        key = _norm_nombre(cli)
        saldo_map[key] = saldo_map.get(key, 0) + (f.get('saldo') or 0)

    def _nc(x):
        return str(x or '').replace('-', '').replace(' ', '').strip()

    # ── Merge obligatorio desde TODOS los archivos de scores del servidor ────────
    # Fuentes: alertas_bcra.json (queries individuales) + alertas_cartera.json (verificación)
    # Prioridad: alertas_cartera sobreescribe alertas_bcra para el mismo CUIT
    scores = {}            # cuit_norm → entry
    scores_by_nombre = {}  # norm_nombre → entry  (plan B)
    alertas_cuits = set()
    _archivos_usados = []

    def _load_score_file(ruta):
        if not os.path.exists(ruta):
            return
        try:
            with open(ruta, 'r', encoding='utf-8') as _f:
                _ad = json.load(_f)
            loaded = 0
            # Estructura estándar: {cartera: [...], alertas: [...]}
            for _c in _ad.get('cartera', []):
                _nc_val = _nc(_c.get('cuit', ''))
                if _nc_val:
                    scores[_nc_val] = _c  # sobreescribe — última fuente cargada gana
                    loaded += 1
                _n = (_c.get('nombre') or '').strip()
                if _n:
                    scores_by_nombre[_norm_nombre(_n)] = _c
            # Estructura plana: {cuit: score_data} (posible formato de alertas_bcra.json)
            if not _ad.get('cartera') and not _ad.get('alertas'):
                for _k, _v in _ad.items():
                    if isinstance(_v, dict) and _v.get('scoreCompleto'):
                        _nc_val = _nc(_k)
                        if _nc_val:
                            scores[_nc_val] = _v
                            loaded += 1
            for _a in _ad.get('alertas', []):
                _nc_val = _nc(_a.get('cuit', ''))
                if _nc_val:
                    alertas_cuits.add(_nc_val)
                _n = (_a.get('nombre') or '').strip()
                if _n and _a.get('scoreCompleto'):
                    scores_by_nombre.setdefault(_norm_nombre(_n), _a)
            _archivos_usados.append(f"{os.path.basename(ruta)}({loaded}sc)")
        except Exception as _e:
            print(f"[cartera-vendedor] Error leyendo {ruta}: {_e}", flush=True)

    # Cargar en orden ascendente de prioridad (cartera sobreescribe bcra)
    for _ruta in list(dict.fromkeys([
        os.path.join(os.getcwd(), 'alertas_bcra.json'), ALERTAS_BCRA_FILE,
        os.path.join(os.getcwd(), 'alertas_cartera.json'), ALERTAS_FILE,
    ])):
        _load_score_file(_ruta)

    print(
        f"[cartera-vendedor] vendedor={v!r} | base={len(base)} | "
        f"archivos={_archivos_usados or ['NINGUNO']} | "
        f"scores={len(scores)} | scores_nombre={len(scores_by_nombre)} | alertas={len(alertas_cuits)}",
        flush=True
    )
    # ──────────────────────────────────────────────────────────────────────────

    result = []
    for cc in base:
        nombre = (cc.get('nombre') or '').strip()
        cuit = (cc.get('cuit') or '').strip()
        cuit_n = _nc(cuit)

        # Plan A: cruce por CUIT normalizado
        sc = scores.get(cuit_n, {})

        # Plan B: cruce por nombre normalizado (si plan A falla)
        if not sc.get('scoreCompleto'):
            nombre_norm_b = _norm_nombre(nombre)
            sc = scores_by_nombre.get(nombre_norm_b, {})

        # Plan C: primeras 2 palabras del nombre
        if not sc.get('scoreCompleto'):
            prim2_b = ' '.join(_norm_nombre(nombre).split()[:2])
            for _k, _sv in scores_by_nombre.items():
                if ' '.join(_k.split()[:2]) == prim2_b:
                    sc = _sv
                    break

        # Buscar saldo: exacto primero, luego por primeras 2 palabras
        nombre_norm = _norm_nombre(nombre)
        total_saldo = saldo_map.get(nombre_norm, 0)
        if total_saldo == 0:
            prim2 = ' '.join(nombre_norm.split()[:2])
            for k, sv in saldo_map.items():
                if ' '.join(k.split()[:2]) == prim2:
                    total_saldo = sv
                    break

        limite_credito = float(cc.get('limiteCredito') or 0)
        cupo_disponible = max(0.0, limite_credito - total_saldo) if limite_credito > 0 else None
        score_val            = sc.get('scoreCompleto') or None
        alerta_temprana      = sc.get('alerta_temprana', False)
        bloquear_oportunidad = sc.get('bloquear_oportunidad', False)
        alerta_log           = sc.get('alerta_logistica', '') or _alerta_logistica(cc.get('ciudad', ''))

        result.append({
            'nombre': nombre,
            'cuit': cuit,
            'ciudad': cc.get('ciudad', ''),
            'vendedor': cc.get('vendedor', ''),
            'email': cc.get('email', ''),
            'total_saldo': total_saldo,
            'limite_credito': limite_credito,
            'cupo_disponible': cupo_disponible,
            'score': score_val,
            'scoreRango': sc.get('scoreRango') or None,
            'scoreColor': sc.get('scoreColor') or None,
            'scoreEmoji': sc.get('scoreEmoji') or None,
            'ultimaSit': sc.get('ultimaSit') or 1,
            'alerta': cuit_n in alertas_cuits or alerta_temprana,
            'alerta_temprana': alerta_temprana,
            'bloquear_oportunidad': bloquear_oportunidad,
            'alerta_logistica': alerta_log,
            'oportunidad': bool(
                score_val and score_val >= 750
                and total_saldo == 0
                and not bloquear_oportunidad
            ),
        })

    # Primero con saldo (desc), luego sin saldo alfabético
    result.sort(key=lambda x: (0 if x['total_saldo'] > 0 else 1, -(x['total_saldo'] or 0), x['nombre']))
    resp = jsonify(result)
    resp.headers['Cache-Control'] = 'no-store'
    return resp

@app.route("/cartera-comercial/<vendedor>")
def get_cartera_comercial(vendedor):
    from urllib.parse import unquote
    v = unquote(vendedor).strip().lower()
    result = [c for c in _cartera_comercial if c.get('vendedor', '').lower() == v]
    # Enriquecer con score y situación desde alertas
    try:
        if os.path.exists(ALERTAS_FILE):
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as f:
                alertas_data = json.load(f)
            cartera_scores = {c['cuit']: c for c in alertas_data.get('cartera', [])}
            for cliente in result:
                if cliente['cuit'] in cartera_scores:
                    sc = cartera_scores[cliente['cuit']]
                    cliente['score'] = sc.get('scoreCompleto')
                    cliente['ultimaSit'] = sc.get('ultimaSit', 1)
    except: pass
    return jsonify(result)

# ── Mapa de supervisión (jefes de equipo comercial) ───────────────────────────
# Clave: CUIT normalizado del supervisor.
# 'supervisa': nombres de vendedor exactamente como aparecen en cartera_comercial.json.
_SUPERVISOR_MAP = {
    '27224289966': {   # Valeria Gutierrez Castex
        'nombre':    'Valeria Gutierrez Castex',
        'supervisa': ['Raul Maza', 'Marcelo Fernandez', 'Ezequiel Mallima'],
    },
    '20207619923': {   # Alejandro Valan
        'nombre':    'Alejandro Valan',
        'supervisa': ['Anabel Borrageiros', 'Sergio Piatanesi', 'Pablo Perticone', 'Adrian Arango'],
    },
}

_cartera_lock = threading.Lock()   # protege lecturas/escrituras de _cartera_comercial


def _sync_cartera_vendedores(saldos: list) -> list:
    """Corrige asignaciones de vendedor en cartera_comercial.json usando saldos como fuente.

    Regla: si un cliente existe en cartera_comercial bajo vendedor A, pero en el reporte
    de saldos aparece bajo vendedor B, y AMBOS están en el mismo equipo de supervisor,
    el cliente se reasigna a vendedor B (el reporte es la fuente más reciente y confiable).
    Solo se hacen cambios intra-equipo para evitar contaminación entre grupos.

    Devuelve lista de strings con los cambios realizados (para logging).
    """
    global _cartera_comercial

    if not saldos or not _cartera_comercial:
        return []

    # Índice inverso: norm_nombre_vendedor → CUIT del supervisor de su equipo.
    # _norm_nombre no está disponible a nivel de módulo en la línea donde se define
    # _SUPERVISOR_MAP, por eso se construye aquí (dentro de la función) donde ya existe.
    _vend_a_equipo: dict = {
        _norm_nombre(nv): sc
        for sc, sd in _SUPERVISOR_MAP.items()
        for nv in ([sd['nombre']] + sd['supervisa'])
    }

    # Construir mapa: norm_nombre_cliente → vendedor_canónico_del_reporte
    saldos_vend: dict = {}
    for fac in saldos:
        if not isinstance(fac, dict):
            continue
        vend = (fac.get('vendedor') or '').strip()
        cli_norm = _norm_nombre(fac.get('cliente') or '')
        if cli_norm and vend and cli_norm not in saldos_vend:
            saldos_vend[cli_norm] = vend

    cambios: list = []
    nueva_cartera = list(_cartera_comercial)

    for i, c in enumerate(nueva_cartera):
        cli_norm    = _norm_nombre(c.get('nombre') or '')
        vend_actual = (c.get('vendedor') or '').strip()
        vend_saldos = saldos_vend.get(cli_norm)

        if not vend_saldos or vend_saldos == vend_actual:
            continue

        # Permitir el cambio solo si ambos vendedores pertenecen al mismo equipo
        equipo_actual = _vend_a_equipo.get(_norm_nombre(vend_actual))
        equipo_saldos = _vend_a_equipo.get(_norm_nombre(vend_saldos))

        if equipo_actual and equipo_saldos and equipo_actual == equipo_saldos:
            nueva_cartera[i] = {**c, 'vendedor': vend_saldos}
            cambios.append(f"{c.get('nombre')}: {vend_actual} → {vend_saldos}")

    if not cambios:
        return []

    # Persistir: escritura atómica (tmp → fsync → rename)
    cc_path = _CC_FILE if os.path.exists(DATA_DIR) else os.path.join(os.getcwd(), 'cartera_comercial.json')
    tmp_path = cc_path + '.sync_tmp'
    try:
        with _cartera_lock:
            with open(tmp_path, 'w', encoding='utf-8') as _f:
                json.dump(nueva_cartera, _f, ensure_ascii=False, indent=2)
                _f.flush()
                os.fsync(_f.fileno())
            os.replace(tmp_path, cc_path)
            _cartera_comercial = nueva_cartera
        print(f"[sync-cartera] {len(cambios)} reasignaciones: {cambios}", flush=True)
    except Exception as _e:
        print(f"[sync-cartera] Error al guardar: {_e}", flush=True)
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except: pass

    return cambios


@app.route("/supervisor")
def supervisor_page():
    resp = send_from_directory('static', 'supervisor.html')
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route("/api/supervisor-cartera/<cuit_supervisor>")
def api_supervisor_cartera(cuit_supervisor):
    """Cartera ampliada para supervisor: sus propios clientes + los de su equipo.
    Incluye aging (días desde fechaFactura de la factura más antigua con saldo > 0),
    score crediticio, saldo total y columna vendedor por cliente."""
    cuit_n = str(cuit_supervisor).replace('-', '').replace(' ', '').strip()
    sup_info = _SUPERVISOR_MAP.get(cuit_n)
    if not sup_info:
        return jsonify({"ok": False, "error": "CUIT no autorizado como supervisor"}), 403

    nombre_propio   = sup_info['nombre']
    nombres_equipo  = [nombre_propio] + sup_info['supervisa']
    # Usar _norm_nombre para comparaciones → tolerante a tildes, espacios y mayúsculas.
    # "Raúl Maza" == "Raul Maza" == "RAUL MAZA" luego de normalizar.
    nombres_norm    = {_norm_nombre(n) for n in nombres_equipo}

    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    def _parse_f(s):
        if not s:
            return None
        s = str(s).strip()
        if '/' in s:
            p = s.split('/')
            try: return datetime(int(p[2]), int(p[1]), int(p[0]))
            except Exception: return None
        if '-' in s and len(s) >= 10:
            p = s.split('-')
            try: return datetime(int(p[0]), int(p[1]), int(p[2]))
            except Exception: return None
        return None

    def _nc(x):
        return str(x or '').replace('-', '').replace(' ', '').strip()

    # ── Base: todos los clientes del equipo desde cartera_comercial ────────────
    base = [c for c in _cartera_comercial
            if _norm_nombre(c.get('vendedor') or '') in nombres_norm]

    # ── Aging, saldos y última factura desde _saldos_gestion ─────────────────
    fuente_s       = _saldos_gestion if _saldos_gestion else _saldos_facturas
    aging_map      = {}   # norm_nombre → max días de factura pendiente con saldo
    saldo_map      = {}   # norm_nombre → saldo total acumulado
    ultima_fac_map = {}   # norm_nombre → datetime de la factura más reciente

    for fac in fuente_s:
        if not isinstance(fac, dict):
            continue
        vend_f = _norm_nombre(fac.get('vendedor') or '')
        if vend_f not in nombres_norm:
            continue
        cli_raw  = (fac.get('cliente') or '').strip()
        cli_norm = _norm_nombre(cli_raw)
        if not cli_norm:
            continue
        saldo_f = float(fac.get('saldo') or 0)
        saldo_map[cli_norm] = saldo_map.get(cli_norm, 0) + saldo_f
        if saldo_f > 0:
            fecha_d = _parse_f(fac.get('fechaFactura'))
            if fecha_d:
                dias = (hoy - fecha_d).days
                aging_map[cli_norm] = max(aging_map.get(cli_norm, 0), dias)
        # Última factura emitida (con o sin saldo) para mostrar en el panel de detalle
        fecha_any = _parse_f(fac.get('fechaFactura'))
        if fecha_any:
            prev = ultima_fac_map.get(cli_norm)
            if not prev or fecha_any > prev:
                ultima_fac_map[cli_norm] = fecha_any

    # ── Fallback: vendedores sin clientes en cartera_comercial → leer de saldos ─
    # Cubre el caso de vendedores (ej. Raúl Maza) que aún no tienen clientes
    # asignados en cartera_comercial.json pero sí aparecen en el reporte de saldos.
    # La comparación usa _norm_nombre para tolerar tildes ("Raúl" == "Raul").
    nombres_con_clientes = {_norm_nombre(c.get('vendedor') or '') for c in base}
    for _nombre_v in nombres_equipo:
        if _norm_nombre(_nombre_v) in nombres_con_clientes:
            continue
        _vistos: set = set()
        for fac in fuente_s:
            if not isinstance(fac, dict):
                continue
            if _norm_nombre(fac.get('vendedor') or '') != _norm_nombre(_nombre_v):
                continue
            cli = (fac.get('cliente') or '').strip()
            if not cli or cli in _vistos:
                continue
            _vistos.add(cli)
            base.append({
                'nombre':        cli,
                'cuit':          '',
                'vendedor':      _nombre_v,
                'ciudad':        '',
                'limiteCredito': 0,
            })
        if _vistos:
            print(f"[sup-cartera] fallback saldos: {_nombre_v} → {len(_vistos)} clientes", flush=True)

    # ── Scores desde archivos de alertas ──────────────────────────────────────
    scores:           dict = {}
    scores_by_nombre: dict = {}
    alertas_cuits:    set  = set()

    def _load_score_file_sup(ruta):
        if not os.path.exists(ruta):
            return
        try:
            with open(ruta, 'r', encoding='utf-8') as _f:
                _ad = json.load(_f)
            for _c in _ad.get('cartera', []):
                _nc_val = _nc(_c.get('cuit', ''))
                if _nc_val:
                    scores[_nc_val] = _c
                _n = (_c.get('nombre') or '').strip()
                if _n:
                    scores_by_nombre[_norm_nombre(_n)] = _c
            if not _ad.get('cartera') and not _ad.get('alertas'):
                for _k, _v in _ad.items():
                    if isinstance(_v, dict) and _v.get('scoreCompleto'):
                        _nc_val = _nc(_k)
                        if _nc_val:
                            scores[_nc_val] = _v
            for _a in _ad.get('alertas', []):
                _nc_val = _nc(_a.get('cuit', ''))
                if _nc_val:
                    alertas_cuits.add(_nc_val)
                _n = (_a.get('nombre') or '').strip()
                if _n and _a.get('scoreCompleto'):
                    scores_by_nombre.setdefault(_norm_nombre(_n), _a)
        except Exception as _e:
            print(f"[sup-cartera] Error cargando scores {ruta}: {_e}", flush=True)

    for _ruta in list(dict.fromkeys([
        os.path.join(os.getcwd(), 'alertas_bcra.json'),   ALERTAS_BCRA_FILE,
        os.path.join(os.getcwd(), 'alertas_cartera.json'), ALERTAS_FILE,
    ])):
        _load_score_file_sup(_ruta)

    # ── Construir resultado ────────────────────────────────────────────────────
    result = []
    for cc in base:
        nombre   = (cc.get('nombre') or '').strip()
        cuit     = (cc.get('cuit')   or '').strip()
        cuit_nc  = _nc(cuit)
        nom_norm = _norm_nombre(nombre)

        sc = scores.get(cuit_nc, {})
        if not sc.get('scoreCompleto'):
            sc = scores_by_nombre.get(nom_norm, {})
        if not sc.get('scoreCompleto'):
            prim2 = ' '.join(nom_norm.split()[:2])
            for _k, _sv in scores_by_nombre.items():
                if ' '.join(_k.split()[:2]) == prim2:
                    sc = _sv
                    break

        total_saldo = saldo_map.get(nom_norm, 0)
        if total_saldo == 0:
            prim2 = ' '.join(nom_norm.split()[:2])
            for k, sv in saldo_map.items():
                if ' '.join(k.split()[:2]) == prim2:
                    total_saldo = sv
                    break

        max_dias = aging_map.get(nom_norm, 0)
        if max_dias == 0:
            prim2 = ' '.join(nom_norm.split()[:2])
            for k, mv in aging_map.items():
                if ' '.join(k.split()[:2]) == prim2:
                    max_dias = mv
                    break

        ultima_fac = ultima_fac_map.get(nom_norm)
        if not ultima_fac:
            prim2 = ' '.join(nom_norm.split()[:2])
            for k, v in ultima_fac_map.items():
                if ' '.join(k.split()[:2]) == prim2:
                    ultima_fac = v
                    break

        limite_credito  = float(cc.get('limiteCredito') or 0)
        cupo_disponible = max(0.0, limite_credito - total_saldo) if limite_credito > 0 else None
        score_val       = sc.get('scoreCompleto') or None
        alerta_temprana = sc.get('alerta_temprana', False)

        result.append({
            'nombre':             nombre,
            'cuit':               cuit,
            'ciudad':             cc.get('ciudad', ''),
            'vendedor':           cc.get('vendedor', ''),
            'email':              cc.get('email', ''),
            'total_saldo':        total_saldo,
            'limite_credito':     limite_credito,
            'cupo_disponible':    cupo_disponible,
            'score':              score_val,
            'scoreRango':         sc.get('scoreRango') or None,
            'scoreColor':         sc.get('scoreColor') or None,
            'ultimaSit':          sc.get('ultimaSit') or 1,
            'alerta':             cuit_nc in alertas_cuits or alerta_temprana,
            'alerta_temprana':    alerta_temprana,
            'max_dias_pendiente': max_dias,
            'ultima_factura':     ultima_fac.strftime('%d/%m/%Y') if ultima_fac else None,
        })

    result.sort(key=lambda x: (0 if x['total_saldo'] > 0 else 1, -(x['total_saldo'] or 0), x['nombre']))

    resp = jsonify({
        "ok":      True,
        "equipo":  nombres_equipo,
        "clientes": result,
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route("/api-v17-scores", methods=["GET"])
@app.route("/scores-cartera", methods=["GET"])   # alias legacy — no rompe clientes viejos
def get_scores_cartera():
    """Devuelve scores de toda la cartera indexados por CUIT normalizado.
    Merge de alertas_bcra.json (queries individuales) + alertas_cartera.json (verificación).
    alertas_cartera tiene prioridad para el mismo CUIT."""
    def _nc2(x):
        return str(x or '').replace('-', '').replace(' ', '').strip()

    scores_out = {}
    archivos_log = []

    for _ruta in list(dict.fromkeys([
        os.path.join(os.getcwd(), 'alertas_bcra.json'), ALERTAS_BCRA_FILE,
        os.path.join(os.getcwd(), 'alertas_cartera.json'), ALERTAS_FILE,
    ])):
        if not os.path.exists(_ruta):
            continue
        try:
            with open(_ruta, 'r', encoding='utf-8') as f:
                _ad = json.load(f)
            antes = len(scores_out)
            # Estructura estándar: {cartera: [...]}
            for c in _ad.get('cartera', []):
                if not c.get('scoreCompleto'):
                    continue
                nc = _nc2(c.get('cuit', ''))
                if nc:
                    scores_out[nc] = {
                        "scoreCompleto":        c.get('scoreCompleto'),
                        "scoreRango":           c.get('scoreRango'),
                        "scoreColor":           c.get('scoreColor'),
                        "scoreEmoji":           c.get('scoreEmoji'),
                        "ultimaSit":            c.get('ultimaSit', 1),
                        "nombre":               c.get('nombre', ''),
                        "alerta_temprana":      c.get('alerta_temprana', False),
                        "bloquear_oportunidad": c.get('bloquear_oportunidad', False),
                        "alerta_logistica":     c.get('alerta_logistica', ''),
                        "inferencia_ingresos":  c.get('inferencia_ingresos'),
                        "fuente_ingresos":      c.get('fuente_ingresos'),
                        "actividad_principal":  c.get('actividad_principal'),
                    }
            # Estructura plana: {cuit: score_data}
            if not _ad.get('cartera') and not _ad.get('alertas'):
                for k, v in _ad.items():
                    if isinstance(v, dict) and v.get('scoreCompleto'):
                        nc = _nc2(k)
                        if nc:
                            scores_out[nc] = {
                                "scoreCompleto":        v.get('scoreCompleto'),
                                "scoreRango":           v.get('scoreRango'),
                                "scoreColor":           v.get('scoreColor'),
                                "scoreEmoji":           v.get('scoreEmoji'),
                                "ultimaSit":            v.get('ultimaSit', 1),
                                "nombre":               v.get('nombre', ''),
                                "alerta_temprana":      v.get('alerta_temprana', False),
                                "bloquear_oportunidad": v.get('bloquear_oportunidad', False),
                                "alerta_logistica":     v.get('alerta_logistica', ''),
                                "inferencia_ingresos":  v.get('inferencia_ingresos'),
                                "fuente_ingresos":      v.get('fuente_ingresos'),
                                "actividad_principal":  v.get('actividad_principal'),
                            }
            nuevos = len(scores_out) - antes
            archivos_log.append(f"{os.path.basename(_ruta)}(+{nuevos})")
        except Exception as e:
            print(f"[scores-cartera] Error {_ruta}: {e}", flush=True)

    # Fallback: score_cache.json — resuelve CUITs cuya escritura en ALERTAS_FILE
    # falló silenciosamente (ej. datos con tipos no serializables en versiones anteriores).
    # Solo agrega CUITs que NO aparecieron en ninguna fuente primaria.
    try:
        with _score_cache_lock:
            _sc_fb = _score_cache_read()
        for _k, _v in _sc_fb.items():
            _nc_k = _nc2(_k)
            if _nc_k and _nc_k not in scores_out and isinstance(_v, dict) and _v.get('score'):
                scores_out[_nc_k] = {
                    "scoreCompleto":        _v.get('score'),
                    "scoreRango":           _v.get('rango'),
                    "scoreColor":           _v.get('color'),
                    "scoreEmoji":           _v.get('emoji'),
                    "ultimaSit":            _v.get('max_sit', 1),
                    "nombre":               _v.get('nombre', ''),
                    "alerta_temprana":      _v.get('alerta_temprana', False),
                    "bloquear_oportunidad": _v.get('bloquear_oportunidad', False),
                    "alerta_logistica":     _v.get('alerta_logistica', ''),
                    "inferencia_ingresos":  _v.get('inferencia_ingresos'),
                    "fuente_ingresos":      _v.get('fuente_ingresos'),
                    "actividad_principal":  _v.get('actividad_principal'),
                }
    except Exception as _sc_e:
        print(f"[scores-cartera] score_cache fallback error: {_sc_e}", flush=True)

    print(f"[scores-cartera] {len(scores_out)} scores totales — {archivos_log}", flush=True)
    if scores_out:
        return jsonify({"ok": True, "scores": scores_out, "total": len(scores_out)})
    return jsonify({"ok": False, "scores": {}, "total": 0})


@app.route("/debug-scores")
def debug_scores():
    """Diagnóstico: muestra qué archivos de scores existen, cuántos tienen score real, y muestra de CUITs."""
    def _nc3(x):
        return str(x or '').replace('-', '').replace(' ', '').strip()

    info = {"data_dir": DATA_DIR, "archivos": {}}
    rutas = {
        "alertas_cartera.json": ALERTAS_FILE,
        "alertas_bcra.json":    ALERTAS_BCRA_FILE,
        "alertas_cartera_cwd":  os.path.join(os.getcwd(), 'alertas_cartera.json'),
        "alertas_bcra_cwd":     os.path.join(os.getcwd(), 'alertas_bcra.json'),
    }
    for nombre, ruta in rutas.items():
        existe = os.path.exists(ruta)
        entry = {"ruta": ruta, "existe": existe}
        if existe:
            try:
                size = os.path.getsize(ruta)
                with open(ruta, 'r', encoding='utf-8') as f:
                    d = json.load(f)
                cartera = d.get('cartera', [])
                con_score = [c for c in cartera if c.get('scoreCompleto')]
                alertas  = d.get('alertas', [])
                entry.update({
                    "size_kb": round(size / 1024, 1),
                    "cartera_total": len(cartera),
                    "cartera_con_score": len(con_score),
                    "alertas_total": len(alertas),
                    "ultima_verif": d.get('ultima_verif', ''),
                    "muestra_cuits_cartera": [_nc3(c.get('cuit','')) for c in cartera[:5]],
                    "muestra_cuits_alertas": [_nc3(a.get('cuit','')) for a in alertas[:5]],
                })
            except Exception as e:
                entry["error"] = str(e)
        info["archivos"][nombre] = entry

    # Muestra de CUITs de cartera_comercial para comparar
    info["cartera_comercial_muestra"] = [
        _nc3(c.get('cuit','')) for c in _cartera_comercial[:5]
    ]
    info["cartera_comercial_total"] = len(_cartera_comercial)
    resp = jsonify(info)
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route("/whatsapp_index.json")
def wsp_index_route():
    return send_from_directory(os.getcwd(), 'whatsapp_index.json')

@app.route("/moras.json")
def moras():
    moras_path = os.path.join(DATA_DIR, 'moras_piattelli.json')
    if os.path.exists(moras_path):
        return send_from_directory(DATA_DIR, 'moras_piattelli.json')
    return send_from_directory(os.getcwd(), 'moras_piattelli.json')

@app.route("/upload-moras", methods=["POST"])
def upload_moras():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Sin archivo"}), 400
        file = request.files['file']
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in ws[1]]
        col_cuit = next((i for i, h in enumerate(headers) if 'cuit' in h), 0)
        col_fecha = next((i for i, h in enumerate(headers) if 'fecha' in h), 1)
        col_saldo = next((i for i, h in enumerate(headers) if 'saldo' in h or 'deuda' in h or 'adeud' in h), 2)
        moras_dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            cuit = str(row[col_cuit] or '').strip().replace('-', '').replace(' ', '')
            if len(cuit) >= 10:
                fecha = str(row[col_fecha] or '').strip()
                saldo = row[col_saldo]
                try:
                    saldo_str = str(saldo).replace('$','').replace(' ','')
                    if ',' in saldo_str and '.' in saldo_str:
                        saldo_str = saldo_str.replace('.','').replace(',','.')
                    elif ',' in saldo_str:
                        saldo_str = saldo_str.replace(',','.')
                    saldo_num = float(saldo_str)
                except:
                    saldo_num = 0
                if cuit in moras_dict:
                    moras_dict[cuit]['saldoAdeudado'] += saldo_num
                    if fecha < moras_dict[cuit]['fechaMora']:
                        moras_dict[cuit]['fechaMora'] = fecha
                else:
                    moras_dict[cuit] = {"fechaMora": fecha, "saldoAdeudado": saldo_num, "enMora": True}
        moras_path = os.path.join(DATA_DIR, 'moras_piattelli.json')
        with open(moras_path, 'w', encoding='utf-8') as f:
            json.dump(moras_dict, f, ensure_ascii=False, indent=2)
        print(f"[moras] Subidas {len(moras_dict)} moras", flush=True)
        return jsonify({"ok": True, "total": len(moras_dict)})
    except Exception as e:
        import traceback
        print(f"[moras] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/cartera_inicial.json")
def cartera_inicial():
    # Sirve _cartera_comercial en memoria — misma fuente que guardar_cliente() escribe.
    # ANTES leía del repo (os.getcwd()) mientras guardar_cliente() escribía en DATA_DIR (/data):
    # eso causaba que los clientes nuevos desaparecieran al recargar.
    resp = jsonify(_cartera_comercial)
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route("/version")
def version():
    import subprocess
    try:
        commit = subprocess.check_output(
            ['git', 'rev-parse', '--short', 'HEAD'], stderr=subprocess.DEVNULL
        ).decode().strip()
    except Exception:
        commit = 'unknown'
    return jsonify({"version": "v77.8", "commit": commit})

@app.route("/datos-bodega", methods=["GET"])
def get_datos_bodega():
    try:
        if os.path.exists(DATOS_FILE):
            with open(DATOS_FILE, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        return jsonify({})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/datos-bodega", methods=["POST"])
def save_datos_bodega():
    try:
        data = request.get_json(force=True)
        with open(DATOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/alertas/limpiar", methods=["POST", "DELETE"])
def limpiar_alertas():
    """Elimina físicamente alertas_cartera.json y alertas_automaticas.json."""
    eliminados = []
    errores = []
    for path, label in [(ALERTAS_FILE, 'alertas_cartera'), (_ALERTAS_AUTO_FILE, 'alertas_automaticas')]:
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"[alertas] Eliminado: {path}", flush=True)
                eliminados.append(label)
        except Exception as e:
            errores.append(f"{label}: {e}")
    if errores:
        return jsonify({"ok": False, "error": "; ".join(errores)}), 500
    msg = "Alertas limpiadas." if eliminados else "No había alertas. La cartera está limpia."
    return jsonify({"ok": True, "mensaje": msg})

@app.route("/alertas", methods=["GET"])
def get_alertas():
    try:
        if os.path.exists(ALERTAS_FILE):
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = {"alertas": [], "ultima_verif": "", "cartera": []}
    except Exception as e:
        data = {"alertas": [], "ultima_verif": "", "cartera": [], "error": str(e)}
    resp = jsonify(data)
    resp.headers['Cache-Control'] = 'no-store'
    return resp

@app.route("/alertas", methods=["POST"])
def save_alertas():
    try:
        data = request.get_json(force=True)
        # Si viene cartera con scores, mergear con la existente
        if data.get('cartera') and os.path.exists(ALERTAS_FILE):
            try:
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as f:
                    existing = json.load(f)
                # Actualizar scores en la cartera existente
                score_map = {c['cuit']: c for c in data['cartera'] if c.get('scoreCompleto')}
                for c in existing.get('cartera', []):
                    if c.get('cuit') in score_map:
                        sc = score_map[c['cuit']]
                        c['scoreCompleto'] = sc.get('scoreCompleto')
                        c['scoreRango'] = sc.get('scoreRango')
                        c['scoreColor'] = sc.get('scoreColor')
                        c['scoreEmoji'] = sc.get('scoreEmoji')
                # Actualizar scores en alertas existentes
                for a in existing.get('alertas', []):
                    if a.get('cuit') in score_map:
                        sc = score_map[a['cuit']]
                        a['scoreCompleto'] = sc.get('scoreCompleto')
                        a['scoreRango'] = sc.get('scoreRango')
                        a['scoreColor'] = sc.get('scoreColor')
                        a['scoreEmoji'] = sc.get('scoreEmoji')
                # Reemplazar alertas si vienen nuevas
                if data.get('alertas') is not None:
                    existing['alertas'] = data['alertas']
                with open(ALERTAS_FILE, 'w', encoding='utf-8') as f:
                    json.dump(existing, f, ensure_ascii=False, indent=2)
                return jsonify({"ok": True})
            except: pass
        with open(ALERTAS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/cartera/eliminar-cliente", methods=["POST"])
def eliminar_cliente_cartera():
    """Elimina un cliente de la cartera y limpia todas sus cachés.
    Body JSON: { "cuit": "30719070481" }
    """
    global _cartera_comercial
    try:
        body = request.get_json(force=True) or {}
        cuit_raw = str(body.get('cuit') or '').replace('-', '').replace(' ', '').strip()
        if len(cuit_raw) < 10:
            return jsonify({"ok": False, "error": "CUIT inválido"}), 400

        _nc = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()

        # 1. Eliminar de cartera_comercial en memoria y en disco
        antes = len(_cartera_comercial)
        _cartera_comercial = [c for c in _cartera_comercial if _nc(c.get('cuit', '')) != cuit_raw]
        if len(_cartera_comercial) == antes:
            return jsonify({"ok": False, "error": "Cliente no encontrado en cartera"}), 404

        _cc_tmp = _CC_FILE + '.tmp'
        with open(_cc_tmp, 'w', encoding='utf-8') as _f:
            json.dump(_cartera_comercial, _f, ensure_ascii=False, indent=2)
            _f.flush(); os.fsync(_f.fileno())
        os.replace(_cc_tmp, _CC_FILE)

        # 2. Limpiar bcra_cache.json
        _limpiados = []
        _bc_path = os.path.join(DATA_DIR, 'bcra_cache.json')
        try:
            if os.path.exists(_bc_path):
                with open(_bc_path, 'r', encoding='utf-8') as _f:
                    _bc = json.load(_f)
                if cuit_raw in _bc:
                    del _bc[cuit_raw]
                    _limpiados.append('bcra_cache')
                    with open(_bc_path, 'w', encoding='utf-8') as _f:
                        json.dump(_bc, _f, ensure_ascii=False)
        except Exception as _e:
            print(f"[eliminar] bcra_cache error: {_e}", flush=True)

        # 3. Limpiar score_cache.json
        with _score_cache_lock:
            _sc = _score_cache_read()
            if cuit_raw in _sc:
                del _sc[cuit_raw]
                _limpiados.append('score_cache')
                _score_cache_write(_sc)

        # 4. Limpiar entrada en ALERTAS_FILE (sección cartera[])
        try:
            with _alertas_file_lock:
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
                    _af = json.load(_f)
                _af['cartera'] = [c for c in (_af.get('cartera') or [])
                                  if _nc(c.get('cuit', '')) != cuit_raw]
                _af_tmp = ALERTAS_FILE + '.tmp'
                with open(_af_tmp, 'w', encoding='utf-8') as _f:
                    json.dump(_af, _f, ensure_ascii=False, default=str)
                    _f.flush(); os.fsync(_f.fileno())
                os.replace(_af_tmp, ALERTAS_FILE)
                _limpiados.append('alertas_cartera')
        except Exception as _e:
            print(f"[eliminar] alertas_file error: {_e}", flush=True)

        # 5. Limpiar session cache en memoria
        _score_session_cache.pop(cuit_raw, None)

        # 6. Limpiar padrón local (bcra_padron.db SQLite) para forzar re-fetch de BCRA
        try:
            if os.path.exists(PADRON_DB_PATH):
                _pl_conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
                _pl_rows = _pl_conn.execute(
                    "DELETE FROM bcra_padron_local WHERE cuit = ?", (cuit_raw,)
                ).rowcount
                _pl_conn.commit()
                _pl_conn.close()
                if _pl_rows:
                    _limpiados.append('padron_local')
        except Exception as _e:
            print(f"[eliminar] padron_local error: {_e}", flush=True)

        print(f"[eliminar] CUIT {cuit_raw} eliminado de cartera. Cachés limpiadas: {_limpiados}", flush=True)
        return jsonify({"ok": True, "cuit": cuit_raw, "caches_limpiadas": _limpiados})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/nombre-custom", methods=["POST"])
def set_nombre_custom():
    """Guarda un nombre personalizado para un CUIT cuando todas las fuentes automáticas fallan.
    Body JSON: { "cuit": "30719070481", "nombre": "BELTE S.R.L" }
    El nombre se persiste en nombres_custom.json (DATA_DIR) y queda como primer resultado
    de /afip/<cuit>, con fuente="custom".
    """
    global _nombres_custom
    try:
        body = request.get_json(force=True) or {}
        cuit_raw = str(body.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        nombre = str(body.get('nombre', '')).strip()
        if not cuit_raw or len(cuit_raw) != 11 or not cuit_raw.isdigit():
            return jsonify({"ok": False, "error": "CUIT inválido"}), 400
        if not nombre:
            # Permite borrar nombre custom si se pasa vacío
            _nombres_custom.pop(cuit_raw, None)
        else:
            _nombres_custom[cuit_raw] = nombre
        with open(NOMBRES_CUSTOM_FILE, 'w', encoding='utf-8') as _nc_w:
            json.dump(_nombres_custom, _nc_w, ensure_ascii=False, indent=2)
        print(f"[nombres_custom] {cuit_raw} → {repr(nombre)}", flush=True)
        return jsonify({"ok": True, "cuit": cuit_raw, "nombre": nombre or None})
    except Exception as _e:
        return jsonify({"ok": False, "error": str(_e)}), 500


@app.route("/verificar-cartera", methods=["POST"])
def verificar_cartera():
    if verificacion_estado["corriendo"]:
        return jsonify({"error": "Ya hay una verificacion en curso"}), 400
    try:
        # Cargar situaciones previas desde db_v17_final.json para comparación correcta de degradación
        _sit_previas: dict = {}
        try:
            if os.path.exists(ALERTAS_FILE):
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
                    _prev = json.load(_f)
                for _c in _prev.get('cartera', []):
                    _nc = str(_c.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
                    if _nc and _c.get('ultimaSit'):
                        _sit_previas[_nc] = {'ultimaSit': _c['ultimaSit'], 'ultimaVerif': _c.get('ultimaVerif')}
        except Exception as _ep:
            print(f"[verif] Advertencia: no se pudo leer sit. previas: {_ep}", flush=True)

        # Leer desde disco — multi-worker: el worker que hizo el upload puede ser distinto a éste
        _cc_base_v = _cc_desde_disco()
        cartera_data = []
        for c in _cc_base_v:
            _cuit = str(c.get("cuit") or "").strip()
            if not _cuit:
                continue
            _nc = _cuit.replace('-', '').replace(' ', '')
            _prev_entry = _sit_previas.get(_nc, {})
            cartera_data.append({
                "cuit":        _cuit,
                "nombre":      str(c.get("nombre") or "").strip(),
                "ciudad":      str(c.get("ciudad") or "").strip(),
                "ultimaSit":   _prev_entry.get('ultimaSit', 1),
                "ultimaVerif": _prev_entry.get('ultimaVerif'),
            })
        if not cartera_data:
            return jsonify({"error": "cartera_comercial.json está vacía o sin CUITs"}), 400
        t = threading.Thread(target=ejecutar_verificacion, args=(cartera_data,), daemon=True)
        t.start()
        return jsonify({"ok": True, "mensaje": f"Verificación iniciada: {len(cartera_data)} clientes desde cartera_comercial.json", "total": len(cartera_data)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _ejecutar_proceso_integral(cartera_data: list, modo_rapido: bool = False):
    import traceback as _tb
    global _proceso_integral_estado

    total       = len(cartera_data)
    _pi_alertas = []   # acumula alertas BCRA detectadas en este proceso

    _cartera_cuits_pi = {
        str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        for c in cartera_data if isinstance(c, dict)
    }

    # Purgar padrón local SQLite — consultar_bcra_cached lo prioriza sobre todo y no tiene TTL.
    # Sin esto, datos históricos (ej: Sit 2 del mes de importación) tapan el estado BCRA actual.
    try:
        if os.path.exists(PADRON_DB_PATH) and _cartera_cuits_pi:
            conn_pi = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
            ph_pi = ','.join('?' * len(_cartera_cuits_pi))
            cur_pi = conn_pi.execute(
                f"DELETE FROM bcra_padron_local WHERE cuit IN ({ph_pi})",
                list(_cartera_cuits_pi)
            )
            conn_pi.commit(); conn_pi.close()
            print(f"[proceso-integral] Padrón local purgado: {cur_pi.rowcount} entradas", flush=True)
    except Exception as _ep:
        print(f"[proceso-integral] Advertencia purga padrón: {_ep}", flush=True)

    # NO se invalida bcra_cache: la TTL de 24h es suficiente para datos BCRA mensuales.
    # Invalidar el caché antes de correr el proceso integral causaba 3900+ requests
    # en ráfaga contra BCRA y rate-limiting después del cliente 4.
    # El caché nocturno (warm-padron) garantiza datos frescos sin golpear el rate-limit.

    # ═══════════════════════════════════════════════════════════════════════
    # FASE 0 — Triage bulk local (historial_detalle + cheques_bcra), sin red.
    # El BCRA en vivo se bloquea/da falsos negativos a partir de la 2da consulta
    # seguida — el bulk local (snapshot diario oficial) es la fuente prioritaria.
    # Solo van a BCRA en vivo los clientes que el bulk no puede resolver con
    # certeza: zona_gris (mora ambigua) y nuevo (sin antecedentes en el bulk).
    # ═══════════════════════════════════════════════════════════════════════
    with _proceso_lock:
        _proceso_integral_estado['mensaje'] = 'Fase 1/2: clasificando cartera con datos bulk (sin red)...'

    _clientes_validos_pi = []
    for c in cartera_data:
        if isinstance(c, dict):
            _cuit_n = str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip()
            if _cuit_n and len(_cuit_n) >= 10:
                _clientes_validos_pi.append((c, _cuit_n))

    _cuits_lista_pi  = [cuit for _, cuit in _clientes_validos_pi]
    _bulk_deudas_pi  = _nomdeu_batch(_cuits_lista_pi)
    _bulk_cheques_pi = _cheques_local_batch(_cuits_lista_pi)

    _categoria_pi: dict = {}
    _cat_count_pi = {'alto_riesgo': 0, 'zona_gris': 0, 'limpio_bulk': 0, 'nuevo': 0}
    for _, cuit in _clientes_validos_pi:
        _cat = _clasificar_bulk(_bulk_deudas_pi.get(cuit))
        _categoria_pi[cuit] = _cat
        _cat_count_pi[_cat] = _cat_count_pi.get(_cat, 0) + 1

    if modo_rapido:
        # Modo rápido: sin BCRA en vivo. Usa datos cacheados tal cual.
        _para_live_pi = []
    else:
        # Modo completo: zona_gris + nuevo + alto_riesgo van a BCRA en vivo.
        # limpio_bulk usa historial_detalle (datos reales per-entidad del bulk mensual BCRA).
        # No se llama BCRA vivo para limpio_bulk: la API bloquea tras pocas consultas.
        _para_live_pi = [
            cuit for cuit in _cuits_lista_pi
            if _categoria_pi.get(cuit) in ('zona_gris', 'nuevo', 'alto_riesgo')
        ]
    print(
        f"[proceso-integral] FASE 0 OK — alto_riesgo={_cat_count_pi['alto_riesgo']} | "
        f"zona_gris={_cat_count_pi['zona_gris']} | limpio_bulk={_cat_count_pi['limpio_bulk']} | "
        f"nuevo={_cat_count_pi['nuevo']} → {len(_para_live_pi)} van a BCRA en vivo"
        + (" (modo rápido: BCRA live omitido)" if modo_rapido else ""),
        flush=True,
    )

    # ═══════════════════════════════════════════════════════════════════════
    # FASE 1 — BCRA en vivo EN PARALELO, solo zona_gris + nuevo + alto_riesgo.
    # ═══════════════════════════════════════════════════════════════════════
    _live_resultados_pi: dict = {}  # {cuit: (bcra_data, cheq_cdi_data, hist_data)}
    if _para_live_pi:
        _N_WORKERS_PI = 8
        with _proceso_lock:
            _proceso_integral_estado['mensaje'] = (
                f'Fase 2/2: BCRA en vivo para {len(_para_live_pi)} clientes ({_N_WORKERS_PI} workers)...'
            )
        print(f"[proceso-integral] FASE 1: BCRA en vivo — {len(_para_live_pi)} clientes", flush=True)

        def _fetch_live_pi(cuit_f):
            try:
                bd, _ = consultar_bcra_cached(cuit_f)
            except Exception as _be:
                print(f'[proceso-integral] BCRA cache fallo {cuit_f}: {_be}', flush=True)
                bd = None
            _sin_datos = (
                not bd or not isinstance(bd, dict) or bd.get('error_bcra') or
                (not (bd.get('results') or {}).get('periodos') and not bd.get('sin_deudas'))
            )
            if _sin_datos:
                _rb, _ = _consultar_respaldo(cuit_f)
                if _rb:
                    bd = _rb
                else:
                    _nomdeu_pi2 = _nomdeu_build_deudas_resp(cuit_f)
                    bd = _nomdeu_pi2 or {}
            try:
                cheq_d, _ = _consultar_bcra_directo(cuit_f, 'cheques')
            except Exception as _ce:
                print(f'[proceso-integral] Cheques CDI fallo {cuit_f}: {_ce}', flush=True)
                cheq_d = None
            if cheq_d and isinstance(cheq_d, dict):
                try:
                    _cheq_path = os.path.join(DATA_DIR, f'cheques_{cuit_f}.json')
                    _tmp_cheq  = _cheq_path + '.tmp'
                    with open(_tmp_cheq, 'w', encoding='utf-8') as _cf:
                        json.dump({'payload': cheq_d, 'ts': time.time()}, _cf, ensure_ascii=False)
                        _cf.flush(); os.fsync(_cf.fileno())
                    os.replace(_tmp_cheq, _cheq_path)
                except Exception:
                    pass
            # Historial 24m en vivo — se busca acá (en paralelo) y NO en el motor de
            # scoring (Paso 2), que de otro modo haría un fetch en vivo SECUENCIAL por
            # cliente y anularía toda la ganancia de la Fase 0/1 (causa real del "se
            # cuelga procesando 1/1283" reportado).
            try:
                hist_d, _ = _consultar_bcra_directo(cuit_f, 'historial', timeout_per_req=8, max_intentos=1)
            except Exception as _he:
                print(f'[proceso-integral] Historial fallo {cuit_f}: {_he}', flush=True)
                hist_d = None
            if hist_d and isinstance(hist_d, dict) and (hist_d.get('results') or {}).get('periodos'):
                try:
                    _hist_path = os.path.join(DATA_DIR, f'historial_{cuit_f}.json')
                    _tmp_hist  = _hist_path + '.tmp'
                    with open(_tmp_hist, 'w', encoding='utf-8') as _hf:
                        json.dump({'payload': hist_d, 'ts': time.time()}, _hf, ensure_ascii=False)
                        _hf.flush(); os.fsync(_hf.fileno())
                    os.replace(_tmp_hist, _hist_path)
                except Exception:
                    pass
            else:
                hist_d = None
            return cuit_f, bd, cheq_d, hist_d

        with ThreadPoolExecutor(max_workers=_N_WORKERS_PI) as _pool_pi:
            _futs_pi = {_pool_pi.submit(_fetch_live_pi, cuit): cuit for cuit in _para_live_pi}
            _done_pi = 0
            for _fut_pi in as_completed(_futs_pi, timeout=1800):
                try:
                    _cuit_r, _bd_r, _cheq_r, _hist_r = _fut_pi.result(timeout=40)
                    _live_resultados_pi[_cuit_r] = (_bd_r, _cheq_r, _hist_r)
                except Exception as _fe_pi:
                    _cuit_r = _futs_pi[_fut_pi]
                    print(f'[proceso-integral] Live fetch fallo {_cuit_r}: {_fe_pi}', flush=True)
                    _live_resultados_pi[_cuit_r] = ({}, None, None)
                _done_pi += 1
                if _done_pi % 20 == 0:
                    print(f"[proceso-integral] FASE 1: {_done_pi}/{len(_para_live_pi)} live", flush=True)
        print(f"[proceso-integral] FASE 1 OK — {len(_live_resultados_pi)} clientes resueltos en vivo", flush=True)
    else:
        print("[proceso-integral] FASE 1: sin clientes para BCRA en vivo — todo resuelto por bulk", flush=True)

    # ═══════════════════════════════════════════════════════════════════════
    # FASE 1.5 — Pre-calentar solvencia (ARCA oficial / TangoFactura) en paralelo.
    # Es una cadena de fetch totalmente independiente de BCRA (Fase 0/1 no la
    # cubre): get_solvency_data hace su propio scraping AFIP con proxies
    # Bright Data/ScraperAPI cuando el caché de 24h está vencido. Sin este
    # pre-calentado, el Paso 3 de la Fase 2 termina haciendo ese fetch en
    # vivo SECUENCIAL por cliente — el motivo real por el que el proceso
    # seguía lento incluso con la cartera ya triada por bulk en Fase 0.
    # get_solvency_data ya valida su propio caché de 24h, así que llamarla
    # para un cliente ya cacheado es gratis (solo lectura de disco).
    # ═══════════════════════════════════════════════════════════════════════
    _para_solvencia_pi = []
    if modo_rapido:
        print("[proceso-integral] FASE 1.5: omitida (modo rápido)", flush=True)
    else:
        for _cuit_sv in _cuits_lista_pi:
            _sv_path_pi = os.path.join(DATA_DIR, f'solvency_{_cuit_sv}.json')
            try:
                if os.path.exists(_sv_path_pi):
                    with open(_sv_path_pi, 'r') as _svf_pi:
                        _sv_cached_pi = json.load(_svf_pi)
                    if time.time() - _sv_cached_pi.get('ts', 0) < 86400:
                        continue
            except Exception:
                pass
            _para_solvencia_pi.append(_cuit_sv)

    if _para_solvencia_pi:
        with _proceso_lock:
            _proceso_integral_estado['mensaje'] = (
                f'Fase 1.5/2: solvencia AFIP para {len(_para_solvencia_pi)} clientes (8 workers)...'
            )
        print(f"[proceso-integral] FASE 1.5: solvencia — {len(_para_solvencia_pi)} clientes sin caché vigente", flush=True)
        with ThreadPoolExecutor(max_workers=8) as _pool_sv:
            _futs_sv = {_pool_sv.submit(get_solvency_data, _cs): _cs for _cs in _para_solvencia_pi}
            _done_sv = 0
            for _fut_sv in as_completed(_futs_sv, timeout=1800):
                try:
                    _fut_sv.result(timeout=40)
                except Exception as _e_sv:
                    print(f'[proceso-integral] Solvencia fallo {_futs_sv[_fut_sv]}: {_e_sv}', flush=True)
                _done_sv += 1
                if _done_sv % 50 == 0:
                    print(f"[proceso-integral] FASE 1.5: {_done_sv}/{len(_para_solvencia_pi)} solvencia", flush=True)
        print("[proceso-integral] FASE 1.5 OK", flush=True)
    else:
        print("[proceso-integral] FASE 1.5: toda la solvencia ya estaba cacheada (<24h)", flush=True)

    # ═══════════════════════════════════════════════════════════════════════
    # FASE 1.8 — Cargar bcra_cache.json una sola vez antes del loop.
    # Para clientes que no van a BCRA en vivo (limpio_bulk / modo_rapido),
    # priorizamos datos reales cacheados de consultas individuales previas
    # sobre los datos sintéticos de _bulk_to_bcra_data().  Esto es lo que
    # hace que el score del proceso integral coincida con el de la consulta
    # individual: ambos usan los mismos datos BCRA.
    # ═══════════════════════════════════════════════════════════════════════
    _bcra_cache_pi: dict = {}
    try:
        _bc_path_pi = os.path.join(DATA_DIR, 'bcra_cache.json')
        if os.path.exists(_bc_path_pi):
            with open(_bc_path_pi, 'r', encoding='utf-8') as _bc_f_pi:
                _bcra_cache_pi = json.load(_bc_f_pi)
            print(f"[proceso-integral] FASE 1.8: bcra_cache.json cargado — {len(_bcra_cache_pi)} CUITs", flush=True)
    except Exception as _bc_load_e:
        print(f"[proceso-integral] FASE 1.8: bcra_cache.json no disponible: {_bc_load_e}", flush=True)

    # ═══════════════════════════════════════════════════════════════════════
    # FASE 1.9 — Cargar score_cache.json para reutilizar scores de consultas
    # individuales recientes. Si un cliente fue consultado en vivo en los últimos
    # 7 días (bcra_cache tiene su entrada con ts fresco), su score ya fue calculado
    # con datos reales del BCRA — el proceso integral lo usa directamente en vez
    # de recalcular con bulk y producir un score diferente.
    # ═══════════════════════════════════════════════════════════════════════
    _score_cache_pi: dict = {}
    try:
        with _score_cache_lock:
            _score_cache_pi = _score_cache_read()
        print(f"[proceso-integral] FASE 1.9: score_cache.json cargado — {len(_score_cache_pi)} CUITs", flush=True)
    except Exception as _sc_load_e:
        print(f"[proceso-integral] FASE 1.9: score_cache.json no disponible: {_sc_load_e}", flush=True)

    _SCORE_CACHE_TTL_PI = 7 * 86400  # 7 días: score de consulta individual es válido

    # ═══════════════════════════════════════════════════════════════════════
    # FASE 2 — Scoring secuencial usando datos ya resueltos (sin I/O BCRA aquí).
    # ═══════════════════════════════════════════════════════════════════════
    for i, c in enumerate(cartera_data):
        if not isinstance(c, dict):
            with _proceso_lock:
                _proceso_integral_estado['procesados'] = i + 1
            continue
        cuit         = str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        nombre       = str(c.get('nombre') or '').strip()
        sit_anterior = int(c.get('ultimaSit', 1) or 1)  # situación previa del cliente

        if not cuit or len(cuit) < 10:
            _proceso_integral_estado['procesados'] = i + 1
            continue

        with _proceso_lock:
            _proceso_integral_estado['cliente_actual'] = nombre or cuit
            _proceso_integral_estado['mensaje'] = f'Procesando {i+1}/{total}: {nombre or cuit}'

        try:
            _score_session_cache.pop(cuit, None)

            # ── Paso 0: Reutilizar score de consulta individual reciente ──────────────
            # _ts en score_cache indica que el score fue calculado por una consulta
            # individual real (datos BCRA en vivo o padron_local fresco). El proceso
            # integral no agrega _ts: sí lo agrega, lo pierde en el siguiente paso.
            # Así se distingue "score individual" de "score bulk anterior".
            _score_cached_pi = _score_cache_pi.get(cuit)
            _score_indiv_ts  = (_score_cached_pi or {}).get('_ts', 0)
            if (_score_indiv_ts and
                    time.time() - _score_indiv_ts < _SCORE_CACHE_TTL_PI and
                    isinstance(_score_cached_pi, dict) and
                    _score_cached_pi.get('score')):
                score_data = _score_cached_pi
                solvency   = get_solvency_data(cuit) or {}
                print(
                    f"[proceso-integral] {nombre or cuit} score reutilizado "
                    f"(consulta individual reciente: {_score_cached_pi.get('score')} {_score_cached_pi.get('rango')})",
                    flush=True,
                )
                with _alertas_file_lock:
                    _actualizar_score_en_cartera(cuit, score_data, solvency)
                with _score_cache_lock:
                    _sc_w = _score_cache_read()
                    _sc_w[cuit] = score_data
                    _score_cache_write(_sc_w)
                with _proceso_lock:
                    _proceso_integral_estado['procesados'] = i + 1
                continue

            # ── Paso 1: BCRA — resuelto por vivo (FASE 1) > caché disco > bulk sintético ──
            # Orden de prioridad:
            #   1. BCRA en vivo (FASE 1)  — datos frescos de la API, 24m de periodos
            #   2. bcra_cache.json        — datos reales de consultas individuales previas
            #   3. _bulk_to_bcra_data()   — sintético, solo 1 periodo; puede divergir del real
            # Usar (2) cuando está disponible elimina la divergencia entre "proceso integral"
            # y "consulta individual": ambos paths terminan usando los mismos datos BCRA.
            if cuit in _live_resultados_pi:
                bcra_data, _cheq_cdi_pi, _hist_live_pi = _live_resultados_pi[cuit]
            else:
                _cheq_cdi_pi  = None
                _hist_live_pi = None
                _deuda_bulk_pi = _bulk_deudas_pi.get(cuit)
                bcra_data = None

                # 1. bcra_cache.json — datos reales de consultas anteriores
                _entry_cache_pi = _bcra_cache_pi.get(cuit)
                if isinstance(_entry_cache_pi, dict):
                    if 'data' in _entry_cache_pi and 'ts' in _entry_cache_pi:
                        _entry_cache_pi = _entry_cache_pi.get('data') or {}
                    if (_entry_cache_pi and
                            isinstance(_entry_cache_pi.get('results'), dict) and
                            not _entry_cache_pi.get('error_bcra')):
                        bcra_data = _entry_cache_pi

                # 2. historial_detalle (bulk real per-entidad) — sit_01/monto_01 por entidad.
                # Datos reales del archivo mensual BCRA: cada entidad tiene su situación
                # y monto real, sin distribución sintética. Misma fuente que _bulk_to_hist_data().
                if bcra_data is None:
                    _hd_bcra_pi = _bulk_bcra_from_historial(cuit, nombre)
                    if _hd_bcra_pi:
                        bcra_data = _hd_bcra_pi

                # 3. Padrón local SQLite — fallback si el CUIT no está en historial_detalle
                if bcra_data is None:
                    _pl_pi = consultar_padron_local(cuit)
                    if (_pl_pi and
                            isinstance((_pl_pi.get('results') or {}).get('periodos'), list) and
                            len((_pl_pi.get('results') or {}).get('periodos') or []) > 0):
                        bcra_data = _pl_pi

                # 4. Bulk sintético — último recurso (agregado, sit_max a todas las entidades)
                if bcra_data is None:
                    bcra_data = _bulk_to_bcra_data(nombre, _deuda_bulk_pi) if _deuda_bulk_pi else {}

            # ── Paso 1.5b: Detectar cheques rechazados activos para alertas ──────────
            _cheq_activos_pi = 0
            _cheq_para_score_pi = _bulk_cheques_pi.get(cuit)  # default seguro si el try de abajo falla
            try:
                _cheq_raw_pi = _cheq_cdi_pi or {}
                # Fallback: caché en disco del CUIT (solo si hubo intento en vivo y falló)
                if not _cheq_raw_pi and cuit in _live_resultados_pi:
                    _cheq_path_pi = os.path.join(DATA_DIR, f'cheques_{cuit}.json')
                    if os.path.exists(_cheq_path_pi):
                        with open(_cheq_path_pi, 'r', encoding='utf-8') as _cpf:
                            _cheq_raw_pi = json.load(_cpf).get('payload') or {}

                _act_live_pi, _, _det_live_pi = _cheques_activos_de(_cheq_raw_pi)
                # Cruzar SIEMPRE contra el bulk local (snapshot diario BCRA, importado
                # por /update-cheques-db) — un "sin_deudas=True" en vivo puede ser un
                # falso negativo, igual que con deudas/historial. Nunca subestimar el
                # riesgo: usar la fuente con más cheques activos.
                _act_bulk_pi, _, _det_bulk_pi = _cheques_activos_de(_bulk_cheques_pi.get(cuit))
                if _act_bulk_pi > _act_live_pi:
                    _cheq_activos_pi, _det_pi = _act_bulk_pi, _det_bulk_pi
                    _cheq_para_score_pi = _bulk_cheques_pi.get(cuit)
                else:
                    _cheq_activos_pi, _det_pi = _act_live_pi, _det_live_pi
                    _cheq_para_score_pi = _cheq_raw_pi

                if _cheq_activos_pi > 0:
                    _pi_alertas.append({
                        'nombre':        nombre,
                        'cuit':          cuit,
                        'tipo':          'cheque',
                        'nroCheques':    _cheq_activos_pi,
                        'totalCheques':  len(_det_pi),
                        'fecha':         time.strftime('%d/%m/%Y'),
                        'scoreCompleto': None,
                        'scoreRango':    None,
                        'scoreColor':    None,
                        'scoreEmoji':    None,
                    })
                    print(
                        f'[proceso-integral] ALERTA CHEQUES: {nombre} ({cuit})'
                        f' — {_cheq_activos_pi}/{len(_det_pi)} cheque(s) activo(s)',
                        flush=True,
                    )
            except Exception as _cal_e:
                print(f'[proceso-integral] Cheques alert parse fallo {cuit}: {_cal_e}', flush=True)

            # ── Paso 2: Score — idéntico a consulta individual ───────────────────────
            # calcular_score_servidor usa exactamente la misma cadena de fuentes que la
            # consulta individual cuit por cuit: caché disco (historial_<cuit>.json)
            # → BCRA vivo 24m (8 s timeout, 1 intento) → bulk fallback (_bulk_to_hist_data).
            # Para cheques: DB local (bcra_nomdeu.db) → BCRA vivo.
            # Esto garantiza score idéntico entre "actualizar toda la cartera" y la
            # consulta individual de un CUIT.
            try:
                score_data = calcular_score_servidor(
                    cuit=cuit, bcra_data=bcra_data or {},
                    ciudad=str(c.get('ciudad', '') or ''),
                )
                if score_data:
                    score_data['bcra_disponible'] = bool(bcra_data) and not bool((bcra_data or {}).get('error_bcra'))
            except Exception as _sce:
                _tb_full = _tb.format_exc().strip()
                _err_msg = f"{type(_sce).__name__}: {str(_sce)[:400]}"
                _err_linea = _tb_full.split('\n')[-1][:200]
                print(
                    f'[proceso-integral] Score falló #{i+1} {cuit} ({nombre})\n'
                    f'  {_err_msg}\n  {_err_linea}\n  traceback:\n{_tb_full}',
                    flush=True,
                )
                with _proceso_lock:
                    _proceso_integral_estado['errores'] += 1
                    _proceso_integral_estado['log_errores'].append({
                        'num': i + 1, 'cuit': cuit, 'nombre': nombre,
                        'tipo': 'ScoreError', 'mensaje': _err_msg, 'linea': _err_linea,
                    })
                score_data = {
                    'score': None, 'rango': 'Error', 'color': '#6b7280', 'emoji': '⚠️',
                    'max_sit': 1, 'bcra_disponible': False,
                    'alerta_temprana': False, 'bloquear_oportunidad': False, 'alerta_logistica': '',
                    '_error_paso2': _err_msg,
                }

            # ── Paso 3: Solvencia — try separado; normalizar si retorna lista ──────────
            try:
                solvency = get_solvency_data(cuit)
                if not isinstance(solvency, dict):
                    solvency = {}
            except Exception as _se:
                print(f'[proceso-integral] Solvencia fallo {cuit}: {_se} — continuando sin AFIP', flush=True)
                solvency = {}

            # ── Contingencia: garantizar score numérico en todos los casos ──────────
            if not score_data.get('score'):
                _ms_c = 1
                if isinstance(bcra_data, dict):
                    _per_c = _safe_periodos((bcra_data.get('results') or {}).get('periodos') or [])
                    if _per_c and isinstance(_per_c[0], dict):
                        _ents_c = [e for e in _per_c[0].get('entidades', []) if isinstance(e, dict)]
                        if _ents_c:
                            _ms_c = max((e.get('situacion', 1) or 1) for e in _ents_c)
                if   _ms_c == 1: _sc_c, _rg_c, _cl_c, _em_c = 650, 'Bueno',       '#ca8a04', '🟡'
                elif _ms_c == 2: _sc_c, _rg_c, _cl_c, _em_c = 480, 'Revisar',     '#ea580c', '🟠'
                else:            _sc_c, _rg_c, _cl_c, _em_c = 320, 'Alto riesgo', '#dc2626', '🔴'
                score_data.update({
                    'score': _sc_c, 'rango': _rg_c, 'color': _cl_c, 'emoji': _em_c,
                    'max_sit': _ms_c, '_contingencia': True,
                })
                print(
                    f'[proceso-integral] Contingencia BCRA sit={_ms_c} → {_sc_c} {_rg_c} '
                    f'({cuit} {nombre})',
                    flush=True,
                )

            # ── Paso 3.5: Detección de alerta BCRA ───────────────────────────────────
            # Usa max_sit ya calculado en score_data — sin llamada extra a la API
            _max_sit_pi = int(score_data.get('max_sit', 1) or 1)
            if score_data.get('score') and (_max_sit_pi > sit_anterior or _max_sit_pi >= 3):
                _pi_alertas.append({
                    'nombre':        nombre,
                    'cuit':          cuit,
                    'sitAnterior':   sit_anterior,
                    'sitActual':     _max_sit_pi,
                    'fecha':         time.strftime('%d/%m/%Y'),
                    'tipo':          'bcra',
                    'scoreCompleto': score_data.get('score'),
                    'scoreRango':    score_data.get('rango'),
                    'scoreColor':    score_data.get('color'),
                    'scoreEmoji':    score_data.get('emoji'),
                })
                print(
                    f'[proceso-integral] ALERTA BCRA: {nombre} sit {sit_anterior}→{_max_sit_pi}',
                    flush=True,
                )

            # Enriquecer alertas de cheques de este cliente con el score ya calculado
            if _cheq_activos_pi and score_data.get('score'):
                for _alerta_ch in _pi_alertas:
                    if (_alerta_ch.get('tipo') == 'cheque' and
                            _alerta_ch.get('cuit') == cuit and
                            not _alerta_ch.get('scoreCompleto')):
                        _alerta_ch['scoreCompleto'] = score_data.get('score')
                        _alerta_ch['scoreRango']    = score_data.get('rango')
                        _alerta_ch['scoreColor']    = score_data.get('color')
                        _alerta_ch['scoreEmoji']    = score_data.get('emoji')

            # ── Paso 4: Persistencia atómica ──────────────────────────────────────────
            with _alertas_file_lock:
                _actualizar_score_en_cartera(cuit, score_data, solvency)

            resp = _score_response(score_data, solvency)

            with _score_cache_lock:
                _sc = _score_cache_read()
                _sc[cuit] = resp
                _score_cache_write(_sc)

            # Flush incremental de alertas_cartera.json cada 50 clientes
            if (i + 1) % 50 == 0:
                print(f'[proceso-integral] Checkpoint {i+1}/{total} — flush disco', flush=True)

        except Exception as e:
            err_tipo = type(e).__name__
            err_msg  = str(e)[:300]
            tb_last  = _tb.format_exc().strip().split('\n')[-1][:200]
            print(
                f'[proceso-integral] ERROR {i+1}/{total} — {cuit} ({nombre})\n'
                f'  {err_tipo}: {err_msg}\n  {tb_last}',
                flush=True,
            )
            with _proceso_lock:
                _proceso_integral_estado['errores'] += 1
                _proceso_integral_estado['log_errores'].append({
                    'num': i + 1, 'cuit': cuit, 'nombre': nombre,
                    'tipo': err_tipo, 'mensaje': err_msg, 'linea': tb_last,
                })

        with _proceso_lock:
            _proceso_integral_estado['procesados'] = i + 1
        # Sin sleep: ya no hay I/O a BCRA en este loop (se resolvió en FASE 0/1,
        # bulk + paralelo). Lo único que corre acá es CPU (scoring) y disco local.

    # ── Merge atómico de alertas BCRA en db_v17_final.json ───────────────────────
    # Preserva alertas tipo 'bodegas' (WhatsApp) del run anterior; reemplaza las 'bcra'
    try:
        with _alertas_file_lock:
            try:
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as _af:
                    _af_data = json.load(_af)
            except Exception:
                _af_data = {}
            # Preservar solo alertas de tipo 'bodegas' (WhatsApp) — regenerar bcra y cheque
            _alertas_prev_otros = [
                a for a in _af_data.get('alertas', [])
                if a.get('tipo') not in ('bcra', 'cheque')
            ]
            _af_data['alertas'] = _alertas_prev_otros + _pi_alertas
            _tmp_pa = ALERTAS_FILE + '.tmp'
            with open(_tmp_pa, 'w', encoding='utf-8') as _af:
                json.dump(_af_data, _af, ensure_ascii=False, default=str)
                _af.flush(); os.fsync(_af.fileno())
            os.replace(_tmp_pa, ALERTAS_FILE)
        _n_bcra_a = sum(1 for a in _pi_alertas if a.get('tipo') == 'bcra')
        _n_cheq_a = sum(1 for a in _pi_alertas if a.get('tipo') == 'cheque')
        print(
            f'[proceso-integral] Alertas guardadas: {_n_bcra_a} BCRA, {_n_cheq_a} cheques '
            f'({len(_alertas_prev_otros)} bodegas preservadas)',
            flush=True,
        )
    except Exception as _ae:
        print(f'[proceso-integral] Error guardando alertas: {_ae}', flush=True)

    # Persistir cartera al finalizar
    try:
        with open(_CC_FILE, 'w', encoding='utf-8') as _f:
            json.dump(_cartera_comercial, _f, ensure_ascii=False, indent=2)
    except Exception as _e:
        print(f'[proceso-integral] Error guardando cartera: {_e}', flush=True)

    with _proceso_lock:
        n_ok       = _proceso_integral_estado['procesados'] - _proceso_integral_estado['errores']
        _n_bcra_f  = sum(1 for a in _pi_alertas if a.get('tipo') == 'bcra')
        _n_cheq_f  = sum(1 for a in _pi_alertas if a.get('tipo') == 'cheque')
        _proceso_integral_estado['corriendo'] = False
        _proceso_integral_estado['mensaje'] = (
            f'Completado — {n_ok} OK, {_proceso_integral_estado["errores"]} errores'
            f' | {total} clientes · {_n_bcra_f} alerta(s) BCRA · {_n_cheq_f} alerta(s) cheques'
        )
    print(f'[proceso-integral] {_proceso_integral_estado["mensaje"]}', flush=True)


@app.route("/proceso-integral", methods=["POST"])
@require_login
def iniciar_proceso_integral():
    global _proceso_integral_estado
    if _proceso_integral_estado.get("corriendo"):
        return jsonify({"error": "Ya hay un proceso en curso — esperá que termine"}), 400
    if verificacion_estado.get("corriendo"):
        return jsonify({"error": "Hay una verificación BCRA en curso — esperá que termine"}), 400
    # Cargar sit. anteriores desde db_v17_final.json para comparación correcta de degradación
    _sit_prev_pi: dict = {}
    try:
        if os.path.exists(ALERTAS_FILE):
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
                _prev_pi = json.load(_f)
            for _c in _prev_pi.get('cartera', []):
                _nc = str(_c.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
                if _nc and _c.get('ultimaSit'):
                    _sit_prev_pi[_nc] = int(_c['ultimaSit'])
    except Exception as _ep:
        print(f"[proceso-integral] Advertencia sit. previas: {_ep}", flush=True)

    # Leer desde disco — multi-worker: el worker que hizo el upload puede ser distinto a éste
    _cc_base_pi = _cc_desde_disco()
    cartera = []
    for c in _cc_base_pi:
        _cuit_pi = str(c.get("cuit", "") or "").strip()
        if not _cuit_pi:
            continue
        _nc_pi = _cuit_pi.replace('-', '').replace(' ', '')
        cartera.append({**c, "ultimaSit": _sit_prev_pi.get(_nc_pi, 1)})
    if not cartera:
        return jsonify({"error": "Cartera vacía — cargá cartera_comercial.json en el servidor"}), 400
    # ── Marcar como corriendo ANTES de lanzar el thread (elimina race condition) ──
    import datetime as _dt
    with _proceso_lock:
        _proceso_integral_estado.update({
            "corriendo": True, "total": len(cartera), "procesados": 0,
            "errores": 0, "cliente_actual": "", "mensaje": "Iniciando proceso...",
            "iniciado_en": _dt.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "log_errores": [],
        })
    _modo_rapido = bool((request.get_json(silent=True) or {}).get('modo_rapido', False))
    t = threading.Thread(target=_ejecutar_proceso_integral, args=(cartera, _modo_rapido), daemon=True)
    t.start()
    _modo_label = " (modo rápido: solo bulk)" if _modo_rapido else ""
    return jsonify({"ok": True, "total": len(cartera), "corriendo": True,
                    "mensaje": f"Proceso integral iniciado{_modo_label}: {len(cartera)} clientes"}), 202


@app.route("/proceso-integral/progreso")
def progreso_proceso_integral():
    with _proceso_lock:
        return jsonify(dict(_proceso_integral_estado))


@app.route("/api/proceso-status")
def proceso_status():
    with _proceso_lock:
        return jsonify(dict(_proceso_integral_estado))


@app.route("/api/reprocesar_vacios", methods=["POST"])
@require_login
def reprocesar_vacios():
    """Reprocesa solo los clientes de la cartera cuyo score esté ausente o sea nulo.
    Usa la misma lógica de _ejecutar_proceso_integral (AFIP SDK + BCRA reintentos v60)."""
    global _proceso_integral_estado
    if _proceso_integral_estado.get("corriendo"):
        return jsonify({"error": "Ya hay un proceso corriendo — esperá que termine"}), 400

    _nc = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()

    # Leer score_cache.json para determinar quién ya tiene score real
    sc = _score_cache_read()

    pendientes = []
    for c in _cartera_comercial:
        cuit = _nc(c.get('cuit'))
        if not cuit or len(cuit) < 10:
            continue
        sd = sc.get(cuit)
        if not sd or not sd.get('score'):
            pendientes.append({
                'cuit':   cuit,
                'nombre': str(c.get('nombre') or '').strip(),
                'ciudad': str(c.get('ciudad') or '').strip(),
            })

    if not pendientes:
        return jsonify({"ok": True, "mensaje": "Todos los clientes ya tienen score calculado.", "total": 0})

    nombres_muestra = [p['nombre'] or p['cuit'] for p in pendientes[:5]]
    print(f"[reprocesar_vacios] {len(pendientes)} sin score: {nombres_muestra}", flush=True)

    # Limpiar cachés de historial/cheques corruptos para los CUITs a reprocesar
    _purgados = 0
    for p in pendientes:
        for _prefix in ('historial_', 'cheques_'):
            _fp = os.path.join(DATA_DIR, f"{_prefix}{p['cuit']}.json")
            if os.path.exists(_fp):
                try:
                    os.remove(_fp)
                    _purgados += 1
                except: pass
    if _purgados:
        print(f"[reprocesar_vacios] Purgados {_purgados} cachés de hist/cheq corruptos", flush=True)

    with _proceso_lock:
        _proceso_integral_estado.update({
            "corriendo":    True,
            "total":        len(pendientes),
            "procesados":   0,
            "errores":      0,
            "cliente_actual": "",
            "mensaje":      f"Reprocesando {len(pendientes)} clientes sin score...",
            "iniciado_en":  datetime.utcnow().isoformat(),
            "log_errores":  [],
        })
    t = threading.Thread(target=_ejecutar_proceso_integral, args=(pendientes,), daemon=True)
    t.start()
    return jsonify({
        "ok":     True,
        "total":  len(pendientes),
        "corriendo": True,
        "mensaje": f"Iniciado: {len(pendientes)} clientes sin score",
        "muestra": nombres_muestra,
    }), 202


@app.route("/recalcular-pendientes", methods=["POST"])
def recalcular_pendientes():
    """Recalcula solo los clientes de cartera_comercial.json sin score o con verificación fallida."""
    if verificacion_estado["corriendo"]:
        return jsonify({"error": "Ya hay una verificación en curso — esperá que termine o usá /verificar-reset"}), 400
    try:
        # Estado actual de alertas_cartera.json
        try:
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
                alertas_data = json.load(_f)
            cartera_status = {
                str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip(): c
                for c in alertas_data.get('cartera', [])
            }
        except Exception:
            cartera_status = {}

        pendientes = []
        for c in _cartera_comercial:
            cuit = str(c.get('cuit') or '').replace('-', '').replace(' ', '').strip()
            if not cuit:
                continue
            st = cartera_status.get(cuit, {})
            sin_score   = not st.get('scoreCompleto')
            pendiente   = st.get('pendiente') is True
            fallido     = st.get('verificacion_fallida') is True
            if sin_score or pendiente or fallido:
                pendientes.append({
                    "cuit":       cuit,
                    "nombre":     str(c.get('nombre') or '').strip(),
                    "ciudad":     str(c.get('ciudad') or '').strip(),
                    "ultimaSit":  c.get('ultimaSit', 1),
                    "ultimaVerif": st.get('ultimaVerif'),
                })

        if not pendientes:
            return jsonify({"ok": True, "mensaje": "Todos los clientes ya tienen score. Nada que recalcular.", "total": 0})

        nombres = [p['nombre'] or p['cuit'] for p in pendientes[:5]]
        print(f"[recalc-pendientes] {len(pendientes)} pendientes: {nombres}{'...' if len(pendientes)>5 else ''}", flush=True)
        t = threading.Thread(target=ejecutar_verificacion, args=(pendientes,), daemon=True)
        t.start()
        return jsonify({
            "ok":      True,
            "mensaje": f"Recálculo iniciado: {len(pendientes)} clientes pendientes.",
            "total":   len(pendientes),
            "muestra": nombres,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/verificar-progreso", methods=["GET"])
def verificar_progreso():
    return jsonify(verificacion_estado)

@app.route("/verificar-reset", methods=["POST", "GET"])
def verificar_reset():
    estaba_corriendo = verificacion_estado["corriendo"]
    verificacion_estado["corriendo"] = False
    verificacion_estado["mensaje"] = "Reset manual. Listo para nueva verificacion."
    print(f"[verif] Reset manual. Estaba corriendo: {estaba_corriendo}", flush=True)
    return jsonify({"ok": True, "estaba_corriendo": estaba_corriendo})


@app.route("/limpiar-solvency", methods=["POST"])
def limpiar_solvency():
    """Elimina todos los solvency_*.json en DATA_DIR para forzar re-scraping en cartera completa."""
    import glob as _glob
    eliminados = 0
    try:
        for _fp in _glob.glob(os.path.join(DATA_DIR, 'solvency_*.json')):
            os.remove(_fp)
            eliminados += 1
        print(f"[limpiar-solvency] {eliminados} archivos eliminados", flush=True)
        return jsonify({"ok": True, "eliminados": eliminados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _score_response(score_data: dict, solvency: dict = None, cheq_data: dict = None) -> dict:
    """Pasamanos transparente: devuelve score_data completo + campos de solvencia + cheques.
    Usa json.dumps(default=str) para serializar sin excepciones de tipo."""
    sol = solvency if isinstance(solvency, dict) else {}
    # Serializar y deserializar con default=str para eliminar cualquier tipo
    # Python no serializable (datetime, Decimal, etc.) antes de enviar al frontend.
    try:
        _safe = json.loads(json.dumps(score_data, default=str))
    except Exception as _se:
        print(f"[score_response] Error serializando score_data: {_se}", flush=True)
        _safe = {"score": score_data.get("score"), "rango": score_data.get("rango")}

    _safe["ok"]                   = True
    _safe["version"]              = _SCORE_VERSION
    _safe["bcra_disponible"]      = score_data.get('bcra_disponible', True)
    _safe["inferencia_ingresos"]  = sol.get('ingresos_anuales')
    _safe["fuente_ingresos"]      = sol.get('fuente_ingresos')
    _safe["actividad_principal"]  = sol.get('actividad_principal')
    # Razón social para rescatar el nombre cuando AFIP y BCRA estuvieron offline
    _safe["razon_social"]         = (sol.get('razon_social') or sol.get('nombre') or '').strip()

    # ── Perfil fiscal — poblado por el Padrón A13 de ARCA ─────────────────────
    # antiguedad_fiscal: años desde inscripción en ARCA (int o None)
    # estado_empleo: 'activo' | 'monotrib' | None (derivado de impuestos activos)
    # juicios_comerciales: pendiente de integrar (int o None)
    _safe["antiguedad_fiscal"]   = sol.get('antiguedad_fiscal')   or sol.get('antiguedad_anos')
    _safe["estado_empleo"]       = sol.get('estado_empleo')       or None
    _safe["juicios_comerciales"] = sol.get('juicios_comerciales') or None
    _safe["estado_clave_afip"]   = sol.get('estado_clave')        or None
    _safe["n_impuestos_activos"] = sol.get('n_impuestos_activos')
    _safe["tiene_iva"]           = sol.get('tiene_iva')
    _safe["fecha_inscripcion"]   = sol.get('fecha_inicio')        or None
    _safe["categoria_monotrib"]  = sol.get('categoria_monotrib')  or None
    _safe["tipo_persona"]        = sol.get('tipo_persona')        or None
    # ── Módulo MiPyME — padrón Min. Producción (datos oficiales SEPYME) ─────────
    _safe["categoria_mipyme"]    = sol.get('categoria_mipyme')    or None
    _safe["sector_mipyme"]       = sol.get('sector_mipyme')       or None
    _safe["empleados_rango"]     = sol.get('empleados_rango')     or None

    _safe.setdefault("override_admin",       False)
    _safe.setdefault("mora_administrativa",  False)
    _safe.setdefault("deuda_90d_interna",    False)
    _safe.setdefault("monto_deuda_90d",      0)
    _safe.setdefault("bloquear_oportunidad", False)
    _safe.setdefault("razonamiento_score",   None)
    _safe.setdefault("mora_tecnica",         False)
    _safe.setdefault("nota_mora_tecnica",    None)
    _safe.setdefault("semaforo",             'verde' if (_safe.get('score') or 0) >= 700 else ('amarillo' if (_safe.get('score') or 0) >= 400 else 'rojo'))

    # Cheques rechazados — incluidos para que el frontend los muestre sin fetch extra
    if isinstance(cheq_data, dict):
        _safe["cheques_data"] = cheq_data

    # LOG DE CONTROL: campos críticos que el frontend necesita
    print(
        f"[score_response] score={_safe.get('score')} rango={_safe.get('rango')} "
        f"override_admin={_safe.get('override_admin')} "
        f"mora_administrativa={_safe.get('mora_administrativa')} "
        f"deuda_90d={_safe.get('deuda_90d_interna')} "
        f"bloquear={_safe.get('bloquear_oportunidad')}",
        flush=True
    )
    return _safe


def _score_cache_read() -> dict:
    """Lee score_cache.json; devuelve {} en cualquier error."""
    try:
        with open(SCORE_CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _score_cache_write(data: dict):
    tmp = SCORE_CACHE_FILE + '.tmp'
    try:
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, default=str)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, SCORE_CACHE_FILE)
    except Exception as e:
        print(f"[score_cache] write error: {e}", flush=True)
        try:
            os.remove(tmp)
        except OSError:
            pass


# ═══════════════════════════════════════════════════════════════════════════
# ALERTAS AUTOMÁTICAS — detección post-upload y cheques diarios
# ═══════════════════════════════════════════════════════════════════════════

_ALERTAS_AUTO_FILE = os.path.join(DATA_DIR, 'alertas_automaticas.json')
_alertas_auto_lock = threading.Lock()
_MAX_ALERTAS        = 300   # máximo histórico en disco


def _alertas_auto_read() -> list:
    try:
        if os.path.exists(_ALERTAS_AUTO_FILE):
            with open(_ALERTAS_AUTO_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return []


def _alertas_auto_write(alertas: list) -> None:
    tmp = _ALERTAS_AUTO_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(alertas[:_MAX_ALERTAS], f, ensure_ascii=False, indent=2)
        f.flush(); os.fsync(f.fileno())
    os.replace(tmp, _ALERTAS_AUTO_FILE)


def _agregar_alertas_auto(nuevas: list) -> None:
    """Prepend nuevas alertas y persiste, manteniendo máximo _MAX_ALERTAS."""
    if not nuevas:
        return
    with _alertas_auto_lock:
        existentes = _alertas_auto_read()
        # Deduplicar: no agregar si mismo cuit + tipo + fecha ya está
        ids_exist = {(a.get('cuit'), a.get('tipo'), a.get('fecha', '')[:10]) for a in existentes}
        filtradas = [a for a in nuevas if (a.get('cuit'), a.get('tipo'), a.get('fecha', '')[:10]) not in ids_exist]
        if filtradas:
            _alertas_auto_write(filtradas + existentes)


def _recalcular_scores_post_upload():
    """
    Background: recalcula score de toda la cartera usando saldos recién subidos
    y BCRA data cacheada en disco — sin llamadas a la API de BCRA.
    Genera alertas si score baja >50 pts o DSO deteriora >15%.
    """
    import time as _t
    _t.sleep(2)   # dejar que el response HTTP salga primero
    try:
        cartera = list(_cartera_comercial)
        if not cartera:
            return

        facturas_mem = list(_saldos_facturas or [])
        if not facturas_mem:
            print("[recalculo] Sin saldos en memoria — abortando", flush=True)
            return

        _nc = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()

        # Cargar bcra_cache.json una sola vez
        bcra_cache_data: dict = {}
        try:
            bc_path = os.path.join(DATA_DIR, 'bcra_cache.json')
            if os.path.exists(bc_path):
                with open(bc_path, 'r', encoding='utf-8') as f:
                    bcra_cache_data = json.load(f)
        except Exception as e:
            print(f"[recalculo] bcra_cache.json: {e}", flush=True)

        # Leer moras internas
        moras_norm: set = set()
        try:
            mp = os.path.join(DATA_DIR, 'moras_piattelli.json')
            if os.path.exists(mp):
                with open(mp, 'r', encoding='utf-8') as f:
                    raw = json.load(f)
                moras_norm = {_nc(str(x)) for x in (raw if isinstance(raw, list) else [])}
        except Exception:
            pass

        # Batch cheques locales para toda la cartera
        cuits_all = [_nc(c.get('cuit', '')) for c in cartera if c.get('cuit')]
        cheques_batch = _cheques_local_batch(cuits_all)

        # Score cache anterior para comparar deltas
        with _score_cache_lock:
            sc_anterior = _score_cache_read()

        # Índice nombre normalizado → cuit (para verificar qué clientes tienen saldos)
        _nn = lambda x: str(x or '').upper().strip()
        nombres_con_saldo: set = {_nn(f.get('cliente', '')) for f in facturas_mem}

        # Índice saldo total por nombre (para detectar exceso de límite de crédito)
        saldo_por_nombre: dict = {}
        for _f in facturas_mem:
            _nn_f = _nn(str(_f.get('cliente', '') or ''))
            _sal_f = float(_f.get('saldo', 0) or 0)
            if _nn_f and _sal_f > 0:
                saldo_por_nombre[_nn_f] = saldo_por_nombre.get(_nn_f, 0) + _sal_f

        alertas_nuevas  = []
        scores_nuevos   = {}
        procesados      = 0

        for cliente in cartera:
            cuit = _nc(cliente.get('cuit', ''))
            if not cuit or len(cuit) < 10:
                continue

            nombre = str(cliente.get('nombre') or cliente.get('cliente') or '')
            # Solo recalcular si este cliente tiene filas en el upload
            if _nn(nombre) not in nombres_con_saldo:
                continue

            bcra_data = bcra_cache_data.get(cuit)
            if not bcra_data:
                continue   # sin BCRA en caché → score incompleto, skip
            # bcra_cache.json guarda en formato {data, error, ts} — extraer el contenido real
            if isinstance(bcra_data, dict) and 'data' in bcra_data and 'ts' in bcra_data:
                bcra_data = bcra_data.get('data') or {}
                if not bcra_data:
                    continue

            hist_data = None
            try:
                h_path = os.path.join(DATA_DIR, f'historial_{cuit}.json')
                if os.path.exists(h_path):
                    with open(h_path, 'r', encoding='utf-8') as _hf:
                        hist_data = json.load(_hf)
            except Exception:
                pass

            cheq_data  = cheques_batch.get(cuit)
            ciudad     = str(cliente.get('ciudad') or '').strip()
            en_mora    = cuit in moras_norm

            # Limpiar session cache para forzar recálculo real
            _score_session_cache.pop(cuit, None)
            try:
                score_nuevo = calcular_rating_predictivo(
                    cuit=cuit, bcra_data=bcra_data,
                    hist_data=hist_data, cheq_data=cheq_data,
                    en_mora=en_mora, ciudad=ciudad,
                )
            except Exception as e:
                print(f"[recalculo] {cuit}: {e}", flush=True)
                continue
            finally:
                _score_session_cache.pop(cuit, None)   # no contaminar otros requests

            if not score_nuevo or not score_nuevo.get('score'):
                continue

            procesados += 1
            score_nuevo_val = int(score_nuevo.get('score', 0))

            _sc_ant      = sc_anterior.get(cuit) or {}
            score_ant_val = _sc_ant.get('scoreCompleto') or _sc_ant.get('score')
            fecha_ahora  = datetime.now().strftime('%Y-%m-%d %H:%M')

            alerta_base = {
                'cuit':       cuit,
                'nombre':     nombre,
                'fecha':      fecha_ahora,
                'rango':      score_nuevo.get('rango', ''),
                'score_nuevo': score_nuevo_val,
                'leida':      False,
            }

            if score_ant_val:
                delta = score_nuevo_val - int(score_ant_val)
                if delta <= -50:
                    alertas_nuevas.append({
                        **alerta_base,
                        'tipo':           'score_drop',
                        'detalle':        f"Score bajó de {score_ant_val} a {score_nuevo_val} ({delta:+d} pts)",
                        'score_anterior': int(score_ant_val),
                        'delta':          delta,
                    })

            if score_nuevo.get('dso_deteriorando'):
                alertas_nuevas.append({
                    **alerta_base,
                    'tipo':           'dso_deteriora',
                    'detalle':        f"DSO deterioró >15% en 60 días · score actual {score_nuevo_val}",
                    'score_anterior': int(score_ant_val) if score_ant_val else None,
                })

            # Límite de crédito superado (≥90% de uso)
            _limite_cred = float(cliente.get('limiteCredito') or 0)
            _saldo_total = saldo_por_nombre.get(_nn(nombre), 0)
            if _limite_cred > 0 and _saldo_total >= _limite_cred * 0.90:
                _uso_pct = int(_saldo_total / _limite_cred * 100)
                alertas_nuevas.append({
                    **alerta_base,
                    'tipo':    'limite_excedido',
                    'detalle': f"Saldo ${_saldo_total:,.0f} ({_uso_pct}% del límite de ${_limite_cred:,.0f})",
                })

            # Deterioro estructural bancario sostenido
            if score_nuevo.get('deterioro_estructural'):
                alertas_nuevas.append({
                    **alerta_base,
                    'tipo':    'deterioro_bcra',
                    'detalle': f"Problemas bancarios sostenidos · score {score_nuevo_val}",
                })

            # Deuda interna antigua (>90 días sin pagar en Odoo)
            _monto_90d = float(score_nuevo.get('monto_deuda_90d') or 0)
            if score_nuevo.get('deuda_90d_interna') and _monto_90d > 50000:
                alertas_nuevas.append({
                    **alerta_base,
                    'tipo':    'deuda_antigua',
                    'detalle': f"Facturas sin cobrar hace más de 90 días · ${_monto_90d:,.0f}",
                })

            resp = _score_response(score_nuevo, None)
            scores_nuevos[cuit] = resp

        # Persistir scores actualizados
        if scores_nuevos:
            with _score_cache_lock:
                sc = _score_cache_read()
                sc.update(scores_nuevos)
                _score_cache_write(sc)

        # Persistir alertas nuevas
        _agregar_alertas_auto(alertas_nuevas)

        print(
            f"[recalculo] OK — {procesados} clientes · {len(alertas_nuevas)} alertas nuevas",
            flush=True
        )

    except Exception:
        import traceback
        print(f"[recalculo] Error inesperado:\n{traceback.format_exc()}", flush=True)


def _check_cheques_cartera_bg():
    """
    Background: verifica cheques rechazados para toda la cartera contra la tabla
    cheques_bcra (bulk local). Compara contra cheques_estado.json (estado anterior).
    Genera alertas para CUITs con cheques nuevos detectados.
    """
    import time as _t
    _t.sleep(3)
    try:
        cartera = list(_cartera_comercial)
        if not cartera:
            return

        _nc = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()
        cuits_all = [_nc(c.get('cuit', '')) for c in cartera if c.get('cuit')]
        if not cuits_all:
            return

        # Leer cheques actuales del bulk local (una sola query batch)
        cheques_actuales = _cheques_local_batch(cuits_all)

        # Estado anterior
        estado_path = os.path.join(DATA_DIR, 'cheques_estado.json')
        estado_anterior: dict = {}
        try:
            if os.path.exists(estado_path):
                with open(estado_path, 'r', encoding='utf-8') as f:
                    estado_anterior = json.load(f)
        except Exception:
            pass

        cartera_idx = {_nc(c.get('cuit', '')): c for c in cartera if c.get('cuit')}
        alertas_nuevas = []
        fecha_ahora = datetime.now().strftime('%Y-%m-%d %H:%M')

        for cuit, cheq in cheques_actuales.items():
            resultados = (cheq.get('results') or {})
            # Contar cheques rechazados actuales
            n_rechazados = (
                len(resultados.get('chequesRechazadosCamaraCompensadora', [])) +
                len(resultados.get('chequesRechazadosDenunciados', []))
            )
            n_anterior = estado_anterior.get(cuit, {}).get('n_rechazados', 0)

            if n_rechazados > n_anterior:
                cliente = cartera_idx.get(cuit, {})
                nombre  = str(cliente.get('nombre') or cliente.get('cliente') or cuit)
                diff    = n_rechazados - n_anterior
                alertas_nuevas.append({
                    'cuit':         cuit,
                    'nombre':       nombre,
                    'fecha':        fecha_ahora,
                    # 'cheque' (no 'cheque_nuevo'): el contador y las tarjetas del
                    # dashboard ("CHEQUES RECH.") filtran por tipo === 'cheque' y
                    # leen nroCheques/totalCheques — con otro tipo quedan invisibles.
                    'tipo':         'cheque',
                    'nroCheques':   n_rechazados,
                    'totalCheques': n_rechazados,
                    'detalle':      f"{diff} cheque(s) rechazado(s) nuevo(s) detectado(s) · total {n_rechazados}",
                    'rango':        '',
                    'score_nuevo':  None,
                    'leida':        False,
                })

            # Actualizar estado
            estado_anterior[cuit] = {
                'n_rechazados': n_rechazados,
                'ultima_check': fecha_ahora,
            }

        # Persistir estado actualizado
        try:
            with open(estado_path, 'w', encoding='utf-8') as f:
                json.dump(estado_anterior, f, ensure_ascii=False)
        except Exception as e:
            print(f"[cheques_check] Error guardando estado: {e}", flush=True)

        _agregar_alertas_auto(alertas_nuevas)
        print(
            f"[cheques_check] OK — {len(cuits_all)} CUITs · {len(alertas_nuevas)} alertas nuevas",
            flush=True
        )

    except Exception:
        import traceback
        print(f"[cheques_check] Error:\n{traceback.format_exc()}", flush=True)


# ── Endpoints de alertas automáticas ─────────────────────────────────────────

@app.route("/api/alertas", methods=["GET"])
def get_alertas_auto():
    """Lista de alertas automáticas. Query params: ?leidas=true para incluir leídas."""
    solo_no_leidas = request.args.get('leidas', 'false').lower() != 'true'
    try:
        alertas = _alertas_auto_read()
        if solo_no_leidas:
            alertas = [a for a in alertas if not a.get('leida')]
        return jsonify({'alertas': alertas, 'total': len(alertas)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route("/api/alertas/marcar-leida", methods=["POST"])
def marcar_alerta_leida():
    """Marca como leída una o todas las alertas. Body: {cuit, tipo, fecha} o {todas: true}."""
    try:
        body = request.get_json(force=True) or {}
        with _alertas_auto_lock:
            alertas = _alertas_auto_read()
            if body.get('todas'):
                for a in alertas:
                    a['leida'] = True
            else:
                for a in alertas:
                    if (a.get('cuit') == body.get('cuit') and
                            a.get('tipo') == body.get('tipo') and
                            a.get('fecha', '')[:10] == (body.get('fecha') or '')[:10]):
                        a['leida'] = True
            _alertas_auto_write(alertas)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route("/api/alertas/verificar-cheques", methods=["POST"])
def verificar_cheques_cartera():
    """Dispara verificación de cheques rechazados para toda la cartera en background."""
    t = threading.Thread(target=_check_cheques_cartera_bg, daemon=True)
    t.start()
    return jsonify({'ok': True, 'mensaje': 'Verificación de cheques iniciada en background'})


@app.route("/save-score-cache", methods=["POST"])
def save_score_cache():
    """Persiste score(s) en score_cache.json. Payload: {cuit: score_data, ...}"""
    try:
        data = request.get_json(force=True) or {}
        if not data:
            return jsonify({"ok": False, "error": "Payload vacío"}), 400
        nc = lambda x: str(x).replace('-', '').replace(' ', '').strip()
        with _score_cache_lock:
            cache = _score_cache_read()
            for cuit_k, score_v in data.items():
                cache[nc(cuit_k)] = score_v
            _score_cache_write(cache)
        print(f"[score_cache] guardados {len(data)} CUITs", flush=True)
        return jsonify({"ok": True, "guardados": len(data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/score-cache-all", methods=["GET"])
def score_cache_all():
    """Devuelve todo el contenido de score_cache.json."""
    with _score_cache_lock:
        return jsonify(_score_cache_read())


def _calcular_score_handler(cuit: str):
    """Lógica compartida por /calcular-score/ y /fetch-score/."""
    from urllib.parse import unquote
    cuit_limpio = str(unquote(cuit)).replace('-', '').replace(' ', '').strip()
    if len(cuit_limpio) < 10:
        return jsonify({"ok": False, "error": "CUIT inválido"}), 400

    # Siempre eliminar de session cache en queries on-demand — evita data leaking entre usuarios
    _score_session_cache.pop(cuit_limpio, None)
    if request.args.get('fresh') == '1':
        # Borrar caché de solvency Y de BCRA para forzar re-consulta completa
        _fp = os.path.join(DATA_DIR, f'solvency_{cuit_limpio}.json')
        if os.path.exists(_fp):
            os.remove(_fp)
        try:
            _bc_path = os.path.join(DATA_DIR, 'bcra_cache.json')
            if os.path.exists(_bc_path):
                with open(_bc_path, 'r', encoding='utf-8') as _f:
                    _bc = json.load(_f)
                if cuit_limpio in _bc:
                    _entry_ts = _bc[cuit_limpio].get('ts', 0) if isinstance(_bc[cuit_limpio], dict) else 0
                    if time.time() - _entry_ts > 300:  # guard: no invalidar si fue cargado hace <5 min
                        del _bc[cuit_limpio]
                        with open(_bc_path, 'w', encoding='utf-8') as _f:
                            json.dump(_bc, _f)
                        print(f"[fetch-score] {cuit_limpio} bcra_cache invalidado", flush=True)
                    else:
                        print(f"[fetch-score] {cuit_limpio} bcra_cache reciente — reutilizando (double-call guard)", flush=True)
        except Exception:
            pass
    else:
        # ── Cache persistente en disco (sobrevive reinicios) ──────────────
        with _score_cache_lock:
            _cached = _score_cache_read().get(cuit_limpio)
        if _cached and _cached.get('score'):
            print(f"[fetch-score] {cuit_limpio} → score_cache.json hit ({_cached['score']})", flush=True)
            return jsonify(_cached)
    try:
        bcra_data, _ = consultar_bcra_cached(cuit_limpio)
        _bcra_denom  = (bcra_data.get('results') or {}).get('denominacion', '').strip()
        _bcra_error  = bool(bcra_data.get('error_bcra')) or bcra_data.get('bcra_disponible') is False

        # ── Validación de CUIT: solo bloquear cuando BCRA devolvió un ERROR real ──
        # Si BCRA respondió con sin_deudas=True, significa que el CUIT ES válido
        # (la persona existe pero no tiene historial crediticio). NO es CUIT inexistente.
        # Solo verificar via ARCA cuando la consulta BCRA falló completamente.
        if _bcra_error and not _bcra_denom:
            _solv_chk  = get_solvency_data(cuit_limpio)
            if not isinstance(_solv_chk, dict): _solv_chk = {}
            _razon_chk = (_solv_chk.get('razon_social') or _solv_chk.get('nombre') or '').strip()
            if not _razon_chk:
                # Re-leer cartera del disco (captura clientes agregados después del último deploy)
                _nc = lambda x: str(x or '').replace('-', '').replace(' ', '').strip()
                _cc_live = _cartera_comercial
                try:
                    _cc_path_live = _CC_FILE if os.path.exists(_CC_FILE) \
                                    else os.path.join(os.getcwd(), 'cartera_comercial.json')
                    with open(_cc_path_live, encoding='utf-8') as _ccf:
                        _cc_live = json.load(_ccf)
                except Exception:
                    pass
                _en_interna = (
                    any(_nc(c.get('cuit', '')) == cuit_limpio for c in _cc_live)
                    or any(_nc(f.get('cuit', '')) == cuit_limpio
                           for f in (_saldos_gestion if _saldos_gestion else _saldos_facturas))
                )
                if not _en_interna:
                    print(f"[fetch-score] {cuit_limpio} CUIT INEXISTENTE — error BCRA + sin AFIP/solvencia", flush=True)
                    return jsonify({"error": "cuit_inexistente", "cuit": cuit_limpio, "score": None}), 200
                print(f"[fetch-score] {cuit_limpio} sin BCRA/AFIP pero en cartera interna — calculando score", flush=True)

        # sin_deudas=True sin error BCRA = cliente válido sin historial crediticio.
        # El motor usa el piso 650 para este caso (_cliente_sin_deuda).
        if bcra_data.get('sin_deudas') and not _bcra_denom and not _bcra_error:
            print(f"[fetch-score] {cuit_limpio} cliente nuevo sin historial BCRA — score base 650", flush=True)

        score_data   = calcular_score_servidor(cuit_limpio, bcra_data or {})
        solvency     = get_solvency_data(cuit_limpio)
        if not isinstance(solvency, dict): solvency = {}
        _actualizar_score_en_cartera(cuit_limpio, score_data, solvency)
        cheq_cached  = _cheques_cache_get(cuit_limpio)
        _resp_indiv  = _score_response(score_data, solvency, cheq_cached)
        # Persistir en score_cache.json con _ts → proceso integral reutiliza este score
        # en lugar de recalcular con datos bulk (evita divergencia individual vs. cartera).
        _resp_indiv['_ts'] = time.time()
        with _score_cache_lock:
            _sc_i = _score_cache_read()
            _sc_i[cuit_limpio] = _resp_indiv
            _score_cache_write(_sc_i)
        return jsonify(_resp_indiv)
    except Exception as e:
        import traceback
        print(f"[score] ERROR {cuit_limpio}: {e}\n{traceback.format_exc()}", flush=True)
        # Fallback: intentar score solo con datos BCRA (sin solvencia/AFIP) para no
        # devolver null al frontend — un score parcial es mejor que un error en blanco.
        try:
            bcra_fb, _ = consultar_bcra_cached(cuit_limpio)
            sd_fb = calcular_rating_predictivo(
                cuit=cuit_limpio, bcra_data=bcra_fb or {},
                solvency_data={},  # evita consultas fiscales en el camino de fallback
            )
            resp_fb = _score_response(sd_fb, {})
            resp_fb['_fallback'] = True
            resp_fb['_error']    = str(e)
            return jsonify(resp_fb)
        except Exception as e2:
            print(f"[score] FALLBACK ERROR {cuit_limpio}: {e2}", flush=True)
        # Último recurso: devolver 200 con score=null para que el frontend
        # pueda usar el score en caché de CARTERA_LOCAL en lugar de null.
        return jsonify({
            "ok": False, "error": str(e),
            "score": None, "rango": "Error", "color": "#6b7280", "emoji": "⚠️",
            "razonamiento_score": None, "mora_administrativa": False,
            "override_admin": False, "bloquear_oportunidad": False,
            "mora_tecnica": False, "nota_mora_tecnica": None,
        })


@app.route("/calcular-score/<cuit>")
def calcular_score_individual(cuit):
    return _calcular_score_handler(cuit)


@app.route("/fetch-score/<cuit>")
def fetch_score_individual(cuit):
    """Alias liviano de /calcular-score/ — misma lógica, ruta limpia."""
    return _calcular_score_handler(cuit)


@app.route("/recalcular-scores", methods=["POST"])
def recalcular_scores():
    """
    Re-calcula scores para toda la cartera usando datos en caché (sin re-scrapear BCRA).
    Botón '↻ Actualizar scores': aplica lógica v9.0 sobre datos ya guardados.
    """
    try:
        with open(ALERTAS_FILE, 'r', encoding='utf-8') as _f:
            existente = json.load(_f)
    except Exception as e:
        return jsonify({"ok": False, "error": f"alertas_cartera.json no disponible: {e}"}), 404

    cartera    = existente.get('cartera', [])
    recalc     = 0
    sin_cache  = 0

    def _cache_load_fresh(fname):
        p = os.path.join(DATA_DIR, fname)
        try:
            with open(p, 'r') as _f:
                d = json.load(_f)
            if time.time() - d.get('ts', 0) < 86400:
                return d.get('payload')
        except: pass
        return None

    for i, c in enumerate(cartera):
        cuit = str(c.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
        if not cuit:
            continue
        bcra_cached, _ = cache_get(cuit)
        if not bcra_cached:
            sin_cache += 1
            continue
        hist_data = _cache_load_fresh(f'historial_{cuit}.json')
        cheq_data = _cache_load_fresh(f'cheques_{cuit}.json')
        solvency  = get_solvency_data(cuit)
        try:
            sd = calcular_rating_predictivo(
                cuit=cuit, bcra_data=bcra_cached,
                hist_data=hist_data, cheq_data=cheq_data,
                en_mora=None, solvency_data=solvency,
                ciudad=str(c.get('ciudad', '') or ''),
            )
            cartera[i].update({
                'scoreCompleto':        sd['score'],
                'scoreRango':           sd['rango'],
                'scoreColor':           sd['color'],
                'scoreEmoji':           sd['emoji'],
                'alerta_temprana':      sd.get('alerta_temprana', False),
                'bloquear_oportunidad': sd.get('bloquear_oportunidad', False),
                'alerta_logistica':     sd.get('alerta_logistica', ''),
                'inferencia_ingresos':  (solvency or {}).get('ingresos_anuales'),
                'fuente_ingresos':      (solvency or {}).get('fuente_ingresos'),
                'actividad_principal':  (solvency or {}).get('actividad_principal'),
                'score_ts':             time.time(),
                'pendiente':            False,
            })
            recalc += 1
        except Exception as e_r:
            print(f"[recalcular] Error {cuit}: {e_r}", flush=True)

    existente['cartera']      = cartera
    existente['ultima_verif'] = time.strftime('%d/%m/%Y %H:%M') + ' (recalc v9.0)'
    try:
        with open(ALERTAS_FILE, 'w', encoding='utf-8') as _f:
            json.dump(existente, _f, ensure_ascii=False)
        print(f"[recalcular] {recalc} scores actualizados, {sin_cache} sin caché BCRA", flush=True)
        return jsonify({"ok": True, "recalculados": recalc, "sin_cache": sin_cache})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/analizar-bodegas", methods=["POST"])
def analizar_bodegas():
    if not GEMINI_KEY:
        return jsonify({"es_negativo": False, "motivo": ""})
    try:
        body = request.get_json(force=True)
        cuit = body.get('cuit', '')
        nombre = body.get('nombre', '')
        mensajes = body.get('mensajes', [])
        es_neg, motivo = analizar_bodegas_server(cuit, nombre, mensajes)
        return jsonify({"es_negativo": es_neg, "motivo": motivo})
    except Exception as e:
        return jsonify({"es_negativo": False, "motivo": str(e)})

# ── PADRÓN OFICIAL BCRA LOCAL (bcra_nomdeu.db) ───────────────────────────────
# Descargado al arrancar: R2 autenticado (prioridad, bucket privado) si están
# configuradas R2_ACCESS_KEY_ID/SECRET/ENDPOINT_URL/BUCKET_NAME, sino BCRA_NOMDEU_URL
# (Google Drive o URL HTTP directa). Soporta archivos >100 MB en Drive via gdown.
# Tablas: denominaciones, entidades, deudas_resumen (legacy), historial_detalle
# (situación y monto reales por CUIT+entidad de los últimos _HIST_DETALLE_MESES meses).

_nomdeu_conn: sqlite3.Connection = None  # type: ignore[assignment]
_mipyme_conn: sqlite3.Connection = None  # type: ignore[assignment]

_GDRIVE_PREFIXES = ('https://drive.google.com', 'https://docs.google.com')


def _extraer_gdrive_id(url: str) -> str | None:
    """Extrae el file ID de cualquier formato de URL de Google Drive."""
    import re
    m = re.search(r'/(?:file/d|d)/([a-zA-Z0-9_-]{25,})', url)
    return m.group(1) if m else None


def _descargar_nomdeu(url: str, dest: str) -> bool:
    """
    Descarga bcra_nomdeu.db. Prioridad:
      1. R2 autenticado (boto3) — si R2_ACCESS_KEY_ID/SECRET/ENDPOINT_URL/BUCKET_NAME
         están configurados, bucket privado, sin exponer el archivo públicamente.
      2. Google Drive: usa gdown que maneja el token de confirmación para archivos grandes.
      3. Otros: requests con streaming directo a BCRA_NOMDEU_URL.
    Retorna True si la descarga fue exitosa.
    """
    if _R2_CONFIGURADO:
        try:
            import boto3
            from botocore.config import Config
            print(f"[nomdeu] Descargando desde R2 (autenticado, bucket={R2_BUCKET_NAME})...", flush=True)
            s3 = boto3.client(
                service_name='s3',
                endpoint_url=R2_ENDPOINT_URL,
                aws_access_key_id=R2_ACCESS_KEY_ID,
                aws_secret_access_key=R2_SECRET_ACCESS_KEY,
                config=Config(signature_version='s3v4'),
            )
            s3.download_file(R2_BUCKET_NAME, 'bcra_nomdeu.db', dest)
            if not os.path.exists(dest) or os.path.getsize(dest) < 1_000_000:
                print("[nomdeu] Descarga R2 falló o archivo demasiado pequeño", flush=True)
                return False
            size_mb = os.path.getsize(dest) / 1_048_576
            print(f"[nomdeu] Descarga R2 completa: {size_mb:.0f} MB", flush=True)
            return True
        except Exception as e:
            print(f"[nomdeu] Error descarga R2: {e}", flush=True)
            return False

    if any(url.startswith(p) for p in _GDRIVE_PREFIXES):
        file_id = _extraer_gdrive_id(url)
        if not file_id:
            print(f"[nomdeu] No se pudo extraer file ID de la URL de Drive: {url}", flush=True)
            return False
        try:
            import gdown
            gdrive_url = f"https://drive.google.com/uc?id={file_id}"
            print(f"[nomdeu] Descargando desde Google Drive (file_id={file_id})...", flush=True)
            # Redirigir TMPDIR al disco persistente — gdown escribe un .part temporal
            # que en Render llena /tmp (1 GB) antes de mover al destino final.
            dest_dir = os.path.dirname(os.path.abspath(dest))
            os.environ['TMPDIR'] = dest_dir
            import tempfile
            tempfile.tempdir = dest_dir
            # quiet=False para ver progreso en logs de Render
            gdown.download(gdrive_url, dest, quiet=False, fuzzy=False)
            if not os.path.exists(dest) or os.path.getsize(dest) < 1_000_000:
                print("[nomdeu] gdown falló o archivo demasiado pequeño", flush=True)
                return False
            size_mb = os.path.getsize(dest) / 1_048_576
            print(f"[nomdeu] Descarga Drive completa: {size_mb:.0f} MB", flush=True)
            return True
        except Exception as e:
            print(f"[nomdeu] Error gdown: {e}", flush=True)
            return False
    else:
        # R2 u otro endpoint HTTP directo
        print(f"[nomdeu] Descargando desde URL directa...", flush=True)
        try:
            r = requests.get(url, stream=True, timeout=600, verify=False)
            r.raise_for_status()
            descargado = 0
            with open(dest, 'wb') as fh:
                for chunk in r.iter_content(chunk_size=8 * 1024 * 1024):
                    if chunk:
                        fh.write(chunk)
                        descargado += len(chunk)
                        if descargado % (100 * 1024 * 1024) < 8 * 1024 * 1024:
                            print(f"[nomdeu]   {descargado / 1_048_576:.0f} MB descargados...", flush=True)
            size_mb = os.path.getsize(dest) / 1_048_576
            print(f"[nomdeu] Descarga directa completa: {size_mb:.0f} MB", flush=True)
            return True
        except Exception as e:
            print(f"[nomdeu] Error descarga directa: {e}", flush=True)
            return False


def _nomdeu_db_valida() -> bool:
    """Verificación rápida: la DB existe y historial_detalle tiene al menos una fila."""
    if not os.path.exists(NOMDEU_DB_PATH):
        return False
    try:
        c = sqlite3.connect(NOMDEU_DB_PATH)
        ok = c.execute("SELECT 1 FROM historial_detalle LIMIT 1").fetchone() is not None
        c.close()
        return ok
    except Exception:
        return False


def _abrir_nomdeu_conn() -> None:
    """Abre la conexión SQLite a nomdeu y expone _nomdeu_conn.
    No hace COUNT(*) — usa SELECT 1 LIMIT 1 para verificar en <1ms."""
    global _nomdeu_conn
    try:
        conn = sqlite3.connect(NOMDEU_DB_PATH, check_same_thread=False)
        conn.execute("PRAGMA query_only = 1")
        conn.execute("PRAGMA cache_size  = -32768")   # 32 MB caché lectura
        if not conn.execute("SELECT 1 FROM historial_detalle LIMIT 1").fetchone():
            conn.close()
            print("[nomdeu] historial_detalle vacío — padrón offline no disponible", flush=True)
            return
        _nomdeu_conn = conn
        print(f"[nomdeu] Listo — historial_detalle disponible ({_HIST_DETALLE_MESES}m reales)", flush=True)
    except Exception as e:
        print(f"[nomdeu] Error abriendo SQLite: {e}", flush=True)


def _descargar_y_abrir_nomdeu(url: str) -> None:
    """Descarga bcra_nomdeu.db y abre la conexión. Corre en hilo daemon."""
    ok = _descargar_nomdeu(url, NOMDEU_DB_PATH)
    if not ok:
        print("[nomdeu] Descarga fallida — padrón offline no disponible", flush=True)
        return
    _abrir_nomdeu_conn()


def _init_nomdeu_db() -> None:
    """Inicializa el padrón offline BCRA (bcra_nomdeu.db).

    Estrategia de dos velocidades:
    - Archivo ya existe en disco (caso habitual: Render persiste /data entre deploys):
      abre la conexión AHORA en el hilo actual (<1s) → nomdeu disponible desde
      la primera request, sin esperar ningún background thread.
    - Archivo no existe o está vencido/inválido: descarga en hilo daemon y abre
      cuando termina (~9 min primer deploy). Gunicorn arranca igual de inmediato
      pero las consultas caen a BCRA live hasta que termine la descarga.
    """
    url = os.environ.get('BCRA_NOMDEU_URL', '').strip()
    if not _R2_CONFIGURADO and not url:
        print("[nomdeu] Ni R2 ni BCRA_NOMDEU_URL configurados — padrón offline desactivado", flush=True)
        return

    periodo_esperado = os.environ.get('BCRA_NOMDEU_PERIODO', '').strip()

    necesita_descarga = True
    if os.path.exists(NOMDEU_DB_PATH):
        edad_dias = (time.time() - os.path.getmtime(NOMDEU_DB_PATH)) / 86400
        if edad_dias < 32 and _nomdeu_db_valida():
            if periodo_esperado:
                try:
                    c = sqlite3.connect(NOMDEU_DB_PATH)
                    _p = c.execute(
                        "SELECT periodo FROM deudas_resumen ORDER BY periodo DESC LIMIT 1"
                    ).fetchone()
                    c.close()
                    periodo_local = _p[0] if _p else ''
                except Exception:
                    periodo_local = ''
                if str(periodo_local) != periodo_esperado:
                    print(
                        f"[nomdeu] Período local ({periodo_local}) ≠ BCRA_NOMDEU_PERIODO ({periodo_esperado})"
                        f" — re-descargando en background", flush=True
                    )
                    os.remove(NOMDEU_DB_PATH)
                else:
                    necesita_descarga = False
                    print(f"[nomdeu] DB existente ({edad_dias:.0f}d) período={periodo_local} — abriendo", flush=True)
            else:
                necesita_descarga = False
                print(f"[nomdeu] DB existente ({edad_dias:.0f}d) — abriendo", flush=True)
        else:
            print(f"[nomdeu] DB stale o sin historial_detalle — re-descargando en background", flush=True)
            try:
                os.remove(NOMDEU_DB_PATH)
            except Exception:
                pass

    if necesita_descarga:
        # Descarga en background — gunicorn arranca igual, nomdeu disponible cuando termine
        threading.Thread(
            target=_descargar_y_abrir_nomdeu, args=(url,),
            daemon=True, name='nomdeu-init'
        ).start()
        return

    # Archivo válido en disco: abrir AHORA (rápido, sin descarga)
    _abrir_nomdeu_conn()


# ── MiPyME: estado de importación ─────────────────────────────────────────────
_mipyme_import_estado: dict = {
    'corriendo':    False,
    'progreso':     0,
    'total':        0,
    'ultimo_paso':  'idle',
    'error':        None,
    'registros_ok': 0,
    'ultima_actualizacion': None,
}


def _import_mipyme_csv(url: str = None) -> bool:
    """
    Descarga el padrón MiPyME oficial (CSV, Min. Producción) e importa a SQLite.

    Estrategia:
      1. Descarga streaming a archivo temporal para no saturar RAM.
      2. Parsing robusto: detecta columnas por nombre (case-insensitive), maneja
         UTF-8 y latin-1, y normaliza CUIT a 11 dígitos sin guiones.
      3. Import por batches de 5000 con INSERT OR REPLACE (idempotente).
      4. Crea índice único por CUIT (PRIMARY KEY) para queries O(log n).

    Mapping de categorías del CSV → clave interna:
      'Micro empresa'            → 'Micro'
      'Pequeña empresa'          → 'Pequeña'
      'Mediana empresa -Tramo 1' → 'Mediana_T1'
      'Mediana empresa -Tramo 2' → 'Mediana_T2'
    """
    global _mipyme_conn, _mipyme_import_estado

    _mipyme_import_estado.update({
        'corriendo': True, 'progreso': 0, 'total': 0,
        'ultimo_paso': 'iniciando', 'error': None, 'registros_ok': 0,
    })

    _url = (url or MIPYME_CSV_URL).strip()
    tmp_path = MIPYME_DB_PATH + '.csv.tmp'

    # ── Paso 1: Descarga streaming ─────────────────────────────────────────────
    try:
        _mipyme_import_estado['ultimo_paso'] = 'descargando_csv'
        print(f"[mipyme] Descargando padrón desde: {_url}", flush=True)
        import requests as _req_mp
        resp = _req_mp.get(_url, stream=True, timeout=120,
                           headers={'User-Agent': 'VendeSeguro/1.0'})
        resp.raise_for_status()
        descargado = 0
        with open(tmp_path, 'wb') as fh:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    fh.write(chunk)
                    descargado += len(chunk)
        size_kb = descargado / 1024
        print(f"[mipyme] CSV descargado: {size_kb:.0f} KB", flush=True)
    except Exception as e:
        _mipyme_import_estado.update({'corriendo': False, 'error': str(e), 'ultimo_paso': 'error_descarga'})
        print(f"[mipyme] Error en descarga: {e}", flush=True)
        return False

    # ── Paso 2: Crear/limpiar DB destino ──────────────────────────────────────
    try:
        _mipyme_import_estado['ultimo_paso'] = 'preparando_db'
        conn_w = sqlite3.connect(MIPYME_DB_PATH, check_same_thread=False)
        conn_w.execute("PRAGMA journal_mode = WAL")
        conn_w.execute("PRAGMA synchronous  = NORMAL")
        conn_w.execute("""
            CREATE TABLE IF NOT EXISTS mipyme_padron (
                cuit        TEXT PRIMARY KEY,
                razon_social TEXT,
                categoria   TEXT,
                sector      TEXT,
                provincia   TEXT,
                fecha_alta  TEXT
            )
        """)
        conn_w.execute("DELETE FROM mipyme_padron")
        conn_w.execute("""
            CREATE TABLE IF NOT EXISTS _mipyme_meta (
                key   TEXT PRIMARY KEY,
                valor TEXT
            )
        """)
        conn_w.commit()
    except Exception as e:
        _mipyme_import_estado.update({'corriendo': False, 'error': str(e), 'ultimo_paso': 'error_db'})
        print(f"[mipyme] Error preparando DB: {e}", flush=True)
        return False

    # Mapping normalizado de categorías CSV → clave interna
    _CAT_MAP = {
        'micro empresa':             'Micro',
        'micro':                     'Micro',
        'pequeña empresa':           'Pequeña',
        'pequeña':                   'Pequeña',
        'pequeña empresa':           'Pequeña',
        'pequena empresa':           'Pequeña',
        'mediana empresa -tramo 1':  'Mediana_T1',
        'mediana empresa - tramo 1': 'Mediana_T1',
        'mediana tramo 1':           'Mediana_T1',
        'mediana_t1':                'Mediana_T1',
        'mediana empresa -tramo 2':  'Mediana_T2',
        'mediana empresa - tramo 2': 'Mediana_T2',
        'mediana tramo 2':           'Mediana_T2',
        'mediana_t2':                'Mediana_T2',
    }

    # ── Paso 3: Parse CSV e INSERT por batches ─────────────────────────────────
    try:
        _mipyme_import_estado['ultimo_paso'] = 'parseando_csv'
        import csv as _csv_mod

        # Intenta UTF-8 primero, fallback a latin-1 (común en datos.gob.ar)
        for _enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                with open(tmp_path, 'r', encoding=_enc, errors='replace') as fh:
                    reader = _csv_mod.DictReader(fh)
                    # Normalizar nombres de columna a minúsculas sin espacios extra
                    _raw_fields = reader.fieldnames or []
                    _fields_lower = [f.lower().strip() for f in _raw_fields]

                    # Detectar columnas por variantes posibles
                    def _col(*candidates):
                        for cand in candidates:
                            for i, fn in enumerate(_fields_lower):
                                if fn == cand or fn.startswith(cand):
                                    return _raw_fields[i]
                        return None

                    col_cuit   = _col('cuit')
                    col_rs     = _col('razon_social', 'denominacion', 'nombre', 'empresa')
                    col_cat    = _col('categoria', 'tramo', 'categoría')
                    col_sec    = _col('sector', 'actividad_principal', 'rubro')
                    col_prov   = _col('provincia')
                    col_alta   = _col('fecha_alta', 'fecha_inscripcion', 'fecha')

                    if not col_cuit or not col_cat:
                        print(f"[mipyme] Encoding {_enc}: columnas no encontradas ({_fields_lower[:8]})", flush=True)
                        continue

                    batch, total_ok = [], 0
                    for row in reader:
                        cuit_raw = str(row.get(col_cuit) or '').replace('-', '').replace(' ', '').strip()
                        if len(cuit_raw) != 11 or not cuit_raw.isdigit():
                            continue
                        cat_raw = (row.get(col_cat) or '').strip().lower()
                        cat_norm = _CAT_MAP.get(cat_raw, '')
                        if not cat_norm:
                            continue  # fila sin categoría reconocible
                        rs    = str(row.get(col_rs) or '').strip()[:200] if col_rs else ''
                        sec   = str(row.get(col_sec) or '').strip()[:80] if col_sec else ''
                        prov  = str(row.get(col_prov) or '').strip()[:60] if col_prov else ''
                        alta  = str(row.get(col_alta) or '').strip()[:20] if col_alta else ''
                        batch.append((cuit_raw, rs, cat_norm, sec, prov, alta))
                        if len(batch) >= 5000:
                            conn_w.executemany(
                                "INSERT OR REPLACE INTO mipyme_padron "
                                "(cuit, razon_social, categoria, sector, provincia, fecha_alta) "
                                "VALUES (?,?,?,?,?,?)", batch
                            )
                            conn_w.commit()
                            total_ok += len(batch)
                            _mipyme_import_estado['registros_ok'] = total_ok
                            batch = []

                    # Flush último batch
                    if batch:
                        conn_w.executemany(
                            "INSERT OR REPLACE INTO mipyme_padron "
                            "(cuit, razon_social, categoria, sector, provincia, fecha_alta) "
                            "VALUES (?,?,?,?,?,?)", batch
                        )
                        conn_w.commit()
                        total_ok += len(batch)

                    _mipyme_import_estado['registros_ok'] = total_ok
                    print(f"[mipyme] Import OK: {total_ok:,} empresas (encoding={_enc})", flush=True)
                    break  # salir del loop de encodings

            except UnicodeDecodeError:
                print(f"[mipyme] Encoding {_enc} falló, probando siguiente...", flush=True)
                continue

        else:
            raise ValueError("No se pudo parsear el CSV con ningún encoding conocido")

    except Exception as e:
        _mipyme_import_estado.update({'corriendo': False, 'error': str(e), 'ultimo_paso': 'error_parse'})
        print(f"[mipyme] Error parseando CSV: {e}", flush=True)
        conn_w.close()
        return False

    # ── Paso 4: Metadata + conectar globalmente ────────────────────────────────
    try:
        conn_w.execute(
            "INSERT OR REPLACE INTO _mipyme_meta (key, valor) VALUES (?,?)",
            ('ultima_actualizacion', time.strftime('%Y-%m-%d %H:%M:%S'))
        )
        conn_w.execute(
            "INSERT OR REPLACE INTO _mipyme_meta (key, valor) VALUES (?,?)",
            ('total_registros', str(_mipyme_import_estado['registros_ok']))
        )
        conn_w.commit()
        conn_w.execute("PRAGMA query_only = 1")
        conn_w.execute("PRAGMA cache_size = -32768")   # 32 MB caché lectura
        _mipyme_conn = conn_w
        _mipyme_import_estado.update({
            'corriendo': False,
            'progreso':  100,
            'ultimo_paso': 'completado',
            'ultima_actualizacion': time.strftime('%Y-%m-%d %H:%M:%S'),
        })
        print(f"[mipyme] DB lista — {_mipyme_import_estado['registros_ok']:,} PyMEs indexadas", flush=True)
    except Exception as e:
        _mipyme_import_estado.update({'corriendo': False, 'error': str(e), 'ultimo_paso': 'error_final'})
        print(f"[mipyme] Error cerrando import: {e}", flush=True)
        return False
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

    return True


def _init_mipyme_db() -> None:
    """Abre conexión de solo lectura a mipyme_padron.db si el archivo existe en disco."""
    global _mipyme_conn
    if not os.path.exists(MIPYME_DB_PATH):
        print("[mipyme] mipyme_padron.db no encontrada — ejecutar POST /update-mipyme-db para importar", flush=True)
        return
    try:
        conn = sqlite3.connect(MIPYME_DB_PATH, check_same_thread=False)
        conn.execute("PRAGMA query_only = 1")
        conn.execute("PRAGMA cache_size = -32768")
        n = conn.execute("SELECT COUNT(*) FROM mipyme_padron").fetchone()[0]
        _mipyme_conn = conn
        print(f"[mipyme] Listo — {n:,} PyMEs en padrón local", flush=True)
    except Exception as e:
        print(f"[mipyme] Error abriendo SQLite: {e}", flush=True)


def _nomdeu_get_nombre(cuit: str):
    """Nombre oficial del BCRA (Nomdeu.txt). Retorna str o None."""
    if _nomdeu_conn is None:
        return None
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    try:
        row = _nomdeu_conn.execute(
            "SELECT nombre FROM denominaciones WHERE cuit = ?", (cuit_limpio,)
        ).fetchone()
        return row[0] if row else None
    except Exception:
        return None


def _nomdeu_get_deuda(cuit: str):
    """
    Resumen de deuda del padrón offline, derivado 100% de historial_detalle
    (situación y monto reales de los últimos _HIST_DETALLE_MESES meses, por
    entidad — ver _nomdeu_agregar_filas). El mes_01 (más reciente) funciona
    como snapshot actual, equivalente a lo que antes daba deudas_resumen.

    Retorna None si el CUIT no tiene ninguna fila en historial_detalle.
    """
    filas = _historial_detalle_rows(cuit)
    if not filas:
        return None
    return _nomdeu_agregar_filas(filas)


@app.route("/admin/arca-diagnostico")
@require_login
def admin_arca_diagnostico():
    """Por qué ARCA está o no operativo, con la causa exacta del fallo.

    No expone material criptográfico: de la clave privada solo informa
    presencia, longitud y si tiene cabecera PEM reconocible.
    Si el módulo no está configurado, reintenta la inicialización para
    capturar y reportar el error concreto del lector de PEM.
    """
    if not _ARCA_MODULO_OK:
        return jsonify({
            "ok": False,
            "arca_modulo": False,
            "error": "arca_ws no pudo importarse (revisar cryptography en el build)",
        }), 503
    try:
        info = arca_ws.diagnostico(DATA_DIR)
        # Si el reintento logró configurarlo, habilitar el canal en caliente
        global ARCA_DISPONIBLE
        if info.get('configurado') and not ARCA_DISPONIBLE:
            ARCA_DISPONIBLE = True
            print("[arca] canal habilitado tras diagnóstico exitoso", flush=True)
        return jsonify({"ok": True, "arca_activo": ARCA_DISPONIBLE, **info})
    except Exception as e:
        return jsonify({"ok": False, "error": f"{type(e).__name__}: {e}"}), 500


@app.route("/admin/nomdeu-lookup/<cuit>")
def admin_nomdeu_lookup(cuit):
    """Diagnóstico: muestra el resumen agregado (_nomdeu_get_deuda) y las
    filas crudas de historial_detalle para un CUIT, más la cobertura total
    de la tabla. Sirve para confirmar si una ausencia es real (CUIT fuera
    del filtro/cobertura del import) o un bug, sin acceso directo al disco."""
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    if _nomdeu_conn is None:
        return jsonify({"ok": False, "error": "nomdeu_conn no inicializada"}), 503
    out = {"ok": True, "cuit": cuit_limpio}
    try:
        out["resumen"] = _nomdeu_get_deuda(cuit_limpio)
    except Exception as e:
        out["resumen"] = f"error: {e}"
    try:
        out["filas_detalle"] = _historial_detalle_rows(cuit_limpio)
    except Exception as e:
        out["filas_detalle"] = f"error: {e}"
    try:
        out["historial_detalle_total"] = _nomdeu_conn.execute(
            "SELECT COUNT(*) FROM historial_detalle"
        ).fetchone()[0]
        out["historial_detalle_cuits"] = _nomdeu_conn.execute(
            "SELECT COUNT(DISTINCT cuit) FROM historial_detalle"
        ).fetchone()[0]
    except Exception as e:
        out["historial_detalle_total"] = f"error: {e}"
    return jsonify(out)


@app.route("/admin/cheques-activos/<cuit>")
def admin_cheques_activos(cuit):
    """Diagnóstico rápido: corre para UN cuit la misma lógica de cruce
    vivo+bulk que usan ejecutar_verificacion/proceso-integral
    (_cheques_activos_de), sin necesidad de esperar la verificación
    completa de cartera (que puede tardar horas). Sirve para confirmar el
    fix de cheques al instante sobre un cliente puntual."""
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    out = {"ok": True, "cuit": cuit_limpio}
    try:
        _cheq_cdi, _ = _consultar_bcra_directo(cuit_limpio, 'cheques')
    except Exception as e:
        _cheq_cdi = None
        out["error_cdi"] = str(e)
    act_live, tot_live, _ = _cheques_activos_de(_cheq_cdi)
    out["activos_vivo"] = act_live
    out["total_vivo"] = tot_live
    act_bulk, tot_bulk, _ = _cheques_activos_de(get_cheques_local(cuit_limpio))
    out["activos_bulk"] = act_bulk
    out["total_bulk"] = tot_bulk
    out["activos_final"] = max(act_live, act_bulk)
    return jsonify(out)


@app.route("/admin/cheques-lookup/<cuit>")
def admin_cheques_lookup(cuit):
    """Diagnóstico temporal: muestra cómo está guardado realmente el cuit en
    cheques_bcra (tipo, longitud, muestras) y si matchea el cuit consultado,
    para descartar un mismatch de formato como el que tuvo historial_bulk."""
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    if not os.path.exists(PADRON_DB_PATH):
        return jsonify({"ok": False, "error": "PADRON_DB_PATH no existe"}), 503
    out = {"ok": True, "cuit": cuit_limpio}
    try:
        conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
        out["filas_exactas"] = conn.execute(
            "SELECT cuit, nro_cheque, fecha_rechazo, monto, estado, tipo, cuit_entidad "
            "FROM cheques_bcra WHERE cuit = ?", (cuit_limpio,)
        ).fetchall()
        out["like_match"] = conn.execute(
            "SELECT cuit FROM cheques_bcra WHERE cuit LIKE ? LIMIT 5",
            (f"%{cuit_limpio}%",)
        ).fetchall()
        out["_debug_cuit_typeof"] = conn.execute(
            "SELECT typeof(cuit), length(cuit) FROM cheques_bcra LIMIT 1"
        ).fetchone()
        out["_debug_cuit_samples"] = [r[0] for r in conn.execute(
            "SELECT cuit FROM cheques_bcra LIMIT 5"
        ).fetchall()]
        out["total_registros"] = conn.execute(
            "SELECT COUNT(*) FROM cheques_bcra"
        ).fetchone()[0]
        conn.close()
    except Exception as e:
        out["error"] = f"{e}"
    return jsonify(out)


# Código BCRA (zero-padded a 5 dígitos) → denominación real del banco. Fuente:
# api.bcra.gob.ar/cheques/v1.0/entidades (endpoint público oficial del BCRA,
# mismo código numérico que usa el bulk de deudores 24DSF.txt). Solo cubre
# bancos habilitados en la cámara de cheques — entidades no bancarias (tarjetas
# de crédito, fintechs, etc., con códigos de 6 dígitos) no están en esta lista
# y siguen mostrando el fallback "Entidad NNNNN".
_ENTIDADES_BCRA = {
    '00005': 'The Royal Bank of Scotland N.V.',
    '00007': 'Banco de Galicia y Buenos Aires S.A.',
    '00011': 'Banco de la Nación Argentina',
    '00014': 'Banco de la Provincia de Buenos Aires',
    '00015': 'Industrial and Commercial Bank of China',
    '00016': 'Citibank N.A.',
    '00017': 'Banco BBVA Argentina S.A.',
    '00018': 'MUFG Bank, Ltd',
    '00020': 'Banco de la Provincia de Córdoba S.A.',
    '00027': 'Banco Supervielle S.A.',
    '00029': 'Banco de la Ciudad de Buenos Aires',
    '00034': 'Banco Patagonia S.A.',
    '00044': 'Banco Hipotecario S.A.',
    '00045': 'Banco de San Juan S.A.',
    '00060': 'Banco del Tucumán S.A.',
    '00065': 'Banco Municipal de Rosario',
    '00072': 'Banco Santander Argentina S.A.',
    '00079': 'Banco Regional de Cuyo S.A.',
    '00083': 'Banco del Chubut S.A.',
    '00086': 'Banco de Santa Cruz S.A.',
    '00093': 'La Pampa S.A.',
    '00094': 'Banco de Corrientes S.A.',
    '00097': 'Banco Provincia del Neuquén S.A.',
    '00147': 'Bibank S.A.',
    '00150': 'Banco GGAL S.A.',
    '00191': 'Banco Credicoop Cooperativo Limitado',
    '00198': 'Banco de Valores S.A.',
    '00247': 'Banco Roela S.A.',
    '00254': 'Banco Mariva S.A.',
    '00259': 'Banco BMA S.A.U.',
    '00266': 'BNP Paribas',
    '00268': 'Banco Provincia de Tierra del Fuego',
    '00277': 'Banco Saenz S.A.',
    '00281': 'Banco Meridian S.A.',
    '00285': 'Banco Macro S.A.',
    '00297': 'Banco Banex S.A.',
    '00299': 'Banco Comafi Sociedad Anónima',
    '00301': 'Banco Piano S.A.',
    '00303': 'Banco Finansur S.A.',
    '00305': 'Banco Julio Sociedad Anónima',
    '00306': 'Banco Privado de Inversiones S.A.',
    '00309': 'Banco Rioja Sociedad Anónima Unipersonal',
    '00310': 'Banco del Sol S.A.',
    '00311': 'Nuevo Banco del Chaco S.A.',
    '00315': 'Banco de Formosa S.A.',
    '00319': 'Banco CMF S.A.',
    '00321': 'Banco de Santiago del Estero S.A.',
    '00322': 'Banco Industrial',
    '00330': 'Nuevo Banco de Santa Fe Sociedad Anónima',
    '00336': 'Banco Bradesco Argentina S.A.U.',
    '00338': 'Banco de Servicios y Transacciones S.A.U.',
    '00341': 'Banco Masventas S.A.',
    '00386': 'Nuevo Banco de Entre Ríos S.A.',
    '00389': 'Banco Columbia S.A.',
    '00426': 'Banco Bica S.A.',
    '00431': 'Banco Coinag S.A.',
    '00432': 'Banco de Comercio S.A.',
    '00435': 'Banco Sucredito Regional S.A.U.',
    '00448': 'Banco Dino S.A.',
    # Proveedores no financieros de crédito (PNFC) — fintechs, billeteras, tarjetas
    # no bancarias, mutuales, etc. Código de ruteo BCRA según nómina provista por
    # el usuario (planilla oficial de entidades), no por endpoint público.
    '70014': 'Financiera Monti S.A. (Monti)',
    '70036': 'Financiera del Sol S.A. (Finandisol)',
    '70044': 'Compañía Financiera Argentina S.A. (Efectivo Sí)',
    '70059': 'Metropolis Compañía Financiera S.A.',
    '70077': 'Financiera Americana S.A. (Americana)',
    '70086': 'Cordial Compañía Financiera S.A. (Tarjeta Walmart / ChangoMás)',
    '70092': 'Financoop Compañía Financiera S.A.',
    '70104': 'GE Compañía Financiera S.A. (GE Money)',
    '70112': 'Siembra Compañía Financiera S.A.',
    '70122': 'TarjetaShopping S.A. (Tarjeta Shopping)',
    '70133': 'Tarjeta Automática S.A. (Carta Automática)',
    '70141': 'Presto S.A. (Tarjeta Presto)',
    '70144': 'Tarjeta Emisora del Sur S.A. (Sur Card)',
    '70155': 'Carta Franca S.A.',
    '70161': 'Fiden S.A. (Garbarino / Compumundo MasterCard)',
    '70172': 'Credil S.A.',
    '70177': 'Tarjeta Mara S.A.',
    '70184': 'Tarjeta Invercred S.A. (Invercred)',
    '70188': 'Credife S.A.',
    '70191': 'Quilmeña S.A. (Tarjeta Quilmeña)',
    '70199': 'Crédito Argentino S.A. (Crediaria)',
    '70200': 'Centro Card S.A. (Centrocard)',
    '70208': 'Préstamos al Instante S.A. (Crédito al Instante)',
    '70211': 'Operadora de Tarjetas de Crédito S.A. (Tarjeta Local)',
    '70215': 'Promotora Financiera del Norte S.A. (Profi Norte)',
    '70221': 'Inversora Regional S.A.',
    '70222': 'Crédito Regional S.A.',
    '70229': 'Créditos del NOA S.A.',
    '70233': 'Tarjeta de Crédito Santa Fe S.A. (Santa Fe Créditos)',
    '70239': 'Fertil Finanzas S.A. (Fértil)',
    '70244': 'Tarjeta Miragro S.A. (Miragro)',
    '70248': 'Favacard S.A. (Favacard / Cabal)',
    '70255': 'Tarjeta Plata S.A.',
    '70256': 'Eurocred S.A.',
    '70266': 'Milenium S.A. (Tarjeta Milenium)',
    '70269': 'Sudamericana de Finanzas S.A.',
    '70275': 'Latin American Finances S.A. (LAF Tarjeta)',
    '70277': 'Patagonia Crédito S.A.',
    '70281': 'Tarjeta Actual S.A.',
    '70288': 'Soluciones Financieras S.A.',
    '70299': 'Nexo Emprendimientos S.A. (Nexo Créditos)',
    '70301': 'Tarjeta Pampeana S.A. (Pampeana)',
    '70305': 'Nueva Card S.A.',
    '70311': 'Grupo Card S.A.',
    '70314': 'Creditar S.R.L.',
    '70318': 'Plaza Crédito S.A.',
    '70322': 'Masventas S.A. (Masventas Fintech)',
    '70329': 'Oasis Crédito S.A. (Oasis)',
    '70330': 'Coppel S.A. (Tarjeta Coppel)',
    '70341': 'Cambio Argentina S.A.',
    '70344': 'Mandataria Central S.A.',
    '70349': 'Centro de Soluciones Financieras S.A.',
    '70355': 'Finandino S.A.',
    '70360': 'Interbanking S.A.',
    '70362': 'Red Link S.A. (Link / Valepei)',
    '70364': 'Prisma Medios de Pago S.A. (Payway)',
    '70366': 'Pagos Mis Cuentas S.A. (PMC Financiero)',
    '70369': 'Sistemas Electrónicos de Pago S.A. (SEPSA Financiero)',
    '70370': 'Siisa S.A. (Servicios Inmobiliarios Computarizados)',
    '70372': 'Cablevisión S.A. (Flow / Telecom)',
    '70377': 'Alstom Argentina S.A.',
    '70381': 'Juan Minetti S.A. (Minetti / Holcim)',
    '70388': 'DirecTV Argentina S.A. (DirecTV / DGO)',
    '70390': 'Bepsa del Paraguay S.A.E.C.A. (Bepsa)',
    '70392': 'MoneyGram Payment Systems Argentina S.A.',
    '70395': 'Transacciones Electrónicas S.A.',
    '70399': 'Processamiento Electrónico de Pagos S.A. (PEP Argentina)',
    '70402': 'Empresa de Transportes de Pasajeros Línea 19 S.R.L.',
    '70404': 'Rapipago S.A. (GIRE / Rapipago Financiero)',
    '70405': 'Servicios Electrónicos de Pago S.A. (SEPSA / Pago Fácil)',
    '70408': 'Mercadolibre S.R.L. (Mercado Pago)',
    '70411': 'Crucero del Norte S.R.L. (Crucero)',
    '70412': 'Servicios de Pago Express S.A. (Pago Express)',
    '70413': 'Total Coin S.A. (TotalCoin)',
    '70415': 'Nuevo Chaco Brokers de Seguros S.A. (Chaco Pagos)',
    '70416': 'Telerecargas S.A.',
    '70417': 'Pluspagos S.A. (Billetera Santa Fe)',
    '70419': 'Pagos Pyme S.A. (Pagos Pyme)',
    '70420': 'Andemar S.A.',
    '70422': 'Epagos S.A. (E-Pagos)',
    '70424': 'Tarjeta PL S.A.',
    '70425': 'Prex Argentina S.R.L. (Prex)',
    '70426': 'Pagos360 S.A. (Pagos 360)',
    '70427': 'Tarjeta Grupar S.A.',
    '70432': 'Uenti S.A. (Ualá)',
    '70433': 'Fábrica de Valores S.A. (Fácil Virtual)',
    '70435': 'Sistema de Pago Integral S.A. (SPI)',
    '70437': 'Webpagos S.A.',
    '70438': 'AJA S.A. (Billetera Came Pagos)',
    '70439': 'Lyra Argentina S.A. (Payzen)',
    '70440': 'Global Payments Argentina S.A.',
    '70441': 'Pagos Digitales S.A.',
    '70442': 'Pei Express S.A.',
    '70443': 'Soluciones Tecnológicas S.A.',
    '70444': 'Pagos Compartidos S.A. (SocioPago)',
    '70445': 'Pagos Asociados S.A.',
    '70446': 'Digital Pay S.A.',
    '70447': 'Tecnologías de Pago S.A. (TecnoPago)',
    '70448': 'Calipso Pagos S.A. (Calipso)',
    '70449': 'Ixe S.A. (Ixe Billetera)',
    '70451': 'Net S.A. (NetPay)',
    '70452': 'Pocket Pay S.A. (PocketPay)',
    '70453': 'Electronic Payment Solutions S.A. (EPS)',
    '70454': 'Wayim S.A.',
    '70455': 'Pagoporinteligencia S.A. (Smart Pay)',
    '70456': 'Aura Pagos S.A. (Aura Pay)',
    '70457': 'Soluciones Digitales S.A.',
    '70458': 'Recarga de Celulares S.A. (RecargaPay)',
    '70459': 'Sampagos S.A. (SamPagos)',
    '70460': 'Moni Online S.A. (Moni)',
    '70461': 'Envíos Online S.R.L.',
    '70462': 'Solución Integral de Pagos S.A.',
    '70464': 'Plataforma Abierta de Servicios Financieros S.A. (PASF)',
    '70465': 'Transferencias AR S.A.',
    '70466': 'Moto Envíos S.A. (Moto Envíos Pay)',
    '70467': 'Tu Billetera S.A.',
    '70468': 'Plata Digital S.R.L.',
    '70469': 'Riel Argentina S.A. (Riel Financiero)',
    '70471': 'S S.A. (Billetera S)',
    '70472': 'Tecnología en Entertainment (TECO) S.A. (Personal Pay)',
    '70473': 'BKR S.A. (Bkr Billetera)',
    '70474': 'Digital Wallet S.A.',
    '70475': 'Yapa S.A.',
    '70477': 'Vaksap S.A.',
    '70479': 'Viumi S.A. (Viumi Macro)',
    '70480': 'MODO S.A.U. (Play Digital / MODO)',
    '70481': 'AMX Argentina S.A. (Claro Pay)',
    '70482': 'Zazu S.A. (Zazu Billetera)',
    '70484': 'Tapp S.A. (Tapp Billetera)',
    '70487': 'Open Coin S.A.',
    '70488': 'Ninin S.A. (Ninin Billetera)',
    '70490': 'Ripio International S.R.L. (Ripio)',
    '70491': 'Zoco S.A. (Zoco Pagos)',
    '70492': 'Fintech Argentina S.R.L. (Fintech Ar)',
    '70493': 'Vanguard Fintech S.A.',
    '70494': 'Bitso Argentina S.R.L. (Bitso)',
    '70495': 'Satoshi Tango S.R.L. (SatoshiTango)',
    '70498': 'Lemontree S.A. (Lemon Cash)',
    '70499': 'Xcoin S.A.',
    '70500': 'Asociación Mutual del Personal de Sancor (Sancor Seguros)',
    '70501': 'Asociación Mutual Sancor de Crédito y Consumo (Mutual SanCor)',
    '70503': 'Ohana Digital S.A. (Ohana)',
    '70505': 'Mutual Alianza (Mutual Alianza Préstamos)',
    '70509': "Let's Bit S.A.",
    '70511': 'Pomelo S.R.L.',
    '70512': 'Mutual de Trabajadores Apícolas (Mutuapis)',
    '70514': 'InvertirOnline S.A. (IOL Pay)',
    '70515': 'BPN Pagos S.A. (Confiable Digital - Neuquén)',
    '70518': 'Belo Argentina S.R.L. (Belo)',
    '70519': 'Dracma S.A.',
    '70521': 'Fondos Online S.A.',
    '70522': 'Portfolio Investment S.A. (Portfolio Pay)',
    '70525': 'Banco Cognitivo S.A. (Cognitivo)',
    '70526': 'Coinbase Argentina S.R.L. (Coinbase)',
    '70529': 'Now X S.A.U. (Ank / Now)',
    '72634': 'Tarjeta Naranja S.A.U. (Naranja X)',
    '72644': 'Mobbex Argentina S.R.L. (Mobbex)',
    '72652': 'Málaga Asset Group S.A. (Málaga Visa)',
    '72660': 'Coop. de Provisión de Serv. de Viv. Créd. y Consumo Cooprofin',
    '72685': 'Cooperativa de Crédito Financiero Palmares Ltda.',
    '72688': 'Tarjeta Tuya S.A. (Nuevo Banco del Chaco)',
    '72692': 'Libercoop Cooperativa de Crédito Ltda.',
    '72701': 'Cooperativa de Crédito Financiero Integral Ltda. (CrediCoop)',
    '72710': 'Sindicato de Luz y Fuerza de Córdoba (Luz y Fuerza Préstamos)',
    '72722': 'Nuevas Cooperativas S.A. (CoopPlus)',
    '72733': 'Cooperativa de Provisión de Servicios Integrales Limitada del Sur (CoopDelSur)',
}


def _nomdeu_get_entidad(codigo: str):
    """Nombre de entidad financiera por código Maeent (5 chars). Retorna str o None.
    Prioriza la lista oficial estática (bancos, vía API pública BCRA) y recurre
    a la tabla 'entidades' de la base offline solo como respaldo futuro."""
    cod5 = str(codigo).strip().zfill(5)
    nombre_estatico = _ENTIDADES_BCRA.get(cod5)
    if nombre_estatico:
        return nombre_estatico
    if _nomdeu_conn is None:
        return None
    try:
        row = _nomdeu_conn.execute(
            "SELECT nombre FROM entidades WHERE codigo = ?",
            (cod5,)
        ).fetchone()
        return row[0] if row else None
    except Exception:
        return None


def _nomdeu_build_deudas_resp(cuit: str):
    """Respuesta BCRA sintética desde deudas_resumen offline.
    Primaria para CUITs en el padrón mensual; las APIs en vivo quedan como
    auxilio solo para CUITs nuevos (ausentes de este padrón)."""
    deuda = _nomdeu_get_deuda(cuit)
    if not deuda:
        return None
    nombre = _nomdeu_get_nombre(cuit) or cuit
    ent_codigos = [c.strip() for c in (deuda['entidades_cod'] or '').split(',') if c.strip()]
    if ent_codigos:
        n_ents = len(ent_codigos)
        entidades_list = []
        for cod in ent_codigos:
            ent_nombre = _nomdeu_get_entidad(cod) or f"Entidad {cod}"
            entidades_list.append({
                'entidad': ent_nombre,
                'situacion': deuda['sit_max'],
                'monto': round(deuda['monto_total'] / n_ents, 1),
            })
    else:
        # Sin entidades en el mes más reciente (deuda histórica ya cerrada/no reportada
        # en el snapshot actual): entidad sintética con el sit_max histórico
        entidades_list = [{
            'entidad': 'Sistema Financiero',
            'situacion': deuda['sit_max'],
            'monto': round(deuda['monto_total'], 1),
        }]
    periodo_str = str(deuda['periodo'])
    periodo_int = int(periodo_str) if periodo_str.isdigit() else 0
    return {
        'results': {
            'denominacion': nombre,
            'periodos': [{'periodo': periodo_int, 'entidades': entidades_list}],
        },
        'sin_deudas': False,
        'bcra_disponible': False,
        'fuente_offline': 'bcra_nomdeu_local',
    }


@app.route("/afip/<cuit>")
def get_afip(cuit):
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    cuit_fmt = cuit_limpio[:2] + '-' + cuit_limpio[2:10] + '-' + cuit_limpio[10:] if len(cuit_limpio) == 11 else cuit

    # 0.5. Nombres custom (admin guardó manualmente vía POST /api/nombre-custom)
    _nc_nombre = _nombres_custom.get(cuit_limpio, '').strip()
    if _nc_nombre:
        return jsonify({"nombre": _nc_nombre, "fuente": "custom"})

    # 1. Cartera comercial Piattelli (O(1), sin red) — respuesta garantizada para clientes propios
    nombre_cc = next(
        (str(c.get('nombre', '')).strip() for c in _cartera_comercial
         if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio),
        None
    )
    if nombre_cc:
        return jsonify({"nombre": nombre_cc, "fuente": "cartera"})

    # 2. Saldos / Facturas (Odoo export) — también O(1), sin red
    fuente_sf = _saldos_gestion if _saldos_gestion else _saldos_facturas
    nombre_sf = next(
        (str(f.get('cliente', '')).strip() for f in fuente_sf
         if str(f.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio),
        None
    )
    if nombre_sf:
        return jsonify({"nombre": nombre_sf, "fuente": "saldos"})

    # 2.5. Padrón oficial BCRA local (Nomdeu.txt mensual — 0ms, 0 red)
    nombre_nomdeu = _nomdeu_get_nombre(cuit_limpio)
    if nombre_nomdeu:
        return jsonify({"nombre": nombre_nomdeu, "fuente": "bcra_nomdeu_local"})

    # 3. Caché BCRA local (solo lectura de disco — sin trigger de consulta BCRA)
    try:
        cached_data, _ = cache_get(cuit_limpio)
        if cached_data:
            den = _norm_bcra_resp(cached_data).get('results', {}).get('denominacion', '').strip()
            if den:
                return jsonify({"nombre": den, "fuente": "bcra_cache"})
    except Exception: pass

    # 4. Historial cacheado en disco (escrito por get_historial — tiene denominacion del legacy)
    try:
        _hp_afip = os.path.join(DATA_DIR, f'historial_{cuit_limpio}.json')
        if os.path.exists(_hp_afip):
            with open(_hp_afip, 'r', encoding='utf-8') as _hf:
                _hc = json.load(_hf)
            _den_raw = (_hc.get('payload', {}).get('results') or {}).get('denominacion')
            den_h = str(_den_raw).strip() if _den_raw else ''
            print(f"[afip] {cuit_limpio} historial_cache den_h={repr(den_h)}", flush=True)
            if den_h: return jsonify({"nombre": den_h, "fuente": "bcra_hist_cache"})
    except Exception as _e:
        print(f"[afip] {cuit_limpio} historial_cache error: {_e}", flush=True)

    # 4.1. Padrón local BCRA (bcra_padron.db) — guarda denominacion cuando BCRA live responde OK.
    # Es la fuente más directa para CUITs con historial bancario que no están en AFIP/TangoFactura.
    try:
        _pl_den = (consultar_padron_local(cuit_limpio) or {})
        _pl_den_str = str((_pl_den.get('results') or {}).get('denominacion') or '').strip()
        if _pl_den_str and not _pl_den_str.isdigit():
            print(f"[afip] {cuit_limpio} padron_local OK: {_pl_den_str}", flush=True)
            return jsonify({"nombre": _pl_den_str, "fuente": "padron_local_bcra"})
    except Exception as _e:
        print(f"[afip] {cuit_limpio} padron_local error: {_e}", flush=True)

    # 4.5. ARCA oficial (WSAA + padrón) — fuente autoritativa de la identidad.
    # Va ANTES de los scrapers de terceros: es el único canal que cubre al CUIT
    # que no figura en el bulk del BCRA (recién inscripto, o sin actividad
    # bancaria), y resuelve tanto razón social como apellido y nombre.
    # Costo cero a partir de la segunda consulta: la identidad queda registrada.
    try:
        _arca_nom = _denominacion_arca(cuit_limpio)
        if _arca_nom:
            print(f"[afip] {cuit_limpio} ARCA oficial OK: {_arca_nom}", flush=True)
            return jsonify({"nombre": _arca_nom, "fuente": "arca_oficial"})
    except Exception as _e:
        print(f"[afip] {cuit_limpio} ARCA oficial error: {_e}", flush=True)

    # 4.6. TangoFactura — razonSocial/apellidoNombre del contribuyente
    try:
        _ua_tf = request.headers.get('User-Agent', 'Mozilla/5.0')
        _tf_r = requests.get(
            f"https://afip.tangofactura.com/Rest/GetContribuyenteFull?cuitContribuyente={cuit_limpio}",
            headers={'User-Agent': _ua_tf, 'Accept': 'application/json'},
            timeout=10, verify=False
        )
        print(f"[afip] {cuit_limpio} tangofactura HTTP={_tf_r.status_code}", flush=True)
        if _tf_r.status_code == 200:
            _contrib_tf = (_tf_r.json().get('Contribuyente') or {})
            # Persona jurídica → razonSocial | Persona física → apellidoNombre o apellido+nombre
            _apellido = str(_contrib_tf.get('apellido') or '').strip()
            _nombre_f = str(_contrib_tf.get('nombre') or '').strip()
            _nombre_fisico = (_apellido + ' ' + _nombre_f).strip() if _apellido else _nombre_f
            _rs_tf = str(
                _contrib_tf.get('razonSocial') or
                _contrib_tf.get('apellidoNombre') or
                _nombre_fisico or
                _contrib_tf.get('denominacion') or ''
            ).strip()
            print(f"[afip] {cuit_limpio} tangofactura campos: razonSocial={repr(_contrib_tf.get('razonSocial'))} apellidoNombre={repr(_contrib_tf.get('apellidoNombre'))} apellido={repr(_apellido)} nombre={repr(_nombre_f)}", flush=True)
            if _rs_tf and not _rs_tf.isdigit():
                print(f"[afip] {cuit_limpio} tangofactura OK: {_rs_tf}", flush=True)
                return jsonify({"nombre": _rs_tf, "fuente": "tangofactura"})
    except Exception as _etf:
        print(f"[afip] {cuit_limpio} tangofactura error: {_etf}", flush=True)

    # 4.65. AfipSDK público — proxy REST gratuito sin certificado, mismo dato que ARCA Padrón
    try:
        _sdk_r = requests.get(
            f"https://app.afipsdk.com/api/v1/afip/persons/{cuit_limpio}",
            headers={'Accept': 'application/json', 'User-Agent': 'VendeSeguro/1.0'},
            timeout=10,
        )
        print(f"[afip] {cuit_limpio} afipsdk HTTP={_sdk_r.status_code}", flush=True)
        if _sdk_r.status_code == 200 and _sdk_r.text.strip():
            _sdk_data = _sdk_r.json()
            _apellido_sdk = str(_sdk_data.get('apellido') or '').strip()
            _nombre_sdk   = str(_sdk_data.get('nombre') or '').strip()
            _fisico_sdk   = (_apellido_sdk + ' ' + _nombre_sdk).strip() if _apellido_sdk else _nombre_sdk
            _rs_sdk = str(
                _sdk_data.get('razonSocial') or
                _sdk_data.get('denominacion') or
                _fisico_sdk or ''
            ).strip()
            if _rs_sdk and not _rs_sdk.isdigit():
                print(f"[afip] {cuit_limpio} afipsdk OK: {_rs_sdk}", flush=True)
                return jsonify({"nombre": _rs_sdk, "fuente": "afipsdk"})
    except Exception as _esdk:
        print(f"[afip] {cuit_limpio} afipsdk error: {_esdk}", flush=True)

    # 4.7. CuitOnline — directorio público (cubre personas físicas y jurídicas que
    # no figuran en BCRA ni en tangofactura). Se valida que el CUIT del resultado
    # coincida exacto con el buscado, para no devolver una coincidencia ajena.
    try:
        import re as _re_co
        _co_r = requests.get(
            "https://www.cuitonline.com/search.php",
            params={"q": cuit_limpio},
            headers={'User-Agent': request.headers.get('User-Agent', 'Mozilla/5.0')},
            timeout=10, verify=True
        )
        print(f"[afip] {cuit_limpio} cuitonline HTTP={_co_r.status_code}", flush=True)
        if _co_r.status_code == 200:
            _co_m = _re_co.search(
                r'href="detalle/(\d{11})/[^"]*"[^>]*title="Ver detalles de ([^"]+)"',
                _co_r.text
            )
            if _co_m and _co_m.group(1) == cuit_limpio:
                import html as _html_co
                _rs_co = _html_co.unescape(_co_m.group(2)).strip()
                if _rs_co and not _rs_co.isdigit():
                    print(f"[afip] {cuit_limpio} cuitonline OK: {_rs_co}", flush=True)
                    return jsonify({"nombre": _rs_co, "fuente": "cuitonline"})
    except Exception as _eco:
        print(f"[afip] {cuit_limpio} cuitonline error: {_eco}", flush=True)

    # 4.8. cuit.com.ar — directorio público, cubre empresas disueltas y antiguas
    try:
        import re as _re_ctc
        _ctc_r = requests.get(
            f"https://www.cuit.com.ar/quees.php?p_cuit={cuit_limpio}",
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'},
            timeout=8,
        )
        print(f"[afip] {cuit_limpio} cuit.com.ar HTTP={_ctc_r.status_code}", flush=True)
        if _ctc_r.status_code == 200:
            _m_ctc = _re_ctc.search(
                r'<h1[^>]*>\s*([A-ZÁÉÍÓÚÑa-záéíóúñ0-9][^<]{2,80}?)\s*</h1>',
                _ctc_r.text
            )
            if _m_ctc:
                import html as _html_ctc
                _nom_ctc = _html_ctc.unescape(_m_ctc.group(1)).strip()
                if _nom_ctc and not _nom_ctc.isdigit() and 'CUIT' not in _nom_ctc.upper():
                    print(f"[afip] {cuit_limpio} cuit.com.ar OK: {_nom_ctc}", flush=True)
                    return jsonify({"nombre": _nom_ctc, "fuente": "cuit_com_ar"})
    except Exception as _ectc:
        print(f"[afip] {cuit_limpio} cuit.com.ar error: {_ectc}", flush=True)

    # 5. API BCRA — historial en vivo (solo si no hay caché)
    try:
        r = requests.get("https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/Historicas/" + cuit_limpio, timeout=12, verify=False)
        if r.status_code == 200:
            den2 = _norm_bcra_resp(r.json()).get('results', {}).get('denominacion', '').strip()
            if den2: return jsonify({"nombre": den2, "fuente": "bcra_hist"})
    except Exception: pass

    # 6. API BCRA — deudas vigentes
    try:
        r = requests.get("https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/" + cuit_limpio, timeout=12, verify=False)
        if r.status_code == 200:
            den3 = _norm_bcra_resp(r.json()).get('results', {}).get('denominacion', '').strip()
            if den3: return jsonify({"nombre": den3, "fuente": "bcra_live"})
    except Exception: pass

    # 7. BORA Sección II — Sociedades (último recurso para jurídicas recién inscriptas)
    _bora_nom = _scrape_bora_razon_social(cuit_limpio)
    if _bora_nom:
        return jsonify({"nombre": _bora_nom, "fuente": "bora_seccion2"})

    # 7.5. Nomdeu SQLite (tabla denominaciones — nombres oficiales del BCRA)
    # Cubre cualquier CUIT que haya tenido deuda en el sistema financiero argentino,
    # incluso si no figura en AFIP/TangoFactura/AfipSDK por ser persona jurídica antigua.
    _nomdeu_nom = _nomdeu_get_nombre(cuit_limpio)
    if _nomdeu_nom and not _nomdeu_nom.isdigit():
        print(f"[afip] {cuit_limpio} nomdeu_sqlite OK: {_nomdeu_nom}", flush=True)
        return jsonify({"nombre": _nomdeu_nom, "fuente": "nomdeu_bcra"})

    # 7.6. Identidad ARCA registrada — el cliente no bancarizado nunca está en
    # Nomdeu, pero sí quedó dado de alta en el padrón local desde el WS oficial.
    _ident_nom = (_ident_arca_get(cuit_limpio) or {}).get('razon_social')
    if _ident_nom and not _ident_nom.isdigit():
        print(f"[afip] {cuit_limpio} identidad ARCA local OK: {_ident_nom}", flush=True)
        return jsonify({"nombre": _ident_nom, "fuente": "arca_oficial"})

    # Ninguna fuente devolvió denominación — puede ser padrón temporalmente offline.
    # El frontend trata fuente=fallback como "sin nombre real" y deja que el score decida.
    print(f"[afip] Sin nombre para CUIT {cuit_limpio} — devolviendo formato", flush=True)
    return jsonify({"nombre": cuit_fmt, "fuente": "fallback", "afip_offline": True})

def _enriquecer_denominacion(payload: dict, cuit: str) -> dict:
    """Completa results.denominacion cuando el BCRA no la trae.

    El BCRA solo conoce a quien está bancarizado: para un CUIT nuevo, o de
    alguien que nunca tomó crédito, devuelve sin_deudas con la denominación en
    blanco y la UI terminaba mostrando "Sin denominación". El padrón fiscal, en
    cambio, tiene a todo el mundo desde el día de la inscripción.

    Orden: fuentes locales (cartera, saldos, Nomdeu, identidad ya registrada) y
    recién después el canal oficial de ARCA. No lanza excepción nunca: si no se
    puede resolver, el payload vuelve igual que como entró.
    """
    if not isinstance(payload, dict):
        return payload
    res = payload.get('results')
    if not isinstance(res, dict):
        return payload
    if str(res.get('denominacion') or '').strip():
        return payload   # el BCRA ya la trajo

    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    try:
        nombre = _denominacion_local(cuit_limpio)
        fuente = 'padron_local'
        if not nombre:
            nombre = _denominacion_arca(cuit_limpio)
            fuente = 'arca_oficial'
        if nombre:
            res['denominacion'] = nombre
            payload['denominacion_fuente'] = fuente
            print(f"[denominacion] {cuit_limpio} completada desde {fuente}: {nombre[:50]}", flush=True)
    except Exception as e:
        print(f"[denominacion] {cuit_limpio} no se pudo completar: {e}", flush=True)
    return payload


@app.route("/deudas/<cuit>")
def get_deudas(cuit):
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    if request.args.get('fresh') == '1':
        try:
            _bc_path = os.path.join(DATA_DIR, 'bcra_cache.json')
            if os.path.exists(_bc_path):
                with open(_bc_path, 'r', encoding='utf-8') as _f:
                    _bc = json.load(_f)
                if cuit_limpio in _bc:
                    del _bc[cuit_limpio]
                    with open(_bc_path, 'w', encoding='utf-8') as _f:
                        json.dump(_bc, _f)
                    print(f"[deudas] {cuit_limpio} bcra_cache invalidado (fresh=1)", flush=True)
        except Exception:
            pass
    _fresh = request.args.get('fresh') == '1'
    try:
        data, error = consultar_bcra_cached(cuit_limpio, skip_padron=_fresh)
        return jsonify(_enriquecer_denominacion(data, cuit_limpio)), 200
    except Exception as e:
        import traceback
        print(f"[deudas] Excepcion {cuit}: {traceback.format_exc()}", flush=True)
        return jsonify({"results": None, "sin_deudas": None, "error_bcra": str(e)}), 200

def _cheques_cache_path(cuit):
    return os.path.join(DATA_DIR, f'cheques_{cuit}.json')

def _cheques_cache_get(cuit):
    try:
        path = _cheques_cache_path(cuit)
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            edad = time.time() - data.get('ts', 0)
            payload = data.get('payload')
            if not isinstance(payload, dict):
                return None
            _causales = (payload.get('results') or {}).get('causales') or []
            if _causales:
                # Tiene cheques reales → caché válido 24h
                if edad < 86400:
                    print(f"[cheques] {cuit} desde caché disco ({len(_causales)} causales)", flush=True)
                    return payload
            else:
                # "Sin antecedentes" → caché 24h. Datos BCRA son estáticos durante el día
                # (actualización mensual). API directa desde Render es el canal confiable.
                if edad < 86400:
                    print(f"[cheques] {cuit} sin antecedentes (caché 24h)", flush=True)
                    return payload
                print(f"[cheques] {cuit} caché 'sin cheques' expirado — reintentando BCRA", flush=True)
    except: pass
    return None

def _cheques_cache_set(cuit, payload):
    try:
        path = _cheques_cache_path(cuit)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump({'payload': payload, 'ts': time.time()}, f, ensure_ascii=False)
    except: pass

@app.route("/deudas/<cuit>/cheques")
def get_cheques(cuit):
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    # Caché en disco primero (TTL 24h con cheques, 1h sin cheques)
    cached = _cheques_cache_get(cuit_limpio)
    if cached:
        print(f"[cheques] {cuit_limpio} desde caché", flush=True)
        return jsonify(cached), 200
    # DB local (snapshot diario BCRA) — fuente autoritativa sin latencia de red
    local_db = get_cheques_local(cuit_limpio)
    if local_db is not None:
        _cheques_cache_set(cuit_limpio, local_db)
        n_ch = len((local_db.get('results') or {}).get('causales') or [])
        print(f"[cheques] {cuit_limpio} desde DB local (causales={n_ch})", flush=True)
        return jsonify(local_db), 200
    # Workers + BCRA en paralelo — el primero que responda gana
    def _fetch_chq(url, tmt, via):
        try:
            r = _bcra_get(url, timeout=tmt) if 'bcra.gob.ar' in url else requests.get(url, timeout=tmt, verify=False)
            if r.status_code == 404:
                # Solo los endpoints BCRA oficiales tienen autoridad para decir
                # que un CUIT no tiene cheques. Los workers Cloudflare devuelven
                # 404 cuando no soportan el sub-path — eso no es "sin cheques".
                if 'bcra.gob.ar' in url:
                    return 'NOT_FOUND', via
                return None, via   # worker 404 = ruta no soportada, no "sin datos"
            if r.status_code == 200 and len(r.text.strip()) > 10:
                d = _norm_bcra_resp(r.json())
                if d.get('results') is not None:
                    return d, via
        except Exception as e:
            print(f"[cheques] {via} error para {cuit_limpio}: {e}", flush=True)
        return None, via

    endpoints_chq = [
        (BCRA_WRAPPER_BASE + '/cheques-rechazados/' + cuit_limpio, 3.5, 'bcra_wrapper'),
        (f"https://api.bcra.gob.ar/CentralDeInformacion/v1.0/ChequesRechazados/{cuit_limpio}", 10, "bcra_cdi"),
        (f"https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/ChequesRechazados/{cuit_limpio}", 10, "bcra_legacy"),
    ]
    got_404_chq = False
    with ThreadPoolExecutor(max_workers=len(endpoints_chq)) as ex:
        futs = {ex.submit(_fetch_chq, url, tmt, via): via for url, tmt, via in endpoints_chq}
        try:
            for fut in as_completed(futs, timeout=20):
                result, via = fut.result()
                if result == 'NOT_FOUND':
                    got_404_chq = True
                elif result:
                    payload = result if result.get('results') else {"results": {"causales": []}, "sin_deudas": True, "error_bcra": None}
                    _cheques_cache_set(cuit_limpio, payload)
                    print(f"[cheques] {cuit_limpio} OK via {via}", flush=True)
                    return jsonify(payload), 200
        except Exception:
            pass
    if got_404_chq:
        payload = {"results": {"causales": []}, "sin_deudas": True, "error_bcra": None}
        _cheques_cache_set(cuit_limpio, payload)
        return jsonify(payload), 200
    print(f"[cheques] {cuit_limpio} sin respuesta — devolviendo vacío", flush=True)
    return jsonify({"results": {"causales": []}, "sin_deudas": True, "error_bcra": None}), 200

@app.route("/deudas/<cuit>/historial")
def get_historial(cuit):
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    hist_path = os.path.join(DATA_DIR, f'historial_{cuit_limpio}.json')
    # Caché disco: solo se sirve si tiene periodos reales (vacío = fetch fallido → re-fetchear)
    try:
        if os.path.exists(hist_path):
            with open(hist_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            _periodos_ok = bool((cached.get('payload', {}).get('results') or {}).get('periodos'))
            if _periodos_ok and time.time() - cached.get('ts', 0) < 86400:
                print(f"[historial] {cuit_limpio} desde caché disco", flush=True)
                return jsonify(cached['payload']), 200
    except: pass
    # Workers + BCRA en paralelo — el primero que responda gana
    def _fetch_hist(url, tmt, via):
        try:
            r = _bcra_get(url, timeout=tmt) if 'bcra.gob.ar' in url else requests.get(url, timeout=tmt, verify=False)
            if r.status_code == 404:
                return 'NOT_FOUND', via
            if r.status_code == 200 and len(r.text.strip()) > 10:
                raw = r.json()
                # CDI v1.0 devuelve 'detalle' (no 'periodos') → necesita _map_detalle_bcra
                _res = raw.get('results') if isinstance(raw, dict) else None
                if isinstance(_res, dict) and 'detalle' in _res:
                    d = _map_detalle_bcra(raw)
                else:
                    d = _norm_bcra_resp(raw)
                if d.get('results') is not None and not d.get('error'):
                    d['sin_deudas'] = len((d.get('results') or {}).get('periodos') or []) == 0
                    return d, via
        except Exception as e:
            print(f"[historial] {via} error para {cuit_limpio}: {e}", flush=True)
        return None, via

    # ScraperAPI (fallback de BrightData) necesita hasta 30s para sortear firewalls BCRA.
    # as_completed timeout = 40s para dar margen cuando ambos endpoints usan ScraperAPI.
    endpoints_hist = [
        (f"https://api.bcra.gob.ar/CentralDeInformacion/v1.0/Deudas/Historicas/{cuit_limpio}", 30, "bcra_cdi"),
        (f"https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/Historicas/{cuit_limpio}",    30, "bcra_legacy"),
    ]
    got_404_hist = False
    with ThreadPoolExecutor(max_workers=len(endpoints_hist)) as ex:
        futs = {ex.submit(_fetch_hist, url, tmt, via): via for url, tmt, via in endpoints_hist}
        try:
            for fut in as_completed(futs, timeout=40):
                result, via = fut.result()
                if result == 'NOT_FOUND':
                    got_404_hist = True   # no retornar aún — otro endpoint puede tener datos
                elif result:
                    try:
                        with open(hist_path, 'w', encoding='utf-8') as f:
                            json.dump({'payload': result, 'ts': time.time()}, f, ensure_ascii=False)
                    except: pass
                    print(f"[historial] {cuit_limpio} OK via {via}", flush=True)
                    return jsonify(result), 200
        except Exception:
            pass
    if got_404_hist:
        # BCRA a veces devuelve 404 como respuesta de rate-limit (falso 404), no como "no hay datos".
        # Si el cliente tiene historial real pero fue consultado 1s después de deudas, caeremos aquí.
        # Último intento: esperar 2s y reintentar CDI v1 directamente antes de declarar "sin datos".
        time.sleep(2)
        _url_retry = f"https://api.bcra.gob.ar/CentralDeInformacion/v1.0/Deudas/Historicas/{cuit_limpio}"
        _retry_data, _ = _fetch_hist(_url_retry, 30, 'bcra_cdi_retry')
        if _retry_data and _retry_data != 'NOT_FOUND':
            try:
                with open(hist_path, 'w', encoding='utf-8') as f:
                    json.dump({'payload': _retry_data, 'ts': time.time()}, f, ensure_ascii=False)
            except: pass
            print(f"[historial] {cuit_limpio} OK en retry post-404 (2s delay)", flush=True)
            return jsonify(_retry_data), 200
        # Antes de declarar "sin datos": un 404 en un endpoint + connection-reset en el otro
        # no es prueba de que no haya historial — probar el padrón offline primero.
        _deuda_bulk_404 = _nomdeu_get_deuda(cuit_limpio)
        if _deuda_bulk_404:
            _offline_hist_404 = _bulk_to_hist_data(cuit_limpio)
            print(f"[historial] {cuit_limpio} fallback historial_detalle ({_HIST_DETALLE_MESES}m reales, post-404)", flush=True)
            return jsonify(_offline_hist_404), 200
        # Sin historial bancario: el CUIT puede ser nuevo o no haber operado nunca
        # con bancos. La identidad igual se resuelve por el padrón fiscal.
        return jsonify(_enriquecer_denominacion(
            {"results": {"denominacion": "", "periodos": []}, "sin_deudas": True, "error_bcra": None},
            cuit_limpio,
        )), 200

    # Fallback offline: si la API falló por completo (sin 404, ej. ConnectionResetError en
    # ambos endpoints), reconstruir el historial real desde historial_detalle — mismo
    # criterio que usa el motor de scoring. Un solo período (_nomdeu_build_deudas_resp)
    # no sirve para un análisis de tendencia real.
    _deuda_bulk = _nomdeu_get_deuda(cuit_limpio)
    if _deuda_bulk:
        _offline_hist = _bulk_to_hist_data(cuit_limpio)
        print(f"[historial] {cuit_limpio} fallback historial_detalle ({_HIST_DETALLE_MESES}m reales)", flush=True)
        return jsonify(_offline_hist), 200

    return jsonify({"results": None, "sin_deudas": None, "error_bcra": "sin_respuesta"}), 200

@app.route("/analizar", methods=["POST"])
def analizar():
    if not GEMINI_KEY and not OPENAI_KEY:
        return jsonify({"error": "API key no configurada"}), 500
    try:
        body = request.get_json()
        prompt = body.get('prompt', '')
        rango_min = int(body.get('rangoMin') or 0)
        rango_max = int(body.get('rangoMax') or 0)
        rango_decision = str(body.get('rangoDecision') or 'denegado')

        def _fmt_pesos(v):
            if v <= 0:
                return '$0'
            s = str(int(v))
            parts, tmp = [], s
            while len(tmp) > 3:
                parts.append(tmp[-3:])
                tmp = tmp[:-3]
            parts.append(tmp)
            return '$' + '.'.join(reversed(parts))

        if rango_decision == 'denegado' or (rango_min == 0 and rango_max == 0):
            rango_label = 'Crédito denegado'
        elif rango_min == rango_max:
            rango_label = f'Límite de crédito sugerido: {_fmt_pesos(rango_min)}'
        else:
            rango_label = f'Límite de crédito sugerido: {_fmt_pesos(rango_min)} – {_fmt_pesos(rango_max)}'

        # El frontend ya incluyó el rango en el prompt; el backend lo refuerza
        # como bloque autorizado para que el system prompt lo tome como referencia.
        # Contexto macro — solo para análisis cualitativo, NUNCA modifica límite
        macro = _fetch_macro_data()
        if macro:
            partes_macro = []
            if macro.get('inflacion') is not None:
                partes_macro.append(f"Inflación interanual: {macro['inflacion']:.1f}%")
            if macro.get('riesgo_pais') is not None:
                partes_macro.append(f"Riesgo país: {int(macro['riesgo_pais'])} bps")
            if macro.get('dolar_blue_venta') is not None:
                partes_macro.append(f"Dólar blue: ${macro.get('dolar_blue_compra','—')} compra / ${macro['dolar_blue_venta']} venta")
            if partes_macro:
                bloque_macro = (
                    "\n\n--- CONTEXTO MACROECONÓMICO (solo para análisis cualitativo — "
                    "NO modifica el límite de crédito ni el score) ---\n"
                    + "\n".join(f"• {p}" for p in partes_macro)
                )
                prompt = prompt + bloque_macro

        payload = {
            "systemInstruction": {"parts": [{"text": CREDIT_ANALYSIS_SYSTEM_PROMPT}]},
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.15, "maxOutputTokens": 1024}
        }
        texto, error = gemini_request(payload, timeout=90, system_prompt=CREDIT_ANALYSIS_SYSTEM_PROMPT)
        if error:
            return jsonify({"error": error}), 500

        # ── Corrección de seguridad post-IA ──────────────────────────────────────
        # Garantiza que RECOMENDACIÓN ESTRATÉGICA siempre cita el rango oficial.
        # Cubre dos casos:
        #   1. El rango oficial no aparece en la sección (IA lo omitió o cambió)
        #   2. La sección contiene "$0" o expresiones de crédito cero/nulo (IA se equivocó)
        if rango_label != 'Crédito denegado' and rango_min > 0:
            import re as _re
            marker = 'RECOMENDACIÓN ESTRATÉGICA:'
            if marker in texto:
                pre, _, section = texto.partition(marker)
                section_clean = section.lstrip(' \n')
                _tiene_cero = bool(_re.search(
                    r'\$\s*0\b'                          # $0 literal
                    r'|\bcero\s+(?:peso|cr[eé]dito|limit)'  # cero pesos/crédito
                    r'|\blímite\b[^.]{0,60}\$\s*0',     # límite ... $0
                    section_clean[:400], _re.IGNORECASE
                ))
                _sin_rango = rango_label not in section_clean
                if _sin_rango or _tiene_cero:
                    # Evitar duplicar si el rango ya encabeza la sección
                    if section_clean.startswith(rango_label):
                        section_clean = section_clean[len(rango_label):].lstrip('. \n')
                    texto = pre + marker + ' ' + rango_label + '. ' + section_clean
                    print(
                        f"[analizar] corrección post-IA: {rango_label} "
                        f"(sin_rango={_sin_rango} cero={_tiene_cero})",
                        flush=True
                    )

        return jsonify({"texto": texto, "rangoMin": rango_min, "rangoMax": rango_max, "rangoDecision": rango_decision})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/procesar-veraz", methods=["POST"])
@app.route("/procesar-informe", methods=["POST"])
def procesar_veraz():
    if not GEMINI_KEY:
        return jsonify({"error": "API key no configurada"}), 500
    try:
        body = request.get_json(force=True)
        if not body:
            return jsonify({"error": "Request body vacio o no es JSON"}), 400
        pdf_base64 = body.get('pdf', '')
        print(f"[procesar-informe] PDF recibido: {len(pdf_base64)} chars, ~{len(pdf_base64)*3//4//1024} KB", flush=True)
        if len(pdf_base64) * 3 // 4 > 20 * 1024 * 1024:
            return jsonify({"error": "PDF demasiado grande (max 20MB)"}), 400
        prompt = (
            "Este puede ser un informe de Veraz/Equifax o de Nosis. Detecta el formato automaticamente y extrae los mismos campos. "
            "Responde SOLO con un objeto JSON valido, sin markdown, sin texto adicional. "
            "Estructura exacta: "
            '{"nombre":"","cuit":"","score":0,"situacion_bcra":"","cheques_rechazados":0,'
            '"monto_cheques":"","saldo_vencido":"","deuda_sistema_financiero":"",'
            '"maximo_atraso":"","entidades_problema":[],"resumen":"",'
            '"socios_directores":[{"nombre":"","cuit_dni":"","cargo":"","score":0,"situacion":""}]} '
            "El array socios_directores debe incluir todos los socios, directores o representantes "
            "legales con su informacion crediticia. Si no hay, dejar array vacio []."
        )
        if not OPENAI_KEY:
            return jsonify({"error": "No hay API key de OpenAI configurada"}), 500
        try:
            import base64 as b64mod
            import fitz
            pdf_bytes = b64mod.b64decode(pdf_base64)
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            imagenes_b64 = []
            for page in list(doc)[:3]:
                mat = fitz.Matrix(2, 2)
                pix = page.get_pixmap(matrix=mat)
                img_bytes = pix.tobytes("png")
                imagenes_b64.append(b64mod.b64encode(img_bytes).decode())
            doc.close()
            print(f"[procesar-informe] PDF convertido a {len(imagenes_b64)} paginas", flush=True)
        except Exception as ex:
            print(f"[procesar-informe] Error convirtiendo PDF: {ex}", flush=True)
            return jsonify({"error": "No se pudo convertir el PDF: " + str(ex)}), 500
        content_oai = [{"type": "text", "text": prompt}]
        for img_b64 in imagenes_b64[:4]:
            content_oai.append({"type": "image_url", "image_url": {"url": "data:image/png;base64," + img_b64}})
        headers_oai = {"Content-Type": "application/json", "Authorization": "Bearer " + OPENAI_KEY}
        body_oai = {"model": "gpt-4o", "max_tokens": 1500, "messages": [{"role": "user", "content": content_oai}]}
        try:
            r_oai = requests.post("https://api.openai.com/v1/chat/completions", headers=headers_oai, json=body_oai, timeout=250)
            d_oai = r_oai.json()
            print(f"[procesar-informe] OpenAI status {r_oai.status_code}", flush=True)
            if r_oai.status_code == 200:
                texto = d_oai["choices"][0]["message"]["content"]
            else:
                msg = d_oai.get("error", {}).get("message", "Error OpenAI")
                return jsonify({"error": "Error OpenAI: " + msg}), 503
        except Exception as ex:
            return jsonify({"error": str(ex)}), 503
        if not texto:
            return jsonify({"error": "No se pudo procesar el PDF."}), 503
        texto_limpio = texto.strip().replace("```json", "").replace("```", "").strip()
        import re as re_mod
        match = re_mod.search(r'\{[\s\S]+\}', texto_limpio)
        if match:
            texto_limpio = match.group(0)
        return jsonify(json.loads(texto_limpio))
    except json.JSONDecodeError as e:
        return jsonify({"error": "Error al parsear respuesta: " + str(e)}), 500
    except Exception as e:
        import traceback
        print(f"[procesar-informe] Exception: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/test-gemini")
def test_gemini():
    if not GEMINI_KEY:
        return jsonify({"error": "No hay API key"}), 500
    payload = {"contents": [{"parts": [{"text": "Responde solo con la palabra OK"}]}]}
    texto, error = gemini_request(payload)
    if error:
        return jsonify({"error": error}), 500
    return jsonify({"ok": True, "respuesta": texto})

@app.route("/cache-stats")
def cache_stats():
    ahora = time.time()
    activos = sum(1 for v in bcra_cache.values() if ahora - v.get('timestamp', 0) < CACHE_TTL)
    return jsonify({"total": len(bcra_cache), "activos": activos, "ttl_horas": CACHE_TTL/3600})

def _fecha_valida(fecha_str, desde):
    try:
        if not fecha_str: return False
        if '/' in str(fecha_str):
            partes = str(fecha_str).split('/')
            from datetime import datetime
            if len(partes[2]) == 2:
                f = datetime(2000+int(partes[2]), int(partes[1]), int(partes[0]))
            else:
                f = datetime(int(partes[2]), int(partes[1]), int(partes[0]))
        else:
            from datetime import datetime
            f = datetime.fromisoformat(str(fecha_str)[:10])
        return f >= desde
    except:
        return True

@app.route("/health")
def health():
    """Estado del servicio + build efectivamente desplegado.

    score_version y los flags de módulos permiten distinguir un deploy nuevo
    de uno viejo sin depender de leer los logs de Render.
    """
    return jsonify({
        "status":         "ok",
        "gemini":         bool(GEMINI_KEY),
        "comercial":      len(_cartera_comercial),
        "score_version":  _SCORE_VERSION,
        "modulos": {
            "scoring_fiscal": SCORING_FISCAL_OK,
            "arca_ws":        _ARCA_MODULO_OK,
            "arca_activo":    ARCA_DISPONIBLE,   # módulo OK + certificado cargado
        },
    })


_warm_estado = {"corriendo": False, "progreso": 0, "total": 0, "ultimo": ""}

@app.route("/warm-padron", methods=["POST", "GET"])
def warm_padron():
    """Precalienta el caché nocturno completo: deudas + historial + cheques.

    Diseñado para ejecutarse via cron a las 3:00 AM (AR) — cron-job.org o similar.
    Con 1300 clientes y 4s entre cada uno, tarda ~1.5h. Al amanecer todos los
    clientes están en caché y no se toca BCRA durante el día.

    Solo refresca los CUITs cuyo caché tiene más de 20h (evita re-queries innecesarios
    si el job se ejecuta varias veces o hay reintentos del cron).
    """
    if _warm_estado["corriendo"]:
        return jsonify({"status": "ya_corriendo", **_warm_estado}), 200

    def _run():
        _warm_estado["corriendo"] = True
        _TTL = 25 * 24 * 3600  # 25 días — datos BCRA se actualizan mensualmente

        try:
            with open(ALERTAS_FILE, 'r', encoding='utf-8') as f:
                cartera = json.load(f)
            cuits = [
                str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip()
                for c in cartera if isinstance(c, dict) and c.get('cuit')
            ]
            cuits = [c for c in cuits if len(c) >= 10]
            _warm_estado["total"] = len(cuits)
            print(f"[warm] Iniciando: {len(cuits)} CUITs", flush=True)

            for i, cuit in enumerate(cuits):
                _warm_estado["progreso"] = i + 1
                _warm_estado["ultimo"] = cuit
                ahora = time.time()

                # ── 1. Deudas ─────────────────────────────────────────────────
                try:
                    local = consultar_padron_local(cuit)
                    _padron_ok = local and local.get('bcra_disponible')
                    _cache_ok  = False
                    _cf = os.path.join(DATA_DIR, 'bcra_cache.json')
                    if not _padron_ok and os.path.exists(_cf):
                        try:
                            _bc = json.load(open(_cf, 'r'))
                            _ent = _bc.get(cuit, {})
                            _cache_ok = ahora - (_ent.get('ts') or 0) < _TTL
                        except Exception:
                            pass
                    if not _padron_ok and not _cache_ok:
                        data, err = consultar_bcra(cuit)
                        if data and not err:
                            _guardar_en_padron_local(cuit, data)
                            print(f"[warm] {i+1}/{len(cuits)} {cuit} deudas OK", flush=True)
                except Exception as e:
                    print(f"[warm] {cuit} deudas error: {e}", flush=True)

                # ── 2. Historial ──────────────────────────────────────────────
                try:
                    _hp = os.path.join(DATA_DIR, f'historial_{cuit}.json')
                    _hist_ok = os.path.exists(_hp) and ahora - os.path.getmtime(_hp) < _TTL
                    if not _hist_ok:
                        _hd, _ = _consultar_bcra_directo(cuit, 'historial')
                        _hd_periodos = (_hd.get('results') or {}).get('periodos') if _hd else None
                        if _hd and _hd_periodos:  # solo cachear si hay datos reales
                            with open(_hp, 'w', encoding='utf-8') as _hf:
                                json.dump({'payload': _hd, 'ts': ahora}, _hf, ensure_ascii=False)
                            print(f"[warm] {i+1}/{len(cuits)} {cuit} historial OK", flush=True)
                except Exception as e:
                    print(f"[warm] {cuit} historial error: {e}", flush=True)

                # ── 3. Cheques ────────────────────────────────────────────────
                try:
                    _cheq_cached = _cheques_cache_get(cuit)
                    if _cheq_cached is None:
                        cheq_local = get_cheques_local(cuit)
                        if cheq_local is not None:
                            _cheques_cache_set(cuit, cheq_local)
                        else:
                            _cd, _ = _consultar_bcra_directo(cuit, 'cheques')
                            if _cd:
                                _cheques_cache_set(cuit, _cd)
                                print(f"[warm] {i+1}/{len(cuits)} {cuit} cheques OK", flush=True)
                except Exception as e:
                    print(f"[warm] {cuit} cheques error: {e}", flush=True)

                # 4s entre clientes — ~1.5h para 1300 CUITs, debajo del rate-limit BCRA
                time.sleep(4)

        except Exception as e:
            print(f"[warm] Error general: {e}", flush=True)
        finally:
            _warm_estado["corriendo"] = False
            print(f"[warm] Finalizado — {len(cuits) if 'cuits' in dir() else '?'} CUITs procesados", flush=True)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "iniciado", "total": _warm_estado["total"]}), 202


@app.route("/warm-padron/estado")
def warm_padron_estado():
    return jsonify(_warm_estado), 200


@app.route("/update-cheques-db", methods=["POST", "GET"])
def update_cheques_db():
    """Actualiza la DB local de cheques rechazados desde el bulk file diario del BCRA.

    Diseñado para ser llamado desde cron-job.org todos los días a las ~10:00 AM AR
    (cuando el BCRA publica el archivo del día). El proceso se ejecuta en background;
    usar /update-cheques-db/estado para monitorear el progreso.

    Parámetro opcional: ?fecha=YYYYMMDD (por defecto: fecha de hoy)
    """
    if _cheques_db_estado.get('corriendo'):
        return jsonify({
            "status":      "ya_corriendo",
            "progreso":    _cheques_db_estado['progreso'],
            "ultimo_paso": _cheques_db_estado['ultimo_paso'],
        }), 202

    fecha = (request.args.get('fecha') or '').strip() or time.strftime('%Y%m%d')

    def _run():
        ok = _import_cheques_zip(fecha)
        print(f"[cheques_db] Actualización {'exitosa' if ok else 'fallida'} para {fecha}", flush=True)
        if ok and list(_cartera_comercial):
            threading.Thread(target=_check_cheques_cartera_bg, daemon=True).start()
            print("[cheques_db] Verificación de cheques de cartera iniciada automáticamente", flush=True)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({
        "status":  "iniciado",
        "fecha":   fecha,
        "message": "Importación en background. Consultar /update-cheques-db/estado.",
    }), 202


@app.route("/update-cheques-db/estado")
def update_cheques_db_estado():
    """Estado actual de la importación de cheques + metadatos de la última importación."""
    meta = {}
    try:
        if os.path.exists(PADRON_DB_PATH):
            conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
            rows = conn.execute("SELECT key, valor FROM _cheques_meta").fetchall()
            conn.close()
            meta = {k: v for k, v in rows}
    except Exception:
        pass
    return jsonify({**_cheques_db_estado, "db_meta": meta}), 200


@app.route("/update-mipyme-db", methods=["POST", "GET"])
def update_mipyme_db():
    """Descarga el padrón MiPyME oficial (datos.gob.ar) e importa a SQLite local.

    El proceso se ejecuta en background (puede tardar 1-3 minutos según red).
    Consultar /update-mipyme-db/estado para monitorear progreso.

    Parámetro opcional: ?url=... para sobreescribir la URL de descarga.
    """
    if _mipyme_import_estado.get('corriendo'):
        return jsonify({
            "status":      "ya_corriendo",
            "progreso":    _mipyme_import_estado['progreso'],
            "ultimo_paso": _mipyme_import_estado['ultimo_paso'],
        }), 202

    url_custom = (request.args.get('url') or '').strip() or None

    def _run():
        ok = _import_mipyme_csv(url_custom)
        print(f"[mipyme] Actualización {'exitosa' if ok else 'fallida'}", flush=True)

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({
        "status":  "iniciado",
        "url":     url_custom or MIPYME_CSV_URL,
        "message": "Importación en background. Consultar /update-mipyme-db/estado.",
    }), 202


@app.route("/update-mipyme-db/estado")
def update_mipyme_db_estado():
    """Estado actual de la importación MiPyME + metadatos de la última importación."""
    meta = {}
    try:
        if os.path.exists(MIPYME_DB_PATH):
            _mc = sqlite3.connect(MIPYME_DB_PATH, check_same_thread=False)
            try:
                rows = _mc.execute("SELECT key, valor FROM _mipyme_meta").fetchall()
                meta = {k: v for k, v in rows}
            except Exception:
                pass
            finally:
                _mc.close()
    except Exception:
        pass
    return jsonify({
        **_mipyme_import_estado,
        "db_cargada": _mipyme_conn is not None,
        "db_meta":    meta,
    }), 200


@app.route("/cache/limpiar/<cuit>", methods=["POST", "GET"])
def limpiar_cache_cuit(cuit):
    cuit_limpio = cuit.replace("-", "")
    eliminados = [key for key in list(bcra_cache.keys()) if key == cuit_limpio]
    for key in eliminados:
        del bcra_cache[key]
    print(f"[cache] Limpiado CUIT {cuit_limpio}: {eliminados}", flush=True)
    return jsonify({"ok": True, "cuit": cuit_limpio, "eliminados": len(eliminados)})

@app.route("/cache/limpiar-todo", methods=["POST", "GET"])
def limpiar_cache_todo():
    total = len(bcra_cache)
    bcra_cache.clear()
    print(f"[cache] Cache completo limpiado: {total} entradas", flush=True)
    return jsonify({"ok": True, "eliminados": total})

@app.route("/dso-ventas/limpiar", methods=["POST"])
def limpiar_dso_ventas():
    try:
        dso_file = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
        if os.path.exists(dso_file):
            os.remove(dso_file)
        return jsonify({"ok": True, "mensaje": "Historial limpiado"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/dso-saldos", methods=["GET"])
def get_dso_saldos():
    try:
        modo = request.args.get('modo', 'actual')
        f_path = os.path.join(DATA_DIR, 'dso_saldos_historico.json' if modo == 'historico' else 'dso_saldos_actual.json')
        if os.path.exists(f_path):
            try:
                with open(f_path, 'r', encoding='utf-8') as f:
                    return jsonify(json.load(f))
            except Exception:
                pass
        # Fallback: build from _saldos_facturas in memory (314 records — SSoT on Render)
        if _saldos_facturas:
            saldos = [
                {"cliente": f.get("cliente", ""), "fecha_factura": f.get("fechaFactura", ""),
                 "fecha_pago": f.get("fechaPago", ""), "saldo": f.get("saldo", 0)}
                for f in _saldos_facturas if (f.get("saldo") or 0) > 0
            ]
            print(f"[dso-saldos] Fallback: {len(saldos)} registros desde saldos_facturas.json", flush=True)
            return jsonify({"saldos": saldos, "ultima_actualizacion": "", "fuente": "facturas"})
        return jsonify({"saldos": [], "ultima_actualizacion": ""})
    except Exception as e:
        print(f"[dso-saldos] Error: {e}", flush=True)
        return jsonify({"saldos": [], "ultima_actualizacion": ""})

@app.route("/dso-saldos", methods=["POST"])
def save_dso_saldos():
    try:
        body = request.get_json(force=True)
        nuevos = body.get('saldos', [])
        if not nuevos:
            return jsonify({"error": "Sin saldos"}), 400
        from datetime import datetime, date as _date
        hoy = datetime.now()

        # La fecha de corte define el "hoy" para el cálculo de días pendientes.
        # Si el body incluye fecha_corte (YYYY-MM-DD), se usa esa fecha en lugar de today().
        # Esto es necesario cuando el reporte corresponde a un cierre anterior (ej: 31/05/2026).
        fecha_corte_str = str(body.get('fecha_corte') or '').strip()
        try:
            if fecha_corte_str and len(fecha_corte_str) >= 10:
                p = fecha_corte_str[:10].split('-')
                corte_d = _date(int(p[0]), int(p[1]), int(p[2]))
            else:
                corte_d = hoy.date()
        except Exception:
            corte_d = hoy.date()

        f_actual = os.path.join(DATA_DIR, 'dso_saldos_actual.json')
        with open(f_actual, 'w', encoding='utf-8') as f:
            json.dump({"saldos": nuevos, "ultima_actualizacion": hoy.strftime('%d/%m/%Y %H:%M'),
                       "fecha_corte": corte_d.strftime('%Y-%m-%d')}, f, ensure_ascii=False)

        # ── Calcular DSO = (Saldo + Cheques) / Ventas × días_período ──────────
        # Fórmula balance-sheet: relaciona el stock de deuda (AR + cheques) con
        # el flujo de ventas de los últimos 3 meses. No usa antigüedad de facturas.

        # 1. Cheques pendientes por cliente (solo positivos = pendientes de cobro)
        _cheques_map: dict = {}   # _norm_dso_match → total_cheques
        _cp = os.path.join(DATA_DIR, 'dso_cheques_actual.json')
        if os.path.exists(_cp):
            try:
                for c in json.load(open(_cp, 'r', encoding='utf-8')).get('cheques', []):
                    tot = float(c.get('total') or 0)
                    if tot > 0:
                        ck = _norm_dso_match(str(c.get('cliente') or ''))
                        if ck:
                            _cheques_map[ck] = _cheques_map.get(ck, 0.0) + tot
            except Exception as _ce:
                print(f"[dso-saldos] Error cargando cheques: {_ce}", flush=True)

        # 2. Ventas por cliente: últimos 3 meses ≤ fecha_corte
        _ventas_nom:   dict = {}   # _norm_nombre(cli) → total_ventas_3m
        _ventas_dso:   dict = {}   # _norm_dso_match(cli) → total_ventas_3m
        _dias_periodo  = 91        # default si no hay datos de ventas
        _corte_ym      = f"{corte_d.year}-{corte_d.month:02d}"
        _vp = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
        if os.path.exists(_vp):
            try:
                _vdata = json.load(open(_vp, 'r', encoding='utf-8'))
                _meses_disp = sorted(
                    [ym for ym in _vdata.get('meses', {}) if ym <= _corte_ym],
                    reverse=True
                )
                _meses_usar = set(_meses_disp[:3])   # máximo 3 meses hacia atrás
                if _meses_usar:
                    _oldest = min(_meses_usar)
                    _oy, _om = int(_oldest[:4]), int(_oldest[5:7])
                    _dias_periodo = (_date(_date.today().year, corte_d.month, corte_d.day)
                                     - _date(_oy, _om, 1)).days + 1
                    # Más simple y correcto:
                    _dias_periodo = (corte_d - _date(_oy, _om, 1)).days + 1
                    for _cli_k, _meses_cli in _vdata.get('por_cliente', {}).items():
                        _vt = sum(float(v) for ym, v in _meses_cli.items() if ym in _meses_usar)
                        if _vt > 0:
                            _ventas_nom[_cli_k] = _ventas_nom.get(_cli_k, 0.0) + _vt
                            _dk = _norm_dso_match(_cli_k)
                            if _dk:
                                _ventas_dso[_dk] = _ventas_dso.get(_dk, 0.0) + _vt
                print(f"[dso-saldos] Ventas: {len(_meses_usar)} meses {sorted(_meses_usar)}, "
                      f"{len(_ventas_nom)} clientes, {_dias_periodo} días", flush=True)
            except Exception as _ve:
                print(f"[dso-saldos] Error cargando ventas: {_ve}", flush=True)

        # 3. Agregar saldo por cliente
        _acc: dict = {}   # _norm_dso_match → {ss: float, nom_original: str}
        for s in nuevos:
            saldo = float(s.get('saldo') or 0)
            if saldo <= 0:
                continue
            _nom_orig = str(s.get('cliente') or '').strip()
            _nk = _norm_dso_match(_nom_orig)
            if _nk:
                if _nk not in _acc:
                    _acc[_nk] = {'ss': 0.0, 'nom_original': _nom_orig}
                _acc[_nk]['ss'] += saldo

        # 4. DSO por cliente
        dso_por_nombre: dict = {}
        _n_con_ventas = 0
        for _nk, _info in _acc.items():
            _saldo_cli   = _info['ss']
            _cheques_cli = _cheques_map.get(_nk, 0.0)
            _nom_n       = _norm_nombre(_info['nom_original'])
            _ventas_cli  = (_ventas_nom.get(_nom_n)
                            or _ventas_dso.get(_nk)
                            or 0.0)
            if _ventas_cli > 0 and _dias_periodo > 0:
                dso_por_nombre[_nk] = round((_saldo_cli + _cheques_cli) / _ventas_cli * _dias_periodo)
                _n_con_ventas += 1
            # Sin ventas: omitir → el cliente aparece con DSO=0 en el portfolio

        f_ind = os.path.join(DATA_DIR, 'dso_individual_actual.json')
        with open(f_ind, 'w', encoding='utf-8') as f:
            json.dump({"por_cuit": {}, "por_nombre": dso_por_nombre,
                       "fecha_corte": corte_d.strftime('%Y-%m-%d'),
                       "dias_periodo": _dias_periodo}, f, ensure_ascii=False)

        total = sum(s.get('saldo', 0) for s in nuevos)
        print(f"[dso-saldos] {len(nuevos)} saldos ${total:,.0f} | corte={corte_d} | "
              f"DSO calculado: {_n_con_ventas}/{len(_acc)} clientes | "
              f"cheques={len(_cheques_map)} | días={_dias_periodo}", flush=True)
        return jsonify({"ok": True, "agregados": len(nuevos), "total": len(nuevos)})
    except Exception as e:
        import traceback
        print(f"[dso-saldos] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/dso-individual-debug")
def dso_individual_debug():
    """Diagnóstico: muestra dso_individual_actual.json.
    ?nombre=X  → busca ese nombre y muestra la clave normalizada vs lo que está en el archivo.
    ?todos=1   → devuelve el dict completo por_nombre."""
    _path = os.path.join(DATA_DIR, 'dso_individual_actual.json')
    if not os.path.exists(_path):
        return jsonify({"existe": False, "mensaje": "Archivo no encontrado"})
    try:
        data = json.load(open(_path, 'r', encoding='utf-8'))
        por_cuit   = data.get('por_cuit', {})
        por_nombre = data.get('por_nombre', {})

        nombre_q = request.args.get('nombre', '').strip()
        if nombre_q:
            norm = _norm_dso_match(nombre_q)
            val  = por_nombre.get(norm)
            # Si piden ?detalle=1 también mostramos las facturas crudas del DSO saldos
            facturas_raw = []
            if request.args.get('detalle') == '1':
                _sp = os.path.join(DATA_DIR, 'dso_saldos_actual.json')
                if os.path.exists(_sp):
                    _saldos_raw = json.load(open(_sp, 'r', encoding='utf-8')).get('saldos', [])
                    for s in _saldos_raw:
                        if _norm_dso_match(str(s.get('cliente') or '')) == norm:
                            facturas_raw.append({
                                'cliente':       s.get('cliente'),
                                'fecha_factura': s.get('fecha_factura') or s.get('fechaFactura'),
                                'saldo':         s.get('saldo'),
                            })
            return jsonify({
                "busqueda_original":  nombre_q,
                "busqueda_norm":      norm,
                "encontrado":         val is not None,
                "dso":                val,
                "claves_similares":   [k for k in por_nombre if norm[:6] in k][:10],
                "facturas_en_archivo": facturas_raw,
            })

        if request.args.get('todos') == '1':
            return jsonify({"por_nombre": por_nombre, "fecha_corte": data.get('fecha_corte')})

        # ?portfolio=1 → muestra los nombres del portfolio y si matchean en el archivo DSO
        if request.args.get('portfolio') == '1':
            _saldos_gestion_desde_disco()
            fuente = _saldos_gestion if _saldos_gestion else _saldos_facturas
            # Construir mapa nombre→cuit como lo hace api_director_data
            clientes_raw = {}
            for f in fuente:
                nombre = str(f.get('cliente') or '').strip()
                cuit   = str(f.get('cuit') or '').replace('-','').replace(' ','').strip()
                key    = cuit or nombre
                if key and key not in clientes_raw:
                    clientes_raw[key] = {'nombre': nombre, 'cuit': cuit}
            resultado = []
            for key, c in list(clientes_raw.items())[:60]:
                nom_norm = _norm_dso_match(c['nombre'])
                dso_val  = por_nombre.get(nom_norm)
                resultado.append({
                    'nombre_portfolio': c['nombre'],
                    'nombre_norm':      nom_norm,
                    'cuit':             c['cuit'],
                    'dso_encontrado':   dso_val,
                })
            return jsonify({"clientes": resultado, "total_portfolio": len(clientes_raw)})

        return jsonify({
            "existe": True,
            "fecha_corte":         data.get('fecha_corte'),
            "clientes_por_cuit":   len(por_cuit),
            "clientes_por_nombre": len(por_nombre),
            "muestra_nombre":      dict(list(por_nombre.items())[:10]),
        })
    except Exception as e:
        return jsonify({"existe": True, "error": str(e)}), 500


@app.route("/dso-cheques", methods=["GET"])
def get_dso_cheques():
    try:
        modo = request.args.get('modo', 'actual')
        f_path = os.path.join(DATA_DIR, 'dso_cheques_historico.json' if modo == 'historico' else 'dso_cheques_actual.json')
        if os.path.exists(f_path):
            with open(f_path, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        return jsonify({"cheques": [], "ultima_actualizacion": ""})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/dso-cheques", methods=["POST"])
def save_dso_cheques():
    try:
        body = request.get_json(force=True)
        nuevos = body.get('cheques', [])
        if not nuevos:
            return jsonify({"error": "Sin cheques"}), 400
        from datetime import datetime
        hoy = datetime.now()
        f_actual = os.path.join(DATA_DIR, 'dso_cheques_actual.json')
        with open(f_actual, 'w', encoding='utf-8') as f:
            json.dump({"cheques": nuevos, "ultima_actualizacion": hoy.strftime('%d/%m/%Y %H:%M')}, f, ensure_ascii=False)
        total_pos = sum(c.get('total', 0) for c in nuevos if c.get('total', 0) > 0)
        total_neg = sum(c.get('total', 0) for c in nuevos if c.get('total', 0) < 0)
        print(f"[dso-cheques] {len(nuevos)} cheques: positivos=${total_pos:,.0f} negativos=${total_neg:,.0f}", flush=True)
        return jsonify({"ok": True, "agregados": len(nuevos), "total": len(nuevos)})
    except Exception as e:
        import traceback
        print(f"[dso-cheques] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/dso-ventas", methods=["GET"])
def get_dso_ventas():
    try:
        dso_file = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
        if os.path.exists(dso_file):
            with open(dso_file, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        return jsonify({"meses": {}, "ultima_actualizacion": ""})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/dso-ventas", methods=["POST"])
def save_dso_ventas():
    """Agrega ventas al historial mensual, tanto totales como por cliente.
    Sobreescribe SOLO los meses incluidos en el upload.
    Formato: {meses: {"YYYY-MM": total}, por_cliente: {"CLIENTE": {"YYYY-MM": total}}}"""
    try:
        body = request.get_json(force=True)
        nuevas = body.get('ventas', [])
        if not nuevas:
            return jsonify({"error": "Sin ventas"}), 400
        from datetime import datetime

        # Acumular totales: global por mes Y por cliente×mes
        meses_nuevos: dict = {}       # ym → total
        clientes_nuevos: dict = {}    # cli_norm → {ym → total}
        sin_fecha = 0
        for v in nuevas:
            fv = _parsear_fecha_dso(v.get('fecha'))
            if fv is None:
                sin_fecha += 1
                continue
            ym  = f"{fv.year}-{fv.month:02d}"
            tot = float(v.get('total', 0) or 0)
            meses_nuevos[ym] = meses_nuevos.get(ym, 0.0) + tot
            # Normalizar nombre de cliente igual que el resto del sistema
            cli = _norm_nombre(v.get('cliente') or '')
            if cli:
                clientes_nuevos.setdefault(cli, {})
                clientes_nuevos[cli][ym] = clientes_nuevos[cli].get(ym, 0.0) + tot

        if not meses_nuevos:
            return jsonify({"error": f"Sin ventas con fecha válida ({sin_fecha} filas sin fecha)"}), 400

        # Cargar historico existente
        dso_file = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
        historico_meses: dict = {}
        historico_clientes: dict = {}
        if os.path.exists(dso_file):
            try:
                prev = json.load(open(dso_file, 'r', encoding='utf-8'))
                historico_meses    = prev.get('meses', {})
                historico_clientes = prev.get('por_cliente', {})
            except Exception:
                pass

        # Sobreescribir meses del upload (wipe+replace por mes)
        historico_meses.update(meses_nuevos)
        meses_validos = sorted(historico_meses.keys(), reverse=True)[:6]
        historico_meses = {k: historico_meses[k] for k in meses_validos}

        # Sobreescribir ventas por cliente en los meses del upload
        for cli, meses_cli in clientes_nuevos.items():
            if cli not in historico_clientes:
                historico_clientes[cli] = {}
            historico_clientes[cli].update(meses_cli)
        # Purgar meses obsoletos por cliente también
        for cli in list(historico_clientes.keys()):
            historico_clientes[cli] = {ym: v for ym, v in historico_clientes[cli].items()
                                       if ym in historico_meses}
            if not historico_clientes[cli]:
                del historico_clientes[cli]

        resultado = {
            "meses": historico_meses,
            "por_cliente": historico_clientes,
            "ultima_actualizacion": datetime.now().strftime('%d/%m/%Y %H:%M')
        }
        with open(dso_file, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        log = ' | '.join(f"{k}=${v:,.0f}" for k, v in sorted(meses_nuevos.items()))
        print(f"[dso-ventas] {log} | {len(clientes_nuevos)} clientes | {sin_fecha} sin fecha", flush=True)
        return jsonify({"ok": True, "meses": meses_nuevos, "total_meses": len(historico_meses)})
    except Exception as e:
        import traceback
        print(f"[dso-ventas] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/test-modelos")
def test_modelos():
    if not GEMINI_KEY:
        return jsonify({"error": "Sin API key"}), 500
    resultados = {}
    combos = [
        ("gemini-1.5-flash-001", "v1beta"), ("gemini-1.5-flash-002", "v1beta"),
        ("gemini-1.5-pro-001", "v1beta"), ("gemini-2.0-flash-001", "v1beta"), ("gemini-2.5-flash", "v1beta"),
    ]
    for modelo, version in combos:
        key = f"{modelo}/{version}"
        url = f"https://generativelanguage.googleapis.com/{version}/models/{modelo}:generateContent?key={GEMINI_KEY}"
        try:
            r = requests.post(url, headers={"Content-Type": "application/json"},
                json={"contents": [{"parts": [{"text": "di OK"}]}]}, timeout=25)
            data = r.json()
            if "candidates" in data: resultados[key] = "OK"
            elif "error" in data: resultados[key] = data["error"].get("message", "error")[:100]
            else: resultados[key] = "respuesta inesperada"
        except Exception as e:
            resultados[key] = str(e)[:100]
    return jsonify(resultados)


# ── Saldos / Facturas por cliente ──
def _load_json_with_fallback(filename: str) -> list:
    """Busca filename en DATA_DIR primero, luego en cwd (para Render donde DATA_DIR=/data)."""
    for base in [DATA_DIR, os.getcwd()]:
        path = os.path.join(base, filename)
        if os.path.exists(path):
            try:
                with open(path, encoding='utf-8') as _f:
                    data = json.load(_f)
                print(f"[saldos] {len(data)} registros cargados desde {path}", flush=True)
                return data
            except Exception as e:
                print(f"[saldos] Error leyendo {path}: {e}", flush=True)
    print(f"[saldos] {filename} no encontrado en DATA_DIR ni en cwd", flush=True)
    return []

_saldos_facturas = _load_json_with_fallback('saldos_facturas.json')

# ── Saldos de Gestión (vista semanal para vendedores — separado del DSO) ──
_saldos_gestion_loaded = _load_json_with_fallback('saldos_gestion_vendedores.json')
_saldos_gestion = _saldos_gestion_loaded if _saldos_gestion_loaded else list(_saldos_facturas)
if not _saldos_gestion_loaded:
    print("[gestion] Usando saldos_facturas como fallback inicial para gestión", flush=True)

# Mtime tracking para detectar uploads en otros workers (gunicorn multi-worker)
_SG_FILE       = os.path.join(DATA_DIR, 'saldos_gestion_vendedores.json')
_SG_MTIME: float = 0.0
_SG_LAST_CHECK: float = 0.0
try:
    _SG_MTIME = os.path.getmtime(_SG_FILE) if os.path.exists(_SG_FILE) else 0.0
except Exception:
    pass

# Caché del DSO global ponderado (Σ dso_i×saldo_i / Σ saldo_i) calculado en /api/dso-todos.
# Lo lee /api/director-data para incluirlo en su respuesta sin necesidad de un segundo fetch.
_dso_global_ponderado_cache: int | None = None

# ── Índices en memoria (O(1) lookup por CUIT y nombre) ──────────────────────
_saldos_idx_cuit:   dict = {}   # cuit_limpio   → [facturas]
_saldos_idx_nombre: dict = {}   # norm_nombre   → [facturas]

def _rebuild_saldos_index():
    """Reconstruye ambos índices desde la fuente vigente. Llamar tras cada carga/upload."""
    global _saldos_idx_cuit, _saldos_idx_nombre
    fuente = _saldos_gestion if _saldos_gestion else _saldos_facturas
    idx_c: dict = {}
    idx_n: dict = {}
    for f in fuente:
        if not isinstance(f, dict):
            continue
        c = str(f.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
        if c and len(c) >= 7:
            idx_c.setdefault(c, []).append(f)
        n = _norm_nombre(f.get('cliente', ''))
        if n:
            idx_n.setdefault(n, []).append(f)
    _saldos_idx_cuit   = idx_c
    _saldos_idx_nombre = idx_n
    print(f"[idx] {len(idx_c)} CUITs · {len(idx_n)} nombres indexados "
          f"({len(fuente)} registros)", flush=True)

def _saldos_gestion_desde_disco() -> None:
    """Detecta si saldos_gestion_vendedores.json cambió en disco (upload en otro worker)
    y recarga _saldos_gestion + índices si es así. Debounce de 10s para no estatear
    el filesystem en cada request. Seguro para gunicorn sync-workers (no hay threading)."""
    global _saldos_gestion, _SG_MTIME, _SG_LAST_CHECK
    now = time.time()
    if now - _SG_LAST_CHECK < 10:
        return                     # revisado hace menos de 10s — no restat
    _SG_LAST_CHECK = now
    if not os.path.exists(_SG_FILE):
        return
    try:
        mtime = os.path.getmtime(_SG_FILE)
        if mtime <= _SG_MTIME:
            return                 # archivo no cambió
        with open(_SG_FILE, 'r', encoding='utf-8') as f:
            fresh = json.load(f)
        if not isinstance(fresh, list) or not fresh:
            return
        _saldos_gestion = fresh
        _SG_MTIME = mtime
        _rebuild_saldos_index()
        print(f"[sg-reload] worker detectó upload externo — {len(fresh)} registros recargados", flush=True)
    except Exception as _e:
        print(f"[sg-reload] error ({_e}), manteniendo memoria", flush=True)

def _buscar_por_nombre_en_idx(nombre: str) -> list:
    """Match estricto de 2 niveles — sin fuzzy ni aproximaciones por palabras genéricas.
    Si no hay coincidencia exacta, devuelve [] (saldo $0). Jamás adivina.
    """
    # Nivel 1: coincidencia exacta sobre nombre normalizado completo
    cn = _norm_nombre(nombre)
    r = _saldos_idx_nombre.get(cn)
    if r:
        print(f"[idx] exacto '{nombre}' → {len(r)} facturas", flush=True)
        return r

    # Nivel 2: primeras 2 palabras significativas (sin sufijos SA/SRL/etc.)
    # Solo aplica si esas 2 palabras tienen más de 3 caracteres c/u (evita 'SA', 'SH', etc.)
    cu    = _norm_ultra(nombre)
    words = [w for w in cu.split() if len(w) > 3]
    if len(words) >= 2:
        prim2 = ' '.join(words[:2])
        r = _saldos_idx_nombre.get(prim2)
        if r:
            print(f"[idx] prim2 '{nombre}' → clave '{prim2}' {len(r)} facturas", flush=True)
            return r
        for k, v in _saldos_idx_nombre.items():
            k_words = [w for w in _norm_ultra(k).split() if len(w) > 3]
            if len(k_words) >= 2 and ' '.join(k_words[:2]) == prim2:
                print(f"[idx] prim2-iter '{nombre}' → clave '{k}' {len(v)} facturas", flush=True)
                return v

    # Sin match → saldo $0, sin facturas. No se adivina.
    print(f"[idx] sin match para '{nombre}' — devuelve []", flush=True)
    return []

def _norm_nombre(s):
    import unicodedata, re
    s = str(s or '').strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

_DSO_SUFIJOS = {'SRL','SRLH','SA','SH','SAS','SCA','SCAL','SAC','SAICF','SACI','SE','SC','CIA','CO','AND'}
def _norm_dso_match(s: str) -> str:
    """Normalización robusta para matching de nombres de clientes entre fuentes distintas.
    Elimina sufijos societarios Y tokens de 1 caracter (fragmentos de abreviaturas con espacios).
    'PANESSIDI S.R.L.' == 'PANESSIDI SRL' == 'PANESSIDI S R L' → todos dan 'PANESSIDI'."""
    import unicodedata, re
    s = str(s or '').strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    tokens = [t for t in s.split() if len(t) > 1 and t not in _DSO_SUFIJOS]
    return ' '.join(tokens)


_SUFIJOS_RE = None
def _norm_ultra(s):
    import unicodedata, re
    global _SUFIJOS_RE
    if _SUFIJOS_RE is None:
        _SUFIJOS_RE = re.compile(
            r'\b(S\.?A\.?|S\.?R\.?L\.?|S\.?H\.?|S\.?A\.?S\.?|S\.?C\.?A\.?|SRLH?)\b',
            re.IGNORECASE
        )
    s = str(s or '').strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = _SUFIJOS_RE.sub(' ', s)
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

try:
    _rebuild_saldos_index()
    # Validación de integridad del índice al inicio
    _test = _buscar_por_nombre_en_idx('Tavella')
    _test_saldo = sum(f.get('saldo', 0) for f in _test)
    print(f"[idx] VALIDACION 'Tavella': {len(_test)} registros, saldo=${_test_saldo:,.2f}", flush=True)
except Exception as e:
    print(f"[idx] Error en indexación inicial: {e}", flush=True)

@app.route("/solvencia/<cuit>")
def get_solvencia_endpoint(cuit):
    """Devuelve datos de solvencia cacheados. Si no hay datos, retorna estado 'no disponible'."""
    from urllib.parse import unquote
    cuit_limpio = str(unquote(cuit)).replace('-', '').replace(' ', '').strip()
    data = get_solvency_data(cuit_limpio)
    if data:
        return jsonify({"ok": True, "data": data})
    return jsonify({"ok": False, "mensaje": "Datos de solvencia temporalmente no disponibles"})

@app.route("/saldos-cliente/<cliente>")
def get_saldos_cliente(cliente):
    _saldos_gestion_desde_disco()
    from urllib.parse import unquote
    nombre_original = unquote(cliente)
    cn = _norm_nombre(nombre_original)
    # Lee de _saldos_gestion (gestión semanal) — _saldos_facturas es solo para DSO/auditoría
    fuente = _saldos_gestion if _saldos_gestion else _saldos_facturas

    # 1. Match exacto
    result = [f for f in fuente if _norm_nombre(f.get('cliente', '')) == cn]

    # 2. Match por primeras 2 palabras con normUltra (strips S.A., S.R.L., etc.)
    if not result:
        cu = _norm_ultra(nombre_original)
        prim2 = ' '.join(cu.split()[:2])
        if len(prim2) > 3:
            result = [f for f in fuente
                      if ' '.join(_norm_ultra(f.get('cliente', '')).split()[:2]) == prim2]
            if result:
                print(f"[match-2p] '{nombre_original}' → prim2='{prim2}' → {len(result)} facturas", flush=True)

    # 3. Match parcial (≥2 palabras significativas en común, longitud >3)
    # Excluye sufijos societarios: "SRL" y "SA" no son palabras distintivas —
    # sin esta exclusión "WINE BAR SRL" matcheaba "VINOTECAS ROMA WINE SRL" (WINE+SRL=2).
    if not result:
        _SUFIJOS_PARCIAL = {'SA', 'SRL', 'SRLH', 'SH', 'SAS', 'SCA', 'SE', 'SC', 'CIA', 'CO', 'AND'}
        palabras = [w for w in cn.split() if len(w) > 3 and w not in _SUFIJOS_PARCIAL]
        if len(palabras) >= 2:
            result = [f for f in fuente
                if sum(1 for p in palabras if p in _norm_nombre(f.get('cliente', '')))
                   >= min(2, len(palabras))]
            if result:
                print(f"[match-parcial] '{nombre_original}' → {len(result)} facturas", flush=True)

    # Audit log: sin match → registrar el string exacto de Odoo para debugging
    if not result:
        clientes_en_sf = list({_norm_nombre(f.get('cliente', '')) for f in fuente})[:5]
        print(f"[match-FAIL] No se encontró match para: '{nombre_original}' (normalizado: '{cn}'). "
              f"Primeros 5 clientes en saldos_facturas: {clientes_en_sf}", flush=True)

    total_saldo = sum(f.get('saldo', 0) for f in result)
    return jsonify({"facturas": result, "total_saldo": total_saldo, "cantidad": len(result)})

def _calc_dso_aging_mensual(cuit_limpio: str, nombre: str = '') -> int | None:
    """DSO aging desde dso_saldos_actual.json (reporte mensual completo).
    Matching primario: CUIT exacto. Fallback: nombre ultra-normalizado (_norm_ultra).
    El reporte semanal (_saldos_gestion) puede omitir facturas viejas; este usa el historial completo.
    Usa Σ(saldo × días_desde_fecha_factura) / Σ(saldo). Retorna None si no hay datos."""
    from datetime import date
    _path = os.path.join(DATA_DIR, 'dso_saldos_actual.json')
    if not os.path.exists(_path):
        return None
    try:
        _saldos = json.load(open(_path, 'r', encoding='utf-8')).get('saldos', [])
    except Exception:
        return None
    hoy = date.today()
    cuit_n  = cuit_limpio.replace('-', '').replace(' ', '').strip()
    nom_key = _norm_dso_match(nombre) if nombre else ''

    # Acumula por método de matching: CUIT primero, nombre como fallback
    acc_cuit = {'sp': 0.0, 'ss': 0.0}
    acc_nom  = {'sp': 0.0, 'ss': 0.0}
    for f in _saldos:
        saldo = float(f.get('saldo') or 0)
        if saldo <= 0:
            continue
        ff_str = str(f.get('fecha_factura') or f.get('fechaFactura') or '').strip()
        dias = 0
        try:
            if '/' in ff_str:
                p = ff_str.split('/')
                ff = date(int(p[2]), int(p[1]), int(p[0]))
                dias = max(0, (hoy - ff).days)
            elif '-' in ff_str and len(ff_str) >= 10:
                p = ff_str.split('-')
                ff = date(int(p[0]), int(p[1]), int(p[2]))
                dias = max(0, (hoy - ff).days)
        except Exception:
            dias = 0
        f_cuit = str(f.get('cuit', '')).replace('-', '').replace(' ', '').strip()
        if cuit_n and f_cuit == cuit_n:
            acc_cuit['sp'] += saldo * dias
            acc_cuit['ss'] += saldo
        elif nom_key and _norm_dso_match(str(f.get('cliente') or '')) == nom_key:
            acc_nom['sp'] += saldo * dias
            acc_nom['ss'] += saldo

    acc = acc_cuit if acc_cuit['ss'] > 0 else acc_nom
    if acc['ss'] <= 0:
        return None
    return round(acc['sp'] / acc['ss'])


_dso_individual_cache: dict = {}

def _get_dso_individual(cuit: str, nombre: str = '') -> int | None:
    """Lookup estático desde dso_individual_actual.json (generado en upload mensual DSO).
    El valor queda congelado al momento del upload y no varía día a día.
    Matching: CUIT exacto primero, luego nombre normalizado con _norm_dso_match."""
    global _dso_individual_cache
    _path = os.path.join(DATA_DIR, 'dso_individual_actual.json')
    if not os.path.exists(_path):
        return None
    try:
        mtime = os.path.getmtime(_path)
        if _dso_individual_cache.get('_mtime') != mtime:
            data = json.load(open(_path, 'r', encoding='utf-8'))
            _dso_individual_cache = {**data, '_mtime': mtime}
    except Exception:
        return None
    cuit_n = cuit.replace('-', '').replace(' ', '').strip()
    if cuit_n:
        val = _dso_individual_cache.get('por_cuit', {}).get(cuit_n)
        if val is not None:
            return int(val)
    if nombre:
        nom_key = _norm_dso_match(nombre)
        val = _dso_individual_cache.get('por_nombre', {}).get(nom_key)
        if val is not None:
            return int(val)
    return None


def _calc_dso_aging(facturas: list) -> int | None:
    """DSO por antigüedad: Σ(saldo × días_desde_fechaFactura) / Σ(saldo).
    Acepta fechaFactura en DD/MM/YYYY o YYYY-MM-DD (ISO).  Retorna None si saldo total = 0."""
    from datetime import date
    hoy = date.today()
    suma_pond, suma_saldo = 0.0, 0.0
    for f in facturas:
        saldo = float(f.get('saldo') or 0)
        if saldo <= 0:
            continue
        ff_str = str(f.get('fechaFactura') or '').strip()
        dias = 0
        try:
            if '/' in ff_str:                   # DD/MM/YYYY
                p = ff_str.split('/')
                ff = date(int(p[2]), int(p[1]), int(p[0]))
                dias = max(0, (hoy - ff).days)
            elif '-' in ff_str and len(ff_str) >= 10:  # YYYY-MM-DD (ISO)
                p = ff_str.split('-')
                ff = date(int(p[0]), int(p[1]), int(p[2]))
                dias = max(0, (hoy - ff).days)
        except Exception:
            dias = 0
        suma_pond  += saldo * dias
        suma_saldo += saldo
    if suma_saldo <= 0:
        return None
    return round(suma_pond / suma_saldo)


@app.route("/saldos-cuit/<cuit>")
def get_saldos_cuit(cuit):
    """Busca facturas por CUIT (prioridad absoluta). Si no hay CUIT en registros, cae a nombre + fuzzy."""
    _saldos_gestion_desde_disco()
    from urllib.parse import unquote
    cuit_limpio = str(unquote(cuit)).replace('-', '').replace(' ', '').strip()
    fuente_g = _saldos_gestion if _saldos_gestion else _saldos_facturas
    # Prioridad 1: CUIT exacto (limpiado de guiones y espacios)
    result = [f for f in fuente_g
              if str(f.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio]
    if result:
        total_saldo = sum(f.get('saldo', 0) for f in result)
        nombre_m = result[0].get('cliente', '')
        enriched, monto_v30, alerta30 = _enrich_con_mora(result)
        dso_aging = _get_dso_individual(cuit_limpio, nombre_m) or _calc_dso_aging(enriched)
        print(f"[saldos-cuit] CUIT {cuit_limpio}: {len(enriched)} facturas, vencido30=${monto_v30:,.0f}, dso_aging={dso_aging}", flush=True)
        return jsonify({"facturas": enriched, "total_saldo": total_saldo, "cantidad": len(enriched),
                        "monto_pendiente_vencido": monto_v30, "alerta_mora_30": alerta30,
                        "dso_aging": dso_aging,
                        "metodo": "cuit", "nombre_match": nombre_m})
    # Prioridad 2: nombre canónico desde cartera_comercial → exact → fuzzy
    nombre_en_cartera = next(
        (str(c.get('nombre', '')).strip() for c in _cartera_comercial
         if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio),
        None
    )
    if nombre_en_cartera:
        cn = _norm_nombre(nombre_en_cartera)
        result = [f for f in fuente_g if _norm_nombre(f.get('cliente', '')) == cn]
        if not result:
            # Fuzzy: match ≥2 palabras significativas
            palabras = [p for p in cn.split() if len(p) > 2]
            if palabras:
                result = [f for f in fuente_g
                          if sum(1 for p in palabras
                                 if p in _norm_nombre(f.get('cliente', ''))) >= min(2, len(palabras))]
                if result:
                    print(f"[saldos-cuit] Fuzzy '{nombre_en_cartera}' → {len(result)} facturas", flush=True)
        total_saldo = sum(f.get('saldo', 0) for f in result)
        enriched, monto_v30, alerta30 = _enrich_con_mora(result)
        dso_aging = _get_dso_individual(cuit_limpio, nombre_en_cartera) or _calc_dso_aging(enriched)
        print(f"[saldos-cuit] Nombre '{nombre_en_cartera}': {len(enriched)} facturas, vencido30=${monto_v30:,.0f}, dso_aging={dso_aging}", flush=True)
        return jsonify({"facturas": enriched, "total_saldo": total_saldo, "cantidad": len(enriched),
                        "monto_pendiente_vencido": monto_v30, "alerta_mora_30": alerta30,
                        "dso_aging": dso_aging,
                        "metodo": "nombre", "nombre_match": nombre_en_cartera})
    print(f"[saldos-cuit] CUIT {cuit_limpio}: sin match en cartera_comercial", flush=True)
    return jsonify({"facturas": [], "total_saldo": 0, "cantidad": 0,
                    "monto_pendiente_vencido": 0, "alerta_mora_30": False, "metodo": "nulo"})


@app.route("/api/cartera-saldos")
@require_login
def api_cartera_saldos():
    """Portfolio de saldos — mismos datos que /api/director-data, accesible a todos los usuarios logueados."""
    return api_director_data.__wrapped__()


@app.route("/api/debug-saldos")
@require_login
def api_debug_saldos():
    """Diagnóstico del estado actual de _saldos_gestion para detectar problemas de columnas."""
    _saldos_gestion_desde_disco()
    fuente = _saldos_gestion if _saldos_gestion else _saldos_facturas
    clientes = list(set(s.get('cliente', '') for s in fuente))
    vendedores = list(set(s.get('vendedor', '') for s in fuente))
    sg_path = os.path.join(DATA_DIR, 'saldos_gestion_vendedores.json')
    sf_path = os.path.join(DATA_DIR, 'saldos_facturas.json')
    return jsonify({
        'total_registros': len(fuente),
        'clientes_unicos': len(clientes),
        'vendedores_unicos': len(vendedores),
        'primeros_3': fuente[:3],
        'usando_gestion': bool(_saldos_gestion),
        'sg_en_disco': os.path.exists(sg_path),
        'sf_en_disco': os.path.exists(sf_path),
        'data_dir': DATA_DIR,
        'muestra_clientes': sorted(clientes)[:10],
        'muestra_vendedores': sorted(vendedores)[:10],
    })


@app.route("/api/facturas/<cuit>")
def api_facturas_por_cuit(cuit):
    """
    Consulta de facturas: CUIT → nombre en cartera → nombre en query string → fuzzy.
    Todos los registros de saldos_facturas tienen cuit='', por lo que el flujo
    normal es siempre por nombre. El CUIT se usa como llave para encontrar el
    nombre canónico en cartera_comercial.
    """
    _saldos_gestion_desde_disco()   # sincroniza si otro worker subió datos
    from urllib.parse import unquote
    cuit_limpio = str(unquote(cuit)).replace('-', '').replace(' ', '').strip()
    nombre_hint = request.args.get('nombre', '').strip()

    # 1. Por CUIT (aplica cuando los registros de gestión incluyen campo cuit)
    result = _saldos_idx_cuit.get(cuit_limpio, [])
    if result:
        total = sum(f.get('saldo', 0) for f in result)
        enriched, monto_v30, alerta30 = _enrich_con_mora(result)
        enriched = _fac_anotar_estado(enriched, cuit_limpio)
        print(f"[facturas] CUIT {cuit_limpio}: {len(enriched)} facturas ${total:,.0f} (método: cuit)", flush=True)
        return jsonify({"facturas": enriched, "total_saldo": total, "cantidad": len(enriched),
                        "monto_pendiente_vencido": monto_v30, "alerta_mora_30": alerta30, "metodo": "cuit"})

    # 2. Nombre canónico desde cartera_comercial (garantiza string exacto del archivo fuente)
    nombre_cartera = next(
        (str(c.get('nombre', '')).strip() for c in _cartera_comercial
         if str(c.get('cuit', '')).replace('-', '').replace(' ', '').strip() == cuit_limpio),
        None
    )
    # 3. Fallback al hint enviado por el frontend (?nombre=)
    nombre = nombre_cartera or nombre_hint

    if nombre:
        result = _buscar_por_nombre_en_idx(nombre)
        if result:
            total = sum(f.get('saldo', 0) for f in result)
            enriched, monto_v30, alerta30 = _enrich_con_mora(result)
            enriched = _fac_anotar_estado(enriched, cuit_limpio)
            metodo = "nombre_cartera" if nombre_cartera else "nombre_hint"
            print(f"[facturas] '{nombre}': {len(enriched)} facturas ${total:,.0f} (método: {metodo})", flush=True)
            return jsonify({"facturas": enriched, "total_saldo": total, "cantidad": len(enriched),
                            "monto_pendiente_vencido": monto_v30, "alerta_mora_30": alerta30, "metodo": metodo})

    print(f"[facturas] CUIT {cuit_limpio} nombre='{nombre}': sin resultados "
          f"(idx_cuit={len(_saldos_idx_cuit)} entradas, idx_nombre={len(_saldos_idx_nombre)} entradas)", flush=True)
    return jsonify({"facturas": [], "total_saldo": 0, "cantidad": 0,
                    "monto_pendiente_vencido": 0, "alerta_mora_30": False, "metodo": "nulo"})


@app.route("/api/facturas/<cuit>/marcar-cobrada", methods=['POST'])
def marcar_factura_cobrada(cuit):
    """Vendedor marca una factura como 'pendiente_validacion' (no reversible por el vendedor)."""
    data = request.get_json(force=True, silent=True) or {}
    nro = str(data.get('nroFactura', '')).strip()
    if not nro:
        return jsonify({'ok': False, 'error': 'nroFactura requerido'}), 400
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    key = _fac_key(cuit_limpio, nro)
    with _facturas_estado_lock:
        estado = _fac_estado_load()
        prev = estado.get(key, {})
        estado[key] = {
            'estado': 'pendiente_validacion',
            'ts': time.time(),
            'cuit': cuit_limpio,
            'nroFactura': nro,
            'enviado_whatsapp': prev.get('enviado_whatsapp', False),
        }
        _fac_estado_save(estado)
    print(f"[fac_cobrada] {cuit_limpio} · {nro} → pendiente_validacion", flush=True)
    return jsonify({'ok': True})


@app.route("/api/facturas/<cuit>/desmarcar-cobrada", methods=['POST'])
def desmarcar_factura_cobrada(cuit):
    """Vendedor revierte un cobro marcado por error (mientras no haya validación de supervisor)."""
    data = request.get_json(force=True, silent=True) or {}
    nro = str(data.get('nroFactura', '')).strip()
    if not nro:
        return jsonify({'ok': False, 'error': 'nroFactura requerido'}), 400
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    key = _fac_key(cuit_limpio, nro)
    with _facturas_estado_lock:
        estado = _fac_estado_load()
        if key in estado:
            prev = estado[key]
            # Conservar enviado_whatsapp; limpiar el estado de cobro
            if prev.get('enviado_whatsapp'):
                estado[key] = {
                    'estado': '',
                    'ts': time.time(),
                    'cuit': cuit_limpio,
                    'nroFactura': nro,
                    'enviado_whatsapp': True,
                    'ts_whatsapp': prev.get('ts_whatsapp'),
                }
            else:
                del estado[key]
        _fac_estado_save(estado)
    print(f"[fac_desmarcar] {cuit_limpio} · {nro} → cobro revertido", flush=True)
    return jsonify({'ok': True})


@app.route("/api/admin/desmarcar-por-nro", methods=['POST'])
def admin_desmarcar_por_nro():
    """Admin: revierte el cobro de una factura buscando solo por nroFactura (sin CUIT)."""
    data = request.get_json(force=True, silent=True) or {}
    nro = str(data.get('nroFactura', '')).strip()
    if not nro:
        return jsonify({'ok': False, 'error': 'nroFactura requerido'}), 400
    with _facturas_estado_lock:
        estado = _fac_estado_load()
        borradas = []
        for key in list(estado.keys()):
            entry = estado[key]
            if str(entry.get('nroFactura', '')) == nro and entry.get('estado') == 'pendiente_validacion':
                if entry.get('enviado_whatsapp'):
                    estado[key] = {k: v for k, v in entry.items() if k != 'estado'}
                    estado[key]['estado'] = ''
                else:
                    del estado[key]
                borradas.append(key)
        _fac_estado_save(estado)
    print(f"[admin_desmarcar] nro={nro} → {len(borradas)} entrada(s) revertidas: {borradas}", flush=True)
    return jsonify({'ok': True, 'revertidas': borradas})


@app.route("/api/facturas/<cuit>/marcar-whatsapp", methods=['POST'])
def marcar_factura_whatsapp(cuit):
    """Vendedor marca una factura como enviada por WhatsApp."""
    data = request.get_json(force=True, silent=True) or {}
    nro = str(data.get('nroFactura', '')).strip()
    if not nro:
        return jsonify({'ok': False, 'error': 'nroFactura requerido'}), 400
    cuit_limpio = str(cuit).replace('-', '').replace(' ', '').strip()
    key = _fac_key(cuit_limpio, nro)
    with _facturas_estado_lock:
        estado = _fac_estado_load()
        if key not in estado:
            estado[key] = {'cuit': cuit_limpio, 'nroFactura': nro}
        estado[key]['enviado_whatsapp'] = True
        estado[key]['ts_whatsapp'] = time.time()
        _fac_estado_save(estado)
    print(f"[fac_wa] {cuit_limpio} · {nro} → enviado_whatsapp", flush=True)
    return jsonify({'ok': True})


@app.route("/api/facturas-estado-resumen")
def facturas_estado_resumen():
    """Resumen: qué CUITs tienen facturas con estado activo (para filtros de cartera)."""
    estado = _fac_estado_load()
    cuits_cobradas, cuits_wa = set(), set()
    for e in estado.values():
        c = e.get('cuit', '')
        if not c:
            continue
        if e.get('estado') == 'pendiente_validacion':
            cuits_cobradas.add(c)
        if e.get('enviado_whatsapp'):
            cuits_wa.add(c)
    return jsonify({'cobradas': list(cuits_cobradas), 'wa_enviadas': list(cuits_wa)})


# ══════════════════════════════════════════════════════════════════════════════
# FACTURAS PDF — ZIP importado desde Odoo
# Flujo: comercial sube el ZIP mensual → se guarda en disco + R2 (backup) →
# endpoint /api/facturas-pdf/<nombre> extrae y sirve el PDF al vuelo →
# link se incluye en el mensaje de WhatsApp de cuenta corriente.
# ══════════════════════════════════════════════════════════════════════════════

_FACTURAS_ZIP_LOCAL   = os.path.join(DATA_DIR, 'facturas_odoo.zip')
_FACTURAS_ZIP_R2_KEY  = 'facturas_odoo.zip'
_FACTURAS_ZIP_META    = os.path.join(DATA_DIR, 'facturas_zip_meta.json')
_FACTURAS_CFG         = os.path.join(DATA_DIR, 'facturas_config.json')
_FACTURAS_CFG_R2_KEY  = 'facturas_config.json'
_FACTURAS_ZIP_LOCK    = threading.Lock()
_FACTURAS_IMPORT_LOCK = threading.Lock()   # evita imports simultáneos

# Estado del último import automático (en memoria, para polling del frontend)
_facturas_import_estado: dict = {
    'corriendo': False,
    'ultimo_paso': '',
    'ultimo_resultado': '',  # 'ok' | 'error' | ''
    'ultima_fecha': '',
}


def _facturas_cfg_read() -> dict:
    """Lee la config de facturas PDF (drive_url, etc.). Intenta disco → R2."""
    try:
        with open(_FACTURAS_CFG, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        pass
    # Fallback: intentar recuperar desde R2
    data = _r2_download_bytes(_FACTURAS_CFG_R2_KEY)
    if data:
        try:
            cfg = json.loads(data.decode('utf-8'))
            with open(_FACTURAS_CFG, 'w', encoding='utf-8') as f:
                f.write(data.decode('utf-8'))
            return cfg
        except Exception:
            pass
    return {}


def _facturas_cfg_write(cfg: dict):
    with open(_FACTURAS_CFG, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False)
    # Backup en R2 en background
    def _bg():
        _r2_upload_bytes(_FACTURAS_CFG_R2_KEY, json.dumps(cfg).encode(), 'application/json')
    threading.Thread(target=_bg, daemon=True).start()


def _facturas_auto_import() -> bool:
    """Descarga el ZIP desde la URL de Drive configurada y hace el merge.
    Retorna True si importó OK, False si no hay URL o falló."""
    if _FACTURAS_IMPORT_LOCK.locked():
        print("[facturas-auto] Ya hay un import en curso, saltando.", flush=True)
        return False
    with _FACTURAS_IMPORT_LOCK:
        cfg = _facturas_cfg_read()
        url = (cfg.get('drive_url') or '').strip()
        if not url:
            print("[facturas-auto] Sin URL de Drive configurada — saltando import.", flush=True)
            return False
        _facturas_import_estado['corriendo']   = True
        _facturas_import_estado['ultimo_paso'] = 'Descargando ZIP desde Drive...'
        _facturas_import_estado['ultimo_resultado'] = ''
        try:
            file_id = _extraer_gdrive_id(url)
            if not file_id:
                raise ValueError("URL de Drive inválida")

            import tempfile
            tmp = tempfile.mktemp(suffix='.zip')
            try:
                import gdown
                gdown.download(f"https://drive.google.com/uc?id={file_id}", tmp, quiet=True)
            except ImportError:
                dl_url = f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t"
                r = requests.get(dl_url, timeout=300, stream=True)
                r.raise_for_status()
                with open(tmp, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=1_048_576):
                        f.write(chunk)

            if not os.path.exists(tmp) or os.path.getsize(tmp) < 1000:
                raise ValueError("Descarga vacía — verificá que el link de Drive sea público")

            _facturas_import_estado['ultimo_paso'] = 'Procesando ZIP...'
            with open(tmp, 'rb') as f:
                raw = f.read()
            try:
                os.remove(tmp)
            except Exception:
                pass

            meta = _procesar_zip_facturas(raw)

            # Actualizar config con fecha del último import automático
            cfg['ultimo_import_auto'] = time.strftime('%Y-%m-%d %H:%M')
            cfg['ultimo_import_total'] = meta['total_pdfs']
            _facturas_cfg_write(cfg)

            _facturas_import_estado['ultimo_resultado'] = 'ok'
            _facturas_import_estado['ultima_fecha']     = time.strftime('%Y-%m-%d %H:%M')
            _facturas_import_estado['ultimo_paso']      = f"OK — {meta['total_pdfs']} PDFs totales (+{meta.get('nuevos_este_import',0)} nuevos)"
            print(f"[facturas-auto] Import automático OK — {meta['total_pdfs']} PDFs", flush=True)
            return True
        except Exception as e:
            _facturas_import_estado['ultimo_resultado'] = 'error'
            _facturas_import_estado['ultimo_paso']      = f"Error: {e}"
            print(f"[facturas-auto] Error: {e}", flush=True)
            return False
        finally:
            _facturas_import_estado['corriendo'] = False


def _facturas_auto_loop():
    """Corre el import automático cada lunes a las 8 AM (hora del servidor, UTC-3 AR)."""
    import datetime as _dt
    time.sleep(120)  # esperar 2 min al arrancar antes del primer check
    while True:
        try:
            ahora = _dt.datetime.utcnow() - _dt.timedelta(hours=3)  # hora Argentina
            # Lunes=0, hora 8 AM
            es_lunes = (ahora.weekday() == 0)
            hora_ok  = (8 <= ahora.hour < 9)

            if es_lunes and hora_ok:
                cfg = _facturas_cfg_read()
                ultima = cfg.get('ultimo_import_auto', '')
                hoy    = ahora.strftime('%Y-%m-%d')
                if not ultima.startswith(hoy):  # no importar dos veces el mismo día
                    print(f"[facturas-auto] Lunes {hoy} — disparando import automático", flush=True)
                    _facturas_auto_import()
                else:
                    print(f"[facturas-auto] Ya importado hoy ({hoy}), saltando.", flush=True)
        except Exception as e:
            print(f"[facturas-auto] Error en loop: {e}", flush=True)
        time.sleep(3600)  # revisar cada hora


def _facturas_zip_meta_read() -> dict:
    try:
        with open(_FACTURAS_ZIP_META, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _r2_upload_bytes(key: str, data: bytes, content_type: str = 'application/octet-stream') -> bool:
    """Sube bytes a R2. Retorna True si OK."""
    if not _R2_CONFIGURADO:
        return False
    try:
        import boto3
        from botocore.config import Config
        import io
        s3 = boto3.client(
            service_name='s3',
            endpoint_url=R2_ENDPOINT_URL,
            aws_access_key_id=R2_ACCESS_KEY_ID,
            aws_secret_access_key=R2_SECRET_ACCESS_KEY,
            config=Config(signature_version='s3v4'),
        )
        s3.upload_fileobj(
            io.BytesIO(data), R2_BUCKET_NAME, key,
            ExtraArgs={'ContentType': content_type},
        )
        return True
    except Exception as e:
        print(f"[r2-upload] Error subiendo {key}: {e}", flush=True)
        return False


def _r2_download_bytes(key: str) -> bytes | None:
    """Descarga un objeto de R2. Retorna bytes o None si no existe/falla."""
    if not _R2_CONFIGURADO:
        return None
    try:
        import boto3
        import io
        from botocore.config import Config
        s3 = boto3.client(
            service_name='s3',
            endpoint_url=R2_ENDPOINT_URL,
            aws_access_key_id=R2_ACCESS_KEY_ID,
            aws_secret_access_key=R2_SECRET_ACCESS_KEY,
            config=Config(signature_version='s3v4'),
        )
        buf = io.BytesIO()
        s3.download_fileobj(R2_BUCKET_NAME, key, buf)
        return buf.getvalue()
    except Exception as e:
        print(f"[r2-download] Error descargando {key}: {e}", flush=True)
        return None


def _facturas_zip_ensure_local() -> bool:
    """Garantiza que el ZIP esté en disco. Si no está, lo descarga desde R2."""
    if os.path.exists(_FACTURAS_ZIP_LOCAL):
        return True
    print("[facturas-zip] ZIP no encontrado en disco — intentando recuperar desde R2", flush=True)
    data = _r2_download_bytes(_FACTURAS_ZIP_R2_KEY)
    if data:
        with _FACTURAS_ZIP_LOCK:
            with open(_FACTURAS_ZIP_LOCAL, 'wb') as f:
                f.write(data)
        print(f"[facturas-zip] ZIP recuperado desde R2 ({len(data)//1024} KB)", flush=True)
        return True
    return False


def _procesar_zip_facturas(raw: bytes) -> dict:
    """Merge acumulativo: agrega los PDFs del nuevo ZIP al archivo maestro existente.
    Los PDFs de importaciones anteriores se conservan — nunca se borran.
    Si un PDF ya existe y viene en el nuevo ZIP, se actualiza (corrección de factura).
    El ZIP maestro resultante se guarda en disco y en R2.
    """
    import zipfile

    # Leer PDFs del nuevo ZIP
    nuevos: dict[str, bytes] = {}
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        for nombre in zf.namelist():
            if nombre.lower().endswith('.pdf'):
                nuevos[nombre] = zf.read(nombre)

    if not nuevos:
        raise ValueError("El ZIP no contiene archivos PDF")

    with _FACTURAS_ZIP_LOCK:
        # Cargar PDFs existentes del ZIP maestro en memoria
        existentes: dict[str, bytes] = {}
        if os.path.exists(_FACTURAS_ZIP_LOCAL):
            try:
                with zipfile.ZipFile(_FACTURAS_ZIP_LOCAL, 'r') as zf_old:
                    for nombre in zf_old.namelist():
                        if nombre.lower().endswith('.pdf'):
                            existentes[nombre] = zf_old.read(nombre)
                print(f"[facturas-zip] ZIP maestro existente: {len(existentes)} PDFs", flush=True)
            except Exception as e:
                print(f"[facturas-zip] ZIP maestro corrupto, descartando: {e}", flush=True)
                existentes = {}

        # Merge: existentes + nuevos (nuevos sobreescriben si mismo nombre)
        merged = {**existentes, **nuevos}
        nuevos_count = len([n for n in nuevos if n not in existentes])
        print(f"[facturas-zip] Merge: {len(existentes)} existentes + {len(nuevos)} nuevos "
              f"({nuevos_count} agregados) = {len(merged)} total", flush=True)

        # Escribir ZIP maestro
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zf_out:
            for nombre, data in merged.items():
                zf_out.writestr(nombre, data)
        merged_bytes = buf.getvalue()

        with open(_FACTURAS_ZIP_LOCAL, 'wb') as f:
            f.write(merged_bytes)

    # Subir ZIP maestro a R2 en background
    def _bg_upload():
        ok = _r2_upload_bytes(_FACTURAS_ZIP_R2_KEY, merged_bytes, 'application/zip')
        print(f"[facturas-zip] R2 upload {'OK' if ok else 'FALLÓ'} ({len(merged_bytes)//1024} KB)", flush=True)
    threading.Thread(target=_bg_upload, daemon=True).start()

    # Meta acumulativo — guardar solo el basename (sin subcarpeta del ZIP)
    nombres_sin_ext = [
        n.split('/')[-1].replace('.pdf', '').replace('.PDF', '')
        for n in merged
    ]
    meta = {
        'fecha_importacion':  time.strftime('%Y-%m-%d %H:%M'),
        'total_pdfs':         len(merged),
        'nuevos_este_import': nuevos_count,
        'nombres':            nombres_sin_ext,
    }
    with open(_FACTURAS_ZIP_META, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False)

    print(f"[facturas-zip] ZIP maestro actualizado — {len(merged)} PDFs totales", flush=True)
    return meta


@app.route("/api/facturas/importar-zip", methods=["POST"])
@require_login
def importar_facturas_zip():
    """Recibe el ZIP mensual de facturas Odoo subido directamente (multipart)."""
    if 'zip' not in request.files:
        return jsonify({"ok": False, "error": "Campo 'zip' requerido"}), 400
    archivo = request.files['zip']
    if not archivo.filename.lower().endswith('.zip'):
        return jsonify({"ok": False, "error": "El archivo debe ser un ZIP"}), 400
    try:
        meta = _procesar_zip_facturas(archivo.read())
        return jsonify({"ok": True, "total_pdfs": meta['total_pdfs'], "fecha": meta['fecha_importacion']})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/facturas/importar-desde-drive", methods=["POST"])
def importar_facturas_desde_drive():
    """Descarga el ZIP de facturas Odoo directamente desde Google Drive.
    El servidor hace la descarga — evita límites de upload del navegador.
    Payload: {"url": "https://drive.google.com/file/d/..."}
    """
    body = request.get_json(force=True) or {}
    url  = (body.get('url') or '').strip()
    if not url:
        return jsonify({"ok": False, "error": "URL requerida"}), 400

    try:
        import tempfile
        print(f"[facturas-zip] Descargando desde Drive: {url[:80]}", flush=True)

        file_id = _extraer_gdrive_id(url)
        if not file_id:
            return jsonify({"ok": False, "error": "URL de Google Drive inválida. Asegurate de compartir con 'Cualquier persona con el link'."}), 400

        # Descargar con gdown (maneja la confirmación de archivos grandes de Drive)
        tmp = tempfile.mktemp(suffix='.zip')
        try:
            import gdown
            gdrive_dl = f"https://drive.google.com/uc?id={file_id}"
            gdown.download(gdrive_dl, tmp, quiet=False)
        except ImportError:
            # Fallback: descarga directa si gdown no está instalado
            dl_url = f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t"
            r = requests.get(dl_url, timeout=300, stream=True)
            r.raise_for_status()
            with open(tmp, 'wb') as f:
                for chunk in r.iter_content(chunk_size=1_048_576):
                    f.write(chunk)

        if not os.path.exists(tmp) or os.path.getsize(tmp) < 1000:
            return jsonify({"ok": False, "error": "La descarga falló o el archivo está vacío. Verificá que el link de Drive sea público."}), 400

        with open(tmp, 'rb') as f:
            raw = f.read()
        try:
            os.remove(tmp)
        except Exception:
            pass

        meta = _procesar_zip_facturas(raw)
        return jsonify({"ok": True, "total_pdfs": meta['total_pdfs'], "fecha": meta['fecha_importacion']})
    except Exception as e:
        print(f"[facturas-zip] Error Drive import: {e}", flush=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/facturas/zip-meta")
def facturas_zip_meta_endpoint():
    """Metadatos del ZIP importado: fecha, total PDFs, config Drive y estado de import."""
    meta  = _facturas_zip_meta_read()
    cfg   = _facturas_cfg_read()
    existe = os.path.exists(_FACTURAS_ZIP_LOCAL)
    return jsonify({
        **meta,
        "zip_disponible":      existe or bool(meta),
        "drive_url_ok":        bool(cfg.get('drive_url')),
        "ultimo_import_auto":  cfg.get('ultimo_import_auto', ''),
        "import_estado":       _facturas_import_estado,
    })


@app.route("/api/facturas/configurar-drive", methods=["POST"])
def facturas_configurar_drive():
    """Guarda la URL de Google Drive del ZIP de facturas. Solo admin/supervisor."""
    body = request.get_json(force=True) or {}
    url  = (body.get('drive_url') or '').strip()
    if not url:
        return jsonify({"ok": False, "error": "URL requerida"}), 400
    if not _extraer_gdrive_id(url):
        return jsonify({"ok": False, "error": "URL de Google Drive inválida"}), 400
    cfg = _facturas_cfg_read()
    cfg['drive_url'] = url
    _facturas_cfg_write(cfg)
    return jsonify({"ok": True})


@app.route("/api/facturas/reimportar", methods=["POST"])
def facturas_reimportar():
    """Dispara el import manual desde la URL de Drive configurada (background)."""
    if _facturas_import_estado.get('corriendo'):
        return jsonify({"ok": False, "error": "Ya hay un import en curso"}), 202
    cfg = _facturas_cfg_read()
    if not cfg.get('drive_url'):
        return jsonify({"ok": False, "error": "URL de Drive no configurada"}), 400
    threading.Thread(target=_facturas_auto_import, daemon=True).start()
    return jsonify({"ok": True, "mensaje": "Import iniciado en background"})


@app.route("/api/facturas/import-estado")
def facturas_import_estado():
    """Polling del estado del import en curso."""
    return jsonify(_facturas_import_estado)


@app.route("/api/facturas-pdf/<path:nombre_pdf>")
def servir_factura_pdf(nombre_pdf):
    """Extrae y sirve un PDF específico del ZIP importado.
    nombre_pdf: nombre sin extensión o con .pdf (ej: 'FA-A 00016-00007069' o 'FA-A 00016-00007069.pdf')
    """
    import zipfile
    from flask import send_file

    # Normalizar nombre
    nombre_limpio = nombre_pdf.strip()
    if not nombre_limpio.lower().endswith('.pdf'):
        nombre_limpio += '.pdf'

    if not _facturas_zip_ensure_local():
        return jsonify({"error": "ZIP de facturas no disponible. Por favor importalo desde la app."}), 404
    try:
        with _FACTURAS_ZIP_LOCK:
            with zipfile.ZipFile(_FACTURAS_ZIP_LOCAL, 'r') as zf:
                nombres_zip = zf.namelist()
                # Buscar match exacto o case-insensitive
                match = next(
                    (n for n in nombres_zip if n.lower() == nombre_limpio.lower()),
                    None,
                )
                if not match:
                    return jsonify({"error": f"PDF '{nombre_limpio}' no encontrado en el ZIP"}), 404
                pdf_bytes = zf.read(match)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=False,
            download_name=nombre_limpio,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ESTADOS DE CUENTA COMPARTIBLES — link único para WhatsApp
# Flujo: el comercial selecciona facturas → POST /api/facturas/crear-lote genera
# un token corto → el mensaje de WhatsApp lleva UN solo link /f/<token> →
# el cliente abre una página mobile-first con el detalle y los PDFs.
# Persistencia: JSON en disco + backup R2 (sobrevive redeploys de Render).
# ══════════════════════════════════════════════════════════════════════════════

_FACTURAS_LOTES_FILE   = os.path.join(DATA_DIR, 'facturas_lotes.json')
_FACTURAS_LOTES_R2_KEY = 'facturas_lotes.json'
_FACTURAS_LOTES_LOCK   = threading.Lock()
_LOTE_TTL_DIAS         = 120   # los links expiran a los 120 días (ciclo de cobranza largo)
_LOTE_MAX_FACTURAS     = 50    # sanidad: nadie reclama más de 50 facturas juntas


def _lotes_read() -> dict:
    """Lee los lotes de facturas compartidos. Disco primero, R2 como fallback."""
    try:
        with open(_FACTURAS_LOTES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        pass
    data = _r2_download_bytes(_FACTURAS_LOTES_R2_KEY)
    if data:
        try:
            lotes = json.loads(data.decode('utf-8'))
            with open(_FACTURAS_LOTES_FILE, 'w', encoding='utf-8') as f:
                f.write(data.decode('utf-8'))
            return lotes
        except Exception:
            pass
    return {}


def _lotes_write(lotes: dict):
    with open(_FACTURAS_LOTES_FILE, 'w', encoding='utf-8') as f:
        json.dump(lotes, f, ensure_ascii=False)

    def _bg():
        _r2_upload_bytes(_FACTURAS_LOTES_R2_KEY, json.dumps(lotes, ensure_ascii=False).encode(), 'application/json')
    threading.Thread(target=_bg, daemon=True).start()


def _lotes_purgar_vencidos(lotes: dict) -> dict:
    """Elimina lotes más viejos que el TTL para que el JSON no crezca sin límite."""
    corte = time.time() - _LOTE_TTL_DIAS * 86400
    return {t: l for t, l in lotes.items() if l.get('ts', 0) >= corte}


@app.route("/api/facturas/crear-lote", methods=["POST"])
def facturas_crear_lote():
    """Crea un lote compartible de facturas reclamadas y devuelve su URL corta.

    Body: { cuit, nombre, facturas: [{nro, fechaFactura, fechaPago, saldo}] }
    La disponibilidad de PDF se resuelve acá (server-side) contra el meta del ZIP,
    así el link nunca promete un PDF que no existe.
    """
    import secrets

    body     = request.get_json(force=True, silent=True) or {}
    nombre   = str(body.get('nombre', '')).strip()[:120]
    cuit     = str(body.get('cuit', '')).replace('-', '').replace(' ', '').strip()[:11]
    facturas = body.get('facturas', [])

    if not isinstance(facturas, list) or not facturas:
        return jsonify({"ok": False, "error": "Lista de facturas requerida"}), 400
    if len(facturas) > _LOTE_MAX_FACTURAS:
        return jsonify({"ok": False, "error": f"Máximo {_LOTE_MAX_FACTURAS} facturas por lote"}), 400

    # PDFs disponibles según el meta del ZIP maestro (basenames sin extensión)
    meta_zip   = _facturas_zip_meta_read()
    pdfs_disp  = {str(n).lower() for n in meta_zip.get('nombres', [])}

    items = []
    for f in facturas:
        if not isinstance(f, dict):
            continue
        nro = str(f.get('nro', '')).strip()[:60]
        try:
            saldo = round(float(f.get('saldo') or 0), 2)
        except (TypeError, ValueError):
            saldo = 0.0
        items.append({
            'nro':           nro,
            'fechaFactura':  str(f.get('fechaFactura', '')).strip()[:12],
            'fechaPago':     str(f.get('fechaPago', '')).strip()[:12],
            'saldo':         saldo,
            'pdf':           bool(nro) and nro.lower() in pdfs_disp,
        })

    if not items:
        return jsonify({"ok": False, "error": "Ninguna factura válida en el lote"}), 400

    token = secrets.token_urlsafe(6)   # ~8 chars URL-safe, impredecible
    with _FACTURAS_LOTES_LOCK:
        lotes = _lotes_purgar_vencidos(_lotes_read())
        lotes[token] = {
            'nombre':   nombre,
            'cuit':     cuit,
            'facturas': items,
            'ts':       time.time(),
            'fecha':    time.strftime('%d/%m/%Y'),
        }
        _lotes_write(lotes)

    return jsonify({"ok": True, "token": token, "url": request.host_url.rstrip('/') + '/f/' + token})


@app.route("/f/<token>")
def ver_lote_facturas(token):
    """Página pública mobile-first con el estado de cuenta del lote.

    Sin login: el cliente final no tiene cuenta. La protección es el token
    aleatorio impredecible + expiración por TTL. Solo expone lo mismo que ya
    viaja en el mensaje de WhatsApp.
    """
    import html as _html
    from datetime import date as _date
    from urllib.parse import quote as _urlquote

    lote = _lotes_read().get(str(token).strip())
    if not lote or lote.get('ts', 0) < time.time() - _LOTE_TTL_DIAS * 86400:
        return (
            "<!doctype html><html lang='es'><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            "<title>Link no disponible</title></head>"
            "<body style='font-family:system-ui,sans-serif;display:flex;align-items:center;"
            "justify-content:center;min-height:100vh;margin:0;background:#f8fafc;color:#334155'>"
            "<div style='text-align:center;padding:24px'><h2>Link no disponible</h2>"
            "<p>Este estado de cuenta expiró o no existe.<br>"
            "Pedile al vendedor que te lo reenvíe.</p></div></body></html>",
            404,
        )

    def _parse_venc(s):
        try:
            d, m, y = str(s).strip().split('/')
            return _date(int(y), int(m), int(d))
        except Exception:
            return None

    def _fmt_monto(v):
        # Formato es-AR: miles con punto, decimales con coma
        return f"{v:,.2f}".replace(',', '§').replace('.', ',').replace('§', '.')

    hoy   = _date.today()
    total = 0.0
    cnt_vencidas = 0
    cards = []
    for f in lote.get('facturas', []):
        nro   = _html.escape(f.get('nro') or 'Sin número')
        saldo = float(f.get('saldo') or 0)
        total += saldo
        venc      = _parse_venc(f.get('fechaPago'))
        dias      = (hoy - venc).days if venc else None
        vencida   = dias is not None and dias > 0
        if vencida:
            cnt_vencidas += 1
        badge = (
            f"<span class='badge venc'>Vencida hace {dias} día{'s' if dias != 1 else ''}</span>"
            if vencida else "<span class='badge ok'>Al día</span>"
        )
        btn_pdf = (
            f"<a class='btn-pdf' href='/api/facturas-pdf/{_html.escape(_urlquote(f['nro']), quote=True)}.pdf' "
            f"target='_blank' rel='noopener'>Ver factura PDF</a>"
            if f.get('pdf') else ""
        )
        cards.append(f"""
      <div class="card{' card-venc' if vencida else ''}">
        <div class="card-head"><span class="nro">{nro}</span>{badge}</div>
        <div class="row"><span>Emitida</span><span>{_html.escape(f.get('fechaFactura') or '-')}</span></div>
        <div class="row"><span>Vencimiento</span><span>{_html.escape(f.get('fechaPago') or '-')}</span></div>
        <div class="row saldo"><span>Saldo</span><span>$ {_fmt_monto(saldo)}</span></div>
        {btn_pdf}
      </div>""")

    n = len(cards)
    nota_venc = (
        f"<div class='alerta'>{cnt_vencidas} factura{'s' if cnt_vencidas != 1 else ''} "
        f"vencida{'s' if cnt_vencidas != 1 else ''}</div>"
        if cnt_vencidas else ""
    )
    nombre_html = _html.escape(lote.get('nombre') or '')

    return f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex,nofollow">
<title>Estado de cuenta — Bodega Piattelli</title>
<style>
  * {{ box-sizing: border-box; margin: 0; }}
  body {{ font-family: system-ui, -apple-system, 'Segoe UI', sans-serif; background: #f1f5f9; color: #1e293b; }}
  .wrap {{ max-width: 480px; margin: 0 auto; padding: 16px 14px 40px; }}
  header {{ text-align: center; padding: 18px 0 14px; }}
  header h1 {{ font-size: 19px; color: #0f172a; }}
  header .sub {{ font-size: 13px; color: #64748b; margin-top: 3px; }}
  .cliente {{ background: #1d4ed8; color: #fff; border-radius: 12px; padding: 14px 16px; margin-bottom: 14px; }}
  .cliente .nom {{ font-size: 15px; font-weight: 700; }}
  .cliente .tot {{ font-size: 24px; font-weight: 800; margin-top: 6px; }}
  .cliente .cnt {{ font-size: 12px; opacity: .85; margin-top: 2px; }}
  .alerta {{ background: #fef2f2; color: #b91c1c; border: 1px solid #fecaca; border-radius: 10px;
             padding: 9px 12px; font-size: 13px; font-weight: 600; margin-bottom: 14px; text-align: center; }}
  .card {{ background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 13px 14px; margin-bottom: 10px; }}
  .card-venc {{ border-left: 4px solid #dc2626; }}
  .card-head {{ display: flex; justify-content: space-between; align-items: center; gap: 8px; margin-bottom: 8px; }}
  .nro {{ font-weight: 700; font-size: 14px; }}
  .badge {{ font-size: 11px; font-weight: 700; padding: 3px 8px; border-radius: 99px; white-space: nowrap; }}
  .badge.venc {{ background: #fef2f2; color: #dc2626; }}
  .badge.ok {{ background: #f0fdf4; color: #16a34a; }}
  .row {{ display: flex; justify-content: space-between; font-size: 13px; color: #475569; padding: 2px 0; }}
  .row.saldo {{ font-weight: 700; color: #0f172a; font-size: 14px; margin-top: 4px; }}
  .btn-pdf {{ display: block; text-align: center; margin-top: 10px; padding: 9px; border-radius: 9px;
              background: #eff6ff; color: #1d4ed8; font-size: 13px; font-weight: 700; text-decoration: none;
              border: 1px solid #bfdbfe; }}
  .pago {{ background: #fff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 14px 16px; margin-top: 16px; }}
  .pago h3 {{ font-size: 13px; color: #0f172a; margin-bottom: 8px; }}
  .pago .row {{ font-size: 12.5px; }}
  .pago .row span:last-child {{ font-weight: 600; color: #1e293b; text-align: right; word-break: break-all; }}
  footer {{ text-align: center; font-size: 11px; color: #94a3b8; margin-top: 22px; }}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>Bodega Piattelli</h1>
    <div class="sub">Estado de cuenta corriente · {_html.escape(lote.get('fecha', ''))}</div>
  </header>
  <div class="cliente">
    <div class="nom">{nombre_html}</div>
    <div class="tot">$ {_fmt_monto(total)}</div>
    <div class="cnt">{n} factura{'s' if n != 1 else ''} pendiente{'s' if n != 1 else ''}</div>
  </div>
  {nota_venc}
  {''.join(cards)}
  <div class="pago">
    <h3>Datos de pago</h3>
    <div class="row"><span>Banco</span><span>Galicia — CC $ 10653-1 081-8</span></div>
    <div class="row"><span>CBU</span><span>0070081820000010653188</span></div>
    <div class="row"><span>Alias</span><span>ARTELINCGALICIA</span></div>
    <div class="row"><span>CUIT</span><span>30-71029502-2</span></div>
  </div>
  <footer>Bodega Piattelli · Ante cualquier consulta respondé el mensaje de WhatsApp</footer>
</div>
</body>
</html>"""


@app.route("/api/alertas-vencimiento")
def alertas_vencimiento_endpoint():
    """Clientes de la cartera con facturas a vencer en los próximos 10 días.

    Itera _cartera_comercial (fuente de CUITs reales) y usa _buscar_por_nombre_en_idx
    — la misma lógica fuzzy que /api/facturas/<cuit> — para obtener las facturas.
    Devuelve el CUIT real de la cartera, no el CUIT vacío de los registros de saldos.
    """
    from datetime import date, timedelta
    hoy    = date.today()
    limite = hoy + timedelta(days=10)

    clientes = []
    for cc in _cartera_comercial:
        nombre   = str(cc.get('nombre', '')).strip()
        cuit_car = str(cc.get('cuit',   '')).replace('-', '').replace(' ', '').strip()
        if not nombre:
            continue

        # Misma búsqueda fuzzy que usa api_facturas_por_cuit
        facturas = _buscar_por_nombre_en_idx(nombre)
        if not facturas:
            continue

        vencen = []
        for f in facturas:
            saldo = float(f.get('saldo') or 0)
            if saldo <= 0:
                continue
            venc = _parse_fecha_venc(f.get('fechaPago', ''))
            if venc and hoy <= venc <= limite:
                vencen.append(saldo)

        if vencen:
            clientes.append({
                'nombre': nombre,
                'cuit':   cuit_car,   # CUIT real de cartera_comercial
                'count':  len(vencen),
                'monto':  round(sum(vencen), 2),
            })

    clientes.sort(key=lambda x: x['monto'], reverse=True)
    return jsonify({'total': len(clientes), 'clientes': clientes})


@app.route("/api/clientes-emision-20d")
def clientes_emision_20d():
    """Clientes de la cartera con al menos una factura emitida hace más de 20 días y saldo > 0.

    Usa _buscar_por_nombre_en_idx (fuzzy match igual que /api/facturas/<cuit>) para
    garantizar que el CUIT retornado es el real de cartera_comercial.
    """
    from datetime import date, timedelta
    hoy    = date.today()
    corte  = hoy - timedelta(days=20)

    def _parse_emision(s):
        """Parsea 'DD/MM/YYYY' a date. fechaFactura en el modelo = fecha de emisión."""
        if not s:
            return None
        try:
            d, m, y = str(s).strip().split('/')
            return date(int(y), int(m), int(d))
        except Exception:
            return None

    clientes = []
    for cc in _cartera_comercial:
        nombre   = str(cc.get('nombre', '')).strip()
        cuit_car = str(cc.get('cuit',   '')).replace('-', '').replace(' ', '').strip()
        if not nombre:
            continue
        facturas = _buscar_por_nombre_en_idx(nombre)
        if not facturas:
            continue

        antiguas = []
        for f in facturas:
            saldo = float(f.get('saldo') or 0)
            if saldo <= 0:
                continue
            em = _parse_emision(f.get('fechaFactura', ''))
            if em and em <= corte:
                antiguas.append(saldo)

        if antiguas:
            clientes.append({
                'nombre': nombre,
                'cuit':   cuit_car,
                'count':  len(antiguas),
                'monto':  round(sum(antiguas), 2),
            })

    clientes.sort(key=lambda x: x['monto'], reverse=True)
    return jsonify({'total': len(clientes), 'clientes': clientes})


@app.route("/upload-saldos-gestion", methods=["POST"])
def upload_saldos_gestion():
    """Saldos semanales de gestión — actualiza la vista comercial sin tocar el DSO de cierre de mes."""
    global _saldos_gestion, _saldos_facturas
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Sin archivo"}), 400
        file = request.files['file']
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        primera = [str(c or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        tiene_header = any(p.upper() in ('VENDEDOR', 'CLIENTE', 'SALDO', 'TOTAL', 'FACTURA', 'NUMERO', 'VENCIM', 'EMISION', 'ADEUDADO', 'PENDIENTE') for p in primera)
        min_row = 2 if tiene_header else 1

        # Detección posicional por nombre de encabezado (tolerante a variantes Odoo)
        if tiene_header:
            hu = [str(c or '').upper().strip() for c in primera]
            _ci = lambda *kws: next((i for i, h in enumerate(hu) if any(k in h for k in kws)), None)
            _col_v  = _ci('VENDEDOR', 'SALESPERSON', 'RESPONSABLE', 'COMERCIAL')
            _col_c  = _ci('CLIENTE', 'PARTNER', 'CONTACTO', 'EMPRESA', 'RAZON')
            _col_nf = _ci('FACTURA', 'NUMERO', 'N°', 'NRO', 'COMPROBANTE', 'INVOICE', 'REFERENCIA')
            # Fecha factura: busca "EMISION" o "FECHA FAC" primero; si no, primera "FECHA" sin "VENCIM"/"PAG"
            _col_ff = _ci('EMISION', 'FECHA FAC', 'FECHA DE FAC') or \
                      next((i for i, h in enumerate(hu) if 'FECHA' in h and 'VENCIM' not in h and 'PAG' not in h and 'DUE' not in h), None)
            _col_fp = _ci('VENCIM', 'DUE DATE', 'FECHA VEN', 'FECHA PAG', 'FECHA PAGO')
            _col_t  = _ci('TOTAL FAC', 'TOTAL COMP', 'IMPORTE TOTAL', 'AMOUNT DUE', 'IMPORTE') or \
                      next((i for i, h in enumerate(hu) if 'TOTAL' in h and 'SALDO' not in h and 'ADEUDADO' not in h and 'PENDIENTE' not in h), None)
            _col_s  = _ci('SALDO PEND', 'IMPORTE ADEUD', 'ADEUDADO', 'PENDIENTE DE COBRO', 'IMPORTE PEND', 'SALDO PENDIENTE') or _ci('SALDO')
            col_v  = _col_v  if _col_v  is not None else 0
            col_c  = _col_c  if _col_c  is not None else 1
            col_nf = _col_nf if _col_nf is not None else 2
            col_ff = _col_ff if _col_ff is not None else 3
            col_fp = _col_fp if _col_fp is not None else 4
            col_t  = _col_t  if _col_t  is not None else 5
            col_s  = _col_s  if _col_s  is not None else 6
            print(f"[upload-gestion] Headers detectados: {hu}", flush=True)
            print(f"[upload-gestion] Cols → vendedor={col_v}('{hu[col_v] if col_v < len(hu) else '?'}') "
                  f"cliente={col_c}('{hu[col_c] if col_c < len(hu) else '?'}') "
                  f"saldo={col_s}('{hu[col_s] if col_s < len(hu) else '?'}')", flush=True)
        else:
            col_v = 0; col_c = 1; col_nf = 2; col_ff = 3; col_fp = 4; col_t = 5; col_s = 6
            print(f"[upload-gestion] Sin header — modo posicional. Primera fila: {primera}", flush=True)

        # Resolver colisión: en exportes Odoo agrupados por vendedor, la columna
        # "Empresa" corresponde al VENDEDOR (no al cliente). La detección de 'EMPRESA'
        # en _col_c queda apuntando al mismo índice que col_v.
        # En ese caso el cliente real está en la columna "Referencia" (col_nf).
        if col_c == col_v:
            old_nf = col_nf
            col_c  = col_nf       # cliente real estaba en la columna "Referencia"
            _usadas = {col_v, col_c, col_ff, col_fp, col_t, col_s}
            col_nf  = next((i for i in range(max(_usadas, default=6) + 3) if i not in _usadas), old_nf)
            print(f"[upload-gestion] Colisión col_c==col_v resuelta → col_c={col_c} col_nf={col_nf}", flush=True)

        # Verificar tipos en la primera fila de datos: si col_nf apunta a un datetime
        # y col_ff a un string, están invertidos (el número de factura quedó en col_ff
        # y la fecha en col_nf). En ese caso los swapeamos.
        _chk_row = ws.cell(row=min_row, column=col_nf + 1).value
        _chk_ff  = ws.cell(row=min_row, column=col_ff + 1).value
        if hasattr(_chk_row, 'strftime') and not hasattr(_chk_ff, 'strftime'):
            col_nf, col_ff = col_ff, col_nf
            print(f"[upload-gestion] Swap nf↔ff detectado (nf era datetime) → col_nf={col_nf} col_ff={col_ff}", flush=True)

        def fmt_fecha(d):
            if not d: return ''
            if hasattr(d, 'strftime'): return d.strftime('%d/%m/%Y')
            s = str(d).strip()
            if len(s) == 10 and s[4] == '-':
                return s[8:] + '/' + s[5:7] + '/' + s[:4]
            return s[:10]
        saldos = []
        for row in ws.iter_rows(min_row=min_row, values_only=True):
            if not row: continue
            vals = list(row) + [None] * 9
            vendedor  = vals[col_v]
            cliente   = vals[col_c]
            nro_fac   = vals[col_nf]
            fecha_fac = vals[col_ff]
            fecha_pago = vals[col_fp]
            total     = vals[col_t]
            saldo     = vals[col_s]
            if not cliente: continue
            try: saldo_f = float(saldo or 0)
            except: saldo_f = 0
            if saldo_f <= 0: continue
            try: total_f = float(total or 0)
            except: total_f = 0
            saldos.append({
                'vendedor': str(vendedor or '').strip(),
                'cliente': str(cliente).strip(),
                'nroFactura': str(nro_fac or '').strip(),
                'fechaFactura': fmt_fecha(fecha_fac),
                'fechaPago': fmt_fecha(fecha_pago),
                'totalFactura': total_f,
                'saldo': saldo_f
            })
        clientes_unicos = len(set(s['cliente'] for s in saldos))
        vendedores_unicos = len(set(s['vendedor'] for s in saldos))
        print(f"[upload-gestion] Parseados: {len(saldos)} registros | "
              f"{clientes_unicos} clientes únicos | {vendedores_unicos} vendedores únicos", flush=True)
        if saldos:
            print(f"[upload-gestion] Primeras 3 filas: {saldos[:3]}", flush=True)
        sg_path = os.path.join(DATA_DIR, 'saldos_gestion_vendedores.json')
        with open(sg_path, 'w', encoding='utf-8') as f:
            json.dump(saldos, f, ensure_ascii=False, indent=2)
        ts_path = os.path.join(DATA_DIR, 'saldos_timestamp.json')
        with open(ts_path, 'w') as f:
            json.dump({'ts': time.time(), 'fecha': time.strftime('%d/%m/%Y %H:%M'), 'tipo': 'gestion'}, f)
        _saldos_gestion  = saldos
        _saldos_facturas = list(saldos)   # sincronizar SSoT para que el índice siempre sea fresco
        _rebuild_saldos_index()
        # Actualizar mtime tracking para que este worker no recargue innecesariamente
        global _SG_MTIME, _SG_LAST_CHECK
        try:
            _SG_MTIME = os.path.getmtime(sg_path)
        except Exception:
            pass
        _SG_LAST_CHECK = time.time()
        # Auto-sync: corregir asignaciones de vendedor en cartera_comercial según el reporte
        try:
            _cambios = _sync_cartera_vendedores(saldos)
        except Exception as _se:
            print(f"[gestion] sync-cartera error (no crítico): {_se}", flush=True)
            _cambios = []
        print(f"[gestion] {len(saldos)} facturas importadas | sync-cartera: {len(_cambios)} cambios", flush=True)
        # Recalcular scores en background usando BCRA cacheado (sin API calls)
        threading.Thread(target=_recalcular_scores_post_upload, daemon=True).start()
        return jsonify({"ok": True, "total": len(saldos), "reasignaciones": len(_cambios), "cambios": _cambios})
    except Exception as e:
        import traceback
        print(f"[gestion] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

# ─── CARTERA: carga masiva desde Odoo ────────────────────────────────────────

# Palabras que indican que una entrada es dirección de entrega/sucursal, no razón social
_CARTERA_DISCARD_KW = [
    # Logística / entrega
    'ENTREGA', ' ENVIO', ' ENVÍO', 'DELIVERY', ' RETIRO', 'EXPEDICION',
    'EXPEDICIÓN', 'DESPACHO', 'PICKING', 'ENVÍOS', 'ENVIOS',
    # Sucursales / depósitos
    ' SUCURSAL', ' SUC.', ' SUC ', 'DEPOSITO', 'DEPÓSITO',
    'ALMACEN', 'ALMACÉN', 'BODEGA', 'PLANTA ', '(DEPOSITO', '(DEPÓSITO',
    # Indicadores de calle / dirección
    ' AV.', ' AV ', 'AVDA.', 'AVDA ', 'AVENIDA ', 'CALLE ', ' C/ ', ' C/',
    ' PASAJE', ' PJE.', ' PJE ', ' RUTA ', ' RN ', 'RN.', ' KM ',
    ' PISO ', ' PISO.', 'DPTO', 'DEPTO', 'OFICINA ', ' OF.', ' OF ',
    ' LOCAL ', ' LOCAL.', '(LOCAL', 'PUERTA',
    # Ordinales / numeración de locales
    ' 1ER ', ' 2DO ', ' 3ER ', ' 4TO ', ' 1° ', ' 2° ', ' 3° ',
    ' (2)', ' (3)', ' #2', ' #3', ' N°2', ' N°3',
    # Direcciones cardinales (sucursales)
    ' NORTE', ' SUR', ' ESTE', ' OESTE', ' CENTRO',
]

_LEGAL_SUFFIXES = [
    'S.R.L', 'SRL', 'S.A.', 'S.A.S', 'SAS', 'S.C.A', 'S.E.',
    'E.I.R.L', 'EIRL', 'LTDA', 'S.A.P.E.M', 'S.E.M', 'U.T.E',
]


def _limpiar_cuit_upload(raw) -> str:
    """Normaliza cualquier variante de CUIT a 11 dígitos sin separadores."""
    if raw is None:
        return ''
    s = str(raw).strip()
    # Excel suele devolver CUITs como float: 30714840203.0 → 30714840203
    if '.' in s:
        s = s.split('.')[0]
    s = s.replace('-', '').replace(' ', '').replace('.', '')
    return s if s.isdigit() and 10 <= len(s) <= 11 else ''


def _score_nombre_cartera(nombre: str) -> int:
    """Score heurístico: 0 = razón social limpia, >100 = dirección/entrega probable."""
    import re as _re
    n = (nombre or '').upper().strip()
    score = 0
    for kw in _CARTERA_DISCARD_KW:
        if kw in n:
            score += 100
    # Número al final = posible número de calle (ej: "EMPRESA 1234")
    if _re.search(r'\s\d{3,}$', n):
        score += 50
    # Paréntesis = aclaraciones de entrega o sucursal
    if '(' in n:
        score += 30
    # Nombres más cortos son preferidos (razón social base)
    score += len(nombre)
    # Bonificación por sufijo legal (S.R.L., S.A., etc.)
    if any(s in n for s in _LEGAL_SUFFIXES):
        score -= 20
    return score


def _razon_descarte(nombre: str) -> str:
    import re as _re
    n = nombre.upper()
    for kw in _CARTERA_DISCARD_KW:
        if kw in n:
            return f'Contiene "{kw.strip()}"'
    if _re.search(r'\s\d{3,}$', n):
        return 'Termina con número de calle'
    if '(' in n:
        return 'Contiene paréntesis (aclaración)'
    return 'Nombre más largo para el mismo CUIT'


@app.route("/upload-cartera", methods=["POST"])
def upload_cartera_endpoint():
    """Parsea archivo Odoo, deduplica por CUIT descartando entradas de entrega."""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Sin archivo"}), 400
        file = request.files['file']
        nombre_arch = (file.filename or '').lower()

        import io, csv as _csv

        # ── Leer filas ────────────────────────────────────────────────────────
        if nombre_arch.endswith('.csv'):
            contenido = file.read().decode('utf-8-sig', errors='replace')
            todas = list(_csv.reader(io.StringIO(contenido)))
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            todas = [[str(v) if v is not None else '' for v in row]
                     for row in ws.iter_rows(values_only=True)]

        if not todas:
            return jsonify({"error": "Archivo vacío"}), 400

        headers = [str(h or '').strip() for h in todas[0]]

        # ── Detección de columnas ─────────────────────────────────────────────
        # Normaliza tildes/acentos para comparación robusta con exports de Odoo
        def _nh(s):
            return (s.upper()
                    .replace('Á','A').replace('É','E').replace('Í','I')
                    .replace('Ó','O').replace('Ú','U').replace('Ü','U')
                    .replace('Ñ','N'))

        headers_n = [_nh(h) for h in headers]

        def find_col(kws):
            for i, hn in enumerate(headers_n):
                if any(_nh(kw) in hn for kw in kws):
                    return i
            return -1

        # CUIT del cliente: busca "CUIT" que NO sea de vendedor/proveedor.
        # Prioriza columnas que contengan "CLIENTE" junto a "CUIT".
        def find_cuit_cliente():
            best = -1
            for i, hn in enumerate(headers_n):
                if 'CUIT' in hn or 'NIF' in hn:
                    if 'VENDEDOR' in hn or 'PROVEEDOR' in hn:
                        continue              # excluir CUIT VENDEDOR
                    if 'CLIENTE' in hn:
                        return i              # match exacto "cuit cliente"
                    if best < 0:
                        best = i              # candidato genérico
            return best

        col_n  = find_col(['CLIENTE', 'NOMBRE', 'NAME', 'RAZON', 'PARTNER', 'EMPRESA', 'COMPANY'])
        col_c  = find_cuit_cliente()
        col_v  = find_col(['VENDEDOR', 'SALESPERSON', 'COMERCIAL', 'REPRESENTANTE'])
        col_ci = find_col(['CIUDAD', 'CITY', 'LOCALIDAD', 'MUNICIPIO', 'POBLACION', 'PROVINCIA'])
        col_e  = find_col(['EMAIL', 'CORREO', 'MAIL'])
        col_l  = find_col(['LIMITE DE CREDITO', 'LIMITE', 'LIMIT', 'CREDITO', 'CREDIT'])
        col_t  = find_col(['TIPO DE CONTACTO', 'TIPO CONTACTO', 'TYPE', 'TIPO'])

        if col_n < 0:
            return jsonify({
                "error": f"No se encontró columna de nombre. Encabezados detectados: {headers[:10]}"
            }), 400

        def gv(row, col):
            if col < 0 or col >= len(row):
                return ''
            v = row[col]
            return str(v).strip() if v is not None else ''

        # ── Procesar filas ────────────────────────────────────────────────────
        grupos   = {}   # cuit → [entries]
        sin_cuit = []
        total_filas = 0

        for row in todas[1:]:
            nombre = gv(row, col_n)
            if not nombre or nombre.upper() in ('FALSE', 'NONE', ''):
                continue
            total_filas += 1

            # Descartar si la columna Tipo indica explícitamente dirección/entrega
            tipo = gv(row, col_t).upper()
            if tipo and any(kw in tipo for kw in ['ENTREGA', 'DELIVERY', 'DIRECCION', 'ENVIO', 'SHIPPING']):
                continue

            cuit = _limpiar_cuit_upload(gv(row, col_c)) if col_c >= 0 else ''
            entry = {
                'nombre':        nombre,
                'cuit':          cuit,
                'vendedor':      gv(row, col_v),
                'ciudad':        gv(row, col_ci),
                'email':         gv(row, col_e),
                'limiteCredito': float(gv(row, col_l) or 0) if gv(row, col_l) else 0.0,
            }

            if cuit:
                grupos.setdefault(cuit, []).append(entry)
            else:
                sin_cuit.append(entry)

        # ── Deduplicar por CUIT ───────────────────────────────────────────────
        resultado   = []
        descartados = []
        duplicados  = 0

        for cuit, entradas in grupos.items():
            if len(entradas) == 1:
                resultado.append(entradas[0])
            else:
                scored = sorted(entradas, key=lambda e: _score_nombre_cartera(e['nombre']))
                resultado.append(scored[0])
                for dup in scored[1:]:
                    duplicados += 1
                    descartados.append({
                        'nombre': dup['nombre'],
                        'cuit':   cuit,
                        'motivo': _razon_descarte(dup['nombre']),
                    })

        resultado.extend(sin_cuit)

        print(
            f"[upload-cartera] {total_filas} filas → {len(resultado)} clientes "
            f"({duplicados} dupl. eliminados, {len(sin_cuit)} sin CUIT)",
            flush=True,
        )
        return jsonify({
            "ok": True,
            "total_filas":          total_filas,
            "total_clientes":       len(resultado),
            "duplicados_eliminados": duplicados,
            "sin_cuit":             len(sin_cuit),
            "clientes":             resultado,
            "descartados":          descartados[:200],
        })

    except Exception as e:
        import traceback
        print(f"[upload-cartera] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500


@app.route("/confirmar-upload-cartera", methods=["POST"])
def confirmar_upload_cartera():
    """Reemplaza cartera_comercial.json con la lista procesada y recarga en memoria."""
    global _cartera_comercial
    try:
        data = request.get_json(force=True)
        clientes = data.get('clientes', [])
        if not clientes:
            return jsonify({"error": "Lista vacía"}), 400

        cartera_nueva = []
        for c in clientes:
            nombre = str(c.get('nombre', '') or '').strip()
            if not nombre:
                continue
            cartera_nueva.append({
                'nombre':        nombre,
                'cuit':          _limpiar_cuit_upload(c.get('cuit', '')) or str(c.get('cuit', '')),
                'vendedor':      str(c.get('vendedor', '') or '').strip(),
                'ciudad':        str(c.get('ciudad', '') or '').strip(),
                'email':         str(c.get('email', '') or '').strip(),
                'limiteCredito': float(c.get('limiteCredito') or 0),
                'plazo':         int(c.get('plazo') or 0),
            })

        tmp = _CC_FILE + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(cartera_nueva, f, ensure_ascii=False, indent=2)
            f.flush(); os.fsync(f.fileno())
        os.replace(tmp, _CC_FILE)
        _cartera_comercial = cartera_nueva

        print(f"[cartera] Importada: {len(cartera_nueva)} clientes", flush=True)
        return jsonify({"ok": True, "total": len(cartera_nueva)})

    except Exception as e:
        import traceback
        print(f"[confirmar-cartera] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500


@app.route("/saldos-timestamp")
def get_saldos_timestamp():
    """Devuelve timestamp de la última carga de saldos para detección de actualizaciones."""
    ts_path = os.path.join(DATA_DIR, 'saldos_timestamp.json')
    try:
        with open(ts_path, 'r') as f:
            data = json.load(f)
        resp = jsonify(data)
        resp.headers['Cache-Control'] = 'no-store'
        return resp
    except:
        return jsonify({"ts": 0, "fecha": None})

def _dso_exhaustion(ar: float, ventas_por_mes: dict, fecha_corte_dt) -> dict:
    """Aplica el método de agotamiento sobre 3 meses hacia atrás desde fecha_corte_dt.
    ventas_por_mes: {(year, month): total}
    Devuelve {dso, breakdown} o {dso: None} si no hay ventas."""
    import calendar
    if ar <= 0:
        return {"dso": 0, "breakdown": []}
    meses = []
    y, m = fecha_corte_dt.year, fecha_corte_dt.month
    for _ in range(3):
        dias  = calendar.monthrange(y, m)[1]
        ventas = ventas_por_mes.get((y, m), 0.0)
        meses.append({"mes": f"{m:02d}/{y}", "dias": dias, "ventas": ventas})
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    ar_rest = ar
    dso_acum = 0.0
    breakdown = []
    for mi in meses:
        if ar_rest <= 0:
            break
        v = mi["ventas"]
        d = mi["dias"]
        if v <= 0:
            breakdown.append({**mi, "dias_dso": 0, "nota": "sin ventas"})
            continue
        if ar_rest >= v:
            dso_acum += d
            ar_rest  -= v
            breakdown.append({**mi, "dias_dso": d, "ar_restante": round(ar_rest)})
        else:
            dp = (ar_rest / v) * d
            dso_acum += dp
            breakdown.append({**mi, "dias_dso": round(dp, 1), "ar_restante": 0})
            ar_rest = 0
    return {"dso": round(dso_acum) if ar > 0 else None, "breakdown": breakdown}


@app.route("/api/dso-todos")
def get_dso_todos():
    """DSO por vendedor y por cliente usando el método de agotamiento.
    Fuentes: dso_saldos_actual, dso_cheques_actual, dso_ventas_historico."""
    import calendar
    from datetime import datetime

    # ── Leer saldos ──────────────────────────────────────────────────────────
    saldos_lista = []
    s_path = os.path.join(DATA_DIR, 'dso_saldos_actual.json')
    if os.path.exists(s_path):
        try:
            saldos_lista = json.load(open(s_path, 'r', encoding='utf-8')).get('saldos', [])
        except Exception:
            pass

    # ── Leer cheques ─────────────────────────────────────────────────────────
    cheques_por_cliente: dict = {}
    c_path = os.path.join(DATA_DIR, 'dso_cheques_actual.json')
    if os.path.exists(c_path):
        try:
            for ch in json.load(open(c_path, 'r', encoding='utf-8')).get('cheques', []):
                tot = float(ch.get('total', 0) or 0)
                if tot <= 0:
                    continue
                cli = _norm_nombre(ch.get('cliente') or '')
                if cli:
                    cheques_por_cliente[cli] = cheques_por_cliente.get(cli, 0.0) + tot
        except Exception:
            pass

    # ── Leer ventas por cliente ───────────────────────────────────────────────
    ventas_globales: dict = {}   # (year, month) → total
    ventas_por_cli: dict  = {}   # cli_norm → {(year, month) → total}
    v_path = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
    if os.path.exists(v_path):
        try:
            vh = json.load(open(v_path, 'r', encoding='utf-8'))
            for ym, tot in vh.get('meses', {}).items():
                try:
                    ventas_globales[(int(ym[:4]), int(ym[5:7]))] = float(tot)
                except Exception:
                    pass
            for cli, meses_cli in vh.get('por_cliente', {}).items():
                cli_norm = _norm_nombre(cli)  # normalizar igual que saldo_por_cli
                ventas_por_cli[cli_norm] = {}
                for ym, tot in meses_cli.items():
                    try:
                        ventas_por_cli[cli_norm][(int(ym[:4]), int(ym[5:7]))] = float(tot)
                    except Exception:
                        pass
        except Exception:
            pass

    # ── Fecha de corte ────────────────────────────────────────────────────────
    fechas = [_parsear_fecha_dso(s.get('fecha_factura')) for s in saldos_lista]
    fechas = [f for f in fechas if f]
    fecha_corte = max(fechas) if fechas else datetime.now()

    # ── Saldo y vendedor por cliente ──────────────────────────────────────────
    saldo_por_cli: dict    = {}
    vendedor_por_cli: dict = {}
    for s in saldos_lista:
        cli = _norm_nombre(s.get('cliente') or '')
        if not cli:
            continue
        saldo_por_cli[cli] = saldo_por_cli.get(cli, 0.0) + float(s.get('saldo', 0) or 0)
        vend = (s.get('vendedor') or '').strip()
        if vend:
            vendedor_por_cli[cli] = vend

    # Enriquecer vendedor_por_cli desde saldos_gestion/facturas (siempre tienen vendedor)
    # Cubre casos donde dso_saldos_actual no tiene el campo vendedor (upload anterior al fix)
    fuente_vend = _saldos_gestion if _saldos_gestion else _saldos_facturas
    for f in fuente_vend:
        cli  = _norm_nombre(f.get('cliente') or '')
        vend = (f.get('vendedor') or '').strip()
        if cli and vend and cli not in vendedor_por_cli:
            vendedor_por_cli[cli] = vend

    # Total saldo para asignación proporcional de ventas a clientes sin datos propios
    total_saldo_global = sum(v for v in saldo_por_cli.values() if v > 0)

    def _ventas_proporcional(cli_saldo: float) -> dict:
        """Asigna ventas proporcionales al saldo del cliente cuando no hay datos per-cliente."""
        if total_saldo_global <= 0 or cli_saldo <= 0:
            return {}
        ratio = cli_saldo / total_saldo_global
        return {ym: tot * ratio for ym, tot in ventas_globales.items()}

    # ── DSO por cliente ───────────────────────────────────────────────────────
    dso_por_cliente: dict = {}
    for cli, saldo in saldo_por_cli.items():
        cheques = cheques_por_cliente.get(cli, 0.0)
        ar      = saldo + cheques
        if ar <= 0:
            dso_por_cliente[cli] = 0
            continue
        # Usar ventas propias si existen, si no asignar proporcionalmente
        ventas_cli = ventas_por_cli.get(cli) or _ventas_proporcional(saldo)
        res = _dso_exhaustion(ar, ventas_cli, fecha_corte)
        dso_por_cliente[cli] = res["dso"]

    # ── DSO por vendedor ──────────────────────────────────────────────────────
    vend_saldo:   dict = {}
    vend_cheques: dict = {}
    vend_ventas:  dict = {}
    for cli, saldo in saldo_por_cli.items():
        vend = vendedor_por_cli.get(cli, '')
        if not vend:
            continue
        vend_saldo[vend]   = vend_saldo.get(vend, 0.0) + saldo
        vend_cheques[vend] = vend_cheques.get(vend, 0.0) + cheques_por_cliente.get(cli, 0.0)
        # Sumar ventas propias del cliente al vendedor
        v_cli = ventas_por_cli.get(cli) or _ventas_proporcional(saldo)
        if vend not in vend_ventas:
            vend_ventas[vend] = {}
        for ym_key, tot in v_cli.items():
            vend_ventas[vend][ym_key] = vend_ventas[vend].get(ym_key, 0.0) + tot

    dso_por_vendedor: dict = {}
    for vend in vend_saldo:
        ar     = vend_saldo[vend] + vend_cheques.get(vend, 0.0)
        ventas = vend_ventas.get(vend, {})
        res    = _dso_exhaustion(ar, ventas, fecha_corte)
        dso_por_vendedor[vend] = {
            "dso":      res["dso"],
            "ar":       round(ar),
            "saldo":    round(vend_saldo[vend]),
            "cheques":  round(vend_cheques.get(vend, 0.0)),
            "breakdown": res["breakdown"],
        }

    # ── DSO ponderado global y por vendedor (promedio ponderado de DSO por cliente) ──
    # Mismo método que usa el panel Director: Σ(dso_i × saldo_i) / Σ(saldo_i)
    # Más preciso que el agotamiento agregado porque usa datos reales de cada cliente.
    _sp_g: float = 0.0
    _ss_g: float = 0.0
    _vend_pond: dict = {}   # vendedor → {sum_pond, sum_saldo}
    for cli, saldo in saldo_por_cli.items():
        dso_cli = dso_por_cliente.get(cli)
        if not dso_cli or saldo <= 0:
            continue
        _sp_g += dso_cli * saldo
        _ss_g += saldo
        vend = vendedor_por_cli.get(cli, '')
        if vend:
            if vend not in _vend_pond:
                _vend_pond[vend] = {'sp': 0.0, 'ss': 0.0}
            _vend_pond[vend]['sp'] += dso_cli * saldo
            _vend_pond[vend]['ss'] += saldo

    dso_global_ponderado = round(_sp_g / _ss_g) if _ss_g > 0 else None
    global _dso_global_ponderado_cache
    if dso_global_ponderado is not None:
        _dso_global_ponderado_cache = dso_global_ponderado
    dso_vendedor_ponderado = {
        v: round(d['sp'] / d['ss'])
        for v, d in _vend_pond.items() if d['ss'] > 0
    }

    print(f"[dso-todos] {len(dso_por_cliente)} clientes | {len(dso_por_vendedor)} vendedores | "
          f"DSO global ponderado={dso_global_ponderado}d | "
          f"fecha_corte={fecha_corte.strftime('%d/%m/%Y')}", flush=True)
    return jsonify({
        "fecha_corte":            fecha_corte.strftime('%d/%m/%Y'),
        "por_vendedor":           dso_por_vendedor,
        "por_cliente":            dso_por_cliente,
        "saldo_por_cliente":      {cli: round(saldo) for cli, saldo in saldo_por_cli.items()},
        "vendedor_por_cliente":   {cli: vendedor_por_cli.get(cli, '') for cli in saldo_por_cli},
        # Valores precalculados server-side (evitan re-calcular en cada cliente frontend)
        "dso_global_ponderado":   dso_global_ponderado,
        "dso_vendedor_ponderado": dso_vendedor_ponderado,
    })


def _parsear_fecha_dso(s):
    """Parsea fecha en formato ISO (YYYY-MM-DD) o argentino (DD/MM/YYYY)."""
    from datetime import datetime
    s = (s or '')[:10].strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s) if '-' in s else datetime(int(s[6:]), int(s[3:5]), int(s[:2]))
    except Exception:
        return None

@app.route("/dso-global-saldos")
def get_dso_global_saldos():
    """DSO global — método de agotamiento (exhaustion method) sobre 3 meses.
    AR = saldos pendientes + cheques (dso_saldos_actual + dso_cheques_actual).
    Ventas = historial mensual agrupado desde dso_ventas_historico.json.
    Algoritmo: restar mes a mes (más reciente primero) hasta agotar el AR."""
    import calendar
    from datetime import datetime

    # ── 1. AR: saldos + cheques del último upload DSO ─────────────────────────
    saldos_lista = []
    s_path = os.path.join(DATA_DIR, 'dso_saldos_actual.json')
    if os.path.exists(s_path):
        try:
            with open(s_path, 'r', encoding='utf-8') as _fs:
                saldos_lista = json.load(_fs).get('saldos', [])
        except Exception as _se:
            print(f"[dso-global] error saldos: {_se}", flush=True)

    if not saldos_lista:
        return jsonify({"dso": None, "saldo_total": 0, "clientes_count": 0,
                        "ventas_3m": 0, "formula": "sin_datos", "breakdown": []})

    saldo_base = sum(float(s.get('saldo', 0) or 0) for s in saldos_lista)
    total_cheques = 0.0
    c_path = os.path.join(DATA_DIR, 'dso_cheques_actual.json')
    if os.path.exists(c_path):
        try:
            with open(c_path, 'r', encoding='utf-8') as _fc:
                # Solo cheques positivos (pendientes de cobro)
                total_cheques = sum(float(ch.get('total', 0) or 0)
                                    for ch in json.load(_fc).get('cheques', [])
                                    if float(ch.get('total', 0) or 0) > 0)
        except Exception as _ce:
            print(f"[dso-global] error cheques: {_ce}", flush=True)
    # AR = Saldos + Cheques (según fórmula del usuario: Deuda Total a Agotar)
    ar_total = saldo_base + total_cheques

    # ── 2. Fecha de corte — último día del período de los saldos ─────────────
    fechas = [_parsear_fecha_dso(s.get('fecha_factura')) for s in saldos_lista]
    fechas = [f for f in fechas if f]
    fecha_corte = max(fechas) if fechas else datetime.now()

    # ── 3. Ventas por mes desde historico mensual ─────────────────────────────
    ventas_por_mes = {}  # {(year, month): total}
    n_filas_ventas = 0
    try:
        v_path = os.path.join(DATA_DIR, 'dso_ventas_historico.json')
        if os.path.exists(v_path):
            with open(v_path, 'r', encoding='utf-8') as _fv:
                vh = json.load(_fv)
            meses_data = vh.get('meses', {})
            n_filas_ventas = len(meses_data)
            for ym, total in meses_data.items():
                try:
                    y2, m2 = int(ym[:4]), int(ym[5:7])
                    ventas_por_mes[(y2, m2)] = float(total)
                except Exception:
                    pass
    except Exception as _ve:
        print(f"[dso-global] error ventas: {_ve}", flush=True)

    # ── 4. 3 meses hacia atrás desde fecha_corte (mes más reciente primero) ──
    meses = []
    y, m = fecha_corte.year, fecha_corte.month
    for _ in range(3):
        dias = calendar.monthrange(y, m)[1]
        ventas = ventas_por_mes.get((y, m), 0.0)
        meses.append({"mes": f"{m:02d}/{y}", "dias": dias, "ventas": ventas})
        m -= 1
        if m == 0:
            m, y = 12, y - 1

    ventas_3m = sum(x['ventas'] for x in meses)

    # ── 5. Método de agotamiento ──────────────────────────────────────────────
    ar_restante = ar_total
    dso_acum    = 0.0
    breakdown   = []
    for mes_info in meses:
        if ar_restante <= 0:
            break
        v = mes_info['ventas']
        d = mes_info['dias']
        if v <= 0:
            breakdown.append({**mes_info, "dias_dso": 0, "ar_restante": round(ar_restante), "nota": "sin ventas"})
            continue
        if ar_restante >= v:
            dso_acum   += d
            ar_restante -= v
            breakdown.append({**mes_info, "dias_dso": d, "ar_restante": round(ar_restante)})
        else:
            dias_parcial = (ar_restante / v) * d
            dso_acum    += dias_parcial
            breakdown.append({**mes_info, "dias_dso": round(dias_parcial, 1), "ar_restante": 0})
            ar_restante  = 0

    dso = round(dso_acum) if ar_total > 0 else None

    clientes_unicos = len({s.get('cliente', '') for s in saldos_lista if s.get('cliente')})
    print(
        f"[dso-global] corte={fecha_corte.strftime('%d/%m/%Y')} "
        f"AR={ar_total:.0f} (saldos={saldo_base:.0f} cheques={total_cheques:.0f}) "
        f"filas_ventas={n_filas_ventas} ventas_3m={ventas_3m:.0f} DSO={dso}d",
        flush=True
    )
    for b in breakdown:
        print(f"  → {b['mes']}: ventas={b['ventas']:.0f} dias_dso={b['dias_dso']} "
              f"ar_restante={b.get('ar_restante',0):.0f}", flush=True)

    return jsonify({
        "dso":           dso,
        "saldo_total":   saldo_base,        # AR para el cálculo = solo saldos
        "saldo_base":    saldo_base,
        "total_cheques": total_cheques,
        "ar_con_cheques": saldo_base + total_cheques,
        "clientes_count": clientes_unicos,
        "ventas_3m":     ventas_3m,
        "fecha_corte":   fecha_corte.strftime('%d/%m/%Y'),
        "formula":       "agotamiento_3m_solo_saldos",
        "breakdown":     breakdown,
        "ultima_actualizacion": time.strftime('%d/%m/%Y')
    })

def _parse_fecha_venc(s):
    """Parsea 'DD/MM/YYYY' a date. fechaPago en el modelo = fecha_vencimiento de Odoo."""
    from datetime import date
    if not s: return None
    try:
        d, m, y = str(s).strip().split('/')
        return date(int(y), int(m), int(d))
    except:
        return None

def _enrich_con_mora(facturas: list) -> tuple:
    """Agrega diasMora a cada factura y retorna (facturas_enriquecidas, monto_vencido_30d, alerta_mora_30)."""
    from datetime import date
    hoy = date.today()
    enriched, monto_v30 = [], 0.0
    for f in facturas:
        fc = dict(f)
        fv = _parse_fecha_venc(fc.get('fechaPago', ''))
        saldo = float(fc.get('saldo') or 0)
        if fv and saldo > 0:
            fc['diasMora'] = max(0, (hoy - fv).days)
            if fc['diasMora'] > 0:
                monto_v30 += saldo
        else:
            fc['diasMora'] = 0
        enriched.append(fc)
    alerta = any(f.get('diasMora', 0) > 30 for f in enriched)
    return enriched, round(monto_v30, 2), alerta


@app.route("/api/alertas-mora")
def get_alertas_mora():
    """Clientes con facturas cuya fecha de vencimiento (fechaPago) supera los 30 días sin cobrar."""
    from datetime import date
    hoy = date.today()
    fuente = _saldos_gestion if _saldos_gestion else _saldos_facturas
    clientes: dict = {}
    for f in fuente:
        saldo = float(f.get('saldo') or 0)
        if saldo <= 0:
            continue
        fv = _parse_fecha_venc(f.get('fechaPago', ''))
        if not fv:
            continue
        dias = (hoy - fv).days
        if dias <= 30:
            continue
        key = str(f.get('cliente', '') or '').strip()
        if not key:
            continue
        if key not in clientes:
            clientes[key] = {
                'nombre': key,
                'cuit': str(f.get('cuit', '') or '').strip(),
                'monto_vencido_30d': 0.0,
                'dias_max_atraso': 0,
                'cantidad_facturas': 0,
            }
        clientes[key]['monto_vencido_30d'] += saldo
        clientes[key]['dias_max_atraso'] = max(clientes[key]['dias_max_atraso'], dias)
        clientes[key]['cantidad_facturas'] += 1
    # Enriquecer CUIT desde cartera_comercial cuando el archivo de saldos no lo trae
    for c in _cartera_comercial:
        nombre_c = str(c.get('nombre', '') or '').strip()
        cuit_c = str(c.get('cuit', '') or '').replace('-', '').replace(' ', '').strip()
        if nombre_c in clientes and not clientes[nombre_c]['cuit'] and cuit_c:
            clientes[nombre_c]['cuit'] = cuit_c
    result = sorted(clientes.values(), key=lambda x: x['monto_vencido_30d'], reverse=True)
    for r in result:
        r['monto_vencido_30d'] = round(r['monto_vencido_30d'], 2)
    total_vencido = sum(r['monto_vencido_30d'] for r in result)
    print(f"[alertas-mora] {len(result)} clientes con vencimiento >30d, total ${total_vencido:,.0f}", flush=True)
    return jsonify({
        'alertas': result,
        'total_clientes': len(result),
        'total_monto_vencido': round(total_vencido, 2),
    })


@app.route("/upload-saldos-facturas", methods=["POST"])
def upload_saldos_facturas():
    """Recibe Excel Odoo: [Vendedor, Cliente, Nro Factura, Fecha Factura, Fecha Pago, Total, Saldo]"""
    global _saldos_facturas, _saldos_gestion
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Sin archivo"}), 400
        file = request.files['file']
        import io, openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active

        # Detectar si primera fila es encabezado textual
        primera = [str(c or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        tiene_header = any(p.upper() in ('VENDEDOR', 'CLIENTE', 'SALDO', 'TOTAL', 'FACTURA', 'NUMERO', 'VENCIM', 'EMISION', 'ADEUDADO', 'PENDIENTE') for p in primera)
        min_row = 2 if tiene_header else 1

        # Detección posicional por nombre de encabezado (tolerante a variantes Odoo)
        if tiene_header:
            hu = [str(c or '').upper().strip() for c in primera]
            _ci = lambda *kws: next((i for i, h in enumerate(hu) if any(k in h for k in kws)), None)
            _col_v  = _ci('VENDEDOR', 'SALESPERSON', 'RESPONSABLE', 'COMERCIAL')
            _col_c  = _ci('CLIENTE', 'PARTNER', 'CONTACTO', 'EMPRESA', 'RAZON')
            _col_nf = _ci('FACTURA', 'NUMERO', 'N°', 'NRO', 'COMPROBANTE', 'INVOICE', 'REFERENCIA')
            _col_ff = _ci('EMISION', 'FECHA FAC', 'FECHA DE FAC') or \
                      next((i for i, h in enumerate(hu) if 'FECHA' in h and 'VENCIM' not in h and 'PAG' not in h and 'DUE' not in h), None)
            _col_fp = _ci('VENCIM', 'DUE DATE', 'FECHA VEN', 'FECHA PAG', 'FECHA PAGO')
            _col_t  = _ci('TOTAL FAC', 'TOTAL COMP', 'IMPORTE TOTAL', 'AMOUNT DUE', 'IMPORTE') or \
                      next((i for i, h in enumerate(hu) if 'TOTAL' in h and 'SALDO' not in h and 'ADEUDADO' not in h and 'PENDIENTE' not in h), None)
            _col_s  = _ci('SALDO PEND', 'IMPORTE ADEUD', 'ADEUDADO', 'PENDIENTE DE COBRO', 'IMPORTE PEND', 'SALDO PENDIENTE') or _ci('SALDO')
            col_v  = _col_v  if _col_v  is not None else 0
            col_c  = _col_c  if _col_c  is not None else 1
            col_nf = _col_nf if _col_nf is not None else 2
            col_ff = _col_ff if _col_ff is not None else 3
            col_fp = _col_fp if _col_fp is not None else 4
            col_t  = _col_t  if _col_t  is not None else 5
            col_s  = _col_s  if _col_s  is not None else 6
        else:
            col_v = 0; col_c = 1; col_nf = 2; col_ff = 3; col_fp = 4; col_t = 5; col_s = 6

        # Resolver colisión: en exportes Odoo agrupados por vendedor, "Empresa" = vendedor.
        # col_c queda apuntando al mismo índice que col_v → el cliente real está en col_nf.
        if col_c == col_v:
            old_nf = col_nf
            col_c  = col_nf
            _usadas = {col_v, col_c, col_ff, col_fp, col_t, col_s}
            col_nf  = next((i for i in range(max(_usadas, default=6) + 3) if i not in _usadas), old_nf)

        # Verificar tipos en la primera fila: si col_nf apunta a datetime y col_ff a string,
        # están invertidos → swapear para que col_nf = número de factura (string).
        _chk_row = ws.cell(row=min_row, column=col_nf + 1).value
        _chk_ff  = ws.cell(row=min_row, column=col_ff + 1).value
        if hasattr(_chk_row, 'strftime') and not hasattr(_chk_ff, 'strftime'):
            col_nf, col_ff = col_ff, col_nf
            print(f"[upload-facturas] Swap nf↔ff detectado (nf era datetime) → col_nf={col_nf} col_ff={col_ff}", flush=True)

        def fmt_fecha(d):
            if not d: return ''
            if hasattr(d, 'strftime'): return d.strftime('%d/%m/%Y')
            s = str(d).strip()
            if len(s) == 10 and s[4] == '-':  # YYYY-MM-DD → DD/MM/YYYY
                return s[8:] + '/' + s[5:7] + '/' + s[:4]
            return s[:10]

        saldos = []
        for row in ws.iter_rows(min_row=min_row, values_only=True):
            if not row: continue
            vals = list(row) + [None] * 9
            vendedor   = vals[col_v]
            cliente    = vals[col_c]
            nro_fac    = vals[col_nf]
            fecha_fac  = vals[col_ff]
            fecha_pago = vals[col_fp]
            total      = vals[col_t]
            saldo      = vals[col_s]
            if not cliente: continue
            try: saldo_f = float(saldo or 0)
            except: saldo_f = 0
            if saldo_f <= 0: continue
            try: total_f = float(total or 0)
            except: total_f = 0
            saldos.append({
                'vendedor': str(vendedor or '').strip(),
                'cliente': str(cliente).strip(),
                'nroFactura': str(nro_fac or '').strip(),
                'fechaFactura': fmt_fecha(fecha_fac),
                'fechaPago': fmt_fecha(fecha_pago),
                'totalFactura': total_f,
                'saldo': saldo_f
            })

        sf_path = os.path.join(DATA_DIR, 'saldos_facturas.json')
        with open(sf_path, 'w', encoding='utf-8') as f:
            json.dump(saldos, f, ensure_ascii=False, indent=2)
        # Calcular fecha_corte = max(fechaFactura) del archivo subido
        _fc_objs = []
        for _s in saldos:
            try:
                _d, _m, _y = _s['fechaFactura'].split('/')
                _fc_objs.append(datetime(int(_y), int(_m), int(_d)))
            except: pass
        _fecha_corte_str = max(_fc_objs).strftime('%d/%m/%Y') if _fc_objs else time.strftime('%d/%m/%Y')
        ts_path = os.path.join(DATA_DIR, 'saldos_timestamp.json')
        with open(ts_path, 'w') as f:
            json.dump({'ts': time.time(), 'fecha': time.strftime('%d/%m/%Y %H:%M'), 'fecha_corte': _fecha_corte_str}, f)
        _saldos_facturas = saldos
        _saldos_gestion  = list(saldos)   # SSoT: facturas siempre sincroniza gestión
        _rebuild_saldos_index()
        # Actualizar mtime tracking (upload_saldos_facturas también persiste gestión a disco)
        global _SG_MTIME, _SG_LAST_CHECK
        try:
            _SG_MTIME = os.path.getmtime(_SG_FILE) if os.path.exists(_SG_FILE) else time.time()
        except Exception:
            pass
        _SG_LAST_CHECK = time.time()
        print(f"[saldos] {len(saldos)} facturas importadas (Odoo positional) — gestión sincronizada", flush=True)
        # Recalcular scores en background usando BCRA cacheado (sin API calls)
        threading.Thread(target=_recalcular_scores_post_upload, daemon=True).start()
        return jsonify({"ok": True, "total": len(saldos)})
    except Exception as e:
        import traceback
        print(f"[saldos] Error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": str(e)}), 500

def _startup_v168():
    """Garantiza que db_v17_final.json existe con motor_version correcto al arrancar."""
    try:
        if os.path.exists(ALERTAS_FILE):
            try:
                with open(ALERTAS_FILE, 'r', encoding='utf-8') as f:
                    doc = json.load(f)
                if doc.get('motor_version') == _MOTOR_VERSION_CARTERA:
                    print(f"[startup] {os.path.basename(ALERTAS_FILE)} OK ({_MOTOR_VERSION_CARTERA})", flush=True)
                    return
            except Exception:
                pass
        with open(ALERTAS_FILE, 'w', encoding='utf-8') as f:
            json.dump({
                'motor_version': _MOTOR_VERSION_CARTERA,
                'alertas':       [],
                'ultima_verif':  f'Init {_MOTOR_VERSION_CARTERA} — {time.strftime("%d/%m/%Y %H:%M")}',
                'cartera':       [],
            }, f, ensure_ascii=False)
        print(f"[startup] {os.path.basename(ALERTAS_FILE)} creado/reseteado ({_MOTOR_VERSION_CARTERA})", flush=True)
    except Exception as e:
        print(f"[startup] Error: {e}", flush=True)

_startup_v168()
_init_padron_db()
_init_nomdeu_db()   # padrón oficial BCRA — abre conexión inmediata si DB existe, descarga en bg si no
_init_mipyme_db()   # padrón PyME (Min. Producción — importado via /update-mipyme-db)

# ── Inicializar ARCA (canal oficial WSAA — requiere certificado) ─────────────
# Reutiliza los helpers R2 existentes para el caché del TA (token 12h).
if ARCA_DISPONIBLE:
    try:
        arca_ws.arca_init(DATA_DIR, _r2_upload_bytes, _r2_download_bytes)
    except FileNotFoundError as e:
        print(f"[init] ARCA sin credenciales ({e}) — se usan las fuentes AFIP fallback", flush=True)
        ARCA_DISPONIBLE = False
    except Exception as e:
        print(f"[init] Error inicializando ARCA: {e}", flush=True)
        ARCA_DISPONIBLE = False

# ── Banner de build: identifica sin ambigüedad qué versión quedó desplegada ──
print(
    "[build] " + " | ".join([
        f"score v{_SCORE_VERSION}",
        f"scoring_fiscal={'OK' if SCORING_FISCAL_OK else 'NO'}",
        f"arca_modulo={'OK' if _ARCA_MODULO_OK else 'NO'}",
        f"arca_activo={'SI' if ARCA_DISPONIBLE else 'NO (sin certificado)'}",
    ]),
    flush=True,
)


def _cheques_auto_update_loop():
    """Verifica y actualiza la DB de cheques rechazados automáticamente cada 24h.
    Al arrancar: si la última importación tiene más de 1 día, dispara actualización.
    Luego duerme 24h y repite. No depende de cron-job.org externo."""
    import datetime as _dt
    # Esperar 60s al arrancar para que la app esté lista antes de descargar
    time.sleep(60)
    while True:
        try:
            last_date_str = ''
            try:
                if os.path.exists(PADRON_DB_PATH):
                    _conn = sqlite3.connect(PADRON_DB_PATH, check_same_thread=False)
                    row = _conn.execute(
                        "SELECT valor FROM _cheques_meta WHERE key = 'last_import_date'"
                    ).fetchone()
                    _conn.close()
                    if row:
                        last_date_str = row[0]
            except Exception:
                pass

            hoy_str = time.strftime('%Y%m%d')
            necesita_update = (last_date_str != hoy_str)

            if necesita_update and not _cheques_db_estado.get('corriendo'):
                print(
                    f"[cheques-auto] última importación={last_date_str or 'nunca'}, "
                    f"hoy={hoy_str} → disparando actualización automática",
                    flush=True,
                )
                ok = _import_cheques_zip(hoy_str)
                if ok and list(_cartera_comercial):
                    threading.Thread(target=_check_cheques_cartera_bg, daemon=True).start()
                    print("[cheques-auto] Verificación de cheques de cartera iniciada", flush=True)
                elif not ok:
                    # Si falla con la fecha de hoy (puede que el BCRA aún no publicó),
                    # probar con ayer
                    ayer = (_dt.date.today() - _dt.timedelta(days=1)).strftime('%Y%m%d')
                    if ayer != last_date_str:
                        print(f"[cheques-auto] Reintentando con ayer={ayer}", flush=True)
                        ok2 = _import_cheques_zip(ayer)
                        if ok2 and list(_cartera_comercial):
                            threading.Thread(target=_check_cheques_cartera_bg, daemon=True).start()
            else:
                print(f"[cheques-auto] DB al día ({last_date_str})", flush=True)
        except Exception as _cau_e:
            print(f"[cheques-auto] Error: {_cau_e}", flush=True)

        time.sleep(86400)  # Volver a verificar en 24 horas


threading.Thread(target=_cheques_auto_update_loop, daemon=True).start()
threading.Thread(target=_facturas_auto_loop, daemon=True).start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, threaded=True)

# Para Gunicorn (Render): 1 worker + 4 threads = eficiente en 512MB RAM
# Comando: gunicorn main:app --workers 1 --threads 4 --timeout 120 --keep-alive 5
